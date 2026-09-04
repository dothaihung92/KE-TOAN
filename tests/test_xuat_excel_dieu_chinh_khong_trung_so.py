import os
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
import datetime, os
src = open(os.path.join(_REPO_ROOT, 'server.py'), encoding='utf-8').read()

def extract_fn(name):
    idx = src.index('def ' + name + '(')
    i = src.index(':', idx)
    lines = src[i+1:].split('\n')
    body = []
    started = False
    for ln in lines:
        if ln.strip() == '' and not started:
            body.append(ln); continue
        if ln and not ln[0].isspace() and started:
            break
        if ln.strip():
            started = True
        body.append(ln)
    return src[idx:i+1] + '\n'.join(body)


class FakeCursor:
    """Mô phỏng MISA ĐÃ có sẵn 1 số bút toán 'DCTR' (đầu vào/phải trả) từ
    lần Xuất/Import TRƯỚC (đúng kịch bản người dùng báo: chạy lại vẫn đếm
    lại từ 1 nên báo 'Số chứng từ đã tồn tại')."""
    def execute(self, sql, params=()):
        if "GLVoucher WHERE RefNoFinance LIKE ?" in sql:
            self._result = [
                ("DCTR001/T12/2025",), ("DCTR002/T12/2025",),
                ("DCTR001/T9/2025",), ("DCTR002/T9/2025",), ("DCTR003/T9/2025",),
            ]
        else:
            self._result = []
        return self

    def fetchone(self):
        return self._result[0] if self._result else None

    def fetchall(self):
        return self._result


class FakeConn:
    def __init__(self, cur):
        self._cur = cur
    def cursor(self):
        return self._cur
    def close(self):
        pass


ns = {'datetime': datetime, 'os': os, 'DOWNLOAD_DIR': '/tmp'}
ns['_to_num'] = lambda v: float(v) if v not in (None, '') else 0
ns['_get_desktop_dir'] = lambda: None
ns['_du_lieu_cty_path'] = lambda cid: None
ns['_GLVOUCHER_HEADERS'] = [
    "Hiển thị trên sổ", "Ngày chứng từ (*)", "Ngày hạch toán (*)", "Số chứng từ (*)",
    "Diễn giải", "Hạn thanh toán", "Loại tiền", "Tỷ giá", "Diễn giải (Hạch toán)",
    "TK Nợ (*)", "TK Có (*)", "Số tiền", "Số tiền quy đổi", "Đối tượng Nợ", "Đối tượng Có",
    "TK ngân hàng", "Khoản mục CP", "Đơn vị", "Đối tượng THCP", "Công trình", "Hợp đồng bán",
    "CP không hợp lý", "Mã thống kê", "Diễn giải (Thuế)", "TK thuế GTGT", "Tiền thuế GTGT",
    "% thuế GTGT", "Tỷ lệ tính thuế (Thuế suất KHAC)", "Giá trị HHDV chưa thuế", "Mẫu số HĐ",
    "Ngày hóa đơn", "Ký hiệu HĐ", "Số hóa đơn", "Nhóm HHDV mua vào", "Mã đối tượng thuế",
    "Tên đối tượng thuế", "Mã số thuế đối tượng thuế",
]
exec(extract_fn('_xuat_excel_dieu_chinh_cong_no'), ns)
_xuat_excel_dieu_chinh_cong_no = ns['_xuat_excel_dieu_chinh_cong_no']

cur = FakeCursor()
ns['_misa_sql_connect'] = lambda cid, database=None: FakeConn(cur)

danh_sach_ncc = [
    {"mst": "0317826028", "ten": "COCOBETE", "inv_no": "234", "inv_date": "2025-10-30", "so_tien": 7875000},
    {"mst": "0999999999", "ten": "NCC KHÁC", "inv_no": "99", "inv_date": "2025-09-11", "so_tien": 1000000},
]
path, so_dong = _xuat_excel_dieu_chinh_cong_no(1, "ncc", danh_sach_ncc, database="TESTDB")

import openpyxl
wb = openpyxl.load_workbook(path)
ws = wb.active
so_ct_1 = ws.cell(2, 4).value
so_ct_2 = ws.cell(3, 4).value

assert so_ct_1 == "DCTR001/T10/2025", f"T10/2025 CHƯA có dòng DCTR nào trong MISA -> phải bắt đầu 001, got {so_ct_1}"
assert so_ct_2 == "DCTR004/T9/2025", (
    f"T9/2025 đã có DCTR001/002/003 trong MISA (từ đợt Import TRƯỚC) -> PHẢI tiếp nối 004, "
    f"KHÔNG được đếm lại từ 001 (đúng lỗi 'Số chứng từ đã tồn tại' người dùng báo) — got {so_ct_2}")
print(f"PASS: số chứng từ mới ({so_ct_1}, {so_ct_2}) tiếp nối ĐÚNG số cao nhất đã có trong MISA theo "
      "từng tháng, không đếm lại từ 1 -> không còn trùng số chứng từ đã Import trước đó.")

# ── Xác nhận tiền tố KHÁC NHAU giữa đầu vào (ncc) và đầu ra (kh) ────────────
cur2 = FakeCursor()
cur2._result = []   # MISA chưa có dòng DCTH nào
ns['_misa_sql_connect'] = lambda cid, database=None: FakeConn(cur2)
danh_sach_kh = [{"mst": "111", "ten": "KH A", "inv_no": "1", "inv_date": "2025-10-30", "so_tien": 500000}]
path2, _ = _xuat_excel_dieu_chinh_cong_no(1, "kh", danh_sach_kh, database="TESTDB")
wb2 = openpyxl.load_workbook(path2)
so_ct_kh = wb2.active.cell(2, 4).value
assert so_ct_kh.startswith("DCTH"), f"Công nợ đầu RA (kh) phải dùng tiền tố KHÁC 'DCTH', got {so_ct_kh}"
assert so_ct_1.startswith("DCTR") and so_ct_kh.startswith("DCTH") and so_ct_1[:4] != so_ct_kh[:4]
print(f"PASS: đầu vào (ncc) dùng tiền tố 'DCTR' ({so_ct_1}), đầu ra (kh) dùng tiền tố 'DCTH' ({so_ct_kh}) "
      "— 2 chiều KHÔNG BAO GIỜ trùng số chứng từ với nhau.")

print("\nALL DONE")
