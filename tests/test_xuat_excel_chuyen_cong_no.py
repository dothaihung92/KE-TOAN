import os
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
import datetime, os
src = open(os.path.join(_REPO_ROOT, 'server.py'), encoding='utf-8').read()

def extract_fn(name):
    idx = src.index('def ' + name + '(')
    i = src.index(':', idx)
    lines = src[i+1:].split('\n')
    body = []
    started = False
    for ln in lines:
        if ln.strip() == '' and not started:
            body.append(ln); continue
        if ln and not ln[0].isspace() and started:
            break
        if ln.strip():
            started = True
        body.append(ln)
    return src[idx:i+1] + '\n'.join(body)


ns = {'datetime': datetime, 'os': os, 'DOWNLOAD_DIR': '/tmp'}
ns['_to_num'] = lambda v: float(v) if v not in (None, '') else 0
ns['_get_desktop_dir'] = lambda: None
ns['_du_lieu_cty_path'] = lambda cid: None
ns['_GLVOUCHER_HEADERS'] = [
    "Hiển thị trên sổ", "Ngày chứng từ (*)", "Ngày hạch toán (*)", "Số chứng từ (*)",
    "Diễn giải", "Hạn thanh toán", "Loại tiền", "Tỷ giá", "Diễn giải (Hạch toán)",
    "TK Nợ (*)", "TK Có (*)", "Số tiền", "Số tiền quy đổi", "Đối tượng Nợ", "Đối tượng Có",
    "TK ngân hàng", "Khoản mục CP", "Đơn vị", "Đối tượng THCP", "Công trình", "Hợp đồng bán",
    "CP không hợp lý", "Mã thống kê", "Diễn giải (Thuế)", "TK thuế GTGT", "Tiền thuế GTGT",
    "% thuế GTGT", "Tỷ lệ tính thuế (Thuế suất KHAC)", "Giá trị HHDV chưa thuế", "Mẫu số HĐ",
    "Ngày hóa đơn", "Ký hiệu HĐ", "Số hóa đơn", "Nhóm HHDV mua vào", "Mã đối tượng thuế",
    "Tên đối tượng thuế", "Mã số thuế đối tượng thuế",
]
exec(extract_fn('_xuat_excel_chuyen_cong_no'), ns)
_xuat_excel_chuyen_cong_no = ns['_xuat_excel_chuyen_cong_no']

danh_sach = [{
    "tu_ma": "0317229511", "tu_ten": "CÔNG TY TNHH CODE LEAP",
    "den_ma": "0319082462", "den_ten": "CÔNG TY TNHH INFINITY ACADEMY",
    "inv_no": "41", "so_tien": 693000, "tt_ngay": "2025-11-13",
}]
path, so_dong = _xuat_excel_chuyen_cong_no(1, "kh", danh_sach, database=None)
assert so_dong == 1

import openpyxl
wb = openpyxl.load_workbook(path)
ws = wb.active
assert ws.cell(2, 10).value == "131" and ws.cell(2, 11).value == "131", "TK Nợ = TK Có = 131"
assert ws.cell(2, 12).value == 693000
assert ws.cell(2, 14).value == "0319082462", "Đối tượng Nợ phải là MÃ đối tượng ĐÍCH (đúng)"
assert ws.cell(2, 15).value == "0317229511", "Đối tượng Có phải là MÃ đối tượng NGUỒN (sai)"
assert ws.cell(2, 4).value.startswith("CDTH001/T11/2025")

print("PASS: Excel 'Chuyển công nợ' ghi đúng — TK Nợ=TK Có=131 (kh), cột 'Đối tượng Nợ' = mã "
      "đối tượng ĐÍCH, 'Đối tượng Có' = mã đối tượng NGUỒN (dùng MÃ, không phải GUID nội bộ — "
      "MISA tự tra khi Import, an toàn hơn ghi thẳng SQL).")

print("\nALL DONE")
