"""
============================================================
  PHẦN MỀM QUẢN LÝ HÓA ĐƠN ĐIỆN TỬ - ĐA CÔNG TY
  Kết nối hoadondientu.gdt.gov.vn (Tổng cục Thuế)
  Backend: FastAPI + SQLite
============================================================
"""
import os
import io
import json
import time
import base64
import sqlite3
import zipfile
import datetime
import concurrent.futures as _cf
from typing import Optional, List

import requests
from fastapi import FastAPI, HTTPException, Body, Request
from fastapi.responses import HTMLResponse, FileResponse, StreamingResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

# ============================================================
#  CẤU HÌNH ĐƯỜNG DẪN
# ============================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, "data")
DOWNLOAD_DIR = os.path.join(BASE_DIR, "downloads")
DB_PATH = os.path.join(DATA_DIR, "hddt.db")
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(DOWNLOAD_DIR, exist_ok=True)

# ============================================================
#  CẤU HÌNH TỐC ĐỘ TẢI (chống bị khóa tạm 429)
#  Mức "balanced" là cân bằng giữa tốc độ và an toàn.
# ============================================================
SPEED_PROFILES = {
    "fast":     {"page": 0.5, "status": 0.5, "month": 0.5, "file": 0.15, "retry_base": 4, "retry_max": 8, "between_loai": 2.5, "song_song": 8},
    "balanced": {"page": 1.0, "status": 1.0, "month": 1.2, "file": 0.6,  "retry_base": 5, "retry_max": 8, "between_loai": 3, "song_song": 5},
    "safe":     {"page": 2.0, "status": 2.0, "month": 2.5, "file": 1.2,  "retry_base": 10, "retry_max": 10, "between_loai": 5, "song_song": 1},
}
CURRENT_SPEED = "fast"  # mặc định nhanh

def SP():
    return SPEED_PROFILES.get(CURRENT_SPEED, SPEED_PROFILES["balanced"])


def _get_desktop_dir():
    """Trả về đường dẫn Desktop của người dùng (Windows/Mac/Linux).
    Hỗ trợ cả Desktop tiếng Việt (OneDrive). Nếu không thấy -> thư mục home."""
    home = os.path.expanduser("~")
    candidates = [
        os.path.join(home, "Desktop"),
        os.path.join(home, "OneDrive", "Desktop"),
        os.path.join(home, "OneDrive", "Máy tính"),
        os.path.join(home, "Máy tính"),
    ]
    # đọc từ registry trên Windows (chính xác nhất)
    try:
        import sys
        if sys.platform.startswith("win"):
            import winreg
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                r"Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders")
            val, _ = winreg.QueryValueEx(key, "Desktop")
            winreg.CloseKey(key)
            if val and os.path.isdir(os.path.expandvars(val)):
                return os.path.expandvars(val)
    except Exception:
        pass
    for c in candidates:
        if os.path.isdir(c):
            return c
    return home


def _resp_xuat(path, fname, extra=None):
    """Trả FileResponse cho file kết xuất. Nếu file ĐÃ CÓ trên Desktop (đã
    được copy ra đó) thì gắn header 'X-Saved-Desktop: 1' để trình duyệt KHÔNG
    tải trùng thêm 1 bản vào thư mục Downloads — người dùng chỉ muốn 1 bản
    ngoài Desktop. File nào KHÔNG có trên Desktop (vd file mẫu) thì không gắn
    header -> trình duyệt vẫn tải bình thường (không mất file)."""
    h = dict(extra or {})
    try:
        d = _get_desktop_dir()
        if d and os.path.isfile(os.path.join(d, fname)):
            h["X-Saved-Desktop"] = "1"
    except Exception:
        pass
    return FileResponse(path, filename=fname, headers=h)

def _open_file_local(path):
    """Mở file bằng ứng dụng mặc định của HĐH (Excel/trình xem XML).
    Chỉ chạy khi server chạy local trên máy người dùng."""
    try:
        import sys, subprocess
        if sys.platform.startswith("win"):
            os.startfile(path)  # type: ignore
        elif sys.platform == "darwin":
            subprocess.Popen(["open", path])
        else:
            subprocess.Popen(["xdg-open", path])
    except Exception:
        pass


def _to_num(v):
    """Chuyển '10.000000' hoặc '807544.000000' thành số; nếu không được giữ nguyên."""
    try:
        f = float(v)
        return int(f) if f == int(f) else f
    except Exception:
        return v if v not in (None,) else ""


def _save_invoice_files(folder, base, zip_bytes):
    """
    Lưu file invoice.zip do TCT trả về, đồng thời giải nén lấy
    invoice.html và file xml ra để dễ đọc.
    folder: thư mục lưu, base: tên file gốc (vd C26TYY_1360_0317483204)
    """
    import io as _io
    import zipfile as _zip
    os.makedirs(folder, exist_ok=True)
    # 1) Lưu nguyên file zip gốc
    zip_path = os.path.join(folder, base + ".zip")
    with open(zip_path, "wb") as f:
        f.write(zip_bytes)
    # 2) Thử giải nén lấy invoice.html và *.xml
    try:
        zf = _zip.ZipFile(_io.BytesIO(zip_bytes))
        for name in zf.namelist():
            low = name.lower()
            data = zf.read(name)
            if low.endswith(".html") or low.endswith(".htm"):
                with open(os.path.join(folder, base + "_invoice.html"), "wb") as f:
                    f.write(data)
            elif low.endswith(".xml"):
                with open(os.path.join(folder, base + ".xml"), "wb") as f:
                    f.write(data)
    except Exception:
        # không phải zip (có thể là xml thuần) -> lưu thẳng .xml
        if zip_bytes[:5] == b"<?xml" or zip_bytes[:5] == b"<HDon":
            with open(os.path.join(folder, base + ".xml"), "wb") as f:
                f.write(zip_bytes)


def _doc_kiem_tra_file_hoadon(fpath):
    """Đọc + thử parse nhanh 1 file hóa đơn đã tải (XML thuần, hoặc ZIP chứa
    XML) để biết file có ĐỌC ĐƯỢC hay không. File hỏng/sai định dạng -> trả
    về [] (rỗng), coi như KHÔNG dùng được (phải tải lại)."""
    try:
        with open(fpath, "rb") as f:
            data = f.read()
        if fpath.lower().endswith(".zip"):
            import zipfile as _zf, io as _io2
            try:
                z = _zf.ZipFile(_io2.BytesIO(data))
                xn = next((n for n in z.namelist()
                           if n.lower().endswith(".xml")), None)
                if xn:
                    data = z.read(xn)
            except Exception:
                pass
        return bool(_parse_xml_invoice(data))
    except Exception:
        return False


# ============================================================
#  GDT API CLIENT  (module gọi API Tổng cục Thuế)
#  --> Nếu TCT đổi endpoint, chỉ cần sửa các URL trong class này
# ============================================================
class GDTClient:
    # Endpoint thật xác nhận từ DevTools: có tiền tố /api
    BASE = "https://hoadondientu.gdt.gov.vn/api"
    BASE_QUERY = "https://hoadondientu.gdt.gov.vn/api/query"

    HEADERS = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/120.0 Safari/537.36",
        "Accept": "application/json, text/plain, */*",
    }

    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(self.HEADERS)
        self.token: Optional[str] = None
        self._last_total = 0
        self.last_query_total = 0
        self.last_query_got = 0
        self._token_dead = False  # bật khi gặp 401 (hết phiên) -> bỏ qua nốt các gọi mạng

    # --- Lấy ảnh captcha ---
    def get_captcha(self):
        url = f"{self.BASE}/captcha"
        r = self.session.get(url, timeout=30)
        r.raise_for_status()
        # Trang trả JSON: {"key": "...", "content": "data:image/svg+xml;base64,..."}
        # (content có thể là chuỗi SVG thô hoặc data-URI base64)
        try:
            data = r.json()
        except Exception:
            # phòng trường hợp trả thẳng SVG/ảnh
            data = {"key": "", "content": r.text}
        return data  # trả nguyên về frontend để hiển thị

    # --- Đăng nhập lấy token ---
    def login(self, username, password, cvalue, ckey):
        url = f"{self.BASE}/security-taxpayer/authenticate"
        payload = {
            "username": username,
            "password": password,
            "cvalue": cvalue,   # mã captcha người dùng nhập
            "ckey": ckey,       # key captcha tương ứng
        }
        r = self.session.post(url, json=payload, timeout=30)
        if r.status_code != 200:
            raise Exception(f"Đăng nhập thất bại ({r.status_code}): {r.text[:200]}")
        data = r.json()
        token = data.get("token")
        if not token:
            raise Exception(f"Không nhận được token: {json.dumps(data)[:200]}")
        self.token = token
        self.session.headers.update({"Authorization": f"Bearer {token}"})
        return token

    def set_token(self, token):
        self.token = token
        self.session.headers.update({"Authorization": f"Bearer {token}"})

    # --- Tra cứu hóa đơn (tự chia nhỏ theo tháng) ---
    # loai: "purchase" (mua vào) hoặc "sold" (bán ra)
    # he_thong: "query" (HĐ điện tử thường) hoặc "sco-query" (HĐ máy tính tiền)
    def query_invoices(self, tu_ngay, den_ngay, loai="purchase", page_size=50,
                       he_thong="query", progress=None):
        """
        Tra cứu hóa đơn trong khoảng ngày bất kỳ (kể cả nhiều tháng/cả năm).
        Trang Thuế giới hạn mỗi lần gọi tối đa ~31 ngày, nên hàm này tự
        CHIA NHỎ khoảng ngày thành từng tháng rồi gộp kết quả lại.
        tu_ngay, den_ngay định dạng dd/mm/yyyy
        he_thong: "query" (hóa đơn điện tử) / "sco-query" (hóa đơn máy tính tiền)
        progress: hàm callback(msg) để báo tiến độ (tùy chọn)
        """
        d_from = datetime.datetime.strptime(tu_ngay, "%d/%m/%Y").date()
        d_to = datetime.datetime.strptime(den_ngay, "%d/%m/%Y").date()
        if d_from > d_to:
            d_from, d_to = d_to, d_from

        all_results = []
        seen = set()  # khử trùng lặp hóa đơn ở ranh giới tháng
        total_expected = 0
        loi_tich_luy = []   # gom LỖI THẬT SỰ (đã thử lại vẫn lỗi) qua các tháng/trạng thái
        cur = d_from
        while cur <= d_to:
            # cuối tháng hiện tại
            if cur.month == 12:
                month_end = datetime.date(cur.year, 12, 31)
            else:
                month_end = datetime.date(cur.year, cur.month + 1, 1) - datetime.timedelta(days=1)
            chunk_end = min(month_end, d_to)

            s_from = cur.strftime("%d/%m/%Y")
            s_to = chunk_end.strftime("%d/%m/%Y")
            if progress:
                progress(f"đang tải {s_from} → {s_to}...")

            self._last_total = 0
            self._loi_rieng_phan = []
            chunk = self._query_one_range(s_from, s_to, loai, page_size, he_thong)
            total_expected += getattr(self, "_last_total", 0) or 0
            for loi in (getattr(self, "_loi_rieng_phan", None) or []):
                loi_tich_luy.append(f"{s_from}-{s_to} {loi}")
            for inv in chunk:
                key = (inv.get("khmshdon"), inv.get("khhdon"),
                       inv.get("shdon"), inv.get("nbmst"))
                if key not in seen:
                    seen.add(key)
                    all_results.append(inv)
            # báo số tờ đã lấy lũy kế
            if progress:
                progress(f"đã lấy {len(all_results)} tờ (đến {s_to})")

            cur = chunk_end + datetime.timedelta(days=1)
            time.sleep(SP()["month"])  # nghỉ giữa các tháng
        # lưu tổng kỳ vọng để báo "đủ chưa"
        self.last_query_total = total_expected
        self.last_query_got = len(all_results)
        # LỖI THẬT SỰ ở 1 phần (vd 1 trạng thái mua vào lỗi hẳn dù đã thử lại) —
        # khác với "0 kết quả" hợp lệ; người gọi (_run_fetch_job) dùng để KHÔNG
        # báo "Hoàn tất" như bình thường mà cảnh báo rõ có thể thiếu dữ liệu.
        self.last_query_errors = loi_tich_luy
        return all_results

    def _query_one_range(self, tu_ngay, den_ngay, loai="purchase",
                         page_size=50, he_thong="query"):
        """
        Tra cứu 1 khoảng ngày (<= 31 ngày).
        Khớp đúng cURL thật từ hoadondientu.gdt.gov.vn:
          HĐ điện tử : /api/query/invoices/{purchase,sold}
          HĐ máy tính tiền: /api/sco-query/invoices/{purchase,sold}
        Mua vào được phân theo trạng thái xử lý (ttxly): 5, 6, 8 -> gọi từng cái và gộp.
        """
        base = f"https://hoadondientu.gdt.gov.vn/api/{he_thong}"
        is_mtt = (he_thong == "sco-query")
        date_filter = (f"tdlap=ge={tu_ngay}T00:00:00;"
                       f"tdlap=le={den_ngay}T23:59:59")

        def _thu_lai(fn, so_lan=2):
            """Gọi fn() tối đa so_lan lần (nghỉ giữa các lần) trước khi bỏ cuộc.
            TOKEN_EXPIRED thì raise ngay, không cần thử lại."""
            last_err = None
            for i in range(so_lan):
                try:
                    return fn()
                except Exception as e:
                    if "TOKEN_EXPIRED" in str(e):
                        raise
                    last_err = e
                    if i < so_lan - 1:
                        time.sleep(SP()["retry_base"])
            raise last_err

        if loai == "purchase":
            url = f"{base}/invoices/purchase"
            action = "Tìm kiếm (hóa đơn %smua vào)" % ("máy tính tiền " if is_mtt else "")
            # Các trạng thái xử lý hóa đơn mua vào (mỗi cái là 1 lần gọi riêng):
            #   5 = tổng hợp/đã cấp mã, 6 = đã nhận không mã, 8 = HĐ có rủi ro...
            results = []
            seen = set()
            total_all = 0
            loi_trang_thai = []   # các trạng thái LỖI THẬT SỰ (không phải rỗng) -> báo cho người dùng biết
            for ttxly in (5, 6, 8):
                search = f"{date_filter};ttxly=={ttxly}"
                try:
                    part, ptotal = _thu_lai(lambda: self._fetch_paginated(
                        url, search, action, page_size, want_total=True))
                    if ptotal:
                        total_all += ptotal
                except Exception as e:
                    # Đã thử lại rồi vẫn lỗi -> GHI NHẬN LÀ LỖI (không coi như "không có"),
                    # để người gọi biết kết quả có thể THIẾU chứ không phải chắc chắn = 0.
                    if "TOKEN_EXPIRED" in str(e):
                        raise
                    part = []
                    loi_trang_thai.append(f"ttxly={ttxly}: {str(e)[:120]}")
                for inv in part:
                    key = (inv.get("khmshdon"), inv.get("khhdon"),
                           inv.get("shdon"), inv.get("nbmst"))
                    if key not in seen:
                        seen.add(key)
                        results.append(inv)
                time.sleep(SP()["status"])  # nghỉ giữa các trạng thái
            self._last_total = total_all
            self._loi_rieng_phan = loi_trang_thai
            return results
        else:
            url = f"{base}/invoices/sold"
            action = "Tìm kiếm (hóa đơn %sbán ra)" % ("máy tính tiền " if is_mtt else "")
            # Thử lại nội bộ trước khi để lỗi lan ra ngoài (bán ra chỉ gọi 1 lần,
            # không có nhiều lượt dự phòng như mua vào theo từng trạng thái, nên
            # dễ bị 1 lần rớt mạng ngẫu nhiên làm mất luôn cả kết quả bán ra).
            results, total = _thu_lai(lambda: self._fetch_paginated(
                url, date_filter, action, page_size, want_total=True))
            self._last_total = total  # lưu để báo lên
            self._loi_rieng_phan = []
            return results

    def _fetch_paginated(self, url, search, action, page_size=50, want_total=False):
        """Gọi 1 endpoint với tham số search, tự phân trang và xử lý 429.
        want_total=True -> trả (results, total_kỳ_vọng) để kiểm tra đủ chưa."""
        from urllib.parse import quote
        extra_headers = {
            "accept-language": "vi",
            "action": quote(action),  # tiếng Việt -> URL-encode (header chỉ nhận latin-1)
            "end-point": "/tra-cuu/tra-cuu-hoa-don",
            "referer": "https://hoadondientu.gdt.gov.vn/tra-cuu/tra-cuu-hoa-don",
        }
        results = []
        state = None
        total_expected = None
        for _ in range(500):  # tối đa 500 trang an toàn
            qs = (f"sort=tdlap:desc&size={page_size}"
                  f"&search={quote(search, safe='=;,:/')}")
            if state:
                qs += f"&state={quote(str(state), safe='')}"
            full_url = f"{url}?{qs}"

            # Gọi có tự động thử lại khi bị 429 (Too Many Requests) HOẶC lỗi mạng
            # thoáng qua (mất kết nối/timeout) — trước đây CHỈ retry khi 429, nên
            # 1 lần rớt mạng ngẫu nhiên giữa chừng làm rớt luôn cả lượt tra cứu
            # (thường gặp nhất ở "bán ra" vì chỉ gọi 1 lần, không có nhiều lượt
            # dự phòng như "mua vào" theo từng trạng thái).
            r = None
            last_net_err = None
            sp = SP()
            # GIỚI HẠN TỔNG THỜI GIAN thử lại 1 trang (không chỉ giới hạn SỐ LẦN):
            # hệ thống hóa đơn máy tính tiền (sco-query) của Tổng cục Thuế phản
            # hồi rất chậm/hay treo — trước đây retry_max=8 lần, mỗi lần chờ
            # timeout=60s + nghỉ tới 60s, cộng dồn có thể tới ~960s CHO 1 TRANG,
            # nhân với 3 trạng thái (ttxly) của "mua vào" x thêm 1 lượt thử lại
            # ở tầng trên -> hàng chục phút đến cả tiếng cho 1 tháng, khiến hóa
            # đơn máy tính tiền gần như không chạy được (cả mua vào lẫn bán ra).
            t_bat_dau_thu = time.time()
            NGUONG_THOI_GIAN_THU_LAI = 75  # giây
            for attempt in range(sp["retry_max"]):
                if time.time() - t_bat_dau_thu > NGUONG_THOI_GIAN_THU_LAI:
                    break
                try:
                    r = self.session.get(full_url, headers=extra_headers, timeout=60)
                except Exception as e:
                    last_net_err = e
                    r = None
                    time.sleep(min(sp["retry_base"] * (attempt + 1), 60))
                    continue
                if r.status_code == 429:
                    ra = r.headers.get("Retry-After")
                    try:
                        wait = int(ra) if ra else sp["retry_base"] * (attempt + 1)
                    except Exception:
                        wait = sp["retry_base"] * (attempt + 1)
                    wait = min(wait, 90)
                    time.sleep(wait)
                    continue
                break

            if r is None:
                raise Exception(
                    f"Lỗi kết nối mạng khi tra cứu (đã thử {sp['retry_max']} lần): "
                    f"{last_net_err}")
            if r.status_code == 401:
                raise Exception("TOKEN_EXPIRED")
            if r.status_code == 429:
                raise Exception(
                    "Trang Thuế tạm chặn do gọi quá nhiều (429). "
                    "Hãy đợi vài phút rồi thử lại, hoặc chuyển sang chế độ 'Chậm & an toàn'.")
            if r.status_code == 400:
                raise Exception(
                    f"Trang Thuế từ chối (400). Thử thu hẹp khoảng ngày "
                    f"(mỗi lần tối đa 1 tháng). Chi tiết: {r.text[:150]}")
            r.raise_for_status()
            data = r.json()
            datas = data.get("datas", []) or []
            results.extend(datas)
            # tổng số hóa đơn kỳ vọng (trang Thuế trả 'total')
            if total_expected is None:
                total_expected = data.get("total")
            state = data.get("state")
            if not state or len(datas) == 0:
                break
            time.sleep(sp["page"])  # nghỉ giữa các trang

        # KIỂM TRA ĐỦ CHƯA: nếu trang Thuế báo total nhiều hơn số lấy được,
        # thử phân trang lại 1 lần nữa (phòng trang bị lỗi giữa chừng).
        # CHỈ áp dụng khi đã lấy được ÍT NHẤT 1 hóa đơn nhưng còn thiếu — tức
        # là dấu hiệu thật sự của "phân trang bị đứt giữa chừng". Nếu lấy
        # được 0 hóa đơn thì KHÔNG chạy lại: trang Thuế thường trả 'total'
        # > 0 dù 'datas' rỗng (total không lọc đúng theo trạng thái/kỳ thực
        # tế đang truy vấn) — trước đây cứ thấy total>0 mà results=0 là chạy
        # lại TOÀN BỘ phân trang (tối đa 500 trang, chờ 429 không giới hạn số
        # lần thử) nên các kỳ KHÔNG có hóa đơn lại chậm hơn hẳn kỳ có hóa đơn.
        if (total_expected and 0 < len(results) < total_expected):
            try:
                retry_results = []
                state2 = None
                so_lan_429 = 0
                for _ in range(500):
                    qs = (f"sort=tdlap:desc&size={page_size}"
                          f"&search={quote(search, safe='=;,:/')}")
                    if state2:
                        qs += f"&state={quote(str(state2), safe='')}"
                    r2 = self.session.get(f"{url}?{qs}", headers=extra_headers, timeout=60)
                    if r2.status_code == 429:
                        so_lan_429 += 1
                        if so_lan_429 > SP()["retry_max"]:
                            break
                        time.sleep(min(SP()["retry_base"] * 2, 90)); continue
                    if r2.status_code != 200:
                        break
                    d2 = r2.json()
                    retry_results.extend(d2.get("datas", []) or [])
                    state2 = d2.get("state")
                    if not state2:
                        break
                    time.sleep(SP()["page"])
                # gộp, khử trùng
                seen = {(x.get("khmshdon"), x.get("khhdon"), x.get("shdon"), x.get("nbmst"))
                        for x in results}
                for inv in retry_results:
                    k = (inv.get("khmshdon"), inv.get("khhdon"), inv.get("shdon"), inv.get("nbmst"))
                    if k not in seen:
                        seen.add(k); results.append(inv)
            except Exception:
                pass

        if want_total:
            return results, total_expected
        return results

    # --- Tải file XML của 1 hóa đơn ---
    def download_xml(self, nbmst, khhdon, khmshdon, shdon, loai="purchase",
                     he_thong="query", max_retry=None):
        """Tải file invoice.zip (chứa XML + invoice.html) của 1 hóa đơn từ TCT.
        Endpoint thật: /api/{he_thong}/invoices/export-xml?nbmst=&khhdon=&shdon=&khmshdon=
        (không có tham số type).
        TRƯỚC ĐÂY hàm này KHÔNG hề thử lại — 1 lần bị 429/rớt mạng là mất file đó
        vĩnh viễn (không lỗi, không log, chỉ đơn giản không có file). Với hàng
        nghìn hóa đơn tải liên tiếp, chỉ cần bị chặn tốc độ giữa chừng là MẤT
        HẲN các file còn lại từ điểm đó. Nay tự thử lại như get_detail."""
        base = f"https://hoadondientu.gdt.gov.vn/api/{he_thong}"
        url = f"{base}/invoices/export-xml"
        params = {
            "nbmst": nbmst, "khhdon": khhdon,
            "shdon": shdon, "khmshdon": khmshdon,
        }
        extra_headers = {
            "accept-language": "vi",
            "action": "Xuat xml",
            "end-point": "/tra-cuu/tra-cuu-hoa-don",
            "referer": "https://hoadondientu.gdt.gov.vn/tra-cuu/tra-cuu-hoa-don",
        }
        sp = SP()
        so_lan = max_retry if max_retry else sp["retry_max"]
        last_err = None
        t_bat_dau_thu = time.time()
        NGUONG_THOI_GIAN_THU_LAI = 150  # giây — chặn cộng dồn khi hệ thống (vd
                                          # máy tính tiền) phản hồi chậm/treo liên tục
        for attempt in range(so_lan):
            if time.time() - t_bat_dau_thu > NGUONG_THOI_GIAN_THU_LAI:
                break
            try:
                r = self.session.get(url, params=params, headers=extra_headers, timeout=90)
            except Exception as e:
                last_err = e
                time.sleep(min(sp["retry_base"] * (attempt + 1), 60))
                continue
            if r.status_code == 200 and r.content[:5] not in (b'{"mes', b'{"err'):
                return r.content  # bytes của file zip
            if r.status_code == 401:
                self._token_dead = True
                return None   # hết phiên, thử lại vô ích
            if r.status_code == 429:
                ra = r.headers.get("Retry-After")
                try:
                    wait = int(ra) if ra else sp["retry_base"] * (attempt + 1)
                except Exception:
                    wait = sp["retry_base"] * (attempt + 1)
                time.sleep(min(wait, 60))
                continue
            last_err = f"status={r.status_code}"
            time.sleep(sp["retry_base"])
        return None

    def get_detail(self, nbmst, khhdon, khmshdon, shdon, he_thong="query",
                   max_retry=None, cho_khi_429=True):
        """Lấy JSON chi tiết đầy đủ 1 hóa đơn.
        max_retry: giới hạn số lần thử (khi xuất Excel hàng loạt -> đặt nhỏ để
        1 hóa đơn lỗi không làm nghẽn cả luồng nhiều phút).
        cho_khi_429: True (mặc định) -> gặp 429 thì CHỜ theo Retry-After rồi thử lại
        (dùng cho tra cứu tương tác, không vội). False -> BỎ QUA NGAY khi bị giới
        hạn tốc độ, không chờ (dùng cho xuất Excel hàng loạt hàng nghìn hóa đơn,
        để 1 lượt rớt không kéo cả tiến trình chờ hàng chục phút)."""
        base = f"https://hoadondientu.gdt.gov.vn/api/{he_thong}"
        url = f"{base}/invoices/detail"
        params = {
            "nbmst": nbmst, "khhdon": khhdon,
            "shdon": shdon, "khmshdon": khmshdon,
        }
        extra_headers = {
            "accept-language": "vi",
            "action": "Xem hoa don",
            "end-point": "/tra-cuu/tra-cuu-hoa-don",
            "referer": "https://hoadondientu.gdt.gov.vn/tra-cuu/tra-cuu-hoa-don",
        }
        sp = SP()
        so_lan = max_retry if max_retry else sp["retry_max"]
        for attempt in range(so_lan):
            try:
                r = self.session.get(url, params=params,
                                     headers=extra_headers, timeout=60)
                if r.status_code == 200:
                    return r.json()
                if r.status_code == 401:
                    self._token_dead = True
                    return None  # token hết hạn, không retry vô ích
                if r.status_code == 429:
                    if not cho_khi_429:
                        return None   # bỏ ngay, không chờ (chế độ nhanh)
                    ra = r.headers.get("Retry-After")
                    try:
                        wait = int(ra) if ra else sp["retry_base"] * (attempt + 1)
                    except Exception:
                        wait = sp["retry_base"] * (attempt + 1)
                    time.sleep(min(wait, 60))
                    continue
                if not cho_khi_429:
                    return None
                # các lỗi khác (5xx...) -> chờ ngắn rồi thử lại
                time.sleep(sp["retry_base"])
            except Exception:
                if not cho_khi_429:
                    return None
                time.sleep(sp["retry_base"])
        return None



# Mỗi công ty giữ 1 client riêng trong RAM (theo phiên token)
CLIENTS = {}  # {company_id: GDTClient}
FETCH_JOBS = {}  # {company_id: {messages, last, running, ...}} - tiến độ tra cứu nền

def get_client(company_id) -> GDTClient:
    if company_id not in CLIENTS:
        CLIENTS[company_id] = GDTClient()
    return CLIENTS[company_id]


# ============================================================
#  DATABASE
# ============================================================
def db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False, timeout=30)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = db()
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS companies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        ten TEXT NOT NULL,
        mst TEXT NOT NULL,
        username TEXT,
        password TEXT,
        ghichu TEXT,
        save_dir TEXT,
        created_at TEXT
    );
    CREATE TABLE IF NOT EXISTS invoices (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        loai TEXT,           -- purchase / sold
        he_thong TEXT,       -- query (HĐ điện tử) / sco-query (HĐ máy tính tiền)
        nbmst TEXT,          -- MST người bán
        nbten TEXT,          -- tên người bán
        nmmst TEXT,          -- MST người mua
        khmshdon TEXT,
        khhdon TEXT,
        shdon TEXT,
        tdlap TEXT,          -- ngày lập
        tgtcthue REAL,       -- tiền chưa thuế
        tgtthue REAL,        -- tiền thuế
        tgtttbso REAL,       -- tổng thanh toán
        tthai TEXT,          -- trạng thái
        raw TEXT,            -- json gốc
        UNIQUE(company_id, khmshdon, khhdon, shdon, loai, he_thong)
    );
    CREATE TABLE IF NOT EXISTS vat_balance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        ky TEXT,                 -- kỳ MM/YYYY
        du_dau_ky REAL,          -- thuế GTGT còn được khấu trừ kỳ trước chuyển sang
        vat_mua REAL,            -- VAT đầu vào trong kỳ
        vat_ban REAL,            -- VAT đầu ra trong kỳ
        phai_nop REAL,           -- số phải nộp (nếu >0)
        du_cuoi_ky REAL,         -- còn được khấu trừ chuyển kỳ sau
        updated_at TEXT,
        UNIQUE(company_id, ky)
    );
    CREATE TABLE IF NOT EXISTS imported_data (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        ky TEXT,                 -- kỳ MM/YYYY (rỗng = áp dụng chung)
        mua_ds REAL, mua_thue REAL,
        ban_ds_0 REAL,
        ban_ds_5 REAL, ban_thue_5 REAL,
        ban_ds_8 REAL, ban_thue_8 REAL,
        ban_ds_10 REAL, ban_thue_10 REAL,
        updated_at TEXT,
        UNIQUE(company_id, ky)
    );
    CREATE TABLE IF NOT EXISTS tokhai_nhap (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        company_id INTEGER,
        so_tk TEXT,              -- số tờ khai
        ngay_dk TEXT,            -- ngày đăng ký (yyyy-mm-dd)
        nguoi_xk TEXT,           -- tên người xuất khẩu
        items_json TEXT,         -- danh sách dòng hàng (JSON)
        updated_at TEXT,
        UNIQUE(company_id, so_tk)
    );
    """)
    # Migration: thêm cột he_thong nếu DB cũ chưa có
    cols = [r[1] for r in conn.execute("PRAGMA table_info(invoices)").fetchall()]
    if "he_thong" not in cols:
        conn.execute("ALTER TABLE invoices ADD COLUMN he_thong TEXT DEFAULT 'query'")
    if "detail_json" not in cols:
        conn.execute("ALTER TABLE invoices ADD COLUMN detail_json TEXT")
    ccols = [r[1] for r in conn.execute("PRAGMA table_info(companies)").fetchall()]
    if "save_dir" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN save_dir TEXT")
    if "nguoi_ky" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN nguoi_ky TEXT")
    if "data_dir" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN data_dir TEXT")
    if "dvc_password" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN dvc_password TEXT")
    if "dvc_password2" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN dvc_password2 TEXT")
    if "last_fetch_error" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN last_fetch_error TEXT")
    if "last_fetch_error_at" not in ccols:
        conn.execute("ALTER TABLE companies ADD COLUMN last_fetch_error_at TEXT")
    # Migration: tách riêng phần HÀNG NHẬP KHẨU (tờ khai NK) trong dữ liệu import
    # để điền đúng chỉ tiêu [23a]/[24a] trên tờ khai 01/GTGT
    icols = [r[1] for r in conn.execute("PRAGMA table_info(imported_data)").fetchall()]
    if "mua_ds_nk" not in icols:
        conn.execute("ALTER TABLE imported_data ADD COLUMN mua_ds_nk REAL DEFAULT 0")
    if "mua_thue_nk" not in icols:
        conn.execute("ALTER TABLE imported_data ADD COLUMN mua_thue_nk REAL DEFAULT 0")
    conn.commit()
    conn.close()

init_db()


# ============================================================
#  FASTAPI APP
# ============================================================
app = FastAPI(title="HDDT Manager")


@app.get("/", response_class=HTMLResponse)
def home():
    with open(os.path.join(BASE_DIR, "static", "index.html"), encoding="utf-8") as f:
        return f.read()


# ---------- TỐC ĐỘ TẢI ----------
@app.get("/api/speed")
def get_speed():
    return {"speed": CURRENT_SPEED}

@app.post("/api/speed")
def set_speed(data: dict = Body(...)):
    global CURRENT_SPEED
    s = data.get("speed", "balanced")
    if s in SPEED_PROFILES:
        CURRENT_SPEED = s
    return {"speed": CURRENT_SPEED}


# ---------- QUẢN LÝ CÔNG TY ----------
@app.get("/api/companies")
def list_companies():
    conn = db()
    rows = conn.execute("SELECT * FROM companies ORDER BY ten").fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        # Đã đăng nhập = có client với token còn hiệu lực
        cl = CLIENTS.get(r["id"])
        d["_online"] = bool(cl and getattr(cl, "token", None))
        d.pop("password", None)  # không gửi mật khẩu ra frontend
        out.append(d)
    return out

@app.get("/api/company/{cid}")
def get_company_detail(cid: int):
    """Lấy chi tiết 1 công ty KÈM mật khẩu (để điền sẵn vào form Sửa).
    App chạy cục bộ trên máy người dùng nên an toàn."""
    conn = db()
    r = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not r:
        raise HTTPException(404, "Không tìm thấy công ty")
    return dict(r)


@app.get("/api/login-status")
def login_status():
    """Trả danh sách company_id đang đăng nhập (token còn hiệu lực)."""
    return {"online": [cid for cid, cl in CLIENTS.items()
                       if getattr(cl, "token", None)]}

@app.post("/api/companies")
def add_company(data: dict = Body(...)):
    conn = db()
    mst = (data.get("mst") or "").strip()
    dup = conn.execute("SELECT ten FROM companies WHERE mst=?", (mst,)).fetchone()
    if dup:
        conn.close()
        raise HTTPException(400, f"MST {mst} đã dùng cho công ty '{dup['ten']}'. "
                                 f"Mỗi công ty phải có MST riêng.")
    conn.execute(
        "INSERT INTO companies (ten, mst, username, password, ghichu, save_dir, data_dir, dvc_password, dvc_password2, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)",
        (data.get("ten"), mst, data.get("username"),
         data.get("password"), data.get("ghichu", ""), data.get("save_dir", ""),
         data.get("data_dir", ""),
         (data.get("dvc_password") or "").strip(),
         (data.get("dvc_password2") or "").strip(),
         datetime.datetime.now().isoformat())
    )
    conn.commit()
    conn.close()
    return {"ok": True}

@app.put("/api/companies/{cid}")
def update_company(cid: int, data: dict = Body(...)):
    conn = db()
    mst = (data.get("mst") or "").strip()
    # Chặn trùng MST với công ty khác (gây lẫn dữ liệu)
    dup = conn.execute("SELECT id, ten FROM companies WHERE mst=? AND id<>?",
                       (mst, cid)).fetchone()
    if dup:
        conn.close()
        raise HTTPException(400, f"MST {mst} đã dùng cho công ty '{dup['ten']}'. "
                                 f"Mỗi công ty phải có MST riêng.")
    # Nếu password để trống -> GIỮ password cũ (không xóa)
    cur = conn.execute("SELECT password, dvc_password, dvc_password2 FROM companies WHERE id=?", (cid,)).fetchone()
    pw = data.get("password")
    if not pw:
        pw = cur["password"] if cur else ""
    # mật khẩu DVC: nếu gửi rỗng -> giữ cũ
    dvc1 = data.get("dvc_password")
    if dvc1 is None or dvc1 == "":
        dvc1 = (cur["dvc_password"] if cur and "dvc_password" in cur.keys() else "") or ""
    dvc2 = data.get("dvc_password2")
    if dvc2 is None or dvc2 == "":
        dvc2 = (cur["dvc_password2"] if cur and "dvc_password2" in cur.keys() else "") or ""
    conn.execute(
        "UPDATE companies SET ten=?, mst=?, username=?, password=?, ghichu=?, save_dir=?, data_dir=?, dvc_password=?, dvc_password2=? WHERE id=?",
        (data.get("ten"), mst, data.get("username"),
         pw, data.get("ghichu", ""), data.get("save_dir", ""),
         data.get("data_dir", ""), dvc1.strip(), dvc2.strip(), cid)
    )
    conn.commit()
    conn.close()
    # Luôn reset phiên đăng nhập cũ sau khi sửa (phòng đổi MST/mật khẩu)
    if cid in CLIENTS:
        del CLIENTS[cid]
    return {"ok": True}

@app.delete("/api/companies/{cid}")
def delete_company(cid: int):
    conn = db()
    conn.execute("DELETE FROM companies WHERE id=?", (cid,))
    conn.execute("DELETE FROM invoices WHERE company_id=?", (cid,))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------- ĐĂNG NHẬP THUẾ ----------
@app.get("/api/captcha/{cid}")
def get_captcha(cid: int):
    client = get_client(cid)
    try:
        data = client.get_captcha()
        return data
    except Exception as e:
        raise HTTPException(500, f"Lỗi lấy captcha: {e}")

@app.post("/api/login/{cid}")
def login(cid: int, body: dict = Body(...)):
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        raise HTTPException(404, "Không tìm thấy công ty")
    client = get_client(cid)
    try:
        token = client.login(
            username=comp["username"] or comp["mst"],
            password=comp["password"],
            cvalue=body.get("cvalue"),
            ckey=body.get("ckey"),
        )
        return {"ok": True, "token": token[:20] + "..."}
    except Exception as e:
        raise HTTPException(401, f"{e}")


# ---------- TỰ ĐỘNG GIẢI CAPTCHA + LOGIN (retry 3 lần) ----------
_DDDDOCR_INSTANCE = None
_DDDDOCR_ERR = ""

def _thu_nap_ddddocr():
    """Thử khởi tạo ddddocr. Trả (instance|None, chuoi_loi)."""
    try:
        import ddddocr
        return ddddocr.DdddOcr(show_ad=False), ""
    except Exception as e:
        return None, f"{type(e).__name__}: {e}"

def _loi_thieu_vcredist(err):
    """Nhận biết lỗi ddddocr do THIẾU Visual C++ Redistributable (onnxruntime
    không nạp được DLL) — để tự cài VC++ rồi thử lại."""
    low = str(err or "").lower()
    return ("dll load failed" in low or "onnxruntime" in low
            or "vcruntime" in low or "pybind11" in low
            or "_pybind11_state" in low)

def _cai_vcredist():
    """Tải & cài 'Microsoft Visual C++ Redistributable (x64)' (im lặng, có xin
    quyền Admin qua UAC). CHỈ chạy trên Windows. Trả True nếu đã CHẠY được
    installer (không chắc người dùng có bấm Yes ở UAC hay không)."""
    import sys as _sys
    if _sys.platform != "win32":
        return False
    import os as _os, tempfile as _tmp, subprocess as _sp
    import urllib.request as _url, ssl as _ssl
    dest = _os.path.join(_tmp.gettempdir(), "vc_redist.x64.exe")
    try:
        ctx = _ssl.create_default_context()
        req = _url.Request("https://aka.ms/vs/17/release/vc_redist.x64.exe",
                           headers={"User-Agent": "Mozilla/5.0"})
        with _url.urlopen(req, timeout=120, context=ctx) as r, open(dest, "wb") as f:
            f.write(r.read())
    except Exception:
        return False
    try:
        # Dùng PowerShell Start-Process -Verb RunAs de HIEN UAC (xin quyen Admin)
        # va -Wait de doi cai xong. subprocess.run thuong khong tu bat UAC duoc.
        ps = ("Start-Process -FilePath '%s' -ArgumentList "
              "'/install','/quiet','/norestart' -Verb RunAs -Wait") % dest
        _sp.run(["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
                 "-Command", ps], timeout=900)
        return True
    except Exception:
        return False
    finally:
        try:
            _os.remove(dest)
        except Exception:
            pass

def _get_ddddocr():
    global _DDDDOCR_INSTANCE, _DDDDOCR_ERR
    if _DDDDOCR_INSTANCE is None:
        inst, err = _thu_nap_ddddocr()
        _DDDDOCR_INSTANCE = inst if inst else False
        _DDDDOCR_ERR = "" if inst else err
    return _DDDDOCR_INSTANCE or None

def _sua_ddddocr_tu_dong():
    """Tự KHẮC PHỤC khi ddddocr không nạp được do thiếu VC++ Redistributable:
    tải & cài VC++ rồi thử nạp lại NGAY trong tiến trình hiện tại. Trả dict
    trạng thái để giao diện hiển thị."""
    global _DDDDOCR_INSTANCE, _DDDDOCR_ERR
    import sys as _sys
    # đã chạy được rồi thì thôi
    if _get_ddddocr() is not None:
        return {"ok": True, "loaded": True, "message": "Bộ giải mã captcha đã sẵn sàng."}
    if not _loi_thieu_vcredist(_DDDDOCR_ERR):
        return {"ok": False, "loaded": False,
                "message": "Lỗi ddddocr không phải do thiếu VC++: " + (_DDDDOCR_ERR or "không rõ")}
    if _sys.platform != "win32":
        return {"ok": False, "loaded": False,
                "message": "Chỉ tự cài được trên Windows. Hãy cài onnxruntime/VC++ thủ công."}
    da_chay = _cai_vcredist()
    if not da_chay:
        return {"ok": False, "loaded": False,
                "message": "Không tải/cài được VC++ Redistributable. Cài tay tại "
                           "https://aka.ms/vs/17/release/vc_redist.x64.exe rồi mở lại phần mềm."}
    # xoá cache import cũ rồi thử nạp lại ngay
    for m in list(_sys.modules):
        if m == "ddddocr" or m.startswith("onnxruntime"):
            _sys.modules.pop(m, None)
    _DDDDOCR_INSTANCE = None
    if _get_ddddocr() is not None:
        return {"ok": True, "loaded": True,
                "message": "Đã cài VC++ Redistributable — bộ giải mã captcha đã chạy được!"}
    return {"ok": True, "loaded": False, "need_restart": True,
            "message": "Đã cài VC++ Redistributable. Vui lòng ĐÓNG và MỞ LẠI phần mềm "
                       "một lần để nhận thư viện mới, rồi tự đăng nhập lại."}

def _svg_to_png(svg_text: str) -> bytes:
    """Rasterize SVG → PNG bytes (pure-Python qua svglib + reportlab). '' nếu thất bại."""
    try:
        from svglib.svglib import svg2rlg
        from reportlab.graphics import renderPM
        import io as _io
        drawing = svg2rlg(_io.StringIO(svg_text))
        if drawing is None:
            return b""
        # phóng to để OCR rõ hơn (mặc định captcha nhỏ ~120x40)
        scale = 4.0
        drawing.width = int((drawing.width or 120) * scale)
        drawing.height = int((drawing.height or 40) * scale)
        drawing.scale(scale, scale)
        buf = _io.BytesIO()
        renderPM.drawToFile(drawing, buf, fmt="PNG", bg=0xFFFFFF)
        return buf.getvalue()
    except Exception:
        return b""


def _svg_to_png_browser(svg_text: str, drv) -> bytes:
    """Rasterize SVG → PNG bằng TRÌNH DUYỆT ẨN THẬT (Chrome headless, qua
    canvas) — CHÍNH XÁC HƠN NHIỀU so với _svg_to_png (svglib chỉ vẽ gần đúng,
    thường sai lệch đúng phần nhiễu/méo mà TCT cố tình thêm vào captcha để
    chống đọc tự động). Đây là cách y hệt lúc người dùng tự bấm nút
    'Tự đăng nhập' trên giao diện (trình duyệt của người dùng vẽ SVG lên
    canvas). '' nếu thất bại (drv=None hoặc lỗi)."""
    if drv is None or not svg_text:
        return b""
    js = """
    var cb = arguments[arguments.length-1];
    var svgB64 = arguments[0];
    try {
      var img = new Image();
      img.onload = function(){
        try {
          var scale = 4;
          var w = img.naturalWidth || 120, h = img.naturalHeight || 40;
          var cv = document.createElement('canvas');
          cv.width = w * scale; cv.height = h * scale;
          var ctx = cv.getContext('2d');
          ctx.fillStyle = '#fff'; ctx.fillRect(0, 0, cv.width, cv.height);
          ctx.drawImage(img, 0, 0, cv.width, cv.height);
          cb({ok: true, data: cv.toDataURL('image/png')});
        } catch(e) { cb({ok: false, err: String(e)}); }
      };
      img.onerror = function(){ cb({ok: false, err: 'img-load-error'}); };
      img.src = 'data:image/svg+xml;base64,' + svgB64;
    } catch(e) { cb({ok: false, err: String(e)}); }
    """
    try:
        b64 = base64.b64encode(svg_text.encode("utf-8")).decode("ascii")
        drv.set_script_timeout(15)
        res = drv.execute_async_script(js, b64)
        if res and res.get("ok") and res.get("data"):
            head, b64png = res["data"].split(",", 1)
            return base64.b64decode(b64png)
    except Exception:
        pass
    return b""

def _preprocess_png(png_bytes: bytes) -> bytes:
    """Làm sạch PNG để ddddocr đoán chuẩn hơn: grayscale + threshold + phóng to."""
    try:
        from PIL import Image, ImageOps, ImageFilter
        import io as _io
        im = Image.open(_io.BytesIO(png_bytes)).convert("L")
        # phóng to nếu nhỏ
        if im.width < 200:
            im = im.resize((im.width * 3, im.height * 3), Image.LANCZOS)
        # tăng tương phản + làm sạch nét nhiễu
        im = ImageOps.autocontrast(im, cutoff=5)
        im = im.filter(ImageFilter.MedianFilter(size=3))
        # nhị phân hóa
        thr = 140
        im = im.point(lambda p: 255 if p > thr else 0)
        buf = _io.BytesIO()
        im.save(buf, format="PNG")
        return buf.getvalue()
    except Exception:
        return png_bytes

def _solve_captcha(content: str, drv=None) -> str:
    """Trả về mã đoán được từ content (data URI / SVG / PNG base64). '' nếu không giải nổi.
    drv: webdriver Chrome ẩn (tùy chọn) — nếu có, dùng để vẽ SVG→PNG CHÍNH XÁC
    như trình duyệt thật (thay vì svglib gần đúng), tăng tỉ lệ OCR đúng."""
    import re as _re_cap
    if not content:
        return ""
    svg_text = ""
    png_bytes = None
    s = content.strip()
    if s.startswith("data:"):
        try:
            head, b64 = s.split(",", 1)
            raw = base64.b64decode(b64)
            if "svg" in head.lower():
                svg_text = raw.decode("utf-8", errors="replace")
            else:
                png_bytes = raw
        except Exception:
            pass
    elif "<svg" in s:
        svg_text = s
    else:
        try:
            png_bytes = base64.b64decode(s)
        except Exception:
            pass

    # 1) Trích <text> trực tiếp từ SVG (chỉ áp dụng khi SVG dùng <text>)
    if svg_text and "<text" in svg_text.lower():
        # bỏ comment + whitespace để regex bắt dễ hơn
        clean = _re_cap.sub(r'<!--.*?-->', '', svg_text, flags=_re_cap.DOTALL)
        # gom mọi nội dung trong <text>...</text> (kể cả <tspan>)
        groups = _re_cap.findall(r'<text\b[^>]*>(.*?)</text>',
                                 clean, flags=_re_cap.DOTALL | _re_cap.IGNORECASE)
        cand = ""
        for g in groups:
            inner = _re_cap.sub(r'<[^>]+>', '', g)  # bỏ tag con
            inner = _re_cap.sub(r'\s+', '', inner)
            cand += inner
        if 4 <= len(cand) <= 10 and cand.isalnum():
            return cand

    # 2) Nếu là SVG: rasterize → PNG để OCR
    if svg_text and not png_bytes:
        # tìm <image href="data:image/...,base64,..."> embed sẵn (đôi khi captcha dạng này)
        m_emb = _re_cap.search(
            r'<image\b[^>]*(?:xlink:)?href\s*=\s*["\']data:image/[^;]+;base64,([^"\']+)["\']',
            svg_text, _re_cap.IGNORECASE)
        if m_emb:
            try:
                png_bytes = base64.b64decode(m_emb.group(1))
            except Exception:
                png_bytes = None
        # ƯU TIÊN vẽ bằng trình duyệt ẩn thật (chính xác hơn hẳn svglib) khi có
        if not png_bytes and drv is not None:
            png_bytes = _svg_to_png_browser(svg_text, drv)
        if not png_bytes:
            png_bytes = _svg_to_png(svg_text)

    # 3) OCR với ddddocr (preprocess + thử nhiều cách)
    if png_bytes:
        ocr = _get_ddddocr()
        if ocr:
            tries = [png_bytes, _preprocess_png(png_bytes)]
            for buf in tries:
                try:
                    ans = (ocr.classification(buf) or "").strip()
                    # giữ alnum, captcha TCT 6 ký tự
                    ans = _re_cap.sub(r'[^A-Za-z0-9]', '', ans)
                    if 4 <= len(ans) <= 10:
                        return ans
                except Exception:
                    pass
    return ""

def _ocr_png(png_bytes: bytes) -> str:
    """OCR 1 ảnh PNG bằng ddddocr, thử cả ảnh gốc và ảnh đã làm sạch. '' nếu fail."""
    import re as _re_o
    ocr = _get_ddddocr()
    if not ocr or not png_bytes:
        return ""
    for buf in (png_bytes, _preprocess_png(png_bytes)):
        try:
            ans = (ocr.classification(buf) or "").strip()
            ans = _re_o.sub(r'[^A-Za-z0-9]', '', ans)
            if 4 <= len(ans) <= 10:
                return ans
        except Exception:
            pass
    return ""


@app.post("/api/fix-ocr")
def fix_ocr():
    """Tự khắc phục lỗi bộ giải mã captcha (ddddocr) không nạp được do thiếu
    'Microsoft Visual C++ Redistributable': tải & cài VC++ (im lặng, có UAC)
    rồi thử nạp lại ngay. Giao diện gọi endpoint này khi gặp lỗi DLL onnxruntime."""
    return _sua_ddddocr_tu_dong()


@app.post("/api/solve-login/{cid}")
def solve_login(cid: int, body: dict = Body(...)):
    """Nhận ảnh PNG (do trình duyệt vẽ từ SVG) → OCR → đăng nhập.
    body: { ckey: str, image: "data:image/png;base64,..." }
    Trình duyệt lặp gọi endpoint này (mỗi lần 1 captcha mới) tới khi thành công.
    """
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        raise HTTPException(404, "Không tìm thấy công ty")
    ckey = body.get("ckey") or ""
    image = body.get("image") or ""
    if image.startswith("data:"):
        image = image.split(",", 1)[1] if "," in image else ""
    try:
        png = base64.b64decode(image)
    except Exception:
        png = b""
    guess = _ocr_png(png)
    if not guess:
        # Phân biệt 2 nguyên nhân để báo cho đúng:
        #  - ddddocr KHÔNG NẠP ĐƯỢC (thường do máy mới thiếu Microsoft Visual
        #    C++ Redistributable mà onnxruntime cần) -> báo rõ để cài, đừng để
        #    người dùng loay hoay tưởng captcha mờ.
        #  - ddddocr nạp OK nhưng đọc không ra -> thử captcha khác (lỗi tạm).
        if _get_ddddocr() is None:
            raise HTTPException(
                500,
                "Chưa dùng được bộ giải mã captcha (ddddocr) trên máy này"
                + (f" — {_DDDDOCR_ERR}" if _DDDDOCR_ERR else "")
                + ". Máy mới thường thiếu 'Microsoft Visual C++ Redistributable "
                "(x64)' — cài đặt gói này rồi khởi động lại phần mềm; hoặc mở "
                f"http://127.0.0.1:8686/api/captcha-debug/{cid} để xem chi tiết lỗi.")
        raise HTTPException(422, "OCR không đọc được mã (ảnh captcha mờ) — thử lại mã khác.")
    client = get_client(cid)
    try:
        token = client.login(
            username=comp["username"] or comp["mst"],
            password=comp["password"],
            cvalue=guess,
            ckey=ckey,
        )
        return {"ok": True, "token": token[:20] + "...", "guess": guess}
    except Exception as e:
        # sai mã / lỗi đăng nhập → trả 401 kèm mã đã đoán để client thử captcha khác
        raise HTTPException(401, f"Sai mã '{guess}': {e}")


def _tu_dong_dang_nhap(cid, so_lan=5, drv=None, progress=None):
    """Tự lấy captcha -> giải -> đăng nhập cho 1 công ty (dùng chung cho endpoint
    /api/auto-login VÀ cho tra cứu hàng loạt — công ty chưa đăng nhập thì tự
    đăng nhập bằng tài khoản/mật khẩu đã lưu thay vì bỏ qua).
    drv: webdriver Chrome ẩn (tùy chọn) để vẽ captcha SVG→PNG chính xác như
    trình duyệt thật (xem _svg_to_png_browser) — tăng tỉ lệ tự đăng nhập thành công.
    progress: callback(text) báo tiến độ từng lần thử (tùy chọn) — để người
    dùng biết đang thử lần thứ mấy, không phải chờ 'im lặng' cả quá trình.
    Trả (ok: bool, message: str, so_lan_thu: int, ma_da_thu: list)."""
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        return False, "Không tìm thấy công ty", 0, []
    if not comp["password"]:
        return False, "Công ty chưa lưu mật khẩu — không thể tự đăng nhập", 0, []
    client = get_client(cid)
    last_err = ""
    tried = []
    for lan in range(1, so_lan + 1):
        if progress:
            progress(f"Đang tự động đăng nhập — lần {lan}/{so_lan} (đang giải captcha)...")
        try:
            cap = client.get_captcha()
        except Exception as e:
            return False, f"Lỗi lấy captcha: {e}", lan, tried
        ckey = cap.get("key") or ""
        cval = _solve_captcha(cap.get("content") or "", drv=drv)
        tried.append(cval or "(không giải được)")
        if not cval:
            last_err = "Không giải được captcha"
            if progress:
                progress(f"Đang tự động đăng nhập — lần {lan}/{so_lan}: không giải được captcha, thử lại...")
            continue
        if progress:
            progress(f"Đang tự động đăng nhập — lần {lan}/{so_lan}: đã đoán mã, đang đăng nhập...")
        try:
            client.login(
                username=comp["username"] or comp["mst"],
                password=comp["password"],
                cvalue=cval,
                ckey=ckey,
            )
            return True, f"Đăng nhập tự động thành công (lần {lan})", lan, tried
        except Exception as e:
            last_err = str(e)
            if progress:
                progress(f"Đang tự động đăng nhập — lần {lan}/{so_lan}: sai mã/lỗi đăng nhập, thử lại...")
            continue
    return False, f"Tự đăng nhập thất bại sau {so_lan} lần. Lỗi cuối: {last_err}", so_lan, tried


def _mo_trinh_duyet_captcha():
    """Mở 1 trình duyệt Chrome ẩn dùng để vẽ captcha SVG→PNG chính xác (xem
    _svg_to_png_browser). Trả None nếu không mở được (máy không có Chrome...)
    — khi đó tự đăng nhập sẽ tự rớt về cách vẽ svglib (kém chính xác hơn)."""
    try:
        return _dvc_make_driver(headless=True, esigner=False)
    except Exception:
        return None


def _dong_trinh_duyet_captcha(drv):
    if drv:
        try:
            drv.quit()
        except Exception:
            pass


@app.post("/api/auto-login/{cid}")
def auto_login(cid: int):
    """Tự lấy captcha → giải → đăng nhập, retry tối đa 5 lần (server-side, fallback).
    Dùng trình duyệt ẩn thật để vẽ captcha (chính xác như lúc người dùng tự bấm
    nút trên giao diện) thay vì svglib để tăng tỉ lệ thành công."""
    drv = _mo_trinh_duyet_captcha()
    try:
        ok, thong_bao, so_lan_thu, ma_da_thu = _tu_dong_dang_nhap(cid, so_lan=8, drv=drv)
    finally:
        _dong_trinh_duyet_captcha(drv)
    if not ok:
        ma_loi = 404 if "Không tìm thấy công ty" in thong_bao else 401
        raise HTTPException(ma_loi, f"{thong_bao}. Mã đã thử: {ma_da_thu}")
    client = get_client(cid)
    return {"ok": True, "token": (client.token or "")[:20] + "...",
            "so_lan_thu": so_lan_thu, "ma_da_thu": ma_da_thu}


@app.get("/api/captcha-debug/{cid}")
def captcha_debug(cid: int):
    """Chẩn đoán vì sao không tự giải được captcha.
    Mở trong trình duyệt: http://127.0.0.1:8686/api/captcha-debug/<id-công-ty>
    """
    import re as _re_dbg
    info = {}

    # 1) Trạng thái thư viện
    info["ddddocr_loaded"] = _get_ddddocr() is not None
    info["ddddocr_error"] = _DDDDOCR_ERR
    try:
        import svglib  # noqa
        info["svglib_loaded"] = True
    except Exception as e:
        info["svglib_loaded"] = False
        info["svglib_error"] = f"{type(e).__name__}: {e}"
    try:
        import reportlab  # noqa
        info["reportlab_loaded"] = True
    except Exception as e:
        info["reportlab_loaded"] = False
        info["reportlab_error"] = f"{type(e).__name__}: {e}"
    try:
        import PIL  # noqa
        info["pillow_loaded"] = True
    except Exception as e:
        info["pillow_loaded"] = False
        info["pillow_error"] = f"{type(e).__name__}: {e}"

    # 2) Lấy 1 captcha thật và phân tích
    try:
        client = get_client(cid)
        cap = client.get_captcha()
    except Exception as e:
        info["captcha_fetch_error"] = f"{type(e).__name__}: {e}"
        return info

    content = cap.get("content") or ""
    info["has_key"] = bool(cap.get("key"))
    info["content_len"] = len(content)
    info["content_head_120"] = content[:120]
    s = content.strip()
    if s.startswith("data:"):
        info["content_type"] = "data-uri: " + s[:40]
    elif "<svg" in s.lower():
        info["content_type"] = "raw-svg"
    else:
        info["content_type"] = "khác (có thể base64 trần)"

    # 3) Có <text> trong SVG không?
    svg_for_check = ""
    if s.startswith("data:") and "svg" in s[:40].lower():
        try:
            svg_for_check = base64.b64decode(s.split(",", 1)[1]).decode("utf-8", "replace")
        except Exception:
            svg_for_check = ""
    elif "<svg" in s.lower():
        svg_for_check = s
    info["co_the_text"] = ("<text" in svg_for_check.lower()) if svg_for_check else None
    info["co_the_path"] = ("<path" in svg_for_check.lower()) if svg_for_check else None
    info["co_the_image"] = ("<image" in svg_for_check.lower()) if svg_for_check else None

    # 4) Thử giải thật
    info["ket_qua_giai"] = _solve_captcha(content) or "(rỗng)"
    return info


# ============================================================
#  DỊCH VỤ CÔNG (dichvucong.gdt.gov.vn)
#  Tự đăng nhập + tải tờ khai / báo cáo ĐÃ NỘP về máy.
#  *** Tính năng thử nghiệm — làm dần, sửa dần ***
#  Cổng có WAF (F5 BIG-IP) + XSRF token + 2 lớp captcha nên có thể
#  cần điều chỉnh sau khi chạy thực tế. Mọi hàm đều trả về thông tin
#  chẩn đoán để dễ dò lỗi.
# ============================================================
DVC_BASE = "https://dichvucong.gdt.gov.vn/tthc"
DVC_CLIENTS = {}   # cid -> DVCClient (giữ phiên đăng nhập)

class DVCClient:
    UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
          "(KHTML, like Gecko) Chrome/149.0.0.0 Safari/537.36")

    SECCH = {
        "sec-ch-ua": '"Google Chrome";v="149", "Chromium";v="149", "Not)A;Brand";v="24"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"Windows"',
    }

    def __init__(self):
        # WAF F5 chặn theo dấu vân tay TLS/JA3 -> dùng curl_cffi giả lập Chrome thật
        # nếu có; không thì lùi về requests thường.
        self.impersonate = False
        try:
            from curl_cffi import requests as _cffi
            self.session = _cffi.Session(impersonate="chrome")
            self.impersonate = True
            # curl_cffi đã tự đặt UA + sec-ch-ua khớp vân tay -> chỉ thêm ngôn ngữ
            self.session.headers.update({
                "Accept-Language": "vi,en-US;q=0.9,en;q=0.8,vi-VN;q=0.7",
            })
        except Exception:
            self.session = requests.Session()
            self.session.headers.update({
                "User-Agent": self.UA,
                "Accept-Language": "vi,en-US;q=0.9,en;q=0.8,vi-VN;q=0.7",
                "Accept": "*/*",
                "Upgrade-Insecure-Requests": "1",
                **self.SECCH,
            })
        self.primed = False
        self.logged_in = False
        self.prime_diag = {}

    def _xsrf(self):
        # Spring/Angular: XSRF-TOKEN nằm trong cookie, gửi lại qua header.
        # requests.Session tự cập nhật cookie mỗi response nên luôn lấy giá trị mới nhất.
        try:
            return self.session.cookies.get("XSRF-TOKEN", "") or ""
        except Exception:
            # cookie có thể trùng tên ở nhiều domain/path
            for c in self.session.cookies:
                if c.name == "XSRF-TOKEN":
                    return c.value or ""
            return ""

    def prime(self):
        """Vào trang chủ rồi trang login như trình duyệt để WAF (F5 BIG-IP) cấp
        cookie TS01.../JSESSIONID/XSRF. Ghi lại chẩn đoán để dò lỗi."""
        nav = {
            "Accept": ("text/html,application/xhtml+xml,application/xml;q=0.9,"
                       "image/avif,image/webp,image/apng,*/*;q=0.8"),
            "Sec-Fetch-Dest": "document",
            "Sec-Fetch-Mode": "navigate",
            "Sec-Fetch-Site": "none",
            "Sec-Fetch-User": "?1",
        }
        diag = {"buoc": []}
        last = None
        for path in ("/homelogin", "/login"):
            try:
                r = self.session.get(DVC_BASE + path, headers=nav, timeout=30,
                                     allow_redirects=True)
                last = r
                diag["buoc"].append({"url": path, "status": r.status_code,
                                     "len": len(r.content or b"")})
                nav = dict(nav); nav["Sec-Fetch-Site"] = "same-origin"
                nav["Referer"] = DVC_BASE + path
            except Exception as e:
                diag["buoc"].append({"url": path, "loi": f"{type(e).__name__}: {e}"})
            time.sleep(0.3)
        try:
            names = [getattr(c, "name", c) for c in self.session.cookies]
            diag["cookies"] = sorted(str(n) for n in names)
        except Exception:
            diag["cookies"] = []
        diag["co_waf"] = any(str(n).startswith("TS") for n in diag["cookies"])
        diag["co_xsrf"] = bool(self._xsrf())
        diag["tls_chrome"] = self.impersonate
        self.prime_diag = diag
        self.primed = True
        return last

    def get_captcha_png(self, referer="/login"):
        """Tải 1 ảnh captcha (PNG bytes). Tự rasterize nếu server trả SVG."""
        ts = int(time.time() * 1000)
        h = {
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
            "Referer": DVC_BASE + referer,
            "Sec-Fetch-Dest": "image",
            "Sec-Fetch-Mode": "no-cors",
            "Sec-Fetch-Site": "same-origin",
        }
        r = self.session.get(f"{DVC_BASE}/getCaptcha?{ts}", headers=h, timeout=30)
        r.raise_for_status()
        ct = (r.headers.get("content-type") or "").lower()
        data = r.content or b""
        head = data[:300].lstrip().lower()
        if "svg" in ct or head.startswith(b"<svg") or b"<svg" in head:
            png = _svg_to_png(data.decode("utf-8", "replace"))
            return png or data
        # đôi khi trả JSON {content: "data:image..."} như cổng hóa đơn
        if "json" in ct:
            try:
                j = r.json()
                cont = j.get("content") or j.get("image") or ""
                if cont.startswith("data:"):
                    b64 = cont.split(",", 1)[1]
                    return base64.b64decode(b64)
            except Exception:
                pass
        return data

    def solve_captcha(self, referer="/login"):
        png = self.get_captcha_png(referer)
        return _ocr_png(png)

    def _post_headers(self, json_body=False, referer="/login"):
        h = {
            "X-Requested-With": "XMLHttpRequest",
            "Origin": "https://dichvucong.gdt.gov.vn",
            "Referer": DVC_BASE + referer,
            "Sec-Fetch-Dest": "empty",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Site": "same-origin",
        }
        tok = self._xsrf()
        if tok:
            h["X-XSRF-TOKEN"] = tok
        h["Content-Type"] = ("application/json"
                             if json_body
                             else "application/x-www-form-urlencoded; charset=UTF-8")
        return h

    def login(self, mst, password, captcha):
        ten_dn = f"{_chuan_mst(mst) or mst}-QL"
        mat_khau = base64.b64encode((password or "").encode("utf-8")).decode("ascii")
        data = {"tenDN": ten_dn, "matKhau": mat_khau,
                "doiTuong": "DN", "captcha": captcha}
        r = self.session.post(DVC_BASE + "/loginLDAP", data=data,
                              headers=self._post_headers(False), timeout=40)
        return r

    def search(self, tu_ngay, den_ngay, captcha):
        """tu_ngay/den_ngay định dạng dd/mm/yyyy. Trả response (HTML bảng hồ sơ)."""
        params = {
            "maNghiepVu": "", "maTTHC": "", "maToKhai": "", "maHoSo": "",
            "tuNgay": tu_ngay, "denNgay": den_ngay,
            "scope_tdt1": "SELF", "mstUyQuyen_tdt1": "", "captcha": captcha,
        }
        r = self.session.get(DVC_BASE + "/ho-so/search", params=params,
                             headers={"X-Requested-With": "XMLHttpRequest",
                                      "Referer": DVC_BASE + "/tchs",
                                      "Sec-Fetch-Dest": "empty",
                                      "Sec-Fetch-Mode": "cors",
                                      "Sec-Fetch-Site": "same-origin"},
                             timeout=60)
        return r

    def download_hoso(self, ma_ho_so):
        h = self._post_headers(True)
        h["Referer"] = f"{DVC_BASE}/tchs/files/detail/{ma_ho_so}?loai="
        r = self.session.post(DVC_BASE + "/tchs/downloadhoso",
                              data=json.dumps({"maHoSo": ma_ho_so}),
                              headers=h, timeout=180)
        return r


def get_dvc_client(cid):
    c = DVC_CLIENTS.get(cid)
    if c is None:
        c = DVCClient()
        DVC_CLIENTS[cid] = c
    return c


def _dvc_parse_ma_ho_so(html):
    """Bóc các mã hồ sơ (maHoSo) từ HTML kết quả tra cứu. Giữ thứ tự, khử trùng."""
    import re as _re
    found, seen = [], set()
    # Mã hồ sơ có nhiều định dạng:
    #   quý : G12.18-260424-00039970            (bắt đầu bằng chữ)
    #   năm : 000.701.18.G12-260330-27012345    (bắt đầu bằng số, nhiều nhóm)
    # -> mẫu chung: các nhóm chữ-số nối bằng dấu chấm, rồi -6số-(nhiều số)
    pats = [
        r'[A-Z0-9]+(?:\.[A-Z0-9]+)+-\d{6}-\d{3,}',
        r"downloadHoSo\(\s*['\"]([^'\"]+)['\"]",
        r'(?:files/detail/)([A-Za-z0-9.\-]{8,60})',
        r'(?:idTKhai|maHoSo)["\'=: ]+([A-Za-z0-9.\-]{8,60})',
    ]
    for i, p in enumerate(pats):
        for m in _re.findall(p, html):
            val = m if isinstance(m, str) else (m[0] if m else "")
            val = val.strip().rstrip("?").split("?")[0]
            if val and val not in seen:
                seen.add(val)
                found.append(val)
        if found and i == 0:
            break   # mẫu chính đã bắt được thì dùng luôn
    return found


def _dvc_parse_id_tbao(html):
    """Bóc idTbao của các thông báo (Tiếp nhận / Xác nhận) trong trang chi tiết hồ sơ.
    Trả về list (idTbao, loaiTBao_đoán)."""
    import re as _re
    ids, seen = [], set()
    pats = [
        r'idTbao["\'\s:=]+["\']?(\d{12,22})',
        r'(?:downloadThongBao|taiThongBao|downloadTB|taiTB|xemThongBao)\(\s*["\']?(\d{12,22})',
        r'data-id-?tbao=["\'](\d{12,22})',
        r'data-id=["\'](\d{14,22})',
    ]
    for p in pats:
        for m in _re.findall(p, html, _re.IGNORECASE):
            if m and m not in seen:
                seen.add(m); ids.append(m)
    return ids


def _dvc_save_folder(cid):
    """Thư mục lưu tờ khai đã nộp: <save_dir|data_dir>/ToKhai_DaNop/."""
    conn = db()
    comp = conn.execute(
        "SELECT mst, save_dir, data_dir FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        return None
    sd = (comp["save_dir"] or "").strip()
    dd = (comp["data_dir"] or "").strip() if "data_dir" in comp.keys() else ""
    base = sd or dd or os.path.join(DATA_DIR, "cong_ty")
    folder = os.path.join(base, "ToKhai_DaNop")
    try:
        os.makedirs(folder, exist_ok=True)
    except Exception:
        return None
    return folder


_DVC_LOI_DN = ["sai", "không đúng", "khong dung", "thất bại", "that bai",
               "lỗi", "error", "tài khoản", "mật khẩu", "khóa", "captcha",
               "mã xác nhận", "ma xac nhan", "chưa đăng nhập"]

def _dvc_login_voi_retry(client, mst, password, so_lan=6):
    """Tự lấy captcha + giải + đăng nhập, thử nhiều lần. Trả (ok, thong_tin)."""
    tried = []
    last = ""
    for lan in range(1, so_lan + 1):
        try:
            cap = client.solve_captcha()
        except Exception as e:
            last = f"Lỗi lấy/giải captcha: {type(e).__name__}: {e}"
            tried.append("(lỗi captcha)")
            continue
        if not cap:
            tried.append("(không đọc được captcha)")
            last = "ddddocr không đọc được mã (rỗng) — có thể chưa cài ddddocr"
            continue
        try:
            r = client.login(mst, password, cap)
        except Exception as e:
            last = f"Lỗi gọi loginLDAP: {e}"
            tried.append(f"{cap}→lỗi")
            continue
        body = (r.text or "")[:400]
        low = body.lower()
        # Heuristic: 200 + không chứa từ khóa lỗi/captcha => coi như thành công.
        captcha_loi = ("captcha" in low or "mã xác nhận" in low or "ma xac nhan" in low)
        co_loi = any(k in low for k in _DVC_LOI_DN)
        if r.status_code == 200 and not co_loi:
            client.logged_in = True
            return True, {"so_lan_thu": lan, "ma_da_thu": tried + [cap],
                          "status": r.status_code, "tra_loi": body}
        last = f"status={r.status_code}, trả lời: {body}"
        tried.append(f"{cap}→{'captcha sai' if captcha_loi else 'từ chối'}")
        if not captcha_loi and not r.ok:
            # lỗi không phải do captcha (vd sai mật khẩu) -> dừng sớm
            break
    return False, {"ma_da_thu": tried, "loi_cuoi": last}


# ============================================================
#  DVC qua TRÌNH DUYỆT THẬT (Selenium + Chrome) — vượt WAF F5
#  Cổng Thuế chặn HTTP thuần (kể cả curl_cffi) vì WAF cần chạy
#  JavaScript. Phần mềm Vsign cũng dùng chromedriver. Ta làm tương tự:
#  mở Chrome (ẩn), để trang tự qua WAF, rồi chạy getCaptcha/loginLDAP/
#  search/downloadhoso NGAY TRONG ngữ cảnh trang (XHR/$.ajax) nên token
#  XSRF/_csrf và cookie được xử lý y như trình duyệt thật.
# ============================================================
_DVC_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
           "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

_ESIGNER_EXT_ID = "ekaaenaocpheoabajfdnkhiibfmfmloo"   # tiện ích ký số GDT (eSigner)

def _find_chrome_extension(ext_id):
    """Tìm thư mục tiện ích Chrome đã cài (để nạp vào trình duyệt tự động).
    Tra trong cac profile Chrome/Chromium tren Windows."""
    import glob
    bases = [
        os.path.expandvars(r"%LOCALAPPDATA%\Google\Chrome\User Data"),
        os.path.expandvars(r"%LOCALAPPDATA%\Chromium\User Data"),
        os.path.expandvars(r"%LOCALAPPDATA%\Microsoft\Edge\User Data"),
    ]
    for base in bases:
        if not os.path.isdir(base):
            continue
        profs = ["Default"] + [os.path.basename(p) for p in glob.glob(os.path.join(base, "Profile *"))]
        for prof in profs:
            extdir = os.path.join(base, prof, "Extensions", ext_id)
            if os.path.isdir(extdir):
                vers = [d for d in glob.glob(os.path.join(extdir, "*")) if os.path.isdir(d)]
                if vers:
                    vers.sort()
                    return vers[-1]
    return None

def _dvc_make_driver(headless=True, esigner=False):
    from selenium import webdriver
    from selenium.webdriver.chrome.options import Options
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1366,920")
    opts.add_argument("--lang=vi-VN")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    # UA chuẩn (tránh chuỗi 'HeadlessChrome' bị WAF nghi ngờ)
    opts.add_argument(f"--user-agent={_DVC_UA}")
    # Nạp tiện ích ký số eSigner đã cài sẵn (để vào được cổng eTax)
    ext_path = _find_chrome_extension(_ESIGNER_EXT_ID) if esigner else None
    if ext_path:
        opts.add_argument(f"--load-extension={ext_path}")
        opts.add_argument(f"--disable-extensions-except={ext_path}")
    try:
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
    except Exception:
        pass
    drv = webdriver.Chrome(options=opts)   # Selenium Manager tự tải chromedriver
    drv._esigner_ext = ext_path
    drv.set_page_load_timeout(70)
    drv.set_script_timeout(150)
    try:
        drv.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument",
            {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"})
    except Exception:
        pass
    return drv

# --- Các đoạn JS chạy trong ngữ cảnh trang dichvucong ---
_JS_GETCAPTCHA = r"""
var cb = arguments[arguments.length-1];
// Dùng chính thẻ <img> captcha của trang + hàm reload của trang (đúng URL trang
// dùng, request kiểu ảnh nên qua WAF). Cùng nguồn -> canvas không bị 'taint'.
try {
  function findImg(){
    var ids=['image-capt','image-capt-cbt','imgCaptcha','captcha','imgcaptcha'];
    for(var i=0;i<ids.length;i++){ var e=document.getElementById(ids[i]);
      if(e && e.tagName==='IMG') return e; }
    var imgs=document.getElementsByTagName('img');
    for(var j=0;j<imgs.length;j++){ if(((imgs[j].src||'').indexOf('getCaptcha'))>=0) return imgs[j]; }
    return null;
  }
  var img = findImg();
  if(!img){ cb({ok:false, err:'no-captcha-img'}); return; }
  var oldsrc = img.src||'';
  // nạp captcha mới: ưu tiên hàm của trang, nếu không thì đổi timestamp trên src cũ
  var reloaded=false;
  try{ if(typeof reloadCaptcha==='function'){ reloadCaptcha(); reloaded=true; } }catch(e){}
  if(!reloaded){ try{ if(typeof reloadCaptchaCbt==='function'){ reloadCaptchaCbt(); reloaded=true; } }catch(e){} }
  if(!reloaded){
    var base=(oldsrc.split('?')[0]) || '/tthc/getCaptcha';
    img.src = base + '?' + Date.now();
  }
  var done=false, tries=0;
  function check(){
    if(done) return;
    tries++;
    if(img.complete && img.naturalWidth>2){
      done=true;
      try{
        var w=img.naturalWidth, h=img.naturalHeight||60, sc=3;
        var c=document.createElement('canvas'); c.width=w*sc; c.height=h*sc;
        var g=c.getContext('2d'); g.fillStyle='#fff'; g.fillRect(0,0,c.width,c.height);
        g.drawImage(img,0,0,c.width,c.height);
        cb({ok:true, w:w, h:h, src:(img.src||'').slice(0,90), png:c.toDataURL('image/png')});
      }catch(e){ cb({ok:false, err:'canvas:'+e, src:(img.src||'').slice(0,90)}); }
      return;
    }
    if(tries>50){ done=true; cb({ok:false, err:'wait', src:(img.src||'').slice(0,90),
                                  complete:img.complete, nw:img.naturalWidth}); return; }
    setTimeout(check,150);
  }
  setTimeout(check, 400);
} catch(e){ cb({ok:false, err:''+e}); }
"""

_JS_LOGIN = r"""
var cb = arguments[arguments.length-1];
var ten = arguments[0], pw = arguments[1], cap = arguments[2];
try {
  $.ajax({ type:'POST', url:'/tthc/loginLDAP',
    data:{ tenDN:ten, matKhau: btoa(unescape(encodeURIComponent(pw))),
           doiTuong:'DN', captcha:cap },
    success:function(d){ cb({ok:true, data:d}); },
    error:function(x){ cb({ok:false, status:x.status, resp:(x.responseText||'').slice(0,300)}); }
  });
} catch(e){ cb({ok:false, err:''+e}); }
"""

_JS_SEARCH = r"""
var cb = arguments[arguments.length-1];
var tu=arguments[0], den=arguments[1], cap=arguments[2];
try {
  $.ajax({ type:'GET', url:'/tthc/ho-so/search', dataType:'html',
    data:{ maNghiepVu:'', maTTHC:'', maToKhai:'', maHoSo:'',
           tuNgay:tu, denNgay:den, scope_tdt1:'SELF', mstUyQuyen_tdt1:'', captcha:cap,
           page:0, size:200 },
    success:function(d){ cb({ok:true, html:d}); },
    error:function(x){ cb({ok:false, status:x.status, resp:(x.responseText||'').slice(0,300)}); }
  });
} catch(e){ cb({ok:false, err:''+e}); }
"""

_JS_DOWNLOAD = r"""
var cb = arguments[arguments.length-1];
var ma = arguments[0];
try {
  $.ajax({ type:'POST', url:'/tthc/tchs/downloadhoso',
    contentType:'application/json', data: JSON.stringify({maHoSo:ma}),
    success:function(d){ cb({ok:true, data:d}); },
    error:function(x){ cb({ok:false, status:x.status, resp:(x.responseText||'').slice(0,200)}); }
  });
} catch(e){ cb({ok:false, err:''+e}); }
"""

# Đọc bảng kết quả tra cứu (HTML fragment) thành mảng dòng×ô bằng DOM trình duyệt.
# Mỗi dòng kèm luôn "ma" (mã hồ sơ) dò trong chính HTML của dòng đó — thử theo đúng thứ tự
# các mẫu mà _dvc_parse_ma_ho_so() dùng trên toàn trang (mẫu ID dạng "A.B.C-260330-0001234"
# thường nằm trong href/data-*, không nhất thiết trong onclick=downloadHoSo(...)).
_JS_PARSE_TABLE = r"""
var cb = arguments[arguments.length-1];
var html = arguments[0];
try {
  var d = document.createElement('div'); d.innerHTML = html;
  var out = [];
  var trs = d.querySelectorAll('table tr');
  var pats = [
    /[A-Z0-9]+(?:\.[A-Z0-9]+)+-\d{6}-\d{3,}/,
    /downloadHoSo\(\s*['"]([^'"]+)['"]/,
    /files\/detail\/([A-Za-z0-9.\-]{8,60})/,
    /(?:idTKhai|maHoSo)["'=: ]+([A-Za-z0-9.\-]{8,60})/
  ];
  for (var i=0;i<trs.length;i++){
    var cells=[]; var cs = trs[i].querySelectorAll('th,td');
    for (var j=0;j<cs.length;j++){ cells.push((cs[j].innerText||cs[j].textContent||'').replace(/\s+/g,' ').trim()); }
    var ma='', rowHtml = trs[i].innerHTML;
    for (var k=0;k<pats.length;k++){
      var m = rowHtml.match(pats[k]);
      if (m){ ma = m[1] || m[0]; break; }
    }
    if (cells.length) out.push({cells:cells, ma:ma});
  }
  cb({ok:true, rows:out});
} catch(e){ cb({ok:false, err:''+e}); }
"""

# Tải 1 thông báo (Tiếp nhận / Xác nhận). body là chuỗi JSON dựng sẵn từ Python
# để giữ nguyên idTbao (số rất lớn, tránh mất chính xác khi qua JS number).
_JS_DOWNLOAD_TB = r"""
var cb = arguments[arguments.length-1];
var body = arguments[0];
try {
  $.ajax({ type:'POST', url:'/tthc/tchs/downloadthongbao',
    contentType:'application/json', data: body,
    success:function(d){ cb({ok:true, data:d}); },
    error:function(x){ cb({ok:false, status:x.status, resp:(x.responseText||'').slice(0,200)}); }
  });
} catch(e){ cb({ok:false, err:''+e}); }
"""

def _dvc_wait_jquery(drv, giay=12):
    import time as _t
    for _ in range(int(giay*2)):
        try:
            if drv.execute_script(
                "return (typeof window.jQuery!=='undefined') && (typeof window.$==='function');"):
                return True
        except Exception:
            pass
        _t.sleep(0.5)
    return False

def _dvc_cap_from_js(res):
    """Từ kết quả _JS_GETCAPTCHA (PNG do trình duyệt vẽ từ canvas) → mã captcha."""
    if not res or not res.get("ok"):
        return ""
    png_uri = res.get("png") or ""
    if png_uri.startswith("data:"):
        try:
            raw = base64.b64decode(png_uri.split(",", 1)[1])
        except Exception:
            return ""
        return _ocr_png(raw)
    # phòng hờ phiên bản cũ trả b64 thô
    b64 = res.get("b64") or ""
    if b64:
        try:
            raw = base64.b64decode(b64)
        except Exception:
            return ""
        ct = (res.get("ct") or "").lower()
        if "svg" in ct or b"<svg" in raw[:200].lower():
            raw = _svg_to_png(raw.decode("utf-8", "replace")) or b""
        return _ocr_png(raw)
    return ""

def _dvc_cap_meta(res):
    """Chuỗi chẩn đoán khi captcha rỗng."""
    if not res:
        return "res=None"
    if not res.get("ok"):
        return (f"ok=False err={res.get('err')} src={res.get('src')} "
                f"complete={res.get('complete')} nw={res.get('nw')}")
    return f"ok=True w={res.get('w')} h={res.get('h')} src={res.get('src')} ocr=rỗng"

def _dvc_norm_data(d):
    if isinstance(d, str):
        try:
            return json.loads(d)
        except Exception:
            return {"_raw": d[:200]}
    return d

def _dvc_browser_login(drv, mst, password, so_lan=8):
    """Mở trang, qua WAF, tự giải captcha + đăng nhập. Trả (ok, info)."""
    import re as _re, time as _t
    drv.get(DVC_BASE + "/homelogin"); _t.sleep(1.5)
    drv.get(DVC_BASE + "/login"); _t.sleep(1.0)
    if not _dvc_wait_jquery(drv, 15):
        return False, {"loi": "Trang login không nạp được jQuery (có thể WAF chặn cả trình duyệt)"}
    digits = _re.sub(r"\D", "", mst or "")
    ten = f"{digits}-QL"
    tried = []
    for lan in range(1, so_lan + 1):
        try:
            capres = drv.execute_async_script(_JS_GETCAPTCHA)
            cap = _dvc_cap_from_js(capres)
        except Exception as e:
            tried.append(f"(lỗi lấy captcha: {e})"); continue
        if not cap:
            tried.append(f"(captcha rỗng: {_dvc_cap_meta(capres)})"); continue
        try:
            res = drv.execute_async_script(_JS_LOGIN, ten, password, cap)
        except Exception as e:
            tried.append(f"{cap}→lỗi gọi login: {e}"); continue
        if res and res.get("ok"):
            d = _dvc_norm_data(res.get("data"))
            st = str(d.get("status")) if isinstance(d, dict) else ""
            if st == "200" or (isinstance(d, dict) and d.get("value")):
                return True, {"so_lan": lan, "ten_dn": ten}
            desc = (d.get("desc") if isinstance(d, dict) else str(d)) or ""
            tried.append(f"{cap}→{str(desc)[:70]}")
        else:
            tried.append(f"{cap}→{str(res)[:80]}")
        _t.sleep(0.4)
    return False, {"ten_dn": ten, "da_thu": tried}

def _dvc_browser_search(drv, tu, den, so_lan=8):
    """Sau khi đăng nhập: vào trang tra cứu, giải captcha lớp 2, tìm hồ sơ."""
    import time as _t
    diag = []
    drv.get(DVC_BASE + "/tchs"); _t.sleep(1.5)
    if not _dvc_wait_jquery(drv, 15):
        return [], ["Trang tra cứu (/tchs) không nạp được — có thể chưa đăng nhập thành công"]
    for lan in range(1, so_lan + 1):
        try:
            cap = _dvc_cap_from_js(drv.execute_async_script(_JS_GETCAPTCHA))
        except Exception as e:
            diag.append(f"(lỗi captcha: {e})"); continue
        if not cap:
            diag.append("(captcha rỗng)"); continue
        try:
            res = drv.execute_async_script(_JS_SEARCH, tu, den, cap)
        except Exception as e:
            diag.append(f"{cap}→lỗi search: {e}"); continue
        if res and res.get("ok"):
            html = res.get("html") or ""
            low = html.lower()
            if ("table-container" in low or "tổng số bản ghi" in low
                    or "totalpage" in low):
                ma = _dvc_parse_ma_ho_so(html)
                diag.append(f"{cap}→OK, {len(ma)} hồ sơ")
                return ma, diag
            diag.append(f"{cap}→chưa ra bảng (len={len(html)})")
        else:
            diag.append(f"{cap}→{str(res)[:80]}")
        _t.sleep(0.4)
    return [], diag

def _dvc_browser_download(drv, ma):
    res = drv.execute_async_script(_JS_DOWNLOAD, ma)
    if not res or not res.get("ok"):
        raise Exception(f"{res}")
    d = _dvc_norm_data(res.get("data"))
    if not isinstance(d, dict):
        raise Exception("phản hồi không hợp lệ")
    content = d.get("content") or ""
    fname = d.get("fileName") or f"{ma}.zip"
    raw = base64.b64decode(content) if content else b""
    return fname, raw

def _dvc_browser_thongbao(drv, ma):
    """Vào trang chi tiết hồ sơ, đọc idTbao rồi tải các thông báo (Tiếp nhận/Xác nhận).
    Trả về (list[(fname, raw)], diag)."""
    import time as _t
    out, diag = [], []
    try:
        drv.get(f"{DVC_BASE}/tchs/files/detail/{ma}?loai=")
        _t.sleep(1.2)
        _dvc_wait_jquery(drv, 10)
        html = drv.page_source or ""
    except Exception as e:
        return out, [f"lỗi mở chi tiết {ma}: {e}"]
    ids = _dvc_parse_id_tbao(html)
    if not ids:
        # dò manh mối để tinh chỉnh sau
        import re as _re
        m = _re.search(r'.{0,40}(?:hongBao|hong báo|Tbao).{0,40}', html)
        diag.append(f"{ma}: không thấy idTbao" + (f" | gợi ý: {m.group(0)[:80]}" if m else ""))
        return out, diag
    for idt in ids:
        body = json.dumps({"idTbao": idt, "loaiTBao": ""})
        try:
            res = drv.execute_async_script(_JS_DOWNLOAD_TB, body)
        except Exception as e:
            diag.append(f"{idt}: lỗi gọi: {e}"); continue
        if not res or not res.get("ok"):
            diag.append(f"{idt}: {str(res)[:80]}"); continue
        d = _dvc_norm_data(res.get("data"))
        if not isinstance(d, dict):
            diag.append(f"{idt}: phản hồi lạ"); continue
        content = d.get("content") or ""
        raw = base64.b64decode(content) if content else b""
        if not raw:
            diag.append(f"{idt}: rỗng"); continue
        fname = d.get("fileName") or f"TB_{ma}_{idt}.zip"
        out.append((fname, raw))
        _t.sleep(0.3)
    diag.append(f"{ma}: {len(out)}/{len(ids)} thông báo")
    return out, diag


def _khong_dau(s):
    import unicodedata
    s = unicodedata.normalize("NFD", str(s or ""))
    return "".join(c for c in s if unicodedata.category(c) != "Mn").lower().strip()

# Các cột báo cáo Tra cứu tờ khai (theo file mẫu)
TRACUU_COLS = ["to_khai", "ky", "loai", "ngay_nop", "lan_nop", "lan_bs", "trang_thai"]
TRACUU_HEADERS = ["Mst", "Tên công ty", "Tờ khai", "Kỳ", "Loại",
                  "Ngày nộp", "Lần nộp", "Lần bổ sung", "Trạng thái"]

# Nhãn gọn cho tên file, theo mã tờ khai (vd '01/GTGT' -> '01GTGT')
_TEN_TO_KHAI_MAP = {
    "01/GTGT": "01GTGT",
    "01/KK-TNCN": "01KK-TNCN",
    "02/KK-TNCN": "02KK-TNCN",
    "03/KK-TNCN": "03KK-TNCN",
    "05/KK-TNCN": "05KK-TNCN",
    "03/TNDN": "QTTNDN_BCTC",
    "05/QTT-TNCN": "QTTNCN",
    "02/TNDN": "TNDN",
}

def _ma_to_khai_tu_ten(ten_to_khai):
    """Lấy MÃ tờ khai (vd '05/KK-TNCN') từ chuỗi đầy đủ 'MÃ - Tên đầy đủ' trong cột Tờ khai."""
    s = str(ten_to_khai or "").strip()
    if not s:
        return ""
    if " - " in s:
        return s.split(" - ", 1)[0].strip()
    return s.split(" ")[0].strip()

def _chuan_ky_ten_file(ky):
    """Chuẩn hoá chuỗi kỳ ('Quý 2/2026','Năm 2025','Tháng 3/2026'...) thành tên gọn để đặt tên file."""
    import re as _re
    kd = _khong_dau(ky)   # đã bỏ dấu + viết thường
    m = _re.search(r'quy\s*(\d)\D+(\d{4})', kd)
    if m:
        return f"QUY{m.group(1)}.{m.group(2)}"
    m = _re.search(r'nam\s*(\d{4})', kd)
    if m:
        return f"NAM{m.group(1)}"
    m = _re.search(r'thang\s*(\d{1,2})\D+(\d{4})', kd)
    if m:
        return f"T{int(m.group(1))}.{m.group(2)}"
    m = _re.match(r'^q\s*(\d)\s*/\s*(\d{4})$', kd)
    if m:
        return f"QUY{m.group(1)}.{m.group(2)}"
    if _re.match(r'^\d{4}$', kd):
        return f"NAM{kd}"
    m = _re.match(r'^(\d{1,2})\s*/\s*(\d{4})$', kd)
    if m:
        return f"T{int(m.group(1))}.{m.group(2)}"
    out = _re.sub(r'[^a-z0-9]', '', kd).upper()
    return out or "KyKhongRo"

def _so_lan_bs(v):
    """Bóc số lần bổ sung từ ô 'Lần bổ sung' (có thể là '0','1','Lần 2',...). 0 = tờ khai gốc."""
    import re as _re
    m = _re.search(r'\d+', str(v or ""))
    return int(m.group(0)) if m else 0

def _ten_file_than_thien(ten_to_khai_full, ky, mst, lan_bs):
    """Ghép tên file thân thiện: {MÃ}_{KỲ}-{MST}[_L{lần bổ sung}].
    Ví dụ: 05KK-TNCN_QUY2.2026-0316429370  hoặc  QTTNDN_BCTC_NAM2025-0316429370_L1"""
    import re as _re
    ma = _ma_to_khai_tu_ten(ten_to_khai_full)
    nhan = _TEN_TO_KHAI_MAP.get(ma.upper()) if ma else None
    if not nhan:
        nhan = _re.sub(r'[^A-Za-z0-9]', '', ma) or "ToKhai"
    ky_s = _chuan_ky_ten_file(ky)
    mst_s = _re.sub(r'[^A-Za-z0-9]', '', str(mst or ""))
    ten = f"{nhan}_{ky_s}-{mst_s}"
    lb = _so_lan_bs(lan_bs)
    if lb > 0:
        ten += f"_L{lb}"
    return ten

def _dvc_map_bang(rows):
    """Từ mảng dòng {cells, ma} (DOM, xem _JS_PARSE_TABLE) → list dict theo TRACUU_COLS + 'ma'.
    Tự dò cột theo tiêu đề. Cũng chấp nhận mảng dòng×ô "thô" (list) để tương thích ngược."""
    if not rows:
        return []
    def _cells_of(r):
        return r.get("cells") or [] if isinstance(r, dict) else (r or [])
    def _ma_of(r):
        return (r.get("ma") or "") if isinstance(r, dict) else ""

    # tìm dòng tiêu đề: dòng chứa 'trang thai' hoặc 'to khai'
    h_idx = -1
    for i, r in enumerate(rows[:5]):
        joined = _khong_dau(" ".join(_cells_of(r)))
        if "trang thai" in joined or "to khai" in joined or "ky tinh" in joined:
            h_idx = i; break
    if h_idx < 0:
        h_idx = 0
    header = [_khong_dau(c) for c in _cells_of(rows[h_idx])]

    def find(*keys, avoid=()):
        for j, h in enumerate(header):
            if any(k in h for k in keys) and not any(a in h for a in avoid):
                return j
        return -1

    col = {
        "trang_thai": find("trang thai"),
        "ngay_nop":   find("ngay nop", "ngay"),
        "ky":         find("ky tinh", avoid=()) if find("ky tinh") >= 0 else find("ky", avoid=("ky thuat",)),
        "lan_bs":     find("bo sung"),
        "lan_nop":    find("lan nop", avoid=("bo sung",)),
        "loai":       find("loai"),
        "to_khai":    find("to khai", "thu tuc", avoid=("loai",)),
    }
    out = []
    for r in rows[h_idx + 1:]:
        cells = _cells_of(r)
        if not cells or all(not c for c in cells):
            continue
        # bỏ dòng phân trang / tổng
        j = _khong_dau(" ".join(cells))
        if "tong so ban ghi" in j or "trang" == j[:5]:
            continue
        rec = {}
        ok_any = False
        for c in TRACUU_COLS:
            idx = col.get(c, -1)
            v = cells[idx] if 0 <= idx < len(cells) else ""
            rec[c] = v
            if v:
                ok_any = True
        if ok_any:
            rec["ma"] = _ma_of(r)
            out.append(rec)
    return out


def _dvc_browser_login_2pass(drv, mst, pw1, pw2):
    """Thử pass1 rồi pass2. Trả (ok, dùng_pass(1/2/0), info)."""
    if pw1:
        ok, info = _dvc_browser_login(drv, mst, pw1)
        if ok:
            return True, 1, info
        info1 = info
    else:
        info1 = {"bo_qua": "pass1 trống"}
    if pw2:
        ok, info = _dvc_browser_login(drv, mst, pw2)
        if ok:
            return True, 2, info
        return False, 0, {"pass1": info1, "pass2": info}
    return False, 0, {"pass1": info1, "pass2": "(không có pass2)"}


def _dvc_browser_tracuu(drv, tu, den, so_lan=8):
    """Tra cứu → trả (rows_struct, ma_list, raw_html, diag)."""
    import time as _t
    diag = []
    drv.get(DVC_BASE + "/tchs"); _t.sleep(1.5)
    if not _dvc_wait_jquery(drv, 15):
        return [], [], "", ["Trang /tchs không nạp được"]
    for lan in range(1, so_lan + 1):
        try:
            cap = _dvc_cap_from_js(drv.execute_async_script(_JS_GETCAPTCHA))
        except Exception as e:
            diag.append(f"(captcha lỗi: {e})"); continue
        if not cap:
            diag.append("(captcha rỗng)"); continue
        try:
            res = drv.execute_async_script(_JS_SEARCH, tu, den, cap)
        except Exception as e:
            diag.append(f"{cap}→lỗi search: {e}"); continue
        if res and res.get("ok"):
            html = res.get("html") or ""
            low = html.lower()
            if ("table-container" in low or "tổng số bản ghi" in low or "totalpage" in low):
                ma = _dvc_parse_ma_ho_so(html)
                rows = []
                try:
                    pr = drv.execute_async_script(_JS_PARSE_TABLE, html)
                    if pr and pr.get("ok"):
                        rows = _dvc_map_bang(pr.get("rows") or [])
                except Exception as e:
                    diag.append(f"parse bảng lỗi: {e}")
                diag.append(f"{cap}→OK: {len(ma)} hồ sơ, {len(rows)} dòng bảng")
                return rows, ma, html, diag
            diag.append(f"{cap}→chưa ra bảng ({len(html)})")
        else:
            diag.append(f"{cap}→{str(res)[:80]}")
        _t.sleep(0.4)
    return [], [], "", diag


def _xuat_tracuu_excel(rows, path):
    """rows: list dict có mst, ten, + TRACUU_COLS. Ghi Excel theo mẫu Tracuu.xls."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook(); ws = wb.active; ws.title = "Tracuutokhai"
    ws["A3"] = "TRA CỨU TỜ KHAI"
    ws["A3"].font = Font(bold=True, size=14)
    hr = 5
    for c, h in enumerate(TRACUU_HEADERS, 1):
        cell = ws.cell(hr, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F5496")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="BBBBBB")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    r = hr + 1
    for rec in rows:
        tt = _khong_dau(rec.get("trang_thai", ""))
        chua_nop = "chua nop" in tt
        do = chua_nop or "khong chap nhan" in tt or "tu choi" in tt
        vals = [rec.get("mst", ""), rec.get("ten", "")] + [rec.get(c, "") for c in TRACUU_COLS]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v); cell.border = bd
            cell.alignment = Alignment(vertical="center", wrap_text=(c == 2 or c == 3))
            if chua_nop:
                # cả dòng đỏ để dễ thấy công ty chưa nộp
                cell.font = Font(color="C00000", bold=True)
            elif c == 9:
                if do:
                    cell.font = Font(color="C00000", bold=True)   # đỏ đậm
                elif "da chap nhan" in tt:
                    cell.font = Font(color="008000")              # xanh
                elif tt:
                    cell.font = Font(color="BF8F00")              # vàng (chờ)
        r += 1
    widths = [16, 42, 40, 10, 12, 13, 9, 11, 16]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, c).column_letter].width = w
    ws.freeze_panes = "A6"
    # bật filter cho các cột (hàng tiêu đề ở dòng 5)
    last_row = max(hr, r - 1)
    ws.auto_filter.ref = f"A{hr}:I{last_row}"
    wb.save(path)

def _xuat_saipass_excel(items, path):
    """items: list dict {mst, ten, ly_do}."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook(); ws = wb.active; ws.title = "SaiMatKhau"
    ws["A1"] = "DANH SÁCH CÔNG TY ĐĂNG NHẬP KHÔNG ĐƯỢC"
    ws["A1"].font = Font(bold=True, size=13)
    heads = ["Mã số thuế", "Tên công ty", "Lý do / Lỗi"]
    for c, h in enumerate(heads, 1):
        cell = ws.cell(3, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = Alignment(horizontal="center")
    r = 4
    for it in items:
        ws.cell(r, 1, it.get("mst", ""))
        ws.cell(r, 2, it.get("ten", ""))
        ws.cell(r, 3, str(it.get("ly_do", ""))[:500])
        r += 1
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 70
    wb.save(path)


@app.post("/api/dvc/test-login/{cid}")
def dvc_test_login(cid: int, body: dict = Body(...)):
    """Thử đăng nhập Dịch vụ công (chỉ kiểm tra, không tải gì).
    body: { matkhau, matkhau2, luu }"""
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        raise HTTPException(404, "Không tìm thấy công ty")
    pw1 = (body.get("matkhau") or (comp["dvc_password"] if "dvc_password" in comp.keys() else "") or "").strip()
    pw2 = (body.get("matkhau2") or (comp["dvc_password2"] if "dvc_password2" in comp.keys() else "") or "").strip()
    if not pw1 and not pw2:
        raise HTTPException(400, "Chưa nhập mật khẩu Dịch vụ công")
    try:
        drv = _dvc_make_driver(headless=True)
    except Exception as e:
        raise HTTPException(500,
            "Không khởi động được trình duyệt Chrome. Hãy chắc máy đã cài Google "
            f"Chrome và có mạng để tải chromedriver lần đầu. Chi tiết: {type(e).__name__}: {e}")
    try:
        ok, dung_pass, info = _dvc_browser_login_2pass(drv, comp["mst"], pw1, pw2)
    except Exception as e:
        import traceback
        raise HTTPException(500, f"Lỗi khi đăng nhập: {type(e).__name__}: {e} | "
                                 f"{traceback.format_exc()[-400:]}")
    finally:
        try: drv.quit()
        except Exception: pass
    if ok and body.get("luu"):
        conn = db()
        conn.execute("UPDATE companies SET dvc_password=?, dvc_password2=? WHERE id=?",
                     (pw1, pw2, cid))
        conn.commit(); conn.close()
    if not ok:
        raise HTTPException(401, f"Đăng nhập thất bại. {json.dumps(info, ensure_ascii=False)}")
    return {"ok": True, "dung_pass": dung_pass, **(info if isinstance(info, dict) else {}),
            "da_luu_mat_khau": bool(body.get("luu"))}


@app.get("/api/etax/explore/{cid}")
def etax_explore(cid: int):
    """CÔNG CỤ DÒ (tạm): đăng nhập dichvucong → mở 'Dịch vụ khác' → vào
    'Thông báo của CQT' (iframe eTax) → ghi lại cấu trúc để xây tự động hóa.
    Mở thẳng trên trình duyệt: http://127.0.0.1:8686/api/etax/explore/<id>"""
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        raise HTTPException(404, "Không tìm thấy công ty")
    pw1 = ((comp["dvc_password"] if "dvc_password" in comp.keys() else "") or "").strip()
    pw2 = (comp["dvc_password2"] if "dvc_password2" in comp.keys() else "") or ""
    pw2 = pw2.strip()
    info = {"buoc": []}
    try:
        drv = _dvc_make_driver(headless=True, esigner=True)
    except Exception as e:
        raise HTTPException(500, f"Không mở được Chrome: {e}")
    info["esigner_ext"] = getattr(drv, "_esigner_ext", None) or "(KHÔNG tìm thấy tiện ích eSigner đã cài)"
    try:
        ok, dp, li = _dvc_browser_login_2pass(drv, comp["mst"], pw1, pw2)
        info["dang_nhap"] = "OK" if ok else f"THẤT BẠI: {json.dumps(li, ensure_ascii=False)[:200]}"
        if not ok:
            return info
        import time as _t
        # 0) TEST trang CQT NATIVE cua dichvucong (khong can eSigner) - POST khong loc ngay
        try:
            drv.get(DVC_BASE + "/tra-cuu-thongbao-cqt"); _t.sleep(3)
            _dvc_wait_jquery(drv, 8)
            _JS_CQT = r"""
              var cb=arguments[arguments.length-1];
              try{
                var bs=document.querySelector('[name="btnSearch"]');
                var f=bs; while(f&&f.tagName!=='FORM') f=f.parentElement;
                if(!f){cb({ok:false,err:'no form'});return;}
                var p=new URLSearchParams();
                f.querySelectorAll('input[name],select[name],textarea[name]').forEach(function(e){
                  var n=e.name,v=e.value;
                  if(n==='tuNgay'||n==='denNgay')v=''; else if(n==='size')v='100'; else if(n==='page')v='0';
                  p.append(n,v==null?'':v);
                });
                fetch(f.getAttribute('action'),{method:'POST',headers:{'Content-Type':'application/x-www-form-urlencoded'},body:p.toString(),credentials:'same-origin'})
                  .then(r=>r.text()).then(t=>cb({ok:true,html:t})).catch(e=>cb({ok:false,err:''+e}));
              }catch(e){cb({ok:false,err:''+e});}
            """
            res = drv.execute_async_script(_JS_CQT)
            html = (res or {}).get("html") or ""
            info["dvc_cqt_len"] = len(html)
            import re as _re0
            try:
                pr = drv.execute_async_script(_JS_PARSE_TABLE, html)
                info["dvc_cqt_bang"] = ((pr or {}).get("rows") or [])[:8]
            except Exception as e:
                info["dvc_cqt_bang_err"] = str(e)
            flat0 = _re0.sub(r"\s+", " ", html)
            info["dvc_cqt_msg"] = [m.group(0)[:110] for m in _re0.finditer(
                r".{0,30}(?:b[ảa]n ghi|kh[ôo]ng c[óo] d[ữu] li[ệe]u|Tr[ạa]ng th[áa]i|S[ốô] th[ôo]ng b[áa]o).{0,30}", flat0, _re0.I)][:8]
            info["dvc_cqt_taive"] = list({o[:150] for o in _re0.findall(
                r'(?:onclick|href)=["\']([^"\']*(?:tai|download|idfile|tepdinhkem|thongbao|tbao|file)[^"\']*)["\']', html, _re0.I)})[:20]
            m0 = _re0.search(r'<table[^>]*>(?:(?!</table>)[\s\S]){0,2200}', html)
            info["dvc_cqt_table_html"] = (m0.group(0)[:2200] if m0 else "(khong thay table)")
        except Exception as e:
            info["dvc_cqt_err"] = str(e)
        # 1) Mở trang Dịch vụ khác
        drv.get(DVC_BASE + "/dich-vu-khac"); _t.sleep(3)
        info["dvk_title"] = drv.title
        info["dvk_url"] = drv.current_url
        # các phần tử có chữ liên quan "Thông báo của CQT"
        try:
            els = drv.find_elements("xpath",
                "//*[contains(translate(text(),'CQTHÔNGBÁO','cqthôngbáo'),'thông báo')]")
            info["el_thongbao"] = [ (e.tag_name + ":" + (e.text or "")[:40]) for e in els[:15] ]
        except Exception as e:
            info["el_thongbao_err"] = str(e)
        # iframe hiện có
        def dump_iframes(tag):
            try:
                fr = drv.find_elements("tag name", "iframe")
                return [{"src": (f.get_attribute("src") or "")[:160]} for f in fr]
            except Exception as e:
                return [f"err: {e}"]
        info["iframes_truoc"] = dump_iframes("truoc")
        import re as _re
        # 1b) DÒ cách 'Thông báo của CQT' được gắn: href, onclick, ngữ cảnh HTML
        full = drv.page_source or ""
        flat = _re.sub(r"\s+", " ", full)
        ctx = []
        for m in _re.finditer(r".{0,120}Th[ôo]ng b[áa]o c[ủu]a CQT.{0,120}", flat):
            ctx.append(m.group(0))
        info["ngu_canh_CQT"] = ctx[:4]
        # các link/onclick liên quan eTax/thuedientu/thong-bao
        info["href_lq"] = list({h for h in _re.findall(r'href=["\']([^"\']+)["\']', full)
                                if any(k in h.lower() for k in ("thuedientu","etaxnnt","thong-bao","tbao","cqt","request"))})[:20]
        info["onclick_lq"] = list({o[:120] for o in _re.findall(r'onclick=["\']([^"\']+)["\']', full)
                                   if any(k in o.lower() for k in ("thuedientu","etaxnnt","thong","tbao","cqt","sso","window.open"))})[:20]
        info["js_open"] = list({o[:140] for o in _re.findall(r'window\.open\([^)]{0,140}', full)})[:15]
        import re as _re
        # 2) Bấm card "Thông báo của CQT" = redirectHandler('360109',...) -> SSO sang eTax
        try:
            drv.execute_script("try{redirectHandler('360109','','','','Y','');}catch(e){}")
        except Exception as e:
            info["rh_err"] = str(e)
        _t.sleep(8)
        info["so_tab"] = len(drv.window_handles)
        # nếu mở tab mới -> sang tab mới nhất
        try:
            if len(drv.window_handles) > 1:
                drv.switch_to.window(drv.window_handles[-1])
                _t.sleep(3)
        except Exception:
            pass
        info["etax_url"] = drv.current_url
        info["etax_title"] = drv.title
        # nếu eTax nằm trong iframe -> vào iframe
        try:
            for f in drv.find_elements("tag name", "iframe"):
                s = (f.get_attribute("src") or "")
                if "thuedientu" in s or "etaxnnt" in s or "Request" in s:
                    info["etax_iframe_src"] = s[:160]
                    drv.switch_to.frame(f); _t.sleep(2)
                    break
        except Exception as e:
            info["etax_fr_err"] = str(e)
        # đọc trang eTax: bảng + nút Tra Cứu + form + link tải
        try:
            _dvc_wait_jquery(drv, 8)
        except Exception:
            pass
        src = drv.page_source or ""
        info["etax_len"] = len(src)
        # bảng kết quả (parse DOM hiện tại)
        try:
            pr = drv.execute_async_script(_JS_PARSE_TABLE, src)
            info["etax_bang"] = ((pr or {}).get("rows") or [])[:6]
        except Exception as e:
            info["etax_bang_err"] = str(e)
        flat = _re.sub(r"\s+", " ", src)
        info["etax_thongdiep"] = [m.group(0)[:120] for m in _re.finditer(
            r".{0,40}(?:b[ảa]n ghi|kh[ôo]ng c[óo] d[ữu] li[ệe]u|Tr[ạa]ng th[áa]i|S[ốô] th[ôo]ng b[áa]o).{0,40}", flat, _re.I)][:8]
        # nút Tra Cứu + form (dse_)
        info["etax_buttons"] = list({(m or "")[:40] for m in _re.findall(r'<(?:button|input)[^>]*value=["\']([^"\']+)["\']', src)})[:20]
        info["etax_forms"] = _re.findall(r'<form[^>]*action=["\']([^"\']*)["\']', src)[:4]
        info["etax_inputs"] = list({n for n in _re.findall(r'<input[^>]+name=["\']([^"\']+)["\']', src)})[:30]
        # link/onclick "Tải thông báo"
        info["etax_taive"] = list({o[:160] for o in _re.findall(
            r'(?:onclick|href)=["\']([^"\']*(?:tai|download|getFile|TBao|dse_)[^"\']*)["\']', src, _re.I)})[:25]
        m = _re.search(r'<table[^>]*>(?:(?!</table>)[\s\S]){0,2600}', src)
        info["etax_table_html"] = (m.group(0)[:2600] if m else "(khong thay table)")
        return info
    except Exception as e:
        import traceback
        info["loi"] = f"{e} | {traceback.format_exc()[-300:]}"
        return info
    finally:
        try: drv.quit()
        except Exception: pass


@app.get("/api/dvc/captcha-debug/{cid}", response_class=HTMLResponse)
def dvc_captcha_debug(cid: int):
    """Trang xem captcha Dịch vụ công thật + thử OCR, để dò vì sao không giải được.
    Mở: http://127.0.0.1:8686/api/dvc/captcha-debug/<id-công-ty>"""
    client = DVCClient()
    DVC_CLIENTS[cid] = client
    try:
        client.prime()
    except Exception as e:
        return f"<h3>Lỗi prime: {e}</h3>"
    ts = int(time.time() * 1000)
    h = {
        "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        "Referer": DVC_BASE + "/login",
        "Sec-Fetch-Dest": "image", "Sec-Fetch-Mode": "no-cors",
        "Sec-Fetch-Site": "same-origin",
    }
    # KHÔNG theo redirect để xem server thực sự trả gì
    r = client.session.get(f"{DVC_BASE}/getCaptcha?{ts}", headers=h, timeout=30,
                           allow_redirects=False)
    redir = {"status": r.status_code, "location": r.headers.get("location"),
             "url_cuoi": r.url}
    ct = (r.headers.get("content-type") or "").lower()
    data = r.content or b""

    # ---- Thử nhiều biến thể request để tìm cái trả về ẢNH ----
    thi_nghiem = []
    variants = [
        ("XHR + referer /login",
         {**h, "X-Requested-With": "XMLHttpRequest"}),
        ("referer /homelogin",
         {**h, "Referer": DVC_BASE + "/homelogin"}),
        ("referer /tchs (captcha lần 2)",
         {**h, "Referer": DVC_BASE + "/tchs", "X-Requested-With": "XMLHttpRequest"}),
        ("không có Sec-Fetch, có XHR",
         {"Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
          "Referer": DVC_BASE + "/login", "X-Requested-With": "XMLHttpRequest"}),
    ]
    for ten, hv in variants:
        try:
            ts2 = int(time.time() * 1000)
            rv = client.session.get(f"{DVC_BASE}/getCaptcha?{ts2}", headers=hv,
                                    timeout=30, allow_redirects=False)
            cv = (rv.headers.get("content-type") or "").lower()
            dv = rv.content or b""
            la_anh = ("image" in cv or "svg" in cv
                      or dv[:5].lstrip().lower().startswith(b"<svg")
                      or dv[:4] in (b"\x89PNG", b"GIF8") or dv[:2] == b"\xff\xd8")
            thi_nghiem.append({"ten": ten, "status": rv.status_code,
                               "ct": cv or "(trống)", "len": len(dv),
                               "la_anh": "✅ ẢNH" if la_anh else "❌ không phải ảnh",
                               "loc": rv.headers.get("location") or ""})
        except Exception as e:
            thi_nghiem.append({"ten": ten, "status": "lỗi", "ct": str(e),
                               "len": 0, "la_anh": "", "loc": ""})
    head = data[:400]
    head_txt = head.decode("utf-8", "replace")
    is_svg = ("svg" in ct or head.lstrip()[:5].lower().startswith(b"<svg")
              or b"<svg" in head.lower())
    # rasterize nếu svg
    png = _svg_to_png(data.decode("utf-8", "replace")) if is_svg else data
    import re as _re
    has_text = ("<text" in head_txt.lower())
    n_path = data.lower().count(b"<path")
    # OCR thử
    ocr = _get_ddddocr()
    ocr_raw = ""
    if ocr:
        try:
            ocr_raw = (ocr.classification(data) or "").strip()
        except Exception as e:
            ocr_raw = f"(lỗi: {e})"
    ocr_png = _ocr_png(png) if png else ""
    # ---- Soi trang /login để tìm URL captcha & form thật ----
    import re as _re2
    soi = {"login": [], "img": [], "form": [], "captcha_refs": []}
    try:
        nav = {"Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
               "Sec-Fetch-Dest": "document", "Sec-Fetch-Mode": "navigate",
               "Sec-Fetch-Site": "same-origin", "Referer": DVC_BASE + "/homelogin"}
        rl = client.session.get(DVC_BASE + "/login", headers=nav, timeout=30,
                                allow_redirects=True)
        ltxt = rl.text or ""
        soi["login"] = [f"status={rl.status_code}", f"len={len(ltxt)}",
                        f"url_cuoi={rl.url}"]
        # làm phẳng khoảng trắng để regex bắt qua nhiều dòng
        flat = _re2.sub(r'\s+', ' ', ltxt)
        for m in _re2.finditer(r'.{0,70}[Cc]aptcha.{0,70}', flat):
            soi["captcha_refs"].append(m.group(0))
        # gộp trùng, giữ tối đa 20
        seen = set(); uniq = []
        for s in soi["captcha_refs"]:
            if s not in seen:
                seen.add(s); uniq.append(s)
        soi["captcha_refs"] = uniq[:20]
        # các lời gọi ajax/fetch/getCaptcha/loginLDAP trong JS
        for kw in ("getCaptcha", "loginLDAP", "checkCaptcha", "$.ajax", "fetch(",
                   "XMLHttpRequest", "/tthc/"):
            for m in _re2.finditer(_re2.escape(kw) + r'.{0,80}', flat):
                soi["img"].append(m.group(0))  # tái dùng list 'img' để hiện
        seen = set(); uniq = []
        for s in soi["img"]:
            if s not in seen:
                seen.add(s); uniq.append(s)
        soi["img"] = uniq[:30]
        # form action + input + giá trị _csrf
        for m in _re2.findall(r'<form[^>]*action=["\']([^"\']*)["\']', ltxt)[:5]:
            soi["form"].append("action=" + m)
        for m in _re2.findall(r'<input[^>]+name=["\']([^"\']+)["\'][^>]*?(?:value=["\']([^"\']*)["\'])?',
                              ltxt)[:25]:
            nm, vl = m
            soi["form"].append(f"input={nm}" + (f" = {vl[:40]}" if vl else ""))
    except Exception as e:
        soi["login"] = [f"lỗi soi login: {e}"]

    raw_b64 = base64.b64encode(data).decode()
    png_b64 = base64.b64encode(png).decode() if png else ""
    mime = ct.split(";")[0] or ("image/svg+xml" if is_svg else "image/png")
    html = f"""<html><head><meta charset='utf-8'><title>DVC Captcha Debug</title>
<style>body{{font-family:system-ui;padding:24px;line-height:1.6}}
img{{border:2px solid #0d6efd;background:#fff;image-rendering:pixelated;margin:6px 0;display:block}}
code{{background:#f0f0f0;padding:2px 5px;border-radius:4px}}
pre{{background:#f6f8fa;padding:10px;border-radius:8px;overflow:auto;max-height:200px}}</style></head><body>
<h2>🔍 Chẩn đoán captcha Dịch vụ công</h2>
<p><b>ddddocr cài đặt:</b> {'✅ có' if ocr else '❌ CHƯA CÀI — đây là nguyên nhân!'} {('('+_DDDDOCR_ERR+')') if _DDDDOCR_ERR else ''}</p>
<p><b>Giả lập TLS Chrome (curl_cffi):</b> {'✅ CÓ — đang vượt vân tay WAF' if client.impersonate else '❌ KHÔNG (chưa cài curl_cffi) — sẽ bị WAF chặn'}</p>
<p><b>HTTP status:</b> {r.status_code} · <b>Content-Type:</b> <code>{ct or '(trống)'}</code> · <b>Kích thước:</b> {len(data)} bytes</p>
<p><b>Redirect (Location):</b> <code>{redir.get('location') or '(không)'}</code></p>
<h3>Thử các biến thể request (tìm cái ra ẢNH):</h3>
<table border="1" cellpadding="6" style="border-collapse:collapse">
<tr><th>Cách gọi</th><th>Status</th><th>Content-Type</th><th>Bytes</th><th>Kết quả</th><th>Location</th></tr>
{''.join(f"<tr><td>{t['ten']}</td><td>{t['status']}</td><td><code>{t['ct']}</code></td><td>{t['len']}</td><td>{t['la_anh']}</td><td>{t['loc']}</td></tr>" for t in thi_nghiem)}
</table>
<p><b>Là SVG:</b> {is_svg} · <b>Có &lt;text&gt;:</b> {has_text} · <b>Số &lt;path&gt;:</b> {n_path}</p>
<p><b>OCR ảnh gốc:</b> <code>{ocr_raw or '(rỗng)'}</code> · <b>OCR sau rasterize:</b> <code>{ocr_png or '(rỗng)'}</code></p>
<h3>🔎 Soi trang /login (tìm URL captcha & form thật):</h3>
<p><b>Trang login:</b> {' · '.join(soi['login'])}</p>
<p><b>Các chỗ chứa "captcha":</b></p>
<pre>{chr(10).join(x.replace('<','&lt;') for x in soi['captcha_refs']) or '(không có — login là SPA/JS render)'}</pre>
<p><b>Lời gọi JS / URL liên quan (getCaptcha, loginLDAP, ajax...):</b></p>
<pre>{chr(10).join(x.replace('<','&lt;') for x in soi['img']) or '(không có)'}</pre>
<p><b>form action / input name:</b></p>
<pre>{chr(10).join(soi['form']) or '(không có)'}</pre>
<h3>Ảnh captcha (gốc):</h3>
<img src="data:{mime};base64,{raw_b64}" width="260">
<h3>Ảnh sau khi chuyển PNG (đưa vào OCR):</h3>
{('<img src="data:image/png;base64,'+png_b64+'" width="260">') if png_b64 else '<i>không rasterize được</i>'}
<h3>400 byte đầu của nội dung:</h3>
<pre>{head_txt.replace('<','&lt;')}</pre>
<p style="color:#666">Bấm F5 để lấy captcha khác. Chụp màn hình trang này gửi lại để tinh chỉnh.</p>
</body></html>"""
    return html


@app.post("/api/dvc/tai-bao-cao/{cid}")
def dvc_tai_bao_cao(cid: int, body: dict = Body(...)):
    """Đăng nhập Dịch vụ công → tra cứu tờ khai/báo cáo ĐÃ NỘP trong khoảng ngày
    → tải file zip về thư mục công ty.
    body: { matkhau, tu_ngay (dd/mm/yyyy), den_ngay (dd/mm/yyyy), luu: bool }"""
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        raise HTTPException(404, "Không tìm thấy công ty")
    pw1 = (body.get("matkhau") or (comp["dvc_password"] if "dvc_password" in comp.keys() else "") or "").strip()
    pw2 = (body.get("matkhau2") or (comp["dvc_password2"] if "dvc_password2" in comp.keys() else "") or "").strip()
    if not pw1 and not pw2:
        raise HTTPException(400, "Chưa nhập mật khẩu Dịch vụ công")
    tu = (body.get("tu_ngay") or "").strip()
    den = (body.get("den_ngay") or "").strip()
    if not tu or not den:
        raise HTTPException(400, "Chưa chọn khoảng ngày (Từ ngày / Đến ngày)")

    folder = _dvc_save_folder(cid)
    if not folder:
        raise HTTPException(400, "Chưa cấu hình thư mục lưu cho công ty này")

    try:
        drv = _dvc_make_driver(headless=True)
    except Exception as e:
        raise HTTPException(500,
            "Không khởi động được trình duyệt Chrome. Hãy chắc máy đã cài Google "
            f"Chrome và có mạng để tải chromedriver lần đầu. Chi tiết: {type(e).__name__}: {e}")

    da_tai, loi_tai, ma_list, search_diag, info = [], [], [], [], {}
    try:
        # 1) Đăng nhập (trình duyệt thật → qua WAF), thử pass1 rồi pass2
        ok, dung_pass, info = _dvc_browser_login_2pass(drv, comp["mst"], pw1, pw2)
        if not ok:
            raise HTTPException(401, f"Đăng nhập thất bại. {json.dumps(info, ensure_ascii=False)}")
        if body.get("luu"):
            conn = db()
            conn.execute("UPDATE companies SET dvc_password=?, dvc_password2=? WHERE id=?",
                         (pw1, pw2, cid))
            conn.commit(); conn.close()

        # 2) Tra cứu hồ sơ đã nộp (captcha lớp 2)
        ma_list, search_diag = _dvc_browser_search(drv, tu, den)
        if not ma_list:
            raise HTTPException(422, "Đăng nhập OK nhưng không tra cứu được danh sách hồ sơ. "
                                     f"Chi tiết: {json.dumps(search_diag, ensure_ascii=False)}")

        # 3) Tải từng hồ sơ: vào trang chi tiết → tải tờ khai + các thông báo
        tai_tb = body.get("tai_thong_bao", True)
        tb_diag = []
        def _luu(fname, raw, pre=""):
            safe = "".join(ch for ch in (pre + fname) if ch not in '\\/:*?"<>|')
            if not safe:
                safe = "file.zip"
            path = os.path.join(folder, safe)
            # tránh ghi đè trùng tên
            if os.path.exists(path):
                base, ext = os.path.splitext(safe)
                k = 2
                while os.path.exists(os.path.join(folder, f"{base}_{k}{ext}")):
                    k += 1
                safe = f"{base}_{k}{ext}"; path = os.path.join(folder, safe)
            with open(path, "wb") as f:
                f.write(raw)
            return safe

        for ma in ma_list:
            # 3a) Mở trang chi tiết hồ sơ (để có ngữ cảnh + đọc idTbao thông báo)
            tb_files = []
            if tai_tb:
                try:
                    tb_files, d = _dvc_browser_thongbao(drv, ma)
                    tb_diag += d
                except Exception as e:
                    tb_diag.append(f"{ma}: lỗi thông báo: {e}")
            # 3b) Tải tờ khai (zip hồ sơ)
            try:
                fname, raw = _dvc_browser_download(drv, ma)
                if not raw:
                    loi_tai.append(f"{ma}: nội dung rỗng")
                else:
                    safe = _luu(fname, raw)
                    da_tai.append({"maHoSo": ma, "file": safe, "kich_thuoc": len(raw)})
            except Exception as e:
                loi_tai.append(f"{ma}: {e}")
            # 3c) Lưu các thông báo của hồ sơ này
            for tb_fname, tb_raw in tb_files:
                try:
                    safe = _luu(tb_fname, tb_raw)
                    da_tai.append({"maHoSo": ma, "file": safe,
                                   "kich_thuoc": len(tb_raw), "loai": "thông báo"})
                except Exception as e:
                    loi_tai.append(f"{ma} (TB): {e}")
            time.sleep(0.4)
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        raise HTTPException(500, f"Lỗi: {type(e).__name__}: {e} | {traceback.format_exc()[-400:]}")
    finally:
        try: drv.quit()
        except Exception: pass

    return {
        "ok": True,
        "thu_muc": folder,
        "so_ho_so": len(ma_list),
        "da_tai": da_tai,
        "loi_tai": loi_tai,
        "chan_doan_dang_nhap": info,
        "chan_doan_tra_cuu": search_diag,
        "chan_doan_thong_bao": tb_diag,
    }


# ============================================================
#  DVC HÀNG LOẠT: đăng nhập + tra cứu + (tùy chọn) tải, cho NHIỀU công ty.
#  Chạy nền (1 trình duyệt dùng lại cho từng công ty), poll trạng thái.
#  Xuất 2 Excel: Tra cứu tờ khai (tổng hợp) + Sai mật khẩu.
# ============================================================
DVC_BATCH = {}        # batch_id -> {...}
_DVC_BATCH_SEQ = {"n": 0}

def _dvc_run_batch(batch_id, cids, body):
    job = DVC_BATCH[batch_id]
    tu = (body.get("tu_ngay") or "").strip()
    den = (body.get("den_ngay") or "").strip()
    ky_label = (body.get("ky_label") or f"{tu} - {den}").strip()
    tai_file = bool(body.get("tai_file"))       # có tải file tờ khai về không
    tai_tb = bool(body.get("tai_thong_bao", True))
    luu_pass = bool(body.get("luu", True))
    tracuu_rows = []      # gom mọi dòng cho Excel tra cứu
    sai_pass = []         # công ty đăng nhập không được
    drv = None
    try:
        try:
            drv = _dvc_make_driver(headless=True)
        except Exception as e:
            job["loi"] = f"Không mở được Chrome: {e}"
            job["running"] = False
            return
        for cid in cids:
            if job.get("cancel"):
                break
            conn = db()
            comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
            conn.close()
            if not comp:
                job["done"] += 1; continue
            ten = comp["ten"]; mst = comp["mst"]
            job["current"] = ten
            item = {"cid": cid, "mst": mst, "ten": ten, "trang_thai": "đang xử lý",
                    "so_dong": 0, "so_file": 0}
            job["items"].append(item)
            pw1 = (comp["dvc_password"] if "dvc_password" in comp.keys() else "") or ""
            pw2 = (comp["dvc_password2"] if "dvc_password2" in comp.keys() else "") or ""
            pw1 = (pw1 or "").strip(); pw2 = (pw2 or "").strip()
            # mật khẩu chung dự phòng (nếu người dùng nhập ở batch)
            if body.get("pass_chung"):
                pw2 = pw2 or (body.get("pass_chung") or "").strip()
            if not pw1 and not pw2:
                item["trang_thai"] = "thiếu mật khẩu"
                sai_pass.append({"mst": mst, "ten": ten, "ly_do": "Chưa lưu mật khẩu Dịch vụ công"})
                job["done"] += 1; continue
            # xoá phiên công ty trước để tránh lẫn đăng nhập
            try:
                drv.get(DVC_BASE + "/homelogin")
                drv.delete_all_cookies()
            except Exception:
                pass
            try:
                ok, dung_pass, info = _dvc_browser_login_2pass(drv, mst, pw1, pw2)
            except Exception as e:
                ok, dung_pass, info = False, 0, {"loi": str(e)}
            if not ok:
                item["trang_thai"] = "sai mật khẩu / lỗi đăng nhập"
                sai_pass.append({"mst": mst, "ten": ten,
                                 "ly_do": json.dumps(info, ensure_ascii=False)[:480]})
                job["done"] += 1; continue
            item["dung_pass"] = dung_pass
            # tra cứu
            try:
                rows, ma_list, raw_html, sdiag = _dvc_browser_tracuu(drv, tu, den)
            except Exception as e:
                rows, ma_list, raw_html, sdiag = [], [], "", [f"lỗi tra cứu: {e}"]
            for rec in rows:
                rec2 = {"mst": mst, "ten": ten}; rec2.update(rec)
                tracuu_rows.append(rec2)
            item["so_dong"] = len(rows)
            # Nếu tra cứu THÀNH CÔNG nhưng KHÔNG có tờ khai nào trong kỳ
            # -> ghi chú "CHƯA NỘP TỜ KHAI" (đỏ). Nếu tra cứu lỗi thì không kết luận.
            if not rows:
                if raw_html:
                    tracuu_rows.append({
                        "mst": mst, "ten": ten,
                        "to_khai": "(Chưa tìm thấy tờ khai trong kỳ)",
                        "ky": ky_label, "loai": "", "ngay_nop": "",
                        "lan_nop": "", "lan_bs": "",
                        "trang_thai": "CHƯA NỘP TỜ KHAI",
                    })
                    item["chua_nop"] = True
                else:
                    item["loi_tra_cuu"] = "; ".join(sdiag)[:200]
            # tải file (tùy chọn)
            if tai_file and ma_list:
                folder = _dvc_save_folder(cid)
                if folder:
                    seen_ma = set()
                    # 1) các hồ sơ dò được đủ dữ liệu dòng (to_khai/ky/lan_bs) -> đặt tên thân thiện
                    for rec in rows:
                        ma = (rec.get("ma") or "").strip()
                        if not ma or ma in seen_ma:
                            continue
                        seen_ma.add(ma)
                        if job.get("cancel"):
                            break
                        ten_goi = _ten_file_than_thien(rec.get("to_khai"), rec.get("ky"), mst, rec.get("lan_bs"))
                        try:
                            if tai_tb:
                                tb_files, _ = _dvc_browser_thongbao(drv, ma)
                                so_tb = len(tb_files)
                                for k, (fn, raw) in enumerate(tb_files, 1):
                                    if raw:
                                        ext = os.path.splitext(fn)[1] or ".xml"
                                        hau_to = f"_ThongBao{k}" if so_tb > 1 else "_ThongBao"
                                        _dvc_luu_file(folder, f"{ten_goi}{hau_to}{ext}", raw); item["so_file"] += 1
                            fn, raw = _dvc_browser_download(drv, ma)
                            if raw:
                                ext = os.path.splitext(fn)[1] or ".zip"
                                _dvc_luu_file(folder, f"{ten_goi}{ext}", raw); item["so_file"] += 1
                        except Exception:
                            pass
                        time.sleep(0.3)
                    # 2) các mã hồ sơ tìm thấy nhưng KHÔNG ghép được dữ liệu dòng -> tải với tên gốc (dự phòng)
                    for ma in ma_list:
                        if ma in seen_ma:
                            continue
                        seen_ma.add(ma)
                        if job.get("cancel"):
                            break
                        try:
                            if tai_tb:
                                tb_files, _ = _dvc_browser_thongbao(drv, ma)
                                for fn, raw in tb_files:
                                    if raw:
                                        _dvc_luu_file(folder, fn, raw); item["so_file"] += 1
                            fn, raw = _dvc_browser_download(drv, ma)
                            if raw:
                                _dvc_luu_file(folder, fn, raw); item["so_file"] += 1
                        except Exception:
                            pass
                        time.sleep(0.3)
            item["trang_thai"] = "xong"
            job["done"] += 1
            time.sleep(0.5)
    finally:
        try:
            if drv: drv.quit()
        except Exception:
            pass
    # xuất Excel
    try:
        out_dir = _dvc_batch_out_dir()
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        if tracuu_rows:
            p = os.path.join(out_dir, f"TraCuuToKhai_{ts}.xlsx")
            _xuat_tracuu_excel(tracuu_rows, p)
            job["file_tracuu"] = p
            _open_file_local(p)   # tự mở file Excel trên máy
        if sai_pass:
            p2 = os.path.join(out_dir, f"SaiMatKhau_{ts}.xlsx")
            _xuat_saipass_excel(sai_pass, p2)
            job["file_saipass"] = p2
    except Exception as e:
        job["loi_xuat"] = str(e)
    job["so_dong_tong"] = len(tracuu_rows)
    job["so_sai_pass"] = len(sai_pass)
    job["running"] = False
    job["current"] = None

def _dvc_luu_file(folder, fname, raw):
    safe = "".join(ch for ch in fname if ch not in '\\/:*?"<>|') or "file.zip"
    path = os.path.join(folder, safe)
    if os.path.exists(path):
        base, ext = os.path.splitext(safe); k = 2
        while os.path.exists(os.path.join(folder, f"{base}_{k}{ext}")):
            k += 1
        path = os.path.join(folder, f"{base}_{k}{ext}")
    with open(path, "wb") as f:
        f.write(raw)
    return path

def _dvc_batch_out_dir():
    """Thư mục lưu 2 file Excel tổng hợp: ưu tiên save_dir/data_dir đầu tiên có."""
    conn = db()
    row = conn.execute("SELECT save_dir, data_dir FROM companies "
                       "WHERE (save_dir IS NOT NULL AND save_dir<>'') "
                       "OR (data_dir IS NOT NULL AND data_dir<>'') LIMIT 1").fetchone()
    conn.close()
    base = ""
    if row:
        base = (row["save_dir"] or row["data_dir"] or "").strip()
    if not base or not os.path.isdir(base):
        base = os.path.join(DATA_DIR, "tracuu_thue")
    os.makedirs(base, exist_ok=True)
    return base


@app.post("/api/dvc/batch")
def dvc_batch_start(body: dict = Body(...)):
    """Bắt đầu chạy hàng loạt. body: { cids:[...], tu_ngay, den_ngay,
       tai_file:bool, tai_thong_bao:bool, pass_chung?:str }. Trả { batch_id }."""
    cids = body.get("cids") or []
    if not cids:
        conn = db()
        cids = [r["id"] for r in conn.execute("SELECT id FROM companies ORDER BY ten").fetchall()]
        conn.close()
    if not (body.get("tu_ngay") and body.get("den_ngay")):
        raise HTTPException(400, "Chưa chọn khoảng ngày")
    _DVC_BATCH_SEQ["n"] += 1
    bid = _DVC_BATCH_SEQ["n"]
    DVC_BATCH[bid] = {"running": True, "total": len(cids), "done": 0,
                      "current": None, "items": [], "cancel": False}
    threading.Thread(target=lambda: _dvc_run_batch(bid, cids, body), daemon=True).start()
    return {"ok": True, "batch_id": bid}

@app.get("/api/dvc/batch-status/{bid}")
def dvc_batch_status(bid: int):
    job = DVC_BATCH.get(bid)
    if not job:
        raise HTTPException(404, "Không tìm thấy phiên")
    return job

@app.post("/api/dvc/batch-cancel/{bid}")
def dvc_batch_cancel(bid: int):
    job = DVC_BATCH.get(bid)
    if job:
        job["cancel"] = True
    return {"ok": True}

@app.get("/api/dvc/tai-excel")
def dvc_tai_excel(path: str):
    """Tải 1 file Excel tổng hợp đã tạo (đường dẫn do batch-status trả về)."""
    from fastapi.responses import FileResponse
    out_dir = os.path.realpath(_dvc_batch_out_dir())
    rp = os.path.realpath(path)
    # an toàn: chỉ cho tải file .xlsx trong thư mục xuất
    if not rp.startswith(out_dir) or not rp.endswith(".xlsx") or not os.path.isfile(rp):
        # cũng cho phép nằm trong DATA_DIR
        if not (rp.startswith(os.path.realpath(DATA_DIR)) and os.path.isfile(rp)):
            raise HTTPException(404, "File không hợp lệ")
    return FileResponse(rp, filename=os.path.basename(rp),
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------- TRA CỨU & TẢI HÓA ĐƠN (streaming tiến độ) ----------
def _run_fetch_job(cid: int, body: dict):
    """Lõi tra cứu + tải hóa đơn cho MỘT công ty (chạy đồng bộ trong thread).

    Giả định: FETCH_JOBS[cid] đã được khởi tạo bởi nơi gọi, và token còn hiệu lực.
    Dùng chung cho tra cứu 1 công ty (/api/fetch) và tra cứu hàng loạt
    (/api/fetch-batch).

    body: {
      tu_ngay, den_ngay,
      loai_list: ["purchase","sold"],
      he_thong_list: ["query","sco-query"],
      dl_buy:  ["xml","pdf"]  # định dạng tải cho mua vào (rỗng = không tải)
      dl_sell: ["xml","pdf"]  # định dạng tải cho bán ra
    }
    """
    client = get_client(cid)

    tu = body.get("tu_ngay")
    den = body.get("den_ngay")
    loai_list = body.get("loai_list", ["purchase", "sold"])
    he_thong_list = body.get("he_thong_list", ["query", "sco-query"])
    lay_ngan_hang = body.get("lay_ngan_hang", False)  # mặc định KHÔNG lấy HĐ ngân hàng
    dl_map = {
        "purchase": body.get("dl_buy", []),
        "sold": body.get("dl_sell", []),
    }

    def la_hoa_don_ngan_hang(inv):
        """Nhận diện hóa đơn ngân hàng/phí dịch vụ NH qua tên người bán."""
        ten = (inv.get("nbten") or "").upper()
        tu_khoa = ["NGAN HANG", "NGÂN HÀNG", "BANK", "TMCP", "THƯƠNG MẠI CỔ PHẦN",
                   "VIETCOMBANK", "VIETINBANK", "AGRIBANK", "BIDV", "TECHCOMBANK",
                   "SACOMBANK", "ACB", "MBBANK", "VPBANK", "TPBANK", "SHB"]
        return any(k in ten for k in tu_khoa)

    conn0 = db()
    comp = conn0.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    save_dir = (comp["save_dir"] or "").strip() if comp else ""
    conn0.close()

    def msg(**kw):
        # Ghi tiến độ vào job thay vì yield (để chạy nền, không phụ thuộc trình duyệt)
        job = FETCH_JOBS.get(cid)
        if job is not None:
            job["messages"].append(kw)
            job["last"] = kw
        # Lưu LỖI vào công ty (DB) để không bị mất khi thông báo/toast tự tắt —
        # người dùng mở lại công ty vẫn thấy đúng lỗi lần tra cứu gần nhất.
        # "done" (thành công, không lỗi) thì xóa ghi chú lỗi cũ đi.
        stage = kw.get("stage")
        if stage == "error":
            _luu_loi_tra_cuu(cid, kw.get("text") or "")
        elif stage == "done":
            _xoa_loi_tra_cuu(cid)
        return None

    def run():
        conn = db()
        try:
            # Ghi nhớ TRẠNG THÁI (tthai) của các hóa đơn ĐÃ CÓ trước khi xóa —
            # dùng để phát hiện hóa đơn bị THAY THẾ/ĐIỀU CHỈNH so với lần tra
            # cứu trước (vd hóa đơn số 10 lần trước "mới" (1), lần này trang
            # Thuế báo đã "bị thay thế" (4) -> PHẢI tải lại, không dùng file cũ).
            old_status = {}
            for r in conn.execute(
                    "SELECT khmshdon, khhdon, shdon, loai, he_thong, tthai "
                    "FROM invoices WHERE company_id=?", (cid,)):
                old_status[(str(r["khmshdon"] or ""), str(r["khhdon"] or ""),
                            str(r["shdon"] or ""), r["loai"], r["he_thong"])] = str(r["tthai"] or "")
            # KHÔNG xóa dữ liệu cũ ngay từ đầu nữa — TRƯỚC ĐÂY xóa sạch hết rồi
            # mới tra cứu lại, nên nếu lượt tra cứu MỚI này lại LỖI (vd 1 trạng
            # thái mua vào lỗi dù đã thử lại) thì dữ liệu CŨ (đã lưu từ lần
            # TRƯỚC thành công) cũng biến mất theo — nhìn như phần mềm "tự mất
            # dữ liệu". Nay chỉ xóa dữ liệu cũ CỦA TỪNG LOẠI (mua/bán x hệ
            # thống) NGAY TRƯỚC KHI lưu kết quả MỚI THÀNH CÔNG của đúng loại
            # đó — loại nào lỗi thì GIỮ NGUYÊN dữ liệu cũ, không bị mất.
            msg(stage="start", text="Bắt đầu tra cứu...")

            total_saved = 0
            file_saved = 0
            file_thieu_tong = 0   # số file XML KHÔNG tải được (sau khi đã thử lại), gộp cả kỳ
            # đếm theo loại để tổng kết: {loai: {"exp": tổng trang Thuế báo, "got": số lấy được}}
            thongke = {"purchase": {"exp": 0, "got": 0}, "sold": {"exp": 0, "got": 0}}
            # Đánh dấu RIÊNG khi 1 loại (mua/bán) bị LỖI THẬT SỰ dù đã thử lại —
            # PHẢI phân biệt với "0 hóa đơn" hợp lệ, để KHÔNG BAO GIỜ báo "Hoàn tất"
            # sạch sẽ khi thực ra chưa tra cứu được (nguyên nhân gốc của việc phần
            # mềm "báo xong" dù bán ra bị lỗi/rớt mạng giữa chừng).
            loai_that_bai = {"purchase": False, "sold": False}
            target_dir = None
            if save_dir:
                try:
                    os.makedirs(save_dir, exist_ok=True)
                    target_dir = save_dir
                except Exception as e:
                    msg(stage="warn", text=f"Không tạo được thư mục lưu: {e}")

            for li, loai in enumerate(loai_list):
                loai_txt = "mua vào" if loai == "purchase" else "bán ra"
                # nghỉ trước khi chuyển từ mua sang bán (mua đã gọi nhiều request)
                if li > 0:
                    time.sleep(SP().get("between_loai", 2))
                for he_thong in he_thong_list:
                    # dừng nếu người dùng chuyển công ty khác
                    _job = FETCH_JOBS.get(cid)
                    if _job and _job.get("cancel"):
                        msg(stage="warn", text="Đã dừng tra cứu (chuyển sang công ty khác)")
                        return
                    ht_txt = " (máy tính tiền)" if he_thong == "sco-query" else ""
                    msg(stage="query",
                        text=f"Đang tra cứu hóa đơn {loai_txt}{ht_txt}...")

                    # Thử tối đa 2 LƯỢT cho toàn bộ khoảng ngày (mỗi lượt bên trong
                    # đã tự thử lại theo tháng/trạng thái) trước khi báo lỗi hẳn.
                    invs = None
                    loi_cuoi = None
                    bo_qua_404 = False
                    for lan_thu in range(2):
                        try:
                            invs = client.query_invoices(
                                tu, den, loai=loai, he_thong=he_thong,
                                progress=lambda t: msg(stage="query", text=f"{loai_txt}{ht_txt}: {t}"))
                            loi_cuoi = None
                            break
                        except Exception as e:
                            es = str(e)
                            if "TOKEN_EXPIRED" in es:
                                msg(stage="error", text="Token hết hạn, cần đăng nhập lại")
                                return
                            if he_thong == "sco-query" and "404" in es:
                                bo_qua_404 = True
                                break
                            loi_cuoi = es
                            if lan_thu == 0:
                                cho = 30 if ("429" in es or "quá nhiều" in es) else 8
                                msg(stage="warn",
                                    text=f"⚠ {loai_txt}{ht_txt}: {es[:140]}. Đang chờ {cho}s rồi thử lại...")
                                time.sleep(cho)

                    if bo_qua_404:
                        msg(stage="warn", text=f"{loai_txt}{ht_txt}: không có (404) — bỏ qua")
                        continue
                    if loi_cuoi is not None:
                        # Đã thử lại vẫn lỗi -> ĐÂY LÀ LỖI THẬT SỰ, không phải "0 hóa đơn".
                        # KHÔNG được continue âm thầm mà không đánh dấu — nếu không, phần
                        # tổng kết cuối cùng không biết loại này đã THẤT BẠI và có thể
                        # báo "Hoàn tất" sạch sẽ dù trang Thuế thực ra có dữ liệu.
                        loai_that_bai[loai] = True
                        msg(stage="error",
                            text=f"✗ {loai_txt}{ht_txt}: LỖI, CHƯA TRA CỨU ĐƯỢC dù đã thử lại — "
                                 f"{loi_cuoi[:140]}. KẾT QUẢ {loai_txt.upper()} CÓ THỂ THIẾU — "
                                 f"NÊN TRA CỨU LẠI riêng kỳ/công ty này (dữ liệu {loai_txt}{ht_txt} "
                                 f"đã lưu từ lần tra cứu trước — NẾU CÓ — vẫn được GIỮ NGUYÊN, "
                                 f"không bị xóa).")
                        continue

                    # Không raise lỗi nhưng vẫn có thể lỗi RIÊNG PHẦN (vd 1 trong 3 trạng
                    # thái mua vào lỗi dù 2 cái kia OK) — vẫn phải cảnh báo, không im lặng.
                    loi_rieng = getattr(client, "last_query_errors", None) or []
                    if loi_rieng:
                        loai_that_bai[loai] = True
                        msg(stage="warn",
                            text=f"⚠ {loai_txt}{ht_txt}: một phần bị lỗi dù đã thử lại "
                                 f"({len(loi_rieng)} lượt) — KẾT QUẢ CÓ THỂ THIẾU. "
                                 f"Chi tiết: {'; '.join(loi_rieng[:3])}")

                    exp0 = getattr(client, "last_query_total", 0) or 0
                    if not invs:
                        if exp0:
                            # Trang Thuế báo CÓ hóa đơn nhưng ta lấy được 0 -> chắc chắn
                            # có vấn đề, TUYỆT ĐỐI không được coi là "không có dữ liệu".
                            loai_that_bai[loai] = True
                            msg(stage="error",
                                text=f"✗ {loai_txt}{ht_txt}: Trang Thuế báo có {exp0} hóa đơn "
                                     f"nhưng KHÔNG lấy được cái nào — LỖI, nên tra cứu LẠI (dữ liệu "
                                     f"{loai_txt}{ht_txt} đã lưu từ lần trước — nếu có — vẫn được "
                                     f"GIỮ NGUYÊN, không bị xóa).")
                        else:
                            # XÁC NHẬN THẬT SỰ 0 hóa đơn (không phải lỗi) -> được phép
                            # xóa dữ liệu CŨ của đúng loại này (phản ánh đúng hiện trạng).
                            conn.execute(
                                "DELETE FROM invoices WHERE company_id=? AND loai=? AND he_thong=?",
                                (cid, loai, he_thong))
                            conn.commit()
                            msg(stage="found",
                                text=f"{loai_txt}{ht_txt}: 0 hóa đơn (không có dữ liệu trong kỳ)",
                                total_saved=total_saved)
                        continue

                    if not lay_ngan_hang:
                        truoc = len(invs)
                        invs = [iv for iv in invs if not la_hoa_don_ngan_hang(iv)]
                        bo = truoc - len(invs)
                        if bo > 0:
                            msg(stage="info",
                                text=f"Đã loại {bo} hóa đơn ngân hàng khỏi {loai_txt}{ht_txt}")

                    # CHỈ xóa dữ liệu CŨ của đúng loại này NGAY TRƯỚC KHI lưu kết quả
                    # MỚI (đã xác nhận tra cứu được) — loại khác lỗi thì không đụng tới.
                    conn.execute(
                        "DELETE FROM invoices WHERE company_id=? AND loai=? AND he_thong=?",
                        (cid, loai, he_thong))
                    for inv in invs:
                        try:
                            # Hóa đơn bán hàng (Mẫu số 2 — hộ/cá nhân kinh doanh
                            # KHÔNG chịu VAT) trang Thuế trả tgtcthue (tiền hàng
                            # CHƯA thuế) = 0/rỗng vì không tách thuế, CHỈ có tổng
                            # tiền thanh toán (tgtttbso). Trước đây lưu thẳng 0,
                            # khiến "Tổng tiền chưa VAT" và các báo cáo tổng hợp
                            # HIỆN 0 dù hóa đơn có giá trị thật — nhìn như phần
                            # mềm "không ghi nhận" hóa đơn dù đã tải được. Với
                            # hóa đơn không VAT (tgtthue=0), tiền CHƯA thuế PHẢI
                            # BẰNG tổng tiền thanh toán -> lấy tgtttbso làm dự
                            # phòng khi tgtcthue rỗng/0.
                            tgtcthue_v = _to_num(inv.get("tgtcthue"))
                            tgtttbso_v = _to_num(inv.get("tgtttbso"))
                            if not tgtcthue_v and tgtttbso_v:
                                tgtcthue_v = tgtttbso_v
                            conn.execute("""
                                INSERT OR IGNORE INTO invoices
                                (company_id, loai, he_thong, nbmst, nbten, nmmst,
                                 khmshdon, khhdon, shdon,
                                 tdlap, tgtcthue, tgtthue, tgtttbso, tthai, raw)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                            """, (
                                cid, loai, he_thong,
                                inv.get("nbmst"), inv.get("nbten"), inv.get("nmmst"),
                                str(inv.get("khmshdon", "")), inv.get("khhdon"), str(inv.get("shdon", "")),
                                inv.get("tdlap"),
                                tgtcthue_v, inv.get("tgtthue"), inv.get("tgtttbso"),
                                str(inv.get("tthai", "")),
                                json.dumps(inv, ensure_ascii=False),
                            ))
                        except Exception:
                            pass
                    conn.commit()
                    total_saved += len(invs)
                    # so sánh với tổng kỳ vọng từ trang Thuế (trước khi lọc ngân hàng)
                    exp = exp0
                    got_raw = getattr(client, "last_query_got", len(invs)) or len(invs)
                    thongke[loai]["exp"] += exp
                    thongke[loai]["got"] += got_raw
                    if exp and got_raw < exp:
                        msg(stage="warn",
                            text=f"⚠ {loai_txt}{ht_txt}: Trang Thuế báo có {exp} HĐ "
                                 f"nhưng chỉ lấy được {got_raw}. Có thể bị thiếu — "
                                 f"nên tra cứu LẠI kỳ này (hoặc chuyển chế độ Chậm & an toàn).")
                    elif exp:
                        msg(stage="info",
                            text=f"✓ {loai_txt}{ht_txt}: Đã lấy đủ {got_raw}/{exp} hóa đơn theo trang Thuế")
                    msg(stage="found",
                        text=f"Tìm thấy {len(invs)} hóa đơn {loai_txt}{ht_txt}",
                        total_saved=total_saved)

                    fmts = dl_map.get(loai, [])
                    if fmts and target_dir and invs:
                      try:
                        sub = os.path.join(target_dir, f"{loai_txt}".replace(" ", "_"))
                        os.makedirs(sub, exist_ok=True)
                        n = len(invs)

                        # INDEX FILE ĐÃ CÓ trong thư mục này (1 lần, tránh quét lại mỗi
                        # hóa đơn) — dùng để BỎ QUA tải hóa đơn đã có file đọc được trên
                        # máy, tránh tải đi tải lại (chỉ cần lấy hóa đơn mới phát sinh).
                        _file_index_dl = {}
                        for fn in os.listdir(sub):
                            low = fn.lower()
                            if not (low.endswith(".xml") or low.endswith(".zip")):
                                continue
                            name = fn.rsplit(".", 1)[0]
                            parts = name.split("_")
                            if len(parts) >= 2:
                                key2 = (parts[0], parts[1].lstrip("0") or "0")
                                prev = _file_index_dl.get(key2)
                                # ƯU TIÊN .xml (đọc trực tiếp, khỏi giải nén)
                                if prev is None or (prev.lower().endswith(".zip") and low.endswith(".xml")):
                                    _file_index_dl[key2] = os.path.join(sub, fn)

                        def _tai_1_file(inv):
                            nbmst = inv.get("nbmst", "")
                            khhdon = inv.get("khhdon", "")
                            khmshdon = inv.get("khmshdon", "")
                            shdon = inv.get("shdon", "")
                            base = f"{khhdon}_{shdon}_{nbmst}"
                            if "xml" not in fmts:
                                return "ok"   # không yêu cầu tải xml -> coi như "xong"

                            # ĐÃ CÓ FILE + ĐỌC ĐƯỢC + TRẠNG THÁI (tthai) KHÔNG ĐỔI so với
                            # lần tra cứu trước -> BỎ QUA, không tải lại trên trang Thuế.
                            # Nếu tthai đã đổi (hóa đơn bị THAY THẾ/ĐIỀU CHỈNH kể từ lần
                            # trước) thì VẪN PHẢI tải lại để lấy đúng bản mới nhất.
                            key = (str(khmshdon or ""), str(khhdon or ""),
                                   str(shdon or ""), loai, he_thong)
                            tthai_cu = old_status.get(key)
                            tthai_moi = str(inv.get("tthai") or "")
                            fkey = (str(khhdon or ""), str(shdon or "").lstrip("0") or "0")
                            fpath = _file_index_dl.get(fkey)
                            if (fpath and (tthai_cu is None or tthai_cu == tthai_moi)
                                    and _doc_kiem_tra_file_hoadon(fpath)):
                                return "skip"

                            try:
                                zdata = client.download_xml(nbmst, khhdon, khmshdon,
                                                            shdon, loai, he_thong)
                            except Exception:
                                zdata = None
                            if zdata:
                                try:
                                    _save_invoice_files(sub, base, zdata)
                                    return "ok"
                                except Exception:
                                    return "fail"
                            return "fail"

                        # TẢI SONG SONG (thay vì từng file 1) để rút ngắn thời gian chờ —
                        # mỗi file là 1 request mạng riêng tới trang Thuế, tải tuần tự
                        # từng cái 1 với hàng chục/hàng trăm hóa đơn khiến tổng thời gian
                        # cộng dồn rất lâu dù mỗi file không lỗi gì. Giới hạn số luồng
                        # đồng thời (không tải ồ ạt) để tránh bị 429 chặn tốc độ.
                        SO_LUONG_SONG_SONG = SP().get("song_song", 4)

                        def _tai_nhieu_file(ds_inv, nhan):
                            """Tải song song danh sách hóa đơn.
                            Trả về (so_thanh_cong, ds_loi, so_bo_qua_da_co_san)."""
                            if getattr(client, "_token_dead", False):
                                return 0, list(ds_inv), 0  # phiên đã hết hạn -> khỏi thử
                            loi = []
                            so_ok = 0
                            so_bo_qua = 0
                            so_xong = 0
                            n_ds = len(ds_inv)
                            with _cf.ThreadPoolExecutor(max_workers=SO_LUONG_SONG_SONG) as ex:
                                futs = {ex.submit(_tai_1_file, inv): inv for inv in ds_inv}
                                for fut in _cf.as_completed(futs):
                                    inv = futs[fut]
                                    try:
                                        kq = fut.result()
                                    except Exception:
                                        kq = "fail"
                                    so_xong += 1
                                    msg(stage="download",
                                        text=f"{nhan} {loai_txt}: {so_xong}/{n_ds} (còn {n_ds - so_xong})",
                                        cur=so_xong, total=n_ds)
                                    if kq == "skip":
                                        so_ok += 1; so_bo_qua += 1
                                    elif kq == "ok":
                                        so_ok += 1
                                    else:
                                        loi.append(inv)
                            return so_ok, loi, so_bo_qua

                        ok_1, loi_file, bo_qua = _tai_nhieu_file(invs, "Đang tải file")
                        file_saved += ok_1

                        # THỬ LẠI 1 LƯỢT các file bị lỗi (thường do bị chặn tốc độ giữa
                        # chừng) — trước đây KHÔNG hề thử lại nên 1 lần vấp là mất file
                        # vĩnh viễn dù dữ liệu hóa đơn (bảng) vẫn tải đủ.
                        if loi_file and not getattr(client, "_token_dead", False):
                            msg(stage="warn",
                                text=f"⚠ {loai_txt}{ht_txt}: {len(loi_file)} file tải chưa được, "
                                     f"đang thử lại...")
                            time.sleep(5)
                            ok_2, loi_file, bo_qua_2 = _tai_nhieu_file(loi_file, "Thử lại file")
                            file_saved += ok_2
                            bo_qua += bo_qua_2

                        if bo_qua:
                            msg(stage="info",
                                text=f"↷ {loai_txt}{ht_txt}: {bo_qua} hóa đơn đã có sẵn file trên máy "
                                     f"(đọc được, trạng thái không đổi) — bỏ qua, không tải lại")

                        if loi_file:
                            file_thieu_tong += len(loi_file)
                            msg(stage="warn",
                                text=f"⚠ {loai_txt}{ht_txt}: KHÔNG tải được {len(loi_file)}/{n} file "
                                     f"(dữ liệu bảng vẫn lưu đủ) — chạy lại tra cứu kỳ này (chế độ "
                                     f"'Chậm & an toàn') để tải nốt, hoặc kết xuất Excel sẽ tự lấy "
                                     f"chi tiết các hóa đơn này qua mạng khi cần.")
                        elif "xml" in fmts:
                            msg(stage="info",
                                text=f"✓ {loai_txt}{ht_txt}: đã tải đủ {n}/{n} file")
                      except Exception as e:
                        msg(stage="warn",
                            text=f"Lỗi tải file {loai_txt} (dữ liệu bảng vẫn lưu): {str(e)[:100]}")

            # ===== TỔNG KẾT số hóa đơn theo trang Thuế (để biết lấy đủ chưa) =====
            tk_mua = thongke["purchase"]
            tk_ban = thongke["sold"]
            dong_tk = []
            if "purchase" in loai_list:
                if loai_that_bai["purchase"]:
                    dong_tk.append(f"✗ Đầu vào (mua): LỖI khi tra cứu — CHƯA CHẮC ĐÃ ĐỦ "
                                   f"(mới lấy được {tk_mua['got']} HĐ) — nên tra cứu LẠI")
                elif tk_mua["exp"]:
                    dau = "✓" if tk_mua["got"] >= tk_mua["exp"] else "⚠"
                    dong_tk.append(f"{dau} Đầu vào (mua): lấy {tk_mua['got']}/{tk_mua['exp']} HĐ trang Thuế báo")
                else:
                    dong_tk.append(f"• Đầu vào (mua): lấy {tk_mua['got']} HĐ")
            if "sold" in loai_list:
                if loai_that_bai["sold"]:
                    dong_tk.append(f"✗ Đầu ra (bán): LỖI khi tra cứu — CHƯA CHẮC ĐÃ ĐỦ "
                                   f"(mới lấy được {tk_ban['got']} HĐ) — nên tra cứu LẠI")
                elif tk_ban["exp"]:
                    dau = "✓" if tk_ban["got"] >= tk_ban["exp"] else "⚠"
                    dong_tk.append(f"{dau} Đầu ra (bán): lấy {tk_ban['got']}/{tk_ban['exp']} HĐ trang Thuế báo")
                else:
                    dong_tk.append(f"• Đầu ra (bán): lấy {tk_ban['got']} HĐ")
            for d in dong_tk:
                msg(stage="info", text=d)

            # cảnh báo nếu thiếu (LỖI THẬT SỰ luôn tính là thiếu, kể cả khi 'got'
            # trùng khớp exp một cách tình cờ — vì exp lúc lỗi có thể không đáng tin)
            co_loi = loai_that_bai["purchase"] or loai_that_bai["sold"]
            thieu = (co_loi or
                     (tk_mua["exp"] and tk_mua["got"] < tk_mua["exp"]) or
                     (tk_ban["exp"] and tk_ban["got"] < tk_ban["exp"]))
            if co_loi:
                # KHÔNG BAO GIỜ báo "Hoàn tất" khi có lỗi — phải là trạng thái LỖI rõ ràng
                loi_ben = []
                if loai_that_bai["purchase"]:
                    loi_ben.append("MUA VÀO")
                if loai_that_bai["sold"]:
                    loi_ben.append("BÁN RA")
                done_text = (f"❌ LỖI khi tra cứu {', '.join(loi_ben)} — DỮ LIỆU CHƯA ĐẦY ĐỦ "
                            f"(hiện có {total_saved} hóa đơn: đầu vào {tk_mua['got']}, "
                            f"đầu ra {tk_ban['got']} — riêng loại LỖI ở trên vẫn đang GIỮ dữ "
                            f"liệu của lần tra cứu THÀNH CÔNG gần nhất, KHÔNG bị mất). "
                            f"BẮT BUỘC tra cứu LẠI công ty này (nên chuyển chế độ "
                            f"'Chậm & an toàn' nếu vẫn lỗi).")
            else:
                done_text = f"Hoàn tất! Đã lưu {total_saved} hóa đơn"
                if "purchase" in loai_list or "sold" in loai_list:
                    done_text += f" (đầu vào: {tk_mua['got']}, đầu ra: {tk_ban['got']})"
                if file_saved:
                    done_text += f", tải {file_saved} file vào: {target_dir}"
                if file_thieu_tong:
                    done_text += f" — ⚠ CÒN {file_thieu_tong} FILE CHƯA TẢI ĐƯỢC (dữ liệu bảng vẫn đủ)"
                if thieu:
                    done_text += " — ⚠ CÓ THỂ THIẾU, nên tra cứu lại (chế độ Chậm & an toàn)"
            msg(stage=("error" if co_loi else "done"), text=done_text,
                total_saved=total_saved, file_saved=file_saved,
                file_thieu=file_thieu_tong,
                tk_mua_got=tk_mua["got"], tk_mua_exp=tk_mua["exp"],
                tk_ban_got=tk_ban["got"], tk_ban_exp=tk_ban["exp"],
                loi_tra_cuu=co_loi)
        except Exception as e:
            msg(stage="error", text=f"Lỗi: {str(e)[:160]}")
        finally:
            conn.close()
            job = FETCH_JOBS.get(cid)
            if job is not None:
                job["running"] = False

    # FETCH_JOBS[cid] đã được khởi tạo bởi nơi gọi (endpoint /api/fetch hoặc batch).
    # Chạy đồng bộ trong thread của nơi gọi.
    run()


def _new_fetch_job():
    return {"messages": [], "last": None, "running": True,
            "cursor": 0, "started": time.time(), "cancel": False}


def _luu_loi_tra_cuu(cid: int, text: str):
    """Lưu LỖI tra cứu gần nhất vào công ty (DB) — để không bị mất khi
    thông báo/toast trên trình duyệt tự tắt. Mở lại công ty vẫn thấy đúng
    lỗi lần tra cứu gần nhất, cho tới khi tra cứu lại THÀNH CÔNG."""
    try:
        cn = db()
        cn.execute(
            "UPDATE companies SET last_fetch_error=?, last_fetch_error_at=? WHERE id=?",
            (text or "", datetime.datetime.now().isoformat(), cid))
        cn.commit()
        cn.close()
    except Exception:
        pass


def _xoa_loi_tra_cuu(cid: int):
    try:
        cn = db()
        cn.execute(
            "UPDATE companies SET last_fetch_error=NULL, last_fetch_error_at=NULL WHERE id=?",
            (cid,))
        cn.commit()
        cn.close()
    except Exception:
        pass


@app.post("/api/fetch/{cid}")
def fetch_invoices(cid: int, body: dict = Body(...)):
    """Tra cứu + tải hóa đơn cho 1 công ty (chạy nền). Trình duyệt poll
    /api/fetch-status/{cid} để lấy tiến độ."""
    client = get_client(cid)
    if not client.token:
        _luu_loi_tra_cuu(cid, "Chưa đăng nhập tài khoản thuế cho công ty này")
        raise HTTPException(401, "Chưa đăng nhập tài khoản thuế cho công ty này")
    _xoa_loi_tra_cuu(cid)   # bắt đầu lượt mới -> xóa ghi chú lỗi lần trước
    FETCH_JOBS[cid] = _new_fetch_job()
    import threading
    threading.Thread(target=lambda: _run_fetch_job(cid, body), daemon=True).start()
    return {"ok": True, "started": True}


# ---------- TRA CỨU HÀNG LOẠT (nhiều công ty cùng lúc) ----------
# Mỗi batch lấy lần lượt từng công ty (tuần tự) để tránh bị Tổng cục Thuế
# chặn tạm (429). Tiến độ từng công ty vẫn ghi vào FETCH_JOBS[cid] như cũ,
# đồng thời BATCH_JOBS[batch_id] tổng hợp trạng thái toàn batch.
BATCH_JOBS = {}      # {batch_id: {...}}
_BATCH_SEQ = {"n": 0}


def _run_batch(batch_id: int, cids: list, body: dict):
    batch = BATCH_JOBS[batch_id]
    # Trình duyệt ẩn dùng CHUNG cho cả batch để vẽ captcha chính xác khi tự
    # đăng nhập (mở LƯỜI — chỉ mở khi thật sự gặp công ty chưa đăng nhập đầu
    # tiên, và dùng lại cho các công ty/lượt thử sau, đỡ tốn thời gian mở lại
    # trình duyệt nhiều lần).
    drv_captcha = [None, False]  # [driver_hoac_None, da_thu_mo_chua]

    def _drv():
        if not drv_captcha[1]:
            drv_captcha[1] = True
            drv_captcha[0] = _mo_trinh_duyet_captcha()
        return drv_captcha[0]

    try:
        for cid in cids:
            if batch.get("cancel"):
                break
            item = batch["items"].get(cid)
            # Đặt "current" + khởi tạo job NGAY TỪ ĐẦU (kể cả trước khi đăng
            # nhập) — trước đây "current" chỉ được gán ngay trước lúc tra cứu,
            # nên suốt bước tự đăng nhập (có thể mất vài chục giây) người dùng
            # không thấy phần mềm đang xử lý công ty nào / tới đâu.
            batch["current"] = cid
            FETCH_JOBS[cid] = _new_fetch_job()

            def _bao_tien_do(t, _cid=cid, _item=item):
                _item["note"] = t
                job = FETCH_JOBS.get(_cid)
                if job is not None:
                    kw = {"stage": "login", "text": t}
                    job["messages"].append(kw)
                    job["last"] = kw

            client = get_client(cid)
            if not client.token:
                # TỰ ĐỘNG ĐĂNG NHẬP (bằng tài khoản/mật khẩu đã lưu) thay vì bỏ
                # qua — trước đây công ty chưa đăng nhập bị SKIP hẳn, người
                # dùng phải tự đăng nhập TỪNG CÔNG TY trước khi tra cứu hàng
                # loạt. Công ty nào chưa lưu mật khẩu / tự đăng nhập thất bại
                # (vd sai mật khẩu, OCR không giải được captcha) thì mới bỏ qua.
                item["status"] = "login"
                _bao_tien_do("Đang tự động đăng nhập...")
                # Dùng trình duyệt ẩn (nếu mở được) để vẽ captcha CHÍNH XÁC như
                # lúc người dùng tự bấm nút trên giao diện, thay vì svglib gần
                # đúng — tăng hẳn tỉ lệ thành công. so_lan=8 vẫn cao hơn mức 5-6
                # lần người dùng thường cần, để dự phòng khi thi thoảng vẫn đọc sai.
                # progress=_bao_tien_do -> báo rõ đang thử lần mấy, không "im
                # lặng" trong lúc chờ (có thể mất vài chục giây).
                ok, thong_bao, _, _ = _tu_dong_dang_nhap(
                    cid, so_lan=8, drv=_drv(), progress=_bao_tien_do)
                if not ok:
                    item["status"] = "skipped"
                    item["note"] = f"Tự đăng nhập thất bại: {thong_bao}"
                    _luu_loi_tra_cuu(cid, item["note"])
                    batch["done"] += 1
                    continue
            item["status"] = "running"
            try:
                _run_fetch_job(cid, body)   # chạy đồng bộ trong thread batch
                last = (FETCH_JOBS.get(cid) or {}).get("last") or {}
                item["total_saved"] = last.get("total_saved", 0)
                if last.get("stage") == "error":
                    item["status"] = "error"
                    item["note"] = last.get("text", "Lỗi")
                else:
                    item["status"] = "done"
                    item["note"] = last.get("text", "Hoàn tất")
            except Exception as e:
                item["status"] = "error"
                item["note"] = str(e)[:160]
                _luu_loi_tra_cuu(cid, item["note"])
            batch["done"] += 1
    finally:
        batch["running"] = False
        batch["current"] = None
        _dong_trinh_duyet_captcha(drv_captcha[0])


@app.post("/api/fetch-batch")
def fetch_batch(body: dict = Body(...)):
    """Tra cứu nhiều công ty một lần.
    body: { cids: [1,2,3], tu_ngay, den_ngay, loai_list, he_thong_list,
            dl_buy, dl_sell, lay_ngan_hang }
    Trả về { batch_id }. Poll /api/fetch-batch-status/{batch_id} để theo dõi."""
    cids = body.get("cids") or []
    cids = [int(c) for c in cids]
    if not cids:
        raise HTTPException(400, "Chưa chọn công ty nào")

    conn = db()
    rows = conn.execute(
        "SELECT id, ten, mst FROM companies WHERE id IN (%s)"
        % ",".join("?" * len(cids)), cids).fetchall()
    conn.close()
    ten_map = {r["id"]: (r["ten"] or r["mst"]) for r in rows}
    # giữ đúng thứ tự người dùng chọn, chỉ lấy công ty có thật
    cids = [c for c in cids if c in ten_map]
    if not cids:
        raise HTTPException(404, "Không tìm thấy công ty đã chọn")

    _BATCH_SEQ["n"] += 1
    batch_id = _BATCH_SEQ["n"]
    BATCH_JOBS[batch_id] = {
        "running": True,
        "cancel": False,
        "total": len(cids),
        "done": 0,
        "current": None,
        "started": time.time(),
        "order": cids,
        "items": {c: {"cid": c, "ten": ten_map[c], "status": "pending",
                      "total_saved": 0, "note": ""} for c in cids},
    }
    import threading
    threading.Thread(target=lambda: _run_batch(batch_id, cids, body),
                     daemon=True).start()
    return {"ok": True, "batch_id": batch_id}


@app.get("/api/fetch-batch-status/{batch_id}")
def fetch_batch_status(batch_id: int):
    batch = BATCH_JOBS.get(batch_id)
    if not batch:
        return {"running": False, "no_job": True}
    # Chi tiết TRỰC TIẾP công ty đang xử lý (đăng nhập/tra cứu/tải file...) —
    # để người dùng thấy phần mềm đang chạy TỚI ĐÂU, không chỉ biết "đang xử
    # lý công ty X" mà không rõ đang ở bước nào bên trong.
    current_detail = ""
    cur_cid = batch["current"]
    if cur_cid is not None:
        job = FETCH_JOBS.get(cur_cid)
        if job:
            current_detail = (job.get("last") or {}).get("text", "")
    return {
        "running": batch["running"],
        "total": batch["total"],
        "done": batch["done"],
        "current": batch["current"],
        "current_detail": current_detail,
        "items": [batch["items"][c] for c in batch["order"]],
    }


@app.post("/api/fetch-batch-cancel/{batch_id}")
def fetch_batch_cancel(batch_id: int):
    batch = BATCH_JOBS.get(batch_id)
    if batch:
        batch["cancel"] = True
        # dừng luôn công ty đang chạy
        cur = batch.get("current")
        if cur and FETCH_JOBS.get(cur):
            FETCH_JOBS[cur]["cancel"] = True
    return {"ok": True}


@app.post("/api/fetch-cancel/{cid}")
def fetch_cancel(cid: int):
    """Dừng tra cứu của công ty (khi người dùng chuyển sang công ty khác)."""
    job = FETCH_JOBS.get(cid)
    if job:
        job["cancel"] = True
    return {"ok": True}


@app.get("/api/fetch-status/{cid}")
def fetch_status(cid: int, cursor: int = 0):
    """Trả về các thông báo tiến độ MỚI kể từ cursor. Trình duyệt poll endpoint này.
    Chạy nền nên thu nhỏ cửa sổ vẫn không mất dữ liệu."""
    job = FETCH_JOBS.get(cid)
    if not job:
        return {"running": False, "messages": [], "cursor": 0, "no_job": True}
    msgs = job["messages"][cursor:]
    return {
        "running": job["running"],
        "messages": msgs,
        "cursor": cursor + len(msgs),
    }


@app.get("/api/invoices/{cid}")
def get_invoices(cid: int, loai: Optional[str] = None):
    conn = db()
    if loai:
        rows = conn.execute(
            "SELECT * FROM invoices WHERE company_id=? AND loai=? ORDER BY tdlap DESC",
            (cid, loai)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM invoices WHERE company_id=? ORDER BY tdlap DESC",
            (cid,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        d = dict(r)
        try:
            raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            raw = {}
        # mô tả trạng thái (Hóa đơn mới / đã bị thay thế / hủy...)
        tthai_ma = str(raw.get("tthai", r["tthai"]) or "").strip()
        d["tthai_mota"] = _mo_ta_trang_thai(tthai_ma)
        d["tthai_ma"] = tthai_ma
        # tên người mua (raw từ trang Thuế có 'nmten'); người bán đã có ở nbten
        d["nmten"] = raw.get("nmten", "") or raw.get("nmtnmua", "") or ""
        # mặt hàng dòng đầu tiên (nếu có detail_json)
        mat_hang = ""
        try:
            dj = r["detail_json"]
            if dj:
                det = json.loads(dj)
                items = _parse_detail_json(det)
                if items:
                    mat_hang = items[0].get("ten_hang", "")
        except Exception:
            pass
        d["mat_hang_dau"] = mat_hang
        out.append(d)
    return out


# ---------- TẢI FILE XML / PDF (đóng gói ZIP) ----------
@app.post("/api/download/{cid}")
def download_files(cid: int, body: dict = Body(...)):
    """
    body: { loai, formats: ["xml","pdf"] }
    Trả về file ZIP chứa toàn bộ XML/PDF của các hóa đơn đã lưu.
    """
    client = get_client(cid)
    if not client.token:
        raise HTTPException(401, "Chưa đăng nhập tài khoản thuế")

    loai = body.get("loai", "purchase")
    formats = body.get("formats", ["xml", "pdf"])

    conn = db()
    rows = conn.execute(
        "SELECT * FROM invoices WHERE company_id=? AND loai=?",
        (cid, loai)).fetchall()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()

    mem = io.BytesIO()
    ok_xml = 0
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in rows:
            base = f"{r['khhdon']}_{r['shdon']}_{r['nbmst']}"
            he_thong = r["he_thong"] or "query"
            if "xml" in formats:
                zdata = client.download_xml(
                    r["nbmst"], r["khhdon"], r["khmshdon"], r["shdon"],
                    loai, he_thong)
                if zdata:
                    # zdata là file invoice.zip của TCT -> giải nén lấy html+xml
                    try:
                        inner = zipfile.ZipFile(io.BytesIO(zdata))
                        for nm in inner.namelist():
                            zf.writestr(f"{base}/{os.path.basename(nm)}", inner.read(nm))
                    except Exception:
                        # không phải zip -> lưu thẳng
                        zf.writestr(f"{base}/{base}.xml", zdata)
                    ok_xml += 1
                time.sleep(SP()["file"])
            if "pdf" in formats:
                pass  # cần endpoint/cURL thật của nút tải PDF để bổ sung
    mem.seek(0)

    if ok_xml == 0 and "xml" in formats:
        raise HTTPException(
            502,
            "Không tải được file nào. Token có thể đã hết hạn — đăng nhập lại rồi thử.")

    fname = f"HoaDon_{comp['mst']}_{loai}.zip"
    return StreamingResponse(
        mem, media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={fname}"})


# ---------- XUẤT EXCEL ----------
# Map mã trạng thái xử lý (ttxly) -> mô tả tiếng Việt
TTXLY_DESC = {
    "4": "Hóa đơn không đủ điều kiện cấp mã",
    "5": "Đã cấp mã hóa đơn",
    "6": "Tổng cục thuế đã nhận không mã",
    "8": "TCT đã nhận HĐ có mã khởi tạo từ máy tính tiền",
}
# Map mã trạng thái hóa đơn (tthai)
TTHAI_DESC = {
    "1": "Hóa đơn mới",
    "2": "Hóa đơn thay thế",
    "3": "Hóa đơn điều chỉnh",
    "4": "Hóa đơn đã bị thay thế",
    "5": "Hóa đơn đã bị điều chỉnh",
    "6": "Hóa đơn hủy",
}

def _mo_ta_trang_thai(tthai):
    s = str(tthai or "").strip()
    return TTHAI_DESC.get(s, s or "")

def _mo_ta_ket_qua(ttxly):
    s = str(ttxly or "").strip()
    return TTXLY_DESC.get(s, s or "")


@app.post("/api/import-companies")
async def import_companies(request: Request):
    """
    Import nhiều công ty từ file Excel. Cột nhận diện theo tiêu đề (dòng 1):
    Tên công ty | MST | Tên đăng nhập | Mật khẩu | Thư mục lưu | Người ký
    (chỉ Tên và MST là bắt buộc). Bỏ qua dòng trùng MST.
    """
    import openpyxl, io as _io
    form = await request.form()
    up = form.get("file")
    if up is None:
        raise HTTPException(400, "Chưa chọn file Excel")
    content = await up.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file Excel: {e}")
    ws = wb.active

    def find_col(*names):
        for c in range(1, ws.max_column + 1):
            v = str(ws.cell(1, c).value or "").strip().lower()
            for n in names:
                if n in v:
                    return c
        return None

    c_ten = find_col("tên công ty", "ten cong ty", "tên")
    c_mst = find_col("mst", "mã số thuế", "ma so thue")
    c_user = find_col("đăng nhập", "dang nhap", "username", "user")
    c_pass = find_col("mật khẩu", "mat khau", "password", "pass")
    c_dir = find_col("thư mục", "thu muc", "save_dir", "lưu")
    c_nk = find_col("người ký", "nguoi ky")
    if not c_ten or not c_mst:
        raise HTTPException(400, "File phải có cột 'Tên công ty' và 'MST'")

    conn = db()
    added = skipped = 0
    errors = []
    for r in range(2, ws.max_row + 1):
        ten = str(ws.cell(r, c_ten).value or "").strip()
        mst = str(ws.cell(r, c_mst).value or "").strip()
        if not ten or not mst:
            continue
        # MST có thể bị Excel đọc thành số -> bỏ .0
        if mst.endswith(".0"):
            mst = mst[:-2]
        dup = conn.execute("SELECT id FROM companies WHERE mst=?", (mst,)).fetchone()
        if dup:
            skipped += 1
            continue
        user = str(ws.cell(r, c_user).value or "").strip() if c_user else ""
        pw = str(ws.cell(r, c_pass).value or "").strip() if c_pass else ""
        sdir = str(ws.cell(r, c_dir).value or "").strip() if c_dir else ""
        nk = str(ws.cell(r, c_nk).value or "").strip() if c_nk else ""
        try:
            conn.execute(
                "INSERT INTO companies (ten, mst, username, password, ghichu, save_dir, nguoi_ky, created_at) "
                "VALUES (?,?,?,?,?,?,?,?)",
                (ten, mst, user, pw, "", sdir, nk, datetime.datetime.now().isoformat()))
            added += 1
        except Exception as e:
            errors.append(f"{ten}: {e}")
    conn.commit()
    conn.close()
    return {"ok": True, "added": added, "skipped": skipped, "errors": errors[:5]}


@app.get("/api/export-companies")
def export_companies():
    """Xuất danh sách công ty hiện có ra file Excel (kèm thông tin đăng nhập)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    conn = db()
    rows = conn.execute("SELECT * FROM companies ORDER BY ten").fetchall()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách công ty"
    headers = ["Tên công ty", "MST", "Tên đăng nhập", "Mật khẩu",
               "Thư mục lưu", "Người ký"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="1F6B4A")
    for r in rows:
        d = dict(r)
        ws.append([d.get("ten", ""), d.get("mst", ""), d.get("username", ""),
                   d.get("password", ""), d.get("save_dir", ""), d.get("nguoi_ky", "")])
    for col, w in zip("ABCDEF", [32, 16, 16, 16, 24, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    fname = "DanhSach_CongTy.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    # lưu ra Desktop cho dễ tìm
    desktop = _get_desktop_dir()
    open_path = path
    if desktop and os.path.isdir(desktop):
        try:
            import shutil
            shutil.copy(path, os.path.join(desktop, fname))
            open_path = os.path.join(desktop, fname)
        except Exception:
            pass
    _open_file_local(open_path)
    return _resp_xuat(path, fname)


@app.get("/api/companies-template")
def companies_template():
    """Tải file Excel mẫu để nhập danh sách công ty."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh sách công ty"
    headers = ["Tên công ty", "MST", "Tên đăng nhập", "Mật khẩu",
               "Thư mục lưu", "Người ký"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="1F6B4A")
    ws.append(["CÔNG TY TNHH ABC", "0301234567", "0301234567", "matkhau123",
               "D:\\HoaDon\\ABC", "Nguyễn Văn A"])
    ws.append(["CÔNG TY TNHH XYZ", "0307654321", "", "", "", ""])
    for col, w in zip("ABCDEF", [32, 16, 16, 16, 22, 18]):
        ws.column_dimensions[col].width = w
    path = os.path.join(DOWNLOAD_DIR, "Mau_DanhSach_CongTy.xlsx")
    wb.save(path)
    return _resp_xuat(path, "Mau_DanhSach_CongTy.xlsx")


def _parse_tokhai_nhap(wb):
    """Đọc 1 file tờ khai hải quan nhập khẩu (.xlsx), trả về dict:
    {so_tk, ngay_dk, nguoi_xk, items:[{ten, sluong, dvt, tri_gia_gtgt, ts_gtgt,
     tien_thue_gtgt, tri_gia_nk, ts_nk, tien_thue_nk}, ...]}
    Cấu trúc: sheet TKN, mỗi trang bắt đầu '<IMP>', mỗi dòng hàng '<NN>'.
    Giá trị lấy theo NHÃN ở cột nhỏ + giá trị ở cột bên phải."""
    if "TKN" in wb.sheetnames:
        ws = wb["TKN"]
    else:
        ws = wb[wb.sheetnames[0]]

    def cell(r, c):
        v = ws.cell(r, c).value
        return str(v).strip() if v is not None else ""

    def to_num(s):
        # "18.320.976,2154" -> 18320976.2154 ; "8%" -> 8
        s = str(s).replace("%", "").strip()
        if not s:
            return 0
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except Exception:
            return 0

    def row_label(r):
        # nhãn của dòng = ô text đầu tiên ở cột C(3) hoặc D(4)
        return cell(r, 3) or cell(r, 4)

    # tìm số tờ khai, ngày đăng ký, người xuất khẩu (trang đầu)
    so_tk = ngay_dk = nguoi_xk = ""
    maxr = ws.max_row
    for r in range(1, min(maxr, 60) + 1):
        lab = cell(r, 3)
        if lab == "Số tờ khai" and not so_tk:
            so_tk = cell(r, 5) or cell(r, 6) or cell(r, 7)
        if lab == "Ngày đăng ký" and not ngay_dk:
            ngay_dk = cell(r, 7) or cell(r, 6) or cell(r, 8)
        if lab == "Người xuất khẩu" and not nguoi_xk:
            # tên thường ở dòng "Tên" 1-2 dòng sau, cột H
            for rr in range(r, min(r + 6, maxr)):
                if cell(rr, 4) == "Tên":
                    nguoi_xk = cell(rr, 8)
                    break

    # chuẩn hóa ngày: "16/01/2026 10:55:16" -> 2026-01-16
    nd = ""
    if ngay_dk:
        d0 = ngay_dk.split()[0]
        if "/" in d0:
            p = d0.split("/")
            if len(p) == 3:
                nd = f"{p[2]}-{int(p[1]):02d}-{int(p[0]):02d}"

    # duyệt các dòng hàng: mỗi khối bắt đầu ô C = '<01>'..'<NN>'
    import re as _re
    items = []
    r = 1
    while r <= maxr:
        cval = cell(r, 3)
        if _re.fullmatch(r"<\d+>", cval):
            blk_start = r
            # tìm hết khối (tới '<NN>' tiếp theo hoặc '<IMP>' hoặc hết)
            r2 = r + 1
            while r2 <= maxr:
                c2 = cell(r2, 3)
                if _re.fullmatch(r"<\d+>", c2) or c2 == "<IMP>":
                    break
                r2 += 1
            blk_end = r2
            it = _parse_hang_block(ws, blk_start, blk_end, cell, to_num)
            if it and (it.get("ten") or it.get("tri_gia_gtgt")):
                items.append(it)
            r = blk_end
        else:
            r += 1

    return {"so_tk": so_tk, "ngay_dk": nd, "nguoi_xk": nguoi_xk, "items": items}


def _parse_hang_block(ws, r0, r1, cell, to_num):
    """Trích 1 dòng hàng từ khối [r0, r1)."""
    it = {"ten": "", "sluong": 0, "dvt": "", "tri_gia_gtgt": 0, "ts_gtgt": "",
          "tien_thue_gtgt": 0, "tri_gia_nk": 0, "ts_nk": "", "tien_thue_nk": 0}
    in_thue_khac = False
    gtgt_found = False
    for r in range(r0, r1):
        lab_c = cell(r, 3)
        lab_d = cell(r, 4)
        # Mô tả hàng hóa -> cột G(7)
        if lab_c == "Mô tả hàng hóa":
            it["ten"] = cell(r, 7) or cell(r, 8)
        # Số lượng (1): nhãn ở cột S(19) -> giá trị V(22), đơn vị AE(31)
        if cell(r, 19) == "Số lượng (1)":
            it["sluong"] = to_num(cell(r, 22))
            it["dvt"] = cell(r, 31)  # AE = đơn vị tính (BAG/PCE...)
        # Thuế nhập khẩu
        if lab_c == "Thuế nhập khẩu":
            in_thue_khac = False
        # Trị giá tính thuế(S) (thuế NK) -> I(9)
        if lab_d == "Trị giá tính thuế(S)":
            it["tri_gia_nk"] = to_num(cell(r, 9))
        # Thuế suất NK -> I(9), chỉ lấy khi CHƯA vào phần thuế khác
        if lab_d == "Thuế suất" and not in_thue_khac and not it["ts_nk"]:
            it["ts_nk"] = cell(r, 9)
        # Số tiền thuế NK -> I(9)
        if lab_d == "Số tiền thuế" and not in_thue_khac and not it["tien_thue_nk"]:
            it["tien_thue_nk"] = to_num(cell(r, 9))
        # Phần "Thuế và thu khác" (chứa GTGT)
        if lab_c == "Thuế và thu khác":
            in_thue_khac = True
        # mục GTGT: Tên = "Thuế GTGT" (cột H=8)
        if in_thue_khac and cell(r, 8) == "Thuế GTGT":
            gtgt_found = True
        if gtgt_found and not it["tri_gia_gtgt"] and lab_d == "Trị giá tính thuế":
            it["tri_gia_gtgt"] = to_num(cell(r, 9))
        if gtgt_found and not it["ts_gtgt"] and lab_d == "Thuế suất":
            it["ts_gtgt"] = cell(r, 9)
        if gtgt_found and not it["tien_thue_gtgt"] and lab_d == "Số tiền thuế":
            it["tien_thue_gtgt"] = to_num(cell(r, 9))
    return it


@app.post("/api/import-tokhai-nhap/{cid}")
async def import_tokhai_nhap(cid: int, request: Request):
    """Import 1 hoặc nhiều file tờ khai hải quan nhập khẩu (.xlsx).
    Mỗi dòng hàng -> 1 dòng trong Chi tiết MUA VÀO."""
    import openpyxl, io as _io
    form = await request.form()
    files = form.getlist("files") or ([form.get("file")] if form.get("file") else [])
    if not files:
        raise HTTPException(400, "Chưa chọn file tờ khai")
    conn = db()
    added = 0
    tong_tk = 0
    loi = []
    for up in files:
        if up is None:
            continue
        try:
            content = await up.read()
            wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
            tk = _parse_tokhai_nhap(wb)
        except Exception as e:
            loi.append(f"{getattr(up,'filename','file')}: {e}")
            continue
        if not tk["so_tk"] or not tk["items"]:
            loi.append(f"{getattr(up,'filename','file')}: không đọc được số tờ khai hoặc dòng hàng")
            continue
        conn.execute("""
            INSERT INTO tokhai_nhap (company_id, so_tk, ngay_dk, nguoi_xk, items_json, updated_at)
            VALUES (?,?,?,?,?,?)
            ON CONFLICT(company_id, so_tk) DO UPDATE SET
                ngay_dk=excluded.ngay_dk, nguoi_xk=excluded.nguoi_xk,
                items_json=excluded.items_json, updated_at=excluded.updated_at
        """, (cid, tk["so_tk"], tk["ngay_dk"], tk["nguoi_xk"],
              json.dumps(tk["items"], ensure_ascii=False),
              datetime.datetime.now().isoformat()))
        added += len(tk["items"])
        tong_tk += 1
    conn.commit()
    conn.close()
    return {"ok": True, "so_to_khai": tong_tk, "so_dong_hang": added, "loi": loi[:5]}


@app.get("/api/tokhai-nhap/{cid}")
def list_tokhai_nhap(cid: int):
    conn = db()
    rows = conn.execute(
        "SELECT so_tk, ngay_dk, nguoi_xk, items_json FROM tokhai_nhap WHERE company_id=? ORDER BY ngay_dk",
        (cid,)).fetchall()
    conn.close()
    out = []
    for r in rows:
        try:
            items = json.loads(r["items_json"]) if r["items_json"] else []
        except Exception:
            items = []
        out.append({"so_tk": r["so_tk"], "ngay_dk": r["ngay_dk"],
                    "nguoi_xk": r["nguoi_xk"], "so_dong": len(items)})
    return out


@app.delete("/api/tokhai-nhap/{cid}")
def del_tokhai_nhap(cid: int, so_tk: str = ""):
    conn = db()
    if so_tk:
        conn.execute("DELETE FROM tokhai_nhap WHERE company_id=? AND so_tk=?", (cid, so_tk))
    else:
        conn.execute("DELETE FROM tokhai_nhap WHERE company_id=?", (cid,))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.post("/api/clear-downloads")
def clear_downloads(scope: str = "temp"):
    """Xóa file đã tải để nhẹ máy.
    CHỈ xóa file BÊN TRONG 'Thư mục lưu file XML/PDF' (save_dir) của mỗi công ty.
    TUYỆT ĐỐI KHÔNG đụng tới thư mục dữ liệu (data_dir) hay bất kỳ thư mục nào khác.
    Trả về số file đã xóa và dung lượng giải phóng (MB)."""
    xoa = 0
    dung_luong = 0
    EXT = (".xml", ".pdf", ".zip", ".html", ".xlsx", ".json")

    def _xoa_file(path):
        nonlocal xoa, dung_luong
        try:
            sz = os.path.getsize(path)
            os.remove(path)
            xoa += 1
            dung_luong += sz
        except Exception:
            pass

    # Chỉ duyệt save_dir của các công ty. KHÔNG xóa data_dir, KHÔNG xóa thư mục khác.
    conn = db()
    rows = conn.execute("SELECT save_dir, data_dir FROM companies").fetchall()
    conn.close()
    # tập data_dir để loại trừ tuyệt đối (phòng khi save_dir == data_dir hoặc lồng nhau)
    data_dirs = set()
    for d in rows:
        dd = (d["data_dir"] or "").strip() if "data_dir" in d.keys() else ""
        if dd:
            data_dirs.add(os.path.realpath(dd))

    seen = set()
    for d in rows:
        sd = (d["save_dir"] or "").strip()
        if not sd or not os.path.isdir(sd):
            continue
        rsd = os.path.realpath(sd)
        if rsd in seen:
            continue
        seen.add(rsd)
        # nếu người dùng đặt save_dir trùng data_dir -> bỏ qua để bảo vệ dữ liệu
        if rsd in data_dirs:
            continue
        for root, _dirs, files in os.walk(sd):
            # không đi vào thư mục data_dir nếu nó nằm lồng trong save_dir
            if os.path.realpath(root) in data_dirs:
                _dirs[:] = []
                continue
            for ff in files:
                if ff.lower().endswith(EXT):
                    _xoa_file(os.path.join(root, ff))

    return {"ok": True, "so_file": xoa, "dung_luong_mb": round(dung_luong / 1024 / 1024, 2)}


@app.get("/api/downloads-size")
def downloads_size():
    """Trả về số file và dung lượng đang chiếm (để hiển thị trước khi xóa)."""
    tong = 0
    sl = 0
    if os.path.isdir(DOWNLOAD_DIR):
        for root, _dirs, files in os.walk(DOWNLOAD_DIR):
            for ff in files:
                if ff.lower().endswith((".xml", ".pdf", ".zip", ".xlsx", ".html")):
                    try:
                        tong += os.path.getsize(os.path.join(root, ff))
                        sl += 1
                    except Exception:
                        pass
    return {"so_file": sl, "dung_luong_mb": round(tong / 1024 / 1024, 2)}


def _doc_sheet_nhap_lieu(wb, sheet_ten, header_marker):
    """Đọc 1 sheet, trả về (header, rows) đã bỏ tiêu đề/tổng/nhóm. None nếu không có sheet."""
    import re as _re2
    ws = None
    for sn in wb.sheetnames:
        if sn.strip().lower() == sheet_ten.lower():
            ws = wb[sn]; break
    if ws is None:
        return None
    hrow = None
    for r in range(1, min(ws.max_row, 15) + 1):
        if str(ws.cell(r, 1).value or "").strip() == header_marker:
            hrow = r; break
    if hrow is None:
        return None
    ncol = ws.max_column
    header = [str(ws.cell(hrow, c).value or "").strip() for c in range(1, ncol + 1)]
    rows = []
    for r in range(hrow + 1, ws.max_row + 1):
        vals = [ws.cell(r, c).value for c in range(1, ncol + 1)]
        if all(v is None or str(v).strip() == "" for v in vals):
            continue
        joined = " ".join(str(v) for v in vals if v is not None).lower().strip()
        if "tổng cộng" in joined or "tổng nhóm" in joined:
            continue
        if _re2.match(r"^\d+\.\s", joined):
            continue
        rows.append([("" if v is None else v) for v in vals])
    return header, rows


# ====== HẠCH TOÁN: nhớ tài khoản Nợ theo MST nhà cung cấp (mỗi cty 1 bộ) ======
NGUONG_5TR = 5_000_000  # hóa đơn >= 5tr -> Có 331 (chuyển khoản), < 5tr -> Có 1111

def _chuan_mst(s):
    """Chuẩn hóa MST: bỏ khoảng trắng, gạch, chấm."""
    return str(s or "").strip().replace("-", "").replace(" ", "").replace(".", "")

def _dinh_dang_mst(s):
    """MST 13 số (đơn vị trực thuộc) -> 10 số + '-' + 3 số. Khác giữ nguyên."""
    s = _chuan_mst(s)
    if len(s) == 13 and s.isdigit():
        return s[:10] + "-" + s[10:]
    return s

def _nam_cua_ngay(ngay):
    """Lấy năm (4 số) từ chuỗi ngày dd/mm/yyyy hoặc yyyy-mm-dd."""
    s = str(ngay or "").strip()
    if "/" in s:
        p = s.split("/")
        if len(p) == 3:
            return p[2][:4]
    if "-" in s:
        p = s.split("-")
        if len(p) == 3:
            return p[0][:4] if len(p[0]) == 4 else p[2][:4]
    return ""

def _so_ct_theo_nam_mst(prefix, ngay, mst):
    """Số chứng từ/phiếu = prefix + năm hóa đơn + MST, cắt tối đa 20 ký tự."""
    return (str(prefix) + _nam_cua_ngay(ngay) + str(mst or "").strip())[:20]

def _so_ct_unique(prefix, ngay, mst, seen):
    """Như _so_ct_theo_nam_mst() nhưng LUÔN DUY NHẤT: nếu trùng với dòng
    trước (cùng NCC, cùng năm) -> cắt bớt để chừa chỗ thêm số thứ tự
    '-2','-3'... vẫn giữ tổng tối đa 20 ký tự (MISA yêu cầu mỗi chứng từ
    phải có số duy nhất). 'seen' là dict đếm dùng chung cho cả lượt xuất."""
    base = _so_ct_theo_nam_mst(prefix, ngay, mst)
    n = seen.get(base, 0) + 1
    seen[base] = n
    if n == 1:
        return base
    suf = f"-{n}"
    return base[: max(0, 20 - len(suf))] + suf

def _so_ct_unique_memo(prefix, ngay, mst, invoice_key, seen, cache):
    """Như _so_ct_unique() nhưng NHỚ theo invoice_key: các dòng cùng 1 hóa
    đơn (cùng invoice_key — vd số HĐ+MST+ngày) dùng lại đúng 1 số chứng từ
    đã cấp, để nhiều dòng hàng của cùng 1 hóa đơn vẫn chung 1 chứng từ."""
    if invoice_key in cache:
        return cache[invoice_key]
    v = _so_ct_unique(prefix, ngay, mst, seen)
    cache[invoice_key] = v
    return v

def _co_theo_tong(tong):
    """Cột Có: >= 5 triệu -> 331 (phải trả NB), còn lại -> 1111 (tiền mặt)."""
    t = _to_num(tong)
    if isinstance(t, (int, float)) and abs(t) >= NGUONG_5TR:
        return "331"
    return "1111"

def _du_lieu_cty_path(cid):
    """File dữ liệu riêng của công ty (hạch toán, sau này thêm hàng hóa...).
    Ưu tiên thư mục data_dir (riêng cho dữ liệu); nếu trống thì dùng save_dir;
    cuối cùng là data/cong_ty/."""
    conn = db()
    comp = conn.execute(
        "SELECT mst, save_dir, data_dir FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    if not comp:
        return None
    # data_dir do người dùng chỉ định riêng -> tạo nếu chưa có
    dd = (comp["data_dir"] or "").strip() if "data_dir" in comp.keys() else ""
    sd = (comp["save_dir"] or "").strip()
    thu_muc = ""
    if dd:
        thu_muc = dd
    elif sd and os.path.isdir(sd):
        thu_muc = sd
    else:
        thu_muc = os.path.join(DATA_DIR, "cong_ty")
    try:
        os.makedirs(thu_muc, exist_ok=True)
    except Exception:
        return None
    mst = _chuan_mst(comp["mst"]) or str(cid)
    return os.path.join(thu_muc, f"DuLieu_{mst}.json")

def _doc_du_lieu_cty(cid):
    p = _du_lieu_cty_path(cid)
    if p and os.path.isfile(p):
        try:
            with open(p, encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}

def _ghi_du_lieu_cty(cid, data):
    p = _du_lieu_cty_path(cid)
    if not p:
        return
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=1)
    except Exception:
        pass

def _init_hach_toan(conn):
    conn.execute("""CREATE TABLE IF NOT EXISTS hach_toan_no(
        company_id INTEGER, mst_ncc TEXT, tk_no TEXT, updated_at TEXT,
        UNIQUE(company_id, mst_ncc))""")

def _get_map_no(cid):
    """Trả về {mst_chuan: tk_no} đã học cho công ty này (DB ưu tiên, fallback file)."""
    conn = db()
    _init_hach_toan(conn)
    rows = conn.execute(
        "SELECT mst_ncc, tk_no FROM hach_toan_no WHERE company_id=?", (cid,)).fetchall()
    conn.close()
    m = {r["mst_ncc"]: r["tk_no"] for r in rows if (r["tk_no"] or "").strip()}
    if not m:
        data = _doc_du_lieu_cty(cid)
        m = {k: v for k, v in (data.get("hach_toan_no", {}) or {}).items() if str(v).strip()}
    return m

def _hoc_map_no(cid, mapping):
    """Học/cập nhật {mst_chuan: tk_no} vào DB + ghi ra file dữ liệu công ty."""
    mapping = {(_chuan_mst(k)): str(v).strip()
               for k, v in (mapping or {}).items()
               if _chuan_mst(k) and str(v).strip()}
    if not mapping:
        return
    conn = db()
    _init_hach_toan(conn)
    now = datetime.datetime.now().isoformat()
    for mst, tk in mapping.items():
        conn.execute("""INSERT INTO hach_toan_no(company_id, mst_ncc, tk_no, updated_at)
            VALUES (?,?,?,?)
            ON CONFLICT(company_id, mst_ncc) DO UPDATE SET
                tk_no=excluded.tk_no, updated_at=excluded.updated_at""",
            (cid, mst, tk, now))
    conn.commit()
    # ghi TOÀN BỘ map từ DB ra file (đổi thư mục lưu vẫn đầy đủ lịch sử)
    full = {r["mst_ncc"]: r["tk_no"] for r in conn.execute(
        "SELECT mst_ncc, tk_no FROM hach_toan_no WHERE company_id=?", (cid,)).fetchall()
        if (r["tk_no"] or "").strip()}
    conn.close()
    data = _doc_du_lieu_cty(cid)
    data["hach_toan_no"] = full
    _ghi_du_lieu_cty(cid, data)


@app.get("/api/hach-toan-no/{cid}")
def hach_toan_no_get(cid: int):
    """Lấy bảng tài khoản Nợ đã học theo MST (để tự điền khi nhập liệu)."""
    return {"map": _get_map_no(cid)}


@app.post("/api/nhap-lieu/import/{cid}")
async def nhap_lieu_import(cid: int, request: Request, loai: str = "in"):
    """Import nhiều file Excel cho Nhập Liệu, GỘP (nối đuôi) thành 1 bảng.
    loai='in'  -> đọc sheet 'Chi tiết MUA VÀO'
    loai='out' -> đọc sheet 'BK Bán ra'
    Đồng thời đọc luôn sheet 'Chi tiết BÁN RA' (nếu có) làm nguồn cho tính
    năng Xuất Kho (Sheet GIATHANH) — khỏi phải import lại riêng.
    Trả về header + tất cả các dòng dữ liệu đã gộp (bỏ tiêu đề và dòng tổng)."""
    import openpyxl, io as _io
    form = await request.form()
    files = form.getlist("files") or ([form.get("file")] if form.get("file") else [])
    if not files:
        raise HTTPException(400, "Chưa chọn file")

    header_in, rows_in = [], []
    header_out, rows_out = [], []
    header_ctbr, rows_ctbr = [], []
    so_file_ok = 0
    loi = []

    for up in files:
        if up is None:
            continue
        fn = getattr(up, "filename", "file")
        try:
            content = await up.read()
            wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        except Exception as e:
            loi.append(f"{fn}: không đọc được ({e})")
            continue

        co_du_lieu = False
        kq_in = _doc_sheet_nhap_lieu(wb, "Chi tiết MUA VÀO", "Ký hiệu")
        if kq_in:
            h, rs = kq_in
            if not header_in:
                header_in = h
            rows_in.extend(rs)
            co_du_lieu = True
        kq_out = _doc_sheet_nhap_lieu(wb, "BK Bán ra", "STT")
        if kq_out:
            h, rs = kq_out
            if not header_out:
                header_out = h
            rows_out.extend(rs)
            co_du_lieu = True
        kq_ctbr = _doc_sheet_nhap_lieu(wb, "Chi tiết BÁN RA", "Ký hiệu")
        if kq_ctbr:
            h, rs = kq_ctbr
            if not header_ctbr:
                header_ctbr = h
            rows_ctbr.extend(rs)
            co_du_lieu = True
        if co_du_lieu:
            so_file_ok += 1
        else:
            loi.append(f"{fn}: không có sheet 'Chi tiết MUA VÀO' / 'BK Bán ra' / 'Chi tiết BÁN RA'")

    conn = db()
    conn.execute("""CREATE TABLE IF NOT EXISTS nhap_lieu (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, loai TEXT,
        header_json TEXT, rows_json TEXT, updated_at TEXT,
        UNIQUE(company_id, loai))""")
    conn.commit()
    conn.close()

    return {"ok": True, "so_file": so_file_ok,
            "in": {"header": header_in, "rows": rows_in, "so_dong": len(rows_in)},
            "out": {"header": header_out, "rows": rows_out, "so_dong": len(rows_out)},
            "ctbr": {"header": header_ctbr, "rows": rows_ctbr, "so_dong": len(rows_ctbr)},
            "loi": loi[:5]}


@app.post("/api/nhap-lieu/save/{cid}")
async def nhap_lieu_save(cid: int, request: Request, loai: str = "in"):
    """Lưu bộ dữ liệu Nhập Liệu của công ty (mỗi công ty 1 bộ/loại, import đè lên)."""
    body = await request.json()
    header = body.get("header", [])
    rows = body.get("rows", [])
    conn = db()
    conn.execute("""CREATE TABLE IF NOT EXISTS nhap_lieu (
        id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, loai TEXT,
        header_json TEXT, rows_json TEXT, updated_at TEXT,
        UNIQUE(company_id, loai))""")
    conn.execute("""INSERT INTO nhap_lieu (company_id, loai, header_json, rows_json, updated_at)
        VALUES (?,?,?,?,?)
        ON CONFLICT(company_id, loai) DO UPDATE SET
            header_json=excluded.header_json, rows_json=excluded.rows_json,
            updated_at=excluded.updated_at""",
        (cid, loai, json.dumps(header, ensure_ascii=False),
         json.dumps(rows, ensure_ascii=False), datetime.datetime.now().isoformat()))
    conn.commit()
    conn.close()

    # HỌC tài khoản Nợ theo MST nhà cung cấp (chỉ với bảng kê ĐẦU VÀO)
    da_hoc = 0
    if loai == "in" and header and rows:
        hlow = [str(h or "").strip().lower() for h in header]
        def _tim_cot(*tu_khoa):
            for i, h in enumerate(hlow):
                if any(k in h for k in tu_khoa):
                    return i
            return -1
        i_mst = _tim_cot("mst bán", "mst ban", "mst")
        i_no = -1
        for i, h in enumerate(hlow):
            if h == "nợ" or h == "no":
                i_no = i; break
        mapping = {}
        if i_mst >= 0 and i_no >= 0:
            for r in rows:
                if i_mst < len(r) and i_no < len(r):
                    mst = _chuan_mst(r[i_mst])
                    tk = str(r[i_no] or "").strip()
                    if mst and tk:
                        mapping[mst] = tk   # dòng sau ghi đè dòng trước (mới nhất thắng)
        if mapping:
            _hoc_map_no(cid, mapping)
            da_hoc = len(mapping)

    # Ghi dữ liệu nhập liệu vào FILE riêng của công ty để dễ tìm/xử lý sau này
    data_cty = _doc_du_lieu_cty(cid)
    data_cty[f"nhap_lieu_{loai}"] = {
        "header": header, "rows": rows,
        "updated_at": datetime.datetime.now().isoformat()}
    _ghi_du_lieu_cty(cid, data_cty)

    file_du_lieu = _du_lieu_cty_path(cid) or ""
    return {"ok": True, "so_dong": len(rows), "da_hoc_no": da_hoc,
            "file_du_lieu": os.path.abspath(file_du_lieu) if file_du_lieu else "",
            "db_path": os.path.abspath(DB_PATH)}


@app.get("/api/nhap-lieu/{cid}")
def nhap_lieu_get(cid: int, loai: str = "in"):
    """Lấy dữ liệu Nhập Liệu đã lưu của công ty."""
    conn = db()
    try:
        r = conn.execute("SELECT header_json, rows_json, updated_at FROM nhap_lieu WHERE company_id=? AND loai=?",
                         (cid, loai)).fetchone()
    except Exception:
        r = None
    conn.close()
    if not r:
        return {"header": [], "rows": [], "updated_at": ""}
    try:
        return {"header": json.loads(r["header_json"]), "rows": json.loads(r["rows_json"]),
                "updated_at": r["updated_at"] or ""}
    except Exception:
        return {"header": [], "rows": [], "updated_at": ""}


@app.delete("/api/nhap-lieu/{cid}")
def nhap_lieu_del(cid: int, loai: str = "in"):
    conn = db()
    try:
        conn.execute("DELETE FROM nhap_lieu WHERE company_id=? AND loai=?", (cid, loai))
        conn.commit()
    except Exception:
        pass
    conn.close()
    return {"ok": True}


@app.post("/api/nhap-lieu/export/{cid}")
async def nhap_lieu_export(cid: int, request: Request, loai: str = "in"):
    """Xuất bảng nhập liệu ra Excel (có dòng tổng cộng)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    body = await request.json()
    header = body.get("header", [])
    rows = body.get("rows", [])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đầu ra" if loai == "out" else "Đầu vào"
    ws.append(header)
    for c in range(1, len(header) + 1):
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    # cột tiền để tính tổng + format
    def la_tien(h):
        t = str(h or "").lower()
        return any(k in t for k in ["doanh số", "thành tiền", "thuế gtgt", "chưa thuế",
                                    "chưa vat", "tiền thuế", "thuế nk", "trị giá", "tổng"])
    tien_idx = [i for i, h in enumerate(header) if la_tien(h)]
    for row in rows:
        vals = []
        for i, v in enumerate(row):
            if i in tien_idx and v not in ("", None):
                vals.append(_to_num(v) or 0)
            else:
                vals.append(v)
        ws.append(vals)
    # dòng tổng cộng
    tong_row = ["" for _ in header]
    if header:
        tong_row[0] = "TỔNG CỘNG"
    for i in tien_idx:
        s = sum(_to_num(r[i]) or 0 for r in rows if i < len(r))
        tong_row[i] = s
    ws.append(tong_row)
    last = ws.max_row
    for c in range(1, len(header) + 1):
        ws.cell(last, c).font = Font(bold=True, color="C00000")
    # format số + độ rộng
    from openpyxl.utils import get_column_letter
    for i in tien_idx:
        for r in range(2, ws.max_row + 1):
            ws.cell(r, i + 1).number_format = "#,##0"
    for c in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{ws.max_row}"

    fname = f"NhapLieu_{'DauRa' if loai=='out' else 'DauVao'}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    desktop = _get_desktop_dir()
    if desktop and os.path.isdir(desktop):
        try:
            import shutil
            shutil.copy(path, os.path.join(desktop, fname))
        except Exception:
            pass
    return _resp_xuat(path, fname)


# ---- Helper CHUNG: ghi 1 workbook theo form MISA + lưu (dùng lại cho xuất
#      lẻ từng chứng từ VÀ cho 'Xuất trọn gói cho MISA') ----
def _viet_wb_misa(headers, rows, sheet, cot_text=None, cot_tien=None, cot_thapphan=None):
    """Tạo openpyxl Workbook theo form MISA: tiêu đề đậm nền xanh, định dạng số
    cho cột tiền/thập phân, dạng chữ cho cột mã/số HĐ. Trả về workbook.
    headers có thể chứa chuỗi rỗng '' cho cột đệm (không tô đậm)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    cot_text = cot_text or set(); cot_tien = cot_tien or set(); cot_thapphan = cot_thapphan or set()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = str(sheet)[:31]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c); cell.value = h
        if h:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E5C8A")
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            if v == "" or v is None:
                continue
            cell = ws.cell(ri, ci); cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_thapphan:
                cell.number_format = "#,##0.####"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    return wb

def _luu_file_misa(wb, fname, cid):
    """Lưu workbook ra DOWNLOAD_DIR + copy sang Desktop và thư mục công ty. Trả path."""
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return path

# Form 'Chứng từ mua dịch vụ' (MISA): 40 cột A..AN + AO đệm + AP 'Lọc' + AQ đệm
MUA_DV_HEADERS = [
    "Hiển thị trên sổ", "Phương thức thanh toán", "Nhận kèm hóa đơn",
    "Là CP mua hàng", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số chứng từ (*)", "Mã nhà cung cấp", "Tên nhà cung cấp", "Địa chỉ",
    "Diễn giải/Lý do chi/Nội dung thanh toán", "NV mua hàng", "Loại tiền",
    "Tỷ giá", "Mã dịch vụ (*)", "Tên dịch vụ", "TK chi phí/TK kho (*)",
    "TK công nợ/TK tiền (*)", "Đối tượng", "ĐVT", "Số lượng", "Đơn giá",
    "Thành tiền", "Thành tiền quy đổi", "Tỷ lệ CK (%)", "Tiền chiết khấu",
    "Tiền chiết khấu quy đổi", "% thuế GTGT", "Tiền thuế GTGT",
    "Tiền thuế GTGT quy đổi", "TK thuế GTGT", "Mẫu số HĐ", "Ký hiệu HĐ",
    "Số hóa đơn", "Ngày hóa đơn", "Nhóm HHDV mua vào", "Mã NCC", "Tên NCC",
    "Mã số thuế NCC", "Địa chỉ NCC", "", "Lọc", ""]
MUA_DV_COT_TEXT = {7, 8, 19, 34, 37, 39, 42, 43}
MUA_DV_COT_TIEN = {23, 29}
MUA_DV_COT_THAPPHAN = {21, 22}

def _gen_mua_hang_dv(cid, header, rows):
    """Từ lưới Nhập Liệu đầu vào -> danh sách dòng 'Chứng từ mua dịch vụ' (chỉ
    dòng Nợ là TK chi phí 6xx). Mỗi dòng là list phẳng dài 43 cột."""
    hlow = [str(h or "").strip().lower() for h in header]
    def tim(eqs=(), contains=()):
        for i, h in enumerate(hlow):
            if h in eqs:
                return i
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    i_so = tim(("số hđ",), ("số hđ", "số hoá đơn", "số hóa đơn"))
    i_ngay = tim(("ngày",), ("ngày",))
    i_nb = tim(("người bán",), ("người bán",))
    i_mst = tim(("mst bán", "mst"), ("mst",))
    i_ten = tim(("tên hàng hóa/dịch vụ",), ("tên hàng",))
    i_ts = tim(("thuế suất",), ("thuế suất",))
    i_tthue = tim(("tiền thuế gtgt",), ("tiền thuế",))
    i_tt = tim(("thành tiền",), ("thành tiền",))
    i_no = tim(("nợ",), ())
    i_co = tim(("có",), ())

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    soct_seen, soct_cache = {}, {}
    def so_chung_tu(sohd, mst_disp, ngay):
        return _so_ct_unique_memo("DV", ngay, mst_disp, (sohd, mst_disp, ngay),
                                  soct_seen, soct_cache)
    out = []
    for r in rows:
        no = str(gv(r, i_no) or "").strip()
        if not no.startswith("6"):          # chỉ lấy Nợ là TK chi phí 6xx
            continue
        co = str(gv(r, i_co) or "").strip()
        mst_disp = _dinh_dang_mst(gv(r, i_mst))
        ngay = str(gv(r, i_ngay) or "")
        sohd = gv(r, i_so)
        ten = gv(r, i_ten)
        soct = so_chung_tu(sohd, mst_disp, ngay)
        tt_val = _to_num(gv(r, i_tt))
        row_vals = {
            1: 0, 2: 1, 3: 1,
            5: ngay, 6: ngay, 7: soct,
            8: mst_disp, 9: gv(r, i_nb),
            11: ten,
            15: "MHDV", 16: ten,
            17: no, 18: co, 19: mst_disp,
            20: "", 21: 1,
            22: tt_val, 23: tt_val,
            28: _chuan_thue_suat(gv(r, i_ts)), 29: _to_num(gv(r, i_tthue)),
            31: "1331",
            34: sohd, 35: ngay, 36: "1",
            37: mst_disp, 38: gv(r, i_nb), 39: mst_disp,
            42: sohd, 43: soct,
        }
        row = [""] * 43
        for c, v in row_vals.items():
            row[c - 1] = v
        out.append(row)
    return out

@app.post("/api/mua-hang-dv/{cid}")
async def mua_hang_dich_vu(cid: int, request: Request):
    """Lọc Bảng kê Mua vào (đã hạch toán) -> kết xuất 'Chứng từ mua dịch vụ'
    theo form MISA. Chỉ lấy dòng có Nợ bắt đầu bằng '6' (chi phí dịch vụ).
    Nhận {header, rows} của lưới Nhập Liệu đầu vào."""
    body = await request.json()
    out = _gen_mua_hang_dv(cid, body.get("header", []), body.get("rows", []))
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst_cty = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"MuaHangDichVu_{mst_cty}.xlsx"
    wb = _viet_wb_misa(MUA_DV_HEADERS, out, "Chứng từ mua dịch vụ",
                       MUA_DV_COT_TEXT, MUA_DV_COT_TIEN, MUA_DV_COT_THAPPHAN)
    path = _luu_file_misa(wb, fname, cid)
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============ DANH MỤC HÀNG HÓA / NVL / VTHH (từ bảng kê mua vào) ============
def _dm_cols(header):
    """Tìm chỉ số cột cần dùng trong lưới Nhập Liệu đầu vào."""
    hlow = [str(h or "").strip().lower() for h in header]
    def find(eqs, contains):
        for i, h in enumerate(hlow):
            if h in eqs:
                return i
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    return {
        "ten": find(("tên hàng hóa/dịch vụ",), ("tên hàng",)),
        "dvt": find(("đvt",), ("đvt", "đơn vị tính")),
        "sl": find(("số lượng",), ("số lượng",)),
        "dgia": find(("đơn giá",), ("đơn giá",)),
        "tt": find(("thành tiền",), ("thành tiền",)),
        "ts": find(("thuế suất",), ("thuế suất",)),
        "sohd": find(("số hđ",), ("số hđ", "số hoá đơn", "số hóa đơn")),
        "ngay": find(("ngày",), ("ngày",)),
        "no": find(("nợ",), ()),
    }

_THUE_HOP_LE = (0, 5, 8, 10)

def _chuan_thue_suat(v):
    """Chuẩn hoá % thuế GTGT về ĐÚNG 1 trong 7 giá trị hợp lệ:
    0, 5, 8, 10, 'KCT', 'KHAC', 'KKKNT'. Nhận chuỗi ('8%','KCT'...) hoặc số.
    Số lẻ (0.1, 0.8, 5.26%, 'KHAC:5.26%'...) được LÀM TRÒN về mức GẦN NHẤT
    trong 0/5/8/10 (chỉ đổi NHÃN hiển thị — tiền thuế GTGT vẫn giữ nguyên
    theo file import, KHÔNG tính lại). 'KHAC' chỉ giữ khi nguồn ghi đúng
    chữ 'KHAC' mà không kèm số, hoặc chuỗi không đọc được thành số."""
    if isinstance(v, (int, float)):
        f = v
    else:
        s = str(v or "").strip().upper().replace(" ", "")
        if s in ("", "KCT", "KHÔNG", "KHONG", "KO", "KHTKKNT"):
            return "KCT"
        if s == "KKKNT":
            return "KKKNT"
        if s == "KHAC":
            return "KHAC"
        if ":" in s:            # 'KHAC:5.26%' -> lấy phần số sau dấu ':'
            s = s.split(":", 1)[1]
        s2 = s.replace("%", "").replace(",", ".")
        try:
            f = float(s2)
        except Exception:
            return "KHAC"
    return min(_THUE_HOP_LE, key=lambda t: abs(f - t))

def _dm_rate(v):
    """Thuế suất GTGT -> 1 trong 7 giá trị hợp lệ (xem _chuan_thue_suat)."""
    return _chuan_thue_suat(v)

def _so_pct(v):
    """Số % thuần (dùng cho thuế NK/TTĐB — KHÔNG giới hạn 0/5/8/10 như VAT).
    '3%'->3, '20%'->20, trống/KCT... -> 0."""
    s = str(v or "").strip().upper().replace("%", "").replace(",", ".")
    if s in ("", "KCT", "KKKNT", "KHTKKNT", "KHÔNG", "KHONG", "KO"):
        return 0
    try:
        f = float(s)
        return int(f) if f == int(f) else f
    except Exception:
        return 0

# Cột Danh mục Hàng hóa/NVL: bỏ cột phụ (J), thêm Thuế suất + Kho
DM_HH_HEADERS = ["Mã Hàng", "Mặt hàng", "ĐVT", "Thuế suất", "Ký tự ", "SL",
                 "Đơn giá", "Thành tiền", "Hoá đơn", "Ngày Hoá Đơn", "Kho"]
# vị trí cột trong DM_HH_HEADERS
DM_I_MA, DM_I_TEN, DM_I_DVT, DM_I_TS, DM_I_KY = 0, 1, 2, 3, 4
DM_I_SL, DM_I_DG, DM_I_TT, DM_I_HD, DM_I_NGAY, DM_I_KHO = 5, 6, 7, 8, 9, 10

# Danh mục TSCĐ / CCDC (mỗi dòng 1 mã, không gộp). 12 cột.
DM_TSCD_HEADERS = ["Mã TSCĐ", "Tên TSCĐ", "ĐVT", "SL", "Đơn giá", "Thành tiền",
                   "HĐ", "Ngày HĐ", "Hạn sử dụng", "Đối tượng phân bổ",
                   "Ngày ghi tăng", "TK Chi phi"]
DM_CCDC_HEADERS = ["Mã CCDC", "Tên CCDC", "ĐVT", "SL", "Đơn giá", "Thành tiền",
                   "HĐ", "Ngày", "Hạn sử dụng", "Đối tượng phân bổ",
                   "Ngày ghi tăng", "TK Chi phi"]

def _dm_la_ts(loai):
    return loai in ("tscd", "ccdc")

def _dm_headers(loai):
    if loai == "tscd":
        return DM_TSCD_HEADERS
    if loai == "ccdc":
        return DM_CCDC_HEADERS
    return DM_HH_HEADERS

def _gen_danh_muc_ts(cid, loai, header, rows):
    """Danh mục TSCĐ (Nợ 2111/211) hoặc CCDC (Nợ 2421/242). MỖI DÒNG = 1 mã
    riêng (TSCD00001, CCDC00001...), nối tiếp. Không gộp theo tên.
    Cột: Mã, Tên, ĐVT, SL, Đơn giá, Thành tiền, HĐ, Ngày, Hạn SD, ĐT phân bổ,
    Ngày ghi tăng, TK Chi phí. Trả (all_rows, so_moi)."""
    prefix = "TSCD" if loai == "tscd" else "CCDC"
    accs = {"2111", "211"} if loai == "tscd" else {"2421", "242"}
    # Giá trị mặc định theo form mẫu
    hsd_md = 60 if loai == "tscd" else 24          # Hạn sử dụng (tháng)
    dtpb_md = "VP"                                  # Đối tượng phân bổ
    tkcp_md = "6424" if loai == "tscd" else "6422"  # TK Chi phí
    data = _doc_du_lieu_cty(cid)
    store = data.get("dm_" + loai, {}) or {}
    next_n = int(store.get("next", 1))
    saved_rows = [list(r) for r in store.get("rows", []) if len(r) >= 12]
    col = _dm_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    def rk(row):
        # khóa dòng: số HĐ | ngày | tên | thành tiền (để không thêm trùng dòng)
        return f"{gv(row,6)}|{gv(row,7)}|{gv(row,1)}|{gv(row,5)}"

    seen = set(rk(r) for r in saved_rows)
    new_rows = []
    for r in rows:
        no = str(gv(r, col["no"]) or "").strip()
        if no not in accs:
            continue
        ten = str(gv(r, col["ten"]) or "").strip()
        if not ten:
            continue
        dvt = str(gv(r, col["dvt"]) or "").strip()
        sl = _to_num(gv(r, col["sl"]))
        dgia = _to_num(gv(r, col["dgia"]))
        tt = _to_num(gv(r, col["tt"]))
        sohd = str(gv(r, col["sohd"]) or "")
        ngay = str(gv(r, col["ngay"]) or "")
        key = f"{sohd}|{ngay}|{ten}|{tt}"
        if key in seen:
            continue
        seen.add(key)
        ma = prefix + str(next_n).zfill(5)
        next_n += 1
        # [Mã, Tên, ĐVT, SL, Đơn giá, Thành tiền, HĐ, Ngày, Hạn SD,
        #  ĐT phân bổ, Ngày ghi tăng, TK Chi phí]
        new_rows.append([ma, ten, dvt, sl, dgia, tt, sohd, ngay,
                         hsd_md, dtpb_md, ngay, tkcp_md])
    # dòng MỚI đẩy lên TRÊN, dòng cũ (đã lưu) xếp phía DƯỚI
    return new_rows + saved_rows, len(new_rows)

def _gen_danh_muc(cid, loai, header, rows):
    """Sinh Danh mục Hàng hóa (loai='hh', Nợ 1561/156) hoặc NVL ('nvl', Nợ 152).
    Mã = base (theo Ký tự=nospaces(Tên)+ĐVT) + '-' + thuế suất (vd HH00001-8).
    Cột Kho mặc định HH/NVL. Nối tiếp + không lặp dòng. Trả (all_rows, so_moi)."""
    if _dm_la_ts(loai):
        return _gen_danh_muc_ts(cid, loai, header, rows)
    prefix = "HH" if loai == "hh" else "NVL"
    accs = {"1561", "156"} if loai == "hh" else {"152"}
    kho = "HH" if loai == "hh" else "NVL"
    data = _doc_du_lieu_cty(cid)
    store = data.get("dm_" + loai, {}) or {}
    keymap = dict(store.get("map", {}))           # Ký tự -> base (HH00001)
    next_n = int(store.get("next", 1))
    # Chỉ giữ dòng đã lưu theo ĐỊNH DẠNG MỚI (11 cột). Dòng cũ (10 cột,
    # chưa có Thuế suất/Kho) bị bỏ để tránh lệch cột; sẽ tự sinh lại từ
    # bảng kê hiện tại với mã giữ nguyên theo bản đồ đã học.
    saved_rows = [list(r) for r in store.get("rows", []) if len(r) >= 11]
    col = _dm_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    def rk(row):
        return f"{gv(row,DM_I_MA)}|{gv(row,DM_I_HD)}|{gv(row,DM_I_NGAY)}|{gv(row,DM_I_TT)}"

    seen = set(rk(r) for r in saved_rows if len(r) >= 10)
    new_rows = []
    for r in rows:
        no = str(gv(r, col["no"]) or "").strip()
        if no not in accs:
            continue
        ten = str(gv(r, col["ten"]) or "").strip()
        dvt = str(gv(r, col["dvt"]) or "").strip()
        if not ten:
            continue
        ky_tu = "".join(ten.split()) + dvt
        if ky_tu in keymap:
            base = keymap[ky_tu]
        else:
            base = prefix + str(next_n).zfill(5)
            keymap[ky_tu] = base
            next_n += 1
        rate = _dm_rate(gv(r, col["ts"]))
        ma = f"{base}-{rate}"
        sl = _to_num(gv(r, col["sl"]))
        dgia = _to_num(gv(r, col["dgia"]))
        tt = _to_num(gv(r, col["tt"]))
        sohd = str(gv(r, col["sohd"]) or "")
        ngay = str(gv(r, col["ngay"]) or "")
        newrow = [ma, ten, dvt, rate, ky_tu, sl, dgia, tt, sohd, ngay, kho]
        if rk(newrow) in seen:
            continue
        seen.add(rk(newrow))
        new_rows.append(newrow)

    # dòng MỚI đẩy lên TRÊN, dòng cũ (đã lưu) xếp phía DƯỚI
    all_rows = new_rows + saved_rows
    return all_rows, len(new_rows)


def _luu_danh_muc(cid, loai, dm_rows):
    """Lưu danh mục đã chỉnh (lưới DM): ghi rows + bộ đếm next để các kỳ sau
    cấp mã nối tiếp. HH/NVL: dựng lại map(Ký tự->base)."""
    prefix = {"hh": "HH", "nvl": "NVL", "tscd": "TSCD", "ccdc": "CCDC"}[loai]
    rows = [list(r) for r in dm_rows if any(str(x).strip() for x in r)]
    maxn = 0
    keymap = {}
    for r in rows:
        ma = str(r[0]) if r else ""
        base = ma.split("-")[0] if ma else ""
        if not _dm_la_ts(loai):
            ky = str(r[DM_I_KY]) if len(r) > DM_I_KY else ""
            if ky and base:
                keymap[ky] = base
        if base.startswith(prefix) and base[len(prefix):].isdigit():
            maxn = max(maxn, int(base[len(prefix):]))
    store = {"next": maxn + 1, "rows": rows}
    if not _dm_la_ts(loai):
        store["map"] = keymap
    data = _doc_du_lieu_cty(cid)
    data["dm_" + loai] = store
    _ghi_du_lieu_cty(cid, data)
    return len(rows)


def _dm_doc_header_ma_ten(wb):
    """File IMPORT danh mục mã hàng do người dùng tự chuẩn bị (không theo mẫu
    bảng kê hóa đơn cố định) -> dò dòng tiêu đề (chứa cả chữ 'mã' và 'tên')
    trong 10 dòng đầu sheet đầu tiên, trả (header, các dòng dữ liệu phía sau)."""
    ws = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = 0
    for i in range(min(10, len(all_rows))):
        cells = [str(c or "").strip().lower() for c in all_rows[i]]
        if any("mã" in c for c in cells) and any("tên" in c for c in cells):
            header_idx = i
            break
    header = list(all_rows[header_idx]) if all_rows else []
    return header, all_rows[header_idx + 1:]


def _dm_import_cols(header):
    """Dò cột trong file IMPORT danh mục (khác _dm_cols dùng cho bảng kê hóa
    đơn) — file này do người dùng tự chuẩn bị nên tên cột có thể khác nhau."""
    hlow = [str(h or "").strip().lower() for h in header]
    def find(contains):
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    return {
        "ma": find(("mã hàng", "mã vật tư", "mã sp", "mã tscđ", "mã ccdc", "mã(*)", "mã (*)", "mã")),
        "ten": find(("tên hàng", "mặt hàng", "tên vật tư", "tên sản phẩm", "tên(*)", "tên (*)", "tên")),
        "dvt": find(("đvt", "đơn vị tính")),
        "ts": find(("thuế suất",)),
        "kho": find(("kho",)),
        "sl": find(("số lượng",)),
        "dgia": find(("đơn giá",)),
        "tt": find(("thành tiền",)),
        "sohd": find(("số hđ", "số hoá đơn", "số hóa đơn")),
        "ngay": find(("ngày",)),
    }


def _nhap_danh_muc(cid, loai, wb):
    """IMPORT danh mục mã hàng (DM Hàng hóa/NVL/TSCĐ/CCDC) từ 1 file Excel
    người dùng tự chuẩn bị: dòng nào CÓ điền cột 'Mã hàng' -> lấy ĐÚNG mã đó;
    dòng KHÔNG điền mã -> tự dò theo TÊN (chuẩn hoá bỏ dấu/hoa-thường/khoảng
    trắng, giống cách so khớp ở Xuất Kho) trong danh mục ĐÃ CÓ SẴN trong phần
    mềm (đã lưu từ trước) để lấy lại đúng mã cũ cho cùng 1 mặt hàng; nếu không
    tìm thấy tên nào khớp thì tự sinh mã mới nối tiếp (giống lúc dò tự động từ
    bảng kê). Trả (all_rows, thống kê) để nạp thẳng vào lưới Danh mục cho
    người dùng xem lại rồi bấm Lưu như bình thường."""
    header, data_rows = _dm_doc_header_ma_ten(wb)
    col = _dm_import_cols(header)
    if col["ten"] < 0:
        raise ValueError("Không tìm thấy cột 'Tên hàng'/'Mặt hàng' trong file")

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else None

    data = _doc_du_lieu_cty(cid)
    store = data.get("dm_" + loai, {}) or {}
    la_ts = _dm_la_ts(loai)
    prefix = {"hh": "HH", "nvl": "NVL", "tscd": "TSCD", "ccdc": "CCDC"}[loai]
    next_n = int(store.get("next", 1))
    so_dien_ma = so_trung_ten = so_moi = 0

    if la_ts:
        saved_rows = [list(r) for r in store.get("rows", []) if len(r) >= 12]
        ten_to_ma = {}
        for r in saved_rows:
            tc = _chuan_ten_hang_xk(r[1])
            if tc:
                ten_to_ma.setdefault(tc, r[0])
        hsd_md = 60 if loai == "tscd" else 24
        dtpb_md = "VP"
        tkcp_md = "6424" if loai == "tscd" else "6422"
        rk = lambda row: f"{row[6]}|{row[7]}|{row[1]}|{row[5]}"
    else:
        saved_rows = [list(r) for r in store.get("rows", []) if len(r) >= 11]
        ten_to_base = {}
        for r in saved_rows:
            tc = _chuan_ten_hang_xk(r[DM_I_TEN])
            base = str(r[DM_I_MA] or "").split("-")[0]
            if tc and base:
                ten_to_base.setdefault(tc, base)
        kho_md = "HH" if loai == "hh" else "NVL"
        rk = lambda row: f"{row[DM_I_MA]}|{row[DM_I_HD]}|{row[DM_I_NGAY]}|{row[DM_I_TT]}"

    seen = set(rk(r) for r in saved_rows)
    new_rows = []
    for r in data_rows:
        ten = str(gv(r, col["ten"]) or "").strip()
        if not ten:
            continue
        ma_dien = str(gv(r, col["ma"]) or "").strip() if col["ma"] >= 0 else ""
        dvt = str(gv(r, col["dvt"]) or "").strip() if col["dvt"] >= 0 else ""
        sl = _to_num(gv(r, col["sl"])) if col["sl"] >= 0 else ""
        dgia = _to_num(gv(r, col["dgia"])) if col["dgia"] >= 0 else ""
        tt = _to_num(gv(r, col["tt"])) if col["tt"] >= 0 else ""
        sohd = str(gv(r, col["sohd"]) or "") if col["sohd"] >= 0 else ""
        ngay = str(gv(r, col["ngay"]) or "") if col["ngay"] >= 0 else ""
        tc = _chuan_ten_hang_xk(ten)

        if la_ts:
            if ma_dien:
                ma = ma_dien
                so_dien_ma += 1
            elif tc in ten_to_ma:
                ma = ten_to_ma[tc]
                so_trung_ten += 1
            else:
                ma = prefix + str(next_n).zfill(5)
                next_n += 1
                so_moi += 1
            ten_to_ma[tc] = ma
            newrow = [ma, ten, dvt, sl, dgia, tt, sohd, ngay, hsd_md, dtpb_md, ngay, tkcp_md]
        else:
            rate = _dm_rate(gv(r, col["ts"]) if col["ts"] >= 0 else "")
            if ma_dien:
                ma = ma_dien
                base = ma_dien.split("-")[0]
                so_dien_ma += 1
            elif tc in ten_to_base:
                base = ten_to_base[tc]
                ma = f"{base}-{rate}"
                so_trung_ten += 1
            else:
                base = prefix + str(next_n).zfill(5)
                next_n += 1
                ma = f"{base}-{rate}"
                so_moi += 1
            ten_to_base[tc] = base
            ky_tu = "".join(ten.split()) + dvt
            kho_val = str(gv(r, col["kho"]) or "").strip() if col["kho"] >= 0 else ""
            newrow = [ma, ten, dvt, rate, ky_tu, sl, dgia, tt, sohd, ngay, kho_val or kho_md]

        key = rk(newrow)
        if key in seen:
            continue
        seen.add(key)
        new_rows.append(newrow)

    # dòng MỚI (import lần này) đẩy lên TRÊN, dòng cũ (đã lưu) xếp phía DƯỚI
    all_rows = new_rows + saved_rows
    return all_rows, {"so_dong": len(new_rows), "so_dien_ma": so_dien_ma,
                       "so_trung_ten": so_trung_ten, "so_moi": so_moi}


def _xuat_dm_excel(rows, ten_sheet, fname, cid, loai="hh"):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    headers = _dm_headers(loai)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ten_sheet
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    if _dm_la_ts(loai):
        cot_tien = {4, 5, 6}          # SL, Đơn giá, Thành tiền
        cot_text = {1, 7}             # Mã, HĐ
    else:
        cot_tien = {DM_I_SL + 1, DM_I_DG + 1, DM_I_TT + 1}
        cot_text = {DM_I_MA + 1, DM_I_KY + 1, DM_I_HD + 1, DM_I_KHO + 1}
    for r, row in enumerate(rows, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c)
            cell.value = v
            if c in cot_text:
                cell.number_format = "@"
            elif c in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return path


@app.post("/api/danh-muc-hang/{cid}")
async def danh_muc_hang(cid: int, request: Request, loai: str = "hh"):
    """Sinh Danh mục Hàng hóa/NVL từ lưới Nhập Liệu (để mở lưới chỉnh sửa)."""
    body = await request.json()
    all_rows, so_moi = _gen_danh_muc(cid, loai, body.get("header", []),
                                     body.get("rows", []))
    return {"headers": _dm_headers(loai), "rows": all_rows,
            "so_dong": len(all_rows), "so_moi": so_moi}


@app.post("/api/danh-muc-hang/import/{cid}")
async def danh_muc_hang_import(cid: int, request: Request, loai: str = "hh"):
    """IMPORT danh mục mã hàng từ 1 file Excel người dùng tự chuẩn bị (không
    phải bảng kê hóa đơn): dòng có điền 'Mã hàng' -> lấy đúng mã đó; dòng
    không điền -> tự dò theo tên trong danh mục đã có sẵn để lấy lại mã cũ,
    không tìm thấy thì tự sinh mã mới. Trả về (headers, rows) để nạp thẳng
    vào lưới Danh mục cho người dùng xem lại rồi bấm Lưu như bình thường."""
    import openpyxl, io as _io
    form = await request.form()
    up = form.get("file")
    if up is None:
        raise HTTPException(400, "Chưa chọn file")
    try:
        content = await up.read()
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file: {e}")
    try:
        all_rows, tk = _nhap_danh_muc(cid, loai, wb)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"headers": _dm_headers(loai), "rows": all_rows, "so_dong": len(all_rows), **tk}


@app.post("/api/danh-muc-hang/save/{cid}")
async def danh_muc_hang_save(cid: int, request: Request, loai: str = "hh"):
    """Lưu danh mục (lưới DM đã chỉnh sửa)."""
    body = await request.json()
    n = _luu_danh_muc(cid, loai, body.get("rows", []))
    return {"ok": True, "so_dong": n}


@app.post("/api/danh-muc-hang-export/{cid}")
async def danh_muc_hang_export(cid: int, request: Request, loai: str = "hh"):
    """Kết xuất Excel Danh mục Hàng hóa/NVL từ lưới DM (đồng thời LƯU lại)."""
    body = await request.json()
    dm_rows = body.get("rows", [])
    _luu_danh_muc(cid, loai, dm_rows)
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    ten = {"hh": "DMHH", "nvl": "DMNVL", "tscd": "DMTSCD", "ccdc": "DMCCDC"}.get(loai, "DM")
    sheet = {"hh": "DMHH", "nvl": "DMNVL", "tscd": "DMTSCD", "ccdc": "DMCCDC"}.get(loai, "DM")
    fname = f"DanhMuc_{ten}_{mst}.xlsx"
    rows_clean = [list(r) for r in dm_rows if any(str(x).strip() for x in r)]
    path = _xuat_dm_excel(rows_clean, sheet, fname, cid, loai)
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(rows_clean))})


VTHH_HEADERS = [
    "Mã (*)", "Tên (*)", "Tính chất", "Đơn vị tính chính", "Tồn tối thiểu",
    "Nhóm VTHH", "Kho ngầm định", "TK kho", "TK doanh thu", "TK chi phí",
    "Tỷ lệ CKMH (%)", "Đơn giá mua gần nhất", "Đơn giá bán ", "Thuế suất GTGT",
    "Loại tiền", "Đơn vị tính", "Đơn giá mua cố định", "Đơn vị chuyển đổi",
    "Tỷ lệ chuyển đổi", "Toán tử", "Mã nguyên vật liệu", "Tên nguyên vật liệu",
    "Đơn vị tính NVL", "Số lượng", "Đặc tính",
    "Theo dõi vật tư hàng hóa theo mã quy cách", "Mã quy cách", "Tên quy cách",
    "Cho phép trùng"]

def _gen_vthh_from_grid(cid, header, rows):
    """Danh mục VTHH từ lưới Nhập Liệu ĐANG MỞ: lấy mọi dòng Nợ 1561/156
    (Hàng hóa) và 152 (NVL), gán mã theo bảng đã học (nhất quán với DM Hàng
    hóa/NVL), lọc trùng theo mã. Trả [Mã, Tên, Tính chất=0, ĐVT]."""
    data = _doc_du_lieu_cty(cid)
    maps = {"hh": dict(data.get("dm_hh", {}).get("map", {})),
            "nvl": dict(data.get("dm_nvl", {}).get("map", {}))}
    nexts = {"hh": int(data.get("dm_hh", {}).get("next", 1)),
             "nvl": int(data.get("dm_nvl", {}).get("next", 1))}
    prefixes = {"hh": "HH", "nvl": "NVL"}
    col = _dm_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    seen_ma = set()
    out = []
    for r in rows:
        no = str(gv(r, col["no"]) or "").strip()
        if no in ("1561", "156"):
            grp = "hh"
        elif no == "152":
            grp = "nvl"
        else:
            continue
        ten = str(gv(r, col["ten"]) or "").strip()
        dvt = str(gv(r, col["dvt"]) or "").strip()
        if not ten:
            continue
        ky = "".join(ten.split()) + dvt
        if ky in maps[grp]:
            base = maps[grp][ky]
        else:
            base = prefixes[grp] + str(nexts[grp]).zfill(5)
            maps[grp][ky] = base
            nexts[grp] += 1
        rate = _dm_rate(gv(r, col["ts"]))
        ma = f"{base}-{rate}"            # mã đầy đủ kèm thuế suất (vd HH00001-8)
        if ma in seen_ma:
            continue                       # lọc trùng theo mã
        seen_ma.add(ma)
        out.append([ma, ten, "0", dvt])

    # Thêm TSCĐ (Nợ 2111/211) và CCDC (Nợ 2421/242): mỗi dòng 1 mã riêng
    for loai_ts in ("tscd", "ccdc"):
        ts_rows, _sm = _gen_danh_muc_ts(cid, loai_ts, header, rows)
        for tr in ts_rows:
            ma = tr[0]
            if ma in seen_ma:
                continue
            seen_ma.add(ma)
            out.append([ma, tr[1], "0", tr[2]])   # Mã, Tên, Tính chất=0, ĐVT

    out.sort(key=lambda x: x[0])
    return out


@app.post("/api/danh-muc-vthh/{cid}")
async def danh_muc_vthh(cid: int, request: Request, export: int = 0):
    """Danh mục VTHH từ dữ liệu đang xử lý trên lưới (Nợ 1561/156/152).
    export=1 -> trả file Excel."""
    body = await request.json()
    out = _gen_vthh_from_grid(cid, body.get("header", []), body.get("rows", []))
    if not export:
        return {"headers": VTHH_HEADERS, "rows": out, "so_dong": len(out)}
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh mục VTHH"
    for c, h in enumerate(VTHH_HEADERS, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    for r, row in enumerate(out, 2):
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c)
            cell.value = v
            if c in (1, 3):
                cell.number_format = "@"
    for c in range(1, len(VTHH_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"DanhMuc_VTHH_{mst}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============ MUA HÀNG NHẬP KHO (form MISA "Mua hàng NK") ============
MUA_NK_HEADERS = [
    "Hiển thị trên sổ", "Hình thức mua hàng", "Phương thức thanh toán",
    "Nhận kèm hóa đơn", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số phiếu nhập (*)", "Số chứng từ thanh toán", "Mẫu số HĐ", "Ký hiệu HĐ",
    "Số hóa đơn", "Ngày hóa đơn", "Mã nhà cung cấp", "Tên nhà cung cấp",
    "Người giao hàng", "Diễn giải", "NV mua hàng", "Loại tiền", "Tỷ giá",
    "Mã hàng (*)", "Tên hàng", "Kho", "Hàng hóa giữ hộ/bán hộ", "TK kho (*)",
    "TK công nợ/TK tiền (*)", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền",
    "Thành tiền quy đổi", "Tỷ lệ CK", "Tiền chiết khấu", "Tiền chiết khấu quy đổi",
    "Phí hàng về kho/Chi phí mua hàng", "% thuế GTGT", "Tiền thuế GTGT",
    "Tiền thuế GTGT quy đổi", "TKĐƯ thuế GTGT", "TK thuế GTGT",
    "Nhóm HHDV mua vào", "Phí trước hải quan", "Giá tính thuế NK", "% thuế NK",
    "Tiền thuế NK", "TK thuế NK", "% thuế TTĐB", "Tiền thuế TTĐB", "TK thuế TTĐB"]

def _nk_cols(header):
    hlow = [str(h or "").strip().lower() for h in header]
    def find(eqs, contains):
        for i, h in enumerate(hlow):
            if h in eqs:
                return i
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    return {
        "kh": find(("ký hiệu",), ("ký hiệu",)),
        "sohd": find(("số hđ",), ("số hđ", "số hoá đơn", "số hóa đơn")),
        "ngay": find(("ngày",), ("ngày",)),
        "nb": find(("người bán",), ("người bán",)),
        "mst": find(("mst bán", "mst"), ("mst",)),
        "ten": find(("tên hàng hóa/dịch vụ",), ("tên hàng",)),
        "dvt": find(("đvt",), ("đvt", "đơn vị tính")),
        "sl": find(("số lượng",), ("số lượng",)),
        "dgia": find(("đơn giá",), ("đơn giá",)),
        "tt": find(("thành tiền",), ("thành tiền",)),
        "ts": find(("thuế suất",), ("thuế suất",)),
        "tthue": find(("tiền thuế gtgt",), ("tiền thuế",)),
        "no": find(("nợ",), ()),
        "co": find(("có",), ()),
        "nk_tg": find(("trị giá tính thuế nk",), ("trị giá tính thuế nk",)),
        "nk_ts": find(("thuế suất nk",), ("thuế suất nk",)),
        "nk_thue": find(("tiền thuế nk",), ("tiền thuế nk",)),
    }

def _gen_mua_hang_nk(cid, header, rows):
    """Lọc dòng Nợ 1561/156 (Hàng hóa, kho HH) và 152 (NVL, kho NVL) -> form
    'Mua hàng NK'. Mã hàng lấy theo DM Hàng hóa/NVL (base theo Ký tự + thuế)."""
    data = _doc_du_lieu_cty(cid)
    maps = {"hh": dict(data.get("dm_hh", {}).get("map", {})),
            "nvl": dict(data.get("dm_nvl", {}).get("map", {}))}
    nexts = {"hh": int(data.get("dm_hh", {}).get("next", 1)),
             "nvl": int(data.get("dm_nvl", {}).get("next", 1))}
    prefixes = {"hh": "HH", "nvl": "NVL"}
    kho_ten = {"hh": "HH", "nvl": "NVL"}
    col = _nk_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    so_phieu_seen, so_phieu_cache = {}, {}
    out = []
    for r in rows:
        no = str(gv(r, col["no"]) or "").strip()
        if no in ("1561", "156"):
            grp = "hh"
        elif no == "152":
            grp = "nvl"
        else:
            continue
        ten = str(gv(r, col["ten"]) or "").strip()
        dvt = str(gv(r, col["dvt"]) or "").strip()
        if not ten:
            continue
        ky = "".join(ten.split()) + dvt
        if ky in maps[grp]:
            base = maps[grp][ky]
        else:
            base = prefixes[grp] + str(nexts[grp]).zfill(5)
            maps[grp][ky] = base
            nexts[grp] += 1
        rate = _dm_rate(gv(r, col["ts"]))
        ma = f"{base}-{rate}"
        kyhieu = str(gv(r, col["kh"]) or "").strip()
        la_nk = kyhieu.upper() == "TKNK"        # tờ khai nhập khẩu
        sohd = str(gv(r, col["sohd"]) or "").strip()
        mst = _dinh_dang_mst(gv(r, col["mst"])) if not la_nk else str(gv(r, col["mst"]) or "").strip()
        ngay = str(gv(r, col["ngay"]) or "")
        mst_raw = gv(r, col["mst"])
        # cùng 1 hóa đơn (số HĐ+MST+ngày) -> dùng lại đúng 1 số phiếu nhập
        so_phieu = _so_ct_unique_memo("NK", ngay, mst_raw, (sohd, mst_raw, ngay),
                                       so_phieu_seen, so_phieu_cache)
        nk_tg = _to_num(gv(r, col["nk_tg"])) or 0
        nk_ts = _so_pct(gv(r, col["nk_ts"])) if col["nk_ts"] >= 0 else 0  # "3%"->3
        nk_thue = _to_num(gv(r, col["nk_thue"])) or 0
        row = [""] * len(MUA_NK_HEADERS)
        row[0] = 0                               # Hiển thị trên sổ
        row[1] = 1 if la_nk else 0               # Hình thức mua hàng (1=NK)
        row[2] = 0                               # Phương thức thanh toán
        row[3] = 1                               # Nhận kèm hóa đơn
        row[4] = ngay; row[5] = ngay             # Ngày HT / Ngày CT
        row[6] = so_phieu                        # Số phiếu nhập
        row[10] = sohd                           # Số hóa đơn
        row[11] = ngay                           # Ngày hóa đơn
        row[12] = mst                            # Mã NCC
        row[13] = str(gv(r, col["nb"]) or "")    # Tên NCC
        row[15] = ten                            # Diễn giải
        row[19] = ma                             # Mã hàng
        row[20] = ten                            # Tên hàng
        row[21] = kho_ten[grp]                   # Kho
        row[23] = no                             # TK kho
        row[24] = str(gv(r, col["co"]) or "")    # TK công nợ/tiền
        row[25] = dvt                            # ĐVT
        row[26] = _to_num(gv(r, col["sl"]))      # Số lượng
        row[27] = _to_num(gv(r, col["dgia"]))    # Đơn giá
        row[28] = _to_num(gv(r, col["tt"]))      # Thành tiền
        row[34] = rate                           # % thuế GTGT
        row[35] = _to_num(gv(r, col["tthue"]))   # Tiền thuế GTGT
        if la_nk:
            row[37] = 1331                        # TKĐƯ thuế GTGT
            row[38] = 33312                       # TK thuế GTGT (nhập khẩu)
        else:
            row[38] = 1331                        # TK thuế GTGT (trong nước)
        row[39] = 1                              # Nhóm HHDV mua vào
        if la_nk:                                # chỉ điền thuế NK khi ĐÚNG là tờ khai NK
            row[41] = nk_tg                      # Giá tính thuế NK
            row[42] = nk_ts                      # % thuế NK (số)
            if nk_thue:
                row[43] = round(nk_thue)         # Tiền thuế NK
                row[44] = "3333"                 # TK thuế NK
        out.append(row)
    return out


@app.post("/api/mua-hang-nk/{cid}")
async def mua_hang_nhap_kho(cid: int, request: Request, export: int = 0):
    """Mua hàng nhập kho từ lưới Nhập Liệu (Nợ 1561/156/152). export=1 -> file."""
    body = await request.json()
    out = _gen_mua_hang_nk(cid, body.get("header", []), body.get("rows", []))
    if not export:
        return {"headers": MUA_NK_HEADERS, "rows": out, "so_dong": len(out)}
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mua hàng NK"
    for c, h in enumerate(MUA_NK_HEADERS, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    cot_text = {7, 9, 10, 11, 13, 20, 22, 24, 25, 38, 39, 45}  # số phiếu, sốHĐ, MST, mã, kho, TK...
    cot_tien = {27, 28, 29, 36, 42, 44}                        # SL, đơn giá, thành tiền, tiền thuế, NK
    for ri, row in enumerate(out, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(MUA_NK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"MuaHangNhapKho_{mst}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============ MUA HÀNG KHÔNG QUA KHO (TSCĐ/CCDC - form MISA) ============
MUA_KQK_HEADERS = [
    "Hiển thị trên sổ", "Hình thức mua hàng", "Phương thức thanh toán",
    "Nhận kèm hóa đơn", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số chứng từ (*)", "Mẫu số HĐ", "Ký hiệu HĐ", "Số hóa đơn", "Ngày hóa đơn",
    "Mã nhà cung cấp", "Tên nhà cung cấp", "Diễn giải", "NV mua hàng",
    "Loại tiền", "Tỷ giá", "Mã hàng (*)", "Tên hàng", "TK chi phí (*)",
    "TK công nợ/TK tiền (*)", "ĐVT", "Số lượng", "Đơn giá", "Thành tiền",
    "Thành tiền quy đổi", "Tỷ lệ CK", "Tiền chiết khấu", "Tiền chiết khấu quy đổi",
    "Chi phí mua hàng", "% thuế GTGT", "Tiền thuế GTGT", "Tiền thuế GTGT quy đổi",
    "TKĐƯ thuế GTGT", "TK thuế GTGT", "Nhóm HHDV mua vào", "Giá tính thuế NK",
    "% thuế NK", "Tiền thuế NK", "TK thuế NK", "% thuế TTĐB", "Tiền thuế TTĐB",
    "TK thuế TTĐB"]

def _ma_ts_map(cid, loai, header, rows):
    """Bản đồ rowkey(sốHĐ|ngày|tên|thành tiền) -> mã TSCĐ/CCDC (nhất quán DM)."""
    cat, _ = _gen_danh_muc_ts(cid, loai, header, rows)
    m = {}
    for row in cat:
        key = f"{row[6]}|{row[7]}|{row[1]}|{row[5]}"
        m[key] = row[0]
    return m

def _gen_mua_hang_kqk(cid, header, rows):
    """Mua hàng KHÔNG qua kho: lọc Nợ 2111/211 (TSCĐ) và 2421/242 (CCDC).
    Mã hàng lấy từ DM TSCĐ/CCDC. Số chứng từ = MHKQK1/<tháng>/<năm>."""
    map_ts = _ma_ts_map(cid, "tscd", header, rows)
    map_cc = _ma_ts_map(cid, "ccdc", header, rows)
    col = _nk_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    so_phieu_seen, so_phieu_cache = {}, {}
    out = []
    for r in rows:
        no = str(gv(r, col["no"]) or "").strip()
        if no in ("2111", "211"):
            mp = map_ts
        elif no in ("2421", "242"):
            mp = map_cc
        else:
            continue
        ten = str(gv(r, col["ten"]) or "").strip()
        if not ten:
            continue
        dvt = str(gv(r, col["dvt"]) or "").strip()
        sl = _to_num(gv(r, col["sl"]))
        dgia = _to_num(gv(r, col["dgia"]))
        tt = _to_num(gv(r, col["tt"]))
        sohd = str(gv(r, col["sohd"]) or "")
        ngay = str(gv(r, col["ngay"]) or "")
        co = str(gv(r, col["co"]) or "").strip()
        kyhieu = str(gv(r, col["kh"]) or "").strip()
        la_nk = kyhieu.upper() == "TKNK"
        mst = _dinh_dang_mst(gv(r, col["mst"])) if not la_nk else str(gv(r, col["mst"]) or "").strip()
        ma = mp.get(f"{sohd}|{ngay}|{ten}|{tt}", "")
        # Số chứng từ: DUY NHẤT theo từng hóa đơn (số HĐ+MST+ngày) — trước đây
        # gộp theo THÁNG (MHKQK1/<tháng>/<năm>) khiến nhiều hóa đơn khác nhau
        # trong cùng 1 tháng bị dồn chung vào 1 chứng từ. Các dòng cùng 1 hóa
        # đơn vẫn dùng chung 1 số (nhớ qua so_phieu_cache).
        mst_raw = gv(r, col["mst"])
        so_ct = _so_ct_unique_memo("MHKQK", ngay, mst_raw, (sohd, mst_raw, ngay),
                                    so_phieu_seen, so_phieu_cache)
        rate = _dm_rate(gv(r, col["ts"]))
        nk_tg = _to_num(gv(r, col["nk_tg"])) or 0
        nk_ts = _so_pct(gv(r, col["nk_ts"])) if col["nk_ts"] >= 0 else 0
        nk_thue = _to_num(gv(r, col["nk_thue"])) or 0
        row_o = [""] * len(MUA_KQK_HEADERS)
        row_o[0] = 0
        row_o[1] = 1 if la_nk else 0                       # Hình thức mua hàng
        row_o[2] = 1 if co.startswith("11") else 0         # PTTT: tiền (11xx)->1
        row_o[3] = 1
        row_o[4] = ngay; row_o[5] = ngay
        row_o[6] = so_ct
        row_o[9] = sohd; row_o[10] = ngay
        row_o[11] = mst; row_o[12] = str(gv(r, col["nb"]) or "")
        row_o[13] = ten
        row_o[17] = ma; row_o[18] = ten
        row_o[19] = no                                      # TK chi phí = Nợ
        row_o[20] = co                                      # TK công nợ/tiền = Có
        row_o[21] = dvt; row_o[22] = sl; row_o[23] = dgia; row_o[24] = tt
        row_o[30] = rate; row_o[31] = _to_num(gv(r, col["tthue"]))
        if la_nk:
            row_o[33] = 1331; row_o[34] = 33312
        else:
            row_o[34] = 1331
        row_o[35] = 1
        if la_nk:                                # chỉ điền thuế NK khi ĐÚNG là tờ khai NK
            row_o[36] = nk_tg; row_o[37] = nk_ts
            if nk_thue:
                row_o[38] = round(nk_thue); row_o[39] = "3333"
        out.append(row_o)
    return out


@app.post("/api/mua-hang-kqk/{cid}")
async def mua_hang_khong_qua_kho(cid: int, request: Request, export: int = 0):
    """Mua hàng không qua kho (Nợ 2111/211/2421/242). export=1 -> file."""
    body = await request.json()
    out = _gen_mua_hang_kqk(cid, body.get("header", []), body.get("rows", []))
    if not export:
        return {"headers": MUA_KQK_HEADERS, "rows": out, "so_dong": len(out)}
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mua hàng không qua Kho"
    for c, h in enumerate(MUA_KQK_HEADERS, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    cot_text = {7, 8, 9, 10, 12, 18, 20, 21, 35, 40}   # số CT, sốHĐ, MST, mã, TK
    cot_tien = {23, 24, 25, 32, 37, 39}                # SL, đơn giá, thành tiền, thuế, NK
    for ri, row in enumerate(out, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(MUA_KQK_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"MuaHangKhongQuaKho_{mst}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============ GHI TĂNG CCDC / TSCĐ (form MISA, từ DM CCDC/TSCĐ) ============
GHITANG_CCDC_HEADERS = [
    "Mã CCDC (*)", "Tên CCDC (*)", "Mã loại CCDC", "Lý do ghi tăng",
    "Số chứng từ ghi tăng", "Ngày ghi tăng (*)", "TK chờ phân bổ",
    "Đơn vị tính", "Số lượng", "Đơn giá", "Thành tiền",
    "Tổng số kỳ PB (tháng)", "Số tiền phân bổ hàng kỳ", "Mã đơn vị sử dụng (*)",
    "Số lượng đơn vị  sử dụng", "Đối tượng phân bổ", "Tỷ lệ phân bổ",
    "TK chi phí", "Khoản mục chi phí"]

GHITANG_TSCD_HEADERS = [
    "Mã tài sản (*)", "Tên tài sản (*)", "Loại tài sản (*)", "Đơn vị sử dụng (*)",
    "Số chứng từ ghi tăng", "Ngày ghi tăng (*)", "Ngày bắt đầu tính KH",
    "TK nguyên giá (*)", "TK khấu hao (*)", "Nguyên giá", "Giá trị tính khấu hao",
    "ĐVT thời gian SD", "Thời gian SD", "Tỷ lệ tính KH tháng (%)",
    "Giá trị KH tháng", "Hao mòn lũy kế", "Đối tượng phân bổ", "Tỷ lệ phân bổ",
    "TK chi phí", "Khoản mục chi phí", "Mã thống kê"]

def _doc_nhap_lieu(cid, loai="in"):
    conn = db()
    try:
        r = conn.execute("SELECT header_json, rows_json FROM nhap_lieu "
                         "WHERE company_id=? AND loai=?", (cid, loai)).fetchone()
    except Exception:
        r = None
    conn.close()
    if not r:
        return [], []
    try:
        return json.loads(r["header_json"]), json.loads(r["rows_json"])
    except Exception:
        return [], []

def _gen_ghi_tang_ccdc(cid, header, rows):
    """Ghi tăng CCDC từ DM CCDC. Lý do ghi tăng = số HĐ; TK chờ phân bổ = 242;
    các cột khác từ DM CCDC. Số tiền PB hàng kỳ = Thành tiền / Tổng số kỳ.
    Số chứng từ ghi tăng = MÃ CCDC (mỗi CCDC 1 số riêng, tăng dần) — đúng quy
    ước MISA (bản ghi nhập tay trên MISA: số chứng từ = mã, vd GTTS00001) và
    tránh lỗi trùng số chứng từ khi ghi sổ."""
    cat, _ = _gen_danh_muc_ts(cid, "ccdc", header, rows)
    out = []
    for d in cat:
        d = (list(d) + [""] * 12)[:12]
        ma, ten, dvt, sl, dgia, tt, sohd, ngay, hsd, dtpb, ngay_ght, tkcp = d
        tt_n = _to_num(tt) or 0
        hsd_n = _to_num(hsd) or 0
        m_pb = round(tt_n / hsd_n) if hsd_n else ""
        ngct = ngay_ght or ngay
        out.append([ma, ten, "", sohd, ma, ngct, 242, dvt, sl, dgia, tt,
                    hsd, m_pb, dtpb, 1, dtpb, 100, tkcp, ""])
    return out

def _gen_ghi_tang_tscd(cid, header, rows):
    """Ghi tăng TSCĐ từ DM TSCĐ. Loại TS=12; TK nguyên giá 2111; TK khấu hao
    2141; Nguyên giá=Giá trị tính KH=thành tiền; ĐVT thời gian SD=0; Tỷ lệ KH
    tháng=100; Thời gian SD=Hạn SD; Giá trị KH tháng=Giá trị KH/Thời gian SD.
    Số chứng từ ghi tăng = MÃ TÀI SẢN (mỗi tài sản 1 số riêng, tăng dần) —
    đúng quy ước MISA và tránh lỗi trùng số chứng từ khi ghi sổ."""
    cat, _ = _gen_danh_muc_ts(cid, "tscd", header, rows)
    out = []
    for d in cat:
        d = (list(d) + [""] * 12)[:12]
        ma, ten, dvt, sl, dgia, tt, sohd, ngay, hsd, dtpb, ngay_ght, tkcp = d
        tt_n = _to_num(tt) or 0
        hsd_n = _to_num(hsd) or 0
        kh_thang = round(tt_n / hsd_n) if hsd_n else ""
        ngct = ngay_ght or ngay
        out.append([ma, ten, 12, dtpb, ma, ngct, ngct, 2111, 2141, tt, tt,
                    0, hsd, 100, kh_thang, "", dtpb, 100, tkcp, "", ""])
    return out

def _xuat_ghitang_excel(headers, rows, sheet, fname, cid, cot_tien, cot_text):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    for ri, row in enumerate(rows, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return path


@app.post("/api/ghi-tang-ccdc/{cid}")
def ghi_tang_ccdc(cid: int):
    header, rows = _doc_nhap_lieu(cid, "in")
    out = _gen_ghi_tang_ccdc(cid, header, rows)
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"GhiTangCCDC_{mst}.xlsx"
    # cột số: SL(9) Đơn giá(10) Thành tiền(11) Số kỳ PB(12) Tiền PB/kỳ(13)
    path = _xuat_ghitang_excel(GHITANG_CCDC_HEADERS, out, "Ghi Tăng CCDC", fname,
                               cid, {9, 10, 11, 12, 13}, {1, 4, 5, 7, 18})
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


@app.post("/api/ghi-tang-tscd/{cid}")
def ghi_tang_tscd(cid: int):
    header, rows = _doc_nhap_lieu(cid, "in")
    out = _gen_ghi_tang_tscd(cid, header, rows)
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"GhiTangTSCD_{mst}.xlsx"
    # cột số: Nguyên giá(10) Giá trị KH(11) Thời gian SD(13) Giá trị KH tháng(15)
    path = _xuat_ghitang_excel(GHITANG_TSCD_HEADERS, out, "Ghi tăng tài sản cố định",
                               fname, cid, {10, 11, 15}, {1, 5, 6, 8, 9, 19})
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============ BÁN HÀNG (form MISA "Chứng từ bán hàng") ============
BAN_HANG_HEADERS = [
    "Hiển thị trên sổ", "Hình thức bán hàng", "Phương thức thanh toán",
    "Kiêm phiếu xuất kho", "Lập kèm hóa đơn", "Đã lập hóa đơn",
    "Ngày hạch toán (*)", "Ngày chứng từ (*)", "Số chứng từ (*)",
    "Số phiếu xuất", "Lý do xuất", "Mẫu số HĐ", "Ký hiệu HĐ", "Số hóa đơn",
    "Ngày hóa đơn", "Mã khách hàng", "Tên khách hàng", "Địa chỉ", "Mã số thuế",
    "Diễn giải", "Nộp vào TK", "NV bán hàng", "Loại tiền", "Tỷ giá",
    "Mã hàng (*)", "Tên hàng", "Hàng khuyến mại", "TK Tiền/Chi phí/Nợ (*)",
    "TK Doanh thu/Có (*)", "ĐVT", "Số lượng", "Đơn giá sau thuế", "Đơn giá",
    "Thành tiền", "Thành tiền quy đổi", "Tỷ lệ CK (%)", "Tiền chiết khấu",
    "Tiền chiết khấu quy đổi", "TK chiết khấu", "Giá tính thuế XK", "% thuế XK",
    "Tiền thuế XK", "TK thuế XK", "% thuế GTGT", "Tiền thuế GTGT",
    "Tiền thuế GTGT quy đổi", "TK thuế GTGT", "HH không TH trên tờ khai thuế GTGT",
    "Kho", "TK giá vốn", "TK Kho", "Đơn giá vốn", "Tiền vốn",
    "Hàng hóa giữ hộ/bán hộ"]

def _bh_cols(header):
    hlow = [str(h or "").strip().lower() for h in header]
    def find(eqs, contains):
        for i, h in enumerate(hlow):
            if h in eqs:
                return i
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    return {
        "maus": find(("ký hiệu mẫu", "mẫu số hđ"), ("ký hiệu mẫu", "mẫu số")),
        "kyhieu": find(("ký hiệu hđ",), ("ký hiệu hđ",)),
        "sohd": find(("số hóa đơn", "số hđ"), ("số hóa đơn", "số hđ")),
        "ngay": find(("ngày lập", "ngày"), ("ngày",)),
        "nguoimua": find(("tên người mua",), ("người mua", "tên người")),
        "mst": find(("mst người mua", "mst mua"), ("mst",)),
        "mathang": find(("mặt hàng",), ("mặt hàng", "tên hàng")),
        "ds": find(("doanh số bán chưa thuế",), ("doanh số", "chưa thuế")),
        "thue": find(("thuế gtgt",), ("thuế gtgt", "tiền thuế")),
    }

def _gen_ban_hang(cid, header, rows):
    """Form MISA 'Chứng từ bán hàng' từ Bảng kê đầu ra. TK Có (doanh thu)=5111;
    TK Nợ (tiền/công nợ) = 131 nếu tổng HĐ >= 5tr, ngược lại 1111."""
    col = _bh_cols(header)

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    # tổng theo từng hóa đơn (ký hiệu + số HĐ) để áp ngưỡng 5tr
    tong = {}
    for r in rows:
        key = f"{gv(r,col['kyhieu'])}|{gv(r,col['sohd'])}"
        t = (_to_num(gv(r, col["ds"])) or 0) + (_to_num(gv(r, col["thue"])) or 0)
        tong[key] = tong.get(key, 0) + t

    seq_thang = {}
    out = []
    for r in rows:
        sohd = str(gv(r, col["sohd"]) or "").strip()
        if not sohd:
            continue
        ds = round(_to_num(gv(r, col["ds"])) or 0)
        thue = round(_to_num(gv(r, col["thue"])) or 0)
        ngay = str(gv(r, col["ngay"]) or "")
        mathang = str(gv(r, col["mathang"]) or "")
        nguoimua = str(gv(r, col["nguoimua"]) or "")
        mst = _dinh_dang_mst(gv(r, col["mst"]))
        kyhieu = str(gv(r, col["kyhieu"]) or "")
        maus = str(gv(r, col["maus"]) or "") if col["maus"] >= 0 else ""
        tong_hd = tong.get(f"{kyhieu}|{sohd}", 0)
        tk_no = 131 if abs(tong_hd) >= NGUONG_5TR else 1111
        # Số chứng từ: BH{seq:03d}/T{tháng}/{năm} (seq theo từng tháng)
        p = ngay.replace("-", "/").split("/")
        thang = nam = ""
        if len(p) == 3:
            thang = str(int(p[1])) if p[1].isdigit() else p[1]
            nam = p[2]
        mk = (thang, nam)
        seq_thang[mk] = seq_thang.get(mk, 0) + 1
        so_ct = f"BH{seq_thang[mk]:03d}/T{thang}/{nam}"[:20]
        rate = _chuan_thue_suat((thue / ds * 100) if ds else 0)
        row = [""] * len(BAN_HANG_HEADERS)
        row[0] = 0; row[1] = 0
        row[2] = 1 if abs(tong_hd) < NGUONG_5TR else 0   # PTTT: <5tr=tiền->1
        row[3] = 0; row[4] = 1; row[5] = 1
        row[6] = ngay; row[7] = ngay; row[8] = so_ct
        row[11] = maus; row[12] = kyhieu               # Mẫu số / Ký hiệu HĐ
        row[13] = sohd; row[14] = ngay
        row[15] = mst; row[16] = nguoimua; row[18] = mst
        row[19] = mathang
        row[24] = "BH"; row[25] = mathang
        row[27] = tk_no; row[28] = 5111                # AB Nợ / AC Có
        row[30] = 1                                     # Số lượng
        row[31] = ds; row[32] = ds; row[33] = ds       # đơn giá / thành tiền
        row[43] = rate; row[44] = thue                 # % thuế / tiền thuế
        row[46] = "33311"                              # TK thuế GTGT
        out.append(row)
    return out


@app.post("/api/ban-hang/{cid}")
async def ban_hang(cid: int, request: Request, export: int = 0):
    """Bán hàng từ lưới Bảng kê đầu ra. export=1 -> file MISA."""
    body = await request.json()
    out = _gen_ban_hang(cid, body.get("header", []), body.get("rows", []))
    if not export:
        return {"headers": BAN_HANG_HEADERS, "rows": out, "so_dong": len(out)}
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chứng từ bán hàng"
    for c, h in enumerate(BAN_HANG_HEADERS, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    cot_text = {9, 13, 14, 16, 19, 25, 28, 29, 47}   # số CT, ký hiệu, sốHĐ, MST, mã, TK
    cot_tien = {32, 33, 34, 45}                       # đơn giá, thành tiền, tiền thuế
    for ri, row in enumerate(out, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(BAN_HANG_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"BanHang_{mst}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(out))})


# ============================================================
#  XUẤT KHO (form MISA "Xuất kho") — dò mã hàng từ TỒN KHO cho
#  từng dòng bán ra (Chi tiết BÁN RA), có trừ tồn tuần tự để không
#  gán quá số hàng thực còn trong kho.
#
#  Quy trình: 1) import file "Tổng hợp tồn kho" (MISA) -> Sheet TON
#             2) import file có sheet "Chi tiết BÁN RA" -> nguồn GIATHANH
#             3) dò mã tự động (đúng tên / gần đúng tên, có trừ tồn)
#             4) người dùng sửa tay các dòng chưa dò được (gợi ý mã+tên+giá)
#             5) xuất file "Xuất kho" (form MISA) từ GIATHANH đã gắn mã
# ============================================================
# điểm giống tối thiểu (0..1) để tự động gắn mã khi KHÔNG có ứng viên "mạnh"
# (trùng y hệt hoặc 1 chuỗi nằm trong chuỗi kia) — trường hợp hiếm, đòi hỏi rất giống.
_XK_NGUONG_FUZZY = 0.95

def _chuan_ten_hang_xk(s):
    """Chuẩn hoá tên hàng để so khớp: bỏ dấu, viết hoa, bỏ nội dung trong
    ngoặc (vd '(N)' — chỉ là ký hiệu phụ/nguồn hàng, không phải tên khác
    mặt hàng), rồi bỏ mọi ký tự không phải chữ/số."""
    import re as _re_xk
    s = _re_xk.sub(r'\([^)]*\)', '', _khong_dau(s).upper())
    return _re_xk.sub(r'[^A-Z0-9]', '', s)

def _diem_giong_ten_xk(a, b):
    """Điểm giống nhau 2 chuỗi ĐÃ chuẩn hoá (0..1), theo tỉ lệ ký tự khớp."""
    import difflib
    if not a or not b:
        return 0.0
    return difflib.SequenceMatcher(None, a, b).ratio()

def _manh_xk(a, b):
    """Ứng viên 'mạnh': tên ĐÃ chuẩn hoá trùng y hệt, hoặc 1 chuỗi nằm trong chuỗi kia
    (vd chỉ lệch hậu tố 'CM', dấu * hay -... — coi như CÙNG một mặt hàng).
    KHÔNG dùng cho các mặt hàng chỉ đơn thuần giống nhau về CHỮ (vd khác kích cỡ
    160/180 ở cùng 1 vị trí) — loại đó phải qua ngưỡng fuzzy riêng, chặt hơn."""
    if not a or not b:
        return False
    return a == b or a in b or b in a

def _kd2(s):
    """_khong_dau() rồi thay nốt 'đ'->'d' (NFD không tách được chữ Đ ra dấu riêng)."""
    return _khong_dau(s).replace("đ", "d")

def _doc_file_ton_kho(wb):
    """Đọc file 'Tổng hợp tồn kho' (báo cáo MISA, tiêu đề gộp ô 2 dòng: nhóm
    Đầu kỳ/Nhập kho/Xuất kho/Cuối kỳ ở dòng trên, Số lượng/Giá trị ở dòng dưới)
    -> list {ma,ten,dvt,ton,gia}. Dò tiêu đề theo CHỮ, không phụ thuộc cột."""
    ws = wb.worksheets[0]
    hdr_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        vals = [_kd2(ws.cell(r, c).value) for c in range(1, ws.max_column + 1)]
        if any("ma hang" in v for v in vals) and any("ten hang" in v for v in vals):
            hdr_row = r
            break
    if not hdr_row:
        return []
    ncol = ws.max_column
    top = [_kd2(ws.cell(hdr_row, c).value) for c in range(1, ncol + 1)]
    sub = [_kd2(ws.cell(hdr_row + 1, c).value) for c in range(1, ncol + 1)]
    top_ff, last = [], ""
    for v in top:                 # dò xuôi qua các ô đã gộp (merge) không có giá trị
        if v:
            last = v
        top_ff.append(last)

    def find_combo(*keys):
        for i in range(ncol):
            c = f"{top_ff[i]} {sub[i]}"
            if all(k in c for k in keys):
                return i
        return -1

    i_ma = next((i for i, v in enumerate(top) if "ma hang" in v), -1)
    i_ten = next((i for i, v in enumerate(top) if "ten hang" in v), -1)
    i_dvt = next((i for i, v in enumerate(top) if v == "dvt" or "don vi tinh" in v), -1)
    i_sl = find_combo("cuoi ky", "so luong")
    i_gt = find_combo("cuoi ky", "gia tri")
    if i_ma < 0 or i_ten < 0:
        return []
    gop = {}
    for r in range(hdr_row + 2, ws.max_row + 1):
        ma = str(ws.cell(r, i_ma + 1).value or "").strip()
        if not ma:               # bỏ dòng 'Tên kho : ...' / dòng tổng
            continue
        ten = str(ws.cell(r, i_ten + 1).value or "").strip()
        dvt = str(ws.cell(r, i_dvt + 1).value or "").strip() if i_dvt >= 0 else ""
        sl = _to_num(ws.cell(r, i_sl + 1).value) if i_sl >= 0 else 0
        gt = _to_num(ws.cell(r, i_gt + 1).value) if i_gt >= 0 else 0
        sl = sl if isinstance(sl, (int, float)) else 0
        gt = gt if isinstance(gt, (int, float)) else 0
        if ma in gop:
            gop[ma]["ton"] += sl
            gop[ma]["gt"] += gt
        else:
            gop[ma] = {"ma": ma, "ten": ten, "dvt": dvt, "ton": sl, "gt": gt}
    out = []
    for it in gop.values():
        gia = round(it["gt"] / it["ton"]) if it["ton"] else 0
        out.append({"ma": it["ma"], "ten": it["ten"], "dvt": it["dvt"],
                    "ton": it["ton"], "gia": gia})
    return out

def _xk_src_cols(header):
    """Tìm cột trong sheet nguồn 'Chi tiết BÁN RA'."""
    hlow = [str(h or "").strip().lower() for h in header]
    def find(eqs, contains):
        for i, h in enumerate(hlow):
            if h in eqs:
                return i
        for i, h in enumerate(hlow):
            if any(k in h for k in contains):
                return i
        return -1
    return {
        "kh": find(("ký hiệu",), ("ký hiệu",)),
        "sohd": find(("số hđ",), ("số hđ", "số hoá đơn", "số hóa đơn")),
        "ngay": find(("ngày",), ("ngày",)),
        "ten": find(("tên hàng hóa/dịch vụ",), ("tên hàng",)),
        "dvt": find(("đvt",), ("đvt", "đơn vị tính")),
        "sl": find(("số lượng",), ("số lượng",)),
        "dgia": find(("đơn giá",), ("đơn giá",)),
        "tt": find(("thành tiền",), ("thành tiền",)),
    }

def _xk_key_ngay(ngay):
    """Khoá sắp xếp theo ngày tăng dần, nhận dd/mm/yyyy hoặc yyyy-mm-dd."""
    s = str(ngay or "").strip().replace("-", "/")
    p = s.split("/")
    if len(p) == 3:
        if len(p[0]) == 4:                       # đã là yyyy/mm/dd
            return f"{p[0]}-{p[1].zfill(2)}-{p[2].zfill(2)}"
        return f"{p[2]}-{p[1].zfill(2)}-{p[0].zfill(2)}"
    return s

def _gen_xk_giathanh(ton_rows, src_header, src_rows, hoc_ma=None):
    """Ghép mã hàng (từ TON) cho từng dòng bán ra (Chi tiết BÁN RA), TRỪ TỒN
    tuần tự theo ngày tăng dần — mặt hàng nào hết tồn (<=0) sẽ KHÔNG gán nữa.

    Nhiều mặt hàng trong TON có TÊN TRÙNG NHAU nhưng khác mã (vd 1 mã do phần
    mềm sinh 'HH00033-8' và 1 mã đã có sẵn trong MISA 'MH90' cho CÙNG 1 sản
    phẩm) -> tự LẤY LUÔN mã còn đủ tồn, KHÔNG hỏi lại: ưu tiên mã đã được HỌC
    (hoc_ma: tên chuẩn hoá -> mã) từ lần người dùng tự gắn trước đó nếu vẫn đủ
    tồn; nếu không, ưu tiên mã XUẤT HIỆN TRƯỚC theo đúng thứ tự trong file tồn
    kho (ton_rows) — dùng HẾT mã đứng trước rồi mới bắt đầu dùng tới mã đứng
    sau, không ưu tiên theo tồn nhiều/ít. Chỉ TÁCH DÒNG (cộng dồn nhiều mã,
    cũng theo đúng thứ tự đó) khi không mã đơn nào đủ cả số lượng cần bán."""
    col = _xk_src_cols(src_header)
    hoc_ma = hoc_ma or {}

    def gv(r, i):
        return r[i] if 0 <= i < len(r) else ""

    ton_list = [dict(it, con_lai=_to_num(it.get("ton")) or 0,
                     ten_chuan=_chuan_ten_hang_xk(it.get("ten")))
                for it in (ton_rows or [])]

    items = []
    for r in (src_rows or []):
        ten = str(gv(r, col["ten"]) or "").strip()
        if not ten:
            continue
        items.append({
            "khhdon": str(gv(r, col["kh"]) or ""), "sohd": str(gv(r, col["sohd"]) or ""),
            "ngay": str(gv(r, col["ngay"]) or ""), "ten_sp": ten,
            "dvt": str(gv(r, col["dvt"]) or ""), "sl": _to_num(gv(r, col["sl"])) or 0,
            "dgia": _to_num(gv(r, col["dgia"])), "tt": _to_num(gv(r, col["tt"])),
        })
    items.sort(key=lambda it: _xk_key_ngay(it["ngay"]))

    def goi_y_cho(ten_chuan, manh):
        manh_ma = {tn["ma"] for tn in manh}
        scored = sorted(
            ({"ma": tn["ma"], "ten": tn["ten"], "dvt": tn["dvt"], "gia": tn["gia"],
              "con_lai": tn["con_lai"], "manh": tn["ma"] in manh_ma,
              "diem": round(_diem_giong_ten_xk(ten_chuan, tn["ten_chuan"]), 3)}
             for tn in ton_list), key=lambda x: (-x["manh"], -x["diem"]))
        return [s for s in scored[:6] if s["diem"] > 0.3 or s["manh"]]

    def dong_trong(it, sl_thieu, ten_chuan, manh):
        rec = dict(it, sl=sl_thieu)
        rec.update(ma="", ten_xk="", dvt_xk="", gia_xk="", mo_ho=True, thieu_ton=True,
                   goi_y=goi_y_cho(ten_chuan, manh))
        return rec

    out = []
    for it in items:
        ten_chuan = _chuan_ten_hang_xk(it["ten_sp"])
        sl_can = it["sl"] if isinstance(it["sl"], (int, float)) else 0
        tt_goc = it.get("tt")
        candidates = [tn for tn in ton_list if tn["con_lai"] > 0]
        manh = [tn for tn in candidates if _manh_xk(ten_chuan, tn["ten_chuan"])]
        pool = manh
        if not pool:                                 # không có ứng viên 'mạnh' -> fuzzy chặt
            best, best_diem = None, 0.0
            for tn in candidates:
                d = _diem_giong_ten_xk(ten_chuan, tn["ten_chuan"])
                if d > best_diem:
                    best, best_diem = tn, d
            if best and best_diem >= _XK_NGUONG_FUZZY:
                pool = [best]
        if not pool:                                  # không tìm thấy tên nào khớp
            out.append(dong_trong(it, sl_can, ten_chuan, manh))
            continue
        # Nhiều mã trùng/gần trùng tên (vd 1 mã do phần mềm sinh + 1 mã cũ có sẵn
        # trong MISA cho cùng sản phẩm) -> LẤY LUÔN mã còn đủ tồn cho số lượng
        # bán, KHÔNG hỏi lại — ưu tiên mã đã HỌC trước đó nếu vẫn đủ tồn, rồi đến
        # mã XUẤT HIỆN TRƯỚC theo đúng thứ tự trong file tồn kho (pool đã giữ
        # nguyên thứ tự đó) — dùng HẾT mã đứng trước rồi mới chuyển sang mã
        # đứng sau, không ưu tiên theo tồn nhiều/ít.
        ma_hoc = hoc_ma.get(ten_chuan)
        pick = next((tn for tn in pool if tn["ma"] == ma_hoc and tn["con_lai"] >= sl_can), None)
        if not pick:
            pick = next((tn for tn in pool if tn["con_lai"] >= sl_can), None)
        if pick:
            pick["con_lai"] -= sl_can
            rec = dict(it)
            rec.update(ma=pick["ma"], ten_xk=pick["ten"], dvt_xk=pick["dvt"],
                       gia_xk=pick["gia"], goi_y=[], mo_ho=False, thieu_ton=False)
            out.append(rec)
            continue
        # KHÔNG mã đơn nào đủ cả số lượng -> TÁCH DÒNG: cộng dồn nhiều mã lại,
        # LẤY THEO ĐÚNG THỨ TỰ xuất hiện trong file tồn kho (dùng hết mã đứng
        # trước mới lấy tới mã đứng sau) cho đến khi đủ số lượng bán; phần còn
        # thiếu (nếu tổng tồn cả nhóm vẫn không đủ) tách thành 1 dòng riêng để
        # trống, kèm gợi ý, cho người dùng tự gắn mã khác.
        can_lay = sl_can
        da_dung_tien = 0
        for tn in pool:
            if can_lay <= 0 or tn["con_lai"] <= 0:
                continue
            lay = min(tn["con_lai"], can_lay)
            tn["con_lai"] -= lay
            can_lay -= lay
            if isinstance(tt_goc, (int, float)) and sl_can:
                tt_phan = round(tt_goc - da_dung_tien) if can_lay <= 0 else round(tt_goc * lay / sl_can)
            else:
                tt_phan = tt_goc
            da_dung_tien += tt_phan if isinstance(tt_phan, (int, float)) else 0
            rec = dict(it, sl=lay, tt=tt_phan)
            rec.update(ma=tn["ma"], ten_xk=tn["ten"], dvt_xk=tn["dvt"], gia_xk=tn["gia"],
                       goi_y=[], mo_ho=False, thieu_ton=False)
            out.append(rec)
        if can_lay > 0:                                # vẫn còn thiếu sau khi gộp cả nhóm
            tt_thieu = round(tt_goc - da_dung_tien) if isinstance(tt_goc, (int, float)) else tt_goc
            rec = dict(it, sl=can_lay, tt=tt_thieu)
            rec.update(ma="", ten_xk="", dvt_xk="", gia_xk="", mo_ho=True, thieu_ton=True,
                       goi_y=goi_y_cho(ten_chuan, manh))
            out.append(rec)
    return out

XK_GIATHANH_XUAT_HEADERS = ["Tồn kho", "Số HĐ", "Ngày", "Tên Sản Phẩm", "ĐVT", "Số lượng",
    "Đơn giá", "Thành tiền", "Mã hàng kho", "Tên hàng xuất kho", "ĐVT kho", "SL kho",
    "Đơn Giá kho", "Thành Tiền kho"]

def _gen_giathanh_export_rows(ton_rows, giathanh_rows):
    """Mirror hệt các cột hiển thị trên lưới màn hình Xuất Kho (giống hàm
    xkRowsToNl ở JS): Tồn kho CHẠY DẦN theo mã (trừ dần từ trên xuống), SL/Đơn
    giá kho fallback về SL/-- bán khi dòng chưa gắn mã — để file xuất khớp
    đúng những gì người dùng thấy trên màn hình."""
    ton_map = {str(t.get("ma") or "").strip(): _to_num(t.get("ton")) or 0 for t in (ton_rows or [])}
    da_dung = {}
    out = []
    for r in (giathanh_rows or []):
        sl_kho = r.get("sl_kho")
        if sl_kho in (None, ""):
            sl_kho = r.get("sl")
        gia_xk = r.get("gia_xk")
        sl_n, gia_n = _to_num(sl_kho), _to_num(gia_xk)
        tt_kho = round(sl_n * gia_n) if isinstance(sl_n, (int, float)) and isinstance(gia_n, (int, float)) else ""
        ma = str(r.get("ma") or "").strip()
        ton = ""
        if ma and ma in ton_map:
            sl_ban = _to_num(r.get("sl"))
            da_dung[ma] = da_dung.get(ma, 0) + (sl_ban if isinstance(sl_ban, (int, float)) else 0)
            ton = ton_map[ma] - da_dung[ma]
        out.append([ton, r.get("sohd", ""), r.get("ngay", ""), r.get("ten_sp", ""), r.get("dvt", ""),
                    r.get("sl", ""), r.get("dgia", ""), r.get("tt", ""),
                    ma, r.get("ten_xk", ""), r.get("dvt_xk", ""), sl_kho, gia_xk, tt_kho])
    return out

def _xk_cuoi_thang(ngay):
    """'dd/mm/yyyy' của NGÀY CUỐI THÁNG chứa ngày truyền vào (dd/mm/yyyy hoặc yyyy-mm-dd)."""
    import calendar
    s = str(ngay or "").strip()
    d = m = y = None
    if "/" in s:
        p = s.split("/")
        if len(p) == 3:
            d, m, y = p
    elif "-" in s:
        p = s.split("-")
        if len(p) == 3:
            (y, m, d) = p if len(p[0]) == 4 else (p[2], p[1], p[0])
    if not (m and y):
        return "", "", ""
    try:
        mi, yi = int(m), int(y)
    except Exception:
        return "", "", ""
    last_day = calendar.monthrange(yi, mi)[1]
    return f"{last_day:02d}/{mi:02d}/{yi}", str(mi), str(yi)

XUAT_KHO_HEADERS = [
    "Hiển thị trên sổ", "Loại xuất kho", "Ngày hạch toán (*)", "Ngày chứng từ (*)",
    "Số chứng từ (*)", "Mẫu số HĐ", "Ký hiệu HĐ", "Mã đối tượng", "Tên đối tượng",
    "Địa chỉ/Bộ phận", "Tên người nhận/Của", "Lý do xuất/Về việc", "Nhân viên bán hàng",
    "Kèm theo", "Số lệnh điều động", "Ngày lệnh điều động", "Người vận chuyển",
    "Tên người vận chuyển", "Hợp đồng số", "Phương tiện vận chuyển", "Xuất tại kho",
    "Địa chỉ kho xuất", "Nhập tại chi nhánh", "Tên chi nhánh", "MST chi nhánh",
    "Nhập tại kho", "Địa chỉ kho nhập", "Mã hàng (*)", "Tên hàng", "Là hàng khuyến mại",
    "Kho (*)", "Hàng hóa giữ hộ/bán hộ", "TK Nợ (*)", "TK Có (*)", "ĐVT", "Số lượng",
    "Đơn giá bán", "Thành tiền", "Đơn giá vốn", "Tiền vốn", "Số lô", "Hạn sử dụng",
    "Đối tượng", "Khoản mục CP", "Đơn vị", "Đối tượng THCP", "Công trình",
    "Đơn đặt hàng", "Hợp đồng bán", "CP không hợp lý", "Mã thống kê"]

def _gen_xuat_kho_rows(giathanh_rows):
    """Từ các dòng GIATHANH đã gắn mã (bỏ dòng chưa gắn) -> mảng dòng form MISA
    'Xuất kho'. Ngày hạch toán/chứng từ = CUỐI THÁNG của hoá đơn cuối cùng (mới
    nhất) trong lô; Số chứng từ = 'XK T{tháng}/{năm}' theo tháng đó."""
    rows = [r for r in (giathanh_rows or []) if str(r.get("ma") or "").strip()]
    if not rows:
        return [], ""
    ngay_cuoi = max(rows, key=lambda r: _xk_key_ngay(r.get("ngay")))["ngay"]
    ngay_ht, thang, nam = _xk_cuoi_thang(ngay_cuoi)
    so_ct = f"XK T{thang}/{nam}"
    out = []
    for r in rows:
        row = [""] * len(XUAT_KHO_HEADERS)
        row[0] = 0
        row[2] = ngay_ht; row[3] = ngay_ht; row[4] = so_ct
        row[27] = r.get("ma", "")                 # AB Mã hàng
        row[28] = r.get("ten_xk") or r.get("ten_sp", "")   # AC Tên hàng
        row[30] = "HH"                             # AE Kho
        row[32] = "632"; row[33] = "1561"          # AG Nợ / AH Có
        row[34] = r.get("dvt_xk") or r.get("dvt", "")      # AI ĐVT
        sl_kho = r.get("sl_kho")
        row[35] = sl_kho if sl_kho not in (None, "", 0) else r.get("sl", "")  # AJ Số lượng
        out.append(row)
    return out, so_ct

@app.post("/api/xk/import-ton/{cid}")
async def xk_import_ton(cid: int, request: Request):
    """Import file 'Tổng hợp tồn kho' (báo cáo MISA) -> lưu làm Sheet TON."""
    import openpyxl, io as _io
    form = await request.form()
    files = form.getlist("files") or ([form.get("file")] if form.get("file") else [])
    if not files:
        raise HTTPException(400, "Chưa chọn file")
    gop = {}
    so_file_ok, loi = 0, []
    for up in files:
        if up is None:
            continue
        fn = getattr(up, "filename", "file")
        try:
            content = await up.read()
            wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
        except Exception as e:
            loi.append(f"{fn}: không đọc được ({e})"); continue
        rows = _doc_file_ton_kho(wb)
        if not rows:
            loi.append(f"{fn}: không thấy cột 'Mã hàng'/'Tên hàng' (không đúng mẫu 'Tổng hợp tồn kho')")
            continue
        for it in rows:
            gop[it["ma"]] = it            # file sau đè file trước (mới nhất thắng)
        so_file_ok += 1
    if not gop:
        raise HTTPException(400, "Không đọc được dữ liệu tồn kho từ file đã chọn. " + "; ".join(loi[:3]))
    ton_rows = list(gop.values())
    data = _doc_du_lieu_cty(cid)
    data["xk_ton"] = ton_rows
    _ghi_du_lieu_cty(cid, data)
    return {"ok": True, "so_file": so_file_ok, "so_dong": len(ton_rows), "loi": loi[:5]}

@app.get("/api/xk/ton/{cid}")
def xk_get_ton(cid: int):
    data = _doc_du_lieu_cty(cid)
    rows = data.get("xk_ton") or []
    return {"rows": rows, "so_dong": len(rows)}

@app.get("/api/xk/banra/{cid}")
def xk_get_banra(cid: int):
    """Số dòng 'Chi tiết BÁN RA' đã có sẵn từ dữ liệu Nhập Liệu (không cần
    import riêng — chỉ cần import & lưu ở màn Nhập Liệu như bình thường)."""
    src = nhap_lieu_get(cid, "ctbr")
    return {"so_dong": len(src.get("rows") or []), "updated_at": src.get("updated_at", "")}

@app.post("/api/xk/tao-giathanh/{cid}")
def xk_tao_giathanh(cid: int):
    """Dò mã hàng tự động cho từng dòng bán ra, dựa theo TON đã import và
    dữ liệu 'Chi tiết BÁN RA' đã lưu sẵn từ màn Nhập Liệu (Import & tách dữ
    liệu -> Lưu cả 2 bảng kê). Áp lại các lựa chọn đã HỌC (từ lần người dùng
    tự gắn tay trước đó) cho những tên hàng có nhiều mã trùng tên (vd 1 mã do
    phần mềm sinh + 1 mã cũ đã có sẵn trong MISA)."""
    data = _doc_du_lieu_cty(cid)
    ton_rows = data.get("xk_ton") or []
    src = nhap_lieu_get(cid, "ctbr")
    if not ton_rows:
        raise HTTPException(400, "Chưa import Sheet TON (Tổng hợp tồn kho)")
    if not src.get("rows"):
        raise HTTPException(400, "Chưa có dữ liệu 'Chi tiết BÁN RA' — vào màn Nhập Liệu, Import & tách dữ liệu rồi Lưu cả 2 bảng kê từ file có sheet 'Chi tiết BÁN RA'")
    hoc_ma = data.get("xk_hoc_ma") or {}
    giathanh = _gen_xk_giathanh(ton_rows, src.get("header") or [], src.get("rows") or [], hoc_ma)
    data["xk_giathanh"] = giathanh
    _ghi_du_lieu_cty(cid, data)
    so_khop = sum(1 for r in giathanh if r.get("ma"))
    return {"rows": giathanh, "so_dong": len(giathanh), "so_khop": so_khop,
            "so_chua_khop": len(giathanh) - so_khop}

@app.get("/api/xk/giathanh/{cid}")
def xk_get_giathanh(cid: int):
    data = _doc_du_lieu_cty(cid)
    rows = data.get("xk_giathanh") or []
    so_khop = sum(1 for r in rows if r.get("ma"))
    return {"rows": rows, "so_dong": len(rows), "so_khop": so_khop,
            "so_chua_khop": len(rows) - so_khop}

@app.post("/api/xk/giathanh-luu/{cid}")
async def xk_luu_giathanh(cid: int, request: Request):
    """Lưu lại GIATHANH sau khi người dùng sửa tay các dòng chưa dò được mã.
    Đồng thời HỌC lại lựa chọn cho các tên hàng từng bị coi là mơ hồ (nhiều mã
    trùng tên), để lần dò mã tự động sau tự áp đúng, không phải hỏi lại."""
    body = await request.json()
    rows = body.get("rows") or []
    data = _doc_du_lieu_cty(cid)
    ton_map = {str(it.get("ma") or "").strip(): it for it in (data.get("xk_ton") or [])}
    hoc_ma = dict(data.get("xk_hoc_ma") or {})
    for r in rows:
        ma = str(r.get("ma") or "").strip()
        if not ma:
            continue
        tn = ton_map.get(ma)
        if tn:                          # bổ sung tên/đvt/giá nếu người dùng chỉ chọn mã
            r["ten_xk"] = r.get("ten_xk") or tn.get("ten", "")
            r["dvt_xk"] = r.get("dvt_xk") or tn.get("dvt", "")
            r["gia_xk"] = r.get("gia_xk") or tn.get("gia", "")
        if r.get("mo_ho"):              # đã tự chọn xong -> học lại cho lần dò sau
            ten_chuan = _chuan_ten_hang_xk(r.get("ten_sp"))
            if ten_chuan:
                hoc_ma[ten_chuan] = ma
        r["mo_ho"] = False
    data["xk_giathanh"] = rows
    data["xk_hoc_ma"] = hoc_ma
    _ghi_du_lieu_cty(cid, data)
    so_khop = sum(1 for r in rows if r.get("ma"))
    return {"ok": True, "so_dong": len(rows), "so_khop": so_khop}

@app.post("/api/xk/export/{cid}")
def xk_export(cid: int):
    """Xuất file 'Xuất kho' (form MISA) từ GIATHANH đã lưu (chỉ lấy dòng đã gắn mã)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    data = _doc_du_lieu_cty(cid)
    giathanh = data.get("xk_giathanh") or []
    out, so_ct = _gen_xuat_kho_rows(giathanh)
    if not out:
        raise HTTPException(400, "Chưa có dòng nào được gắn mã hàng để xuất")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Xuất kho"
    for c, h in enumerate(XUAT_KHO_HEADERS, 1):
        ws.cell(1, c).value = h
        ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
        ws.cell(1, c).fill = PatternFill("solid", fgColor="2E5C8A")
    cot_text = {5, 28, 31, 33, 34, 35}     # số CT, mã hàng, kho, TK nợ/có, ĐVT
    cot_tien = {36, 37, 38, 39, 40}
    for ri, row in enumerate(out, 2):
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_text:
                cell.number_format = "@"
            elif ci in cot_tien:
                cell.number_format = "#,##0"
    for c in range(1, len(XUAT_KHO_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.freeze_panes = "A2"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    thang_ten = so_ct.replace("XK T", "T").replace("/", ".")
    fname = f"XuatKho_{thang_ten}_{mst}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    tong = len(giathanh)
    so_bo_qua = tong - len(out)
    return _resp_xuat(path, fname,
                      {"X-So-Dong": str(len(out)), "X-Bo-Qua": str(so_bo_qua)})

@app.post("/api/xk/export-giathanh/{cid}")
def xk_export_giathanh(cid: int):
    """Kết xuất giá thành: xuất TOÀN BỘ bảng GIATHANH (đủ cột BÁN & KHO cạnh
    nhau, kể cả dòng CHƯA gắn mã — tô đỏ nhạt để dễ nhận biết) ra file Excel
    để xem/lưu trữ/gửi đối chiếu — khác với 'Xuất kho' (chỉ lấy dòng đã gắn
    mã, đúng 51 cột form nhập MISA)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    data = _doc_du_lieu_cty(cid)
    giathanh = data.get("xk_giathanh") or []
    if not giathanh:
        raise HTTPException(400, "Chưa có dữ liệu giá thành — hãy bấm \"Dò mã hàng tự động\" trước")
    ton_rows = data.get("xk_ton") or []
    rows = _gen_giathanh_export_rows(ton_rows, giathanh)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giá thành"
    for c, h in enumerate(XK_GIATHANH_XUAT_HEADERS, 1):
        cell = ws.cell(1, c)
        cell.value = h
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1A7A7A")
    cot_tien = {7, 8, 13, 14}
    do_nhat = PatternFill("solid", fgColor="FDE8E8")
    for ri, row in enumerate(rows, 2):
        chua_gan = not str(giathanh[ri - 2].get("ma") or "").strip()
        for ci, v in enumerate(row, 1):
            cell = ws.cell(ri, ci)
            cell.value = v
            if ci in cot_tien and isinstance(v, (int, float)):
                cell.number_format = "#,##0"
            if chua_gan:
                cell.fill = do_nhat
    for c in range(1, len(XK_GIATHANH_XUAT_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.freeze_panes = "A2"
    # dòng tổng kết cuối bảng — giống hệt tổng kết hiển thị trên màn hình (chỉ
    # cộng SL/Thành tiền KHO của những dòng ĐÃ gắn mã, không fallback về SL bán)
    ri_tong = len(rows) + 2
    t_sl = sum(v for r in giathanh if isinstance((v := _to_num(r.get("sl"))), (int, float)))
    t_ban = sum(v for r in giathanh if isinstance((v := _to_num(r.get("tt"))), (int, float)))
    t_sl_gan, t_kho = 0, 0
    for r, row in zip(giathanh, rows):
        if not str(r.get("ma") or "").strip():
            continue
        if isinstance(row[11], (int, float)):
            t_sl_gan += row[11]
        if isinstance(row[13], (int, float)):
            t_kho += row[13]
    ws.cell(ri_tong, 4).value = "TỔNG CỘNG"
    ws.cell(ri_tong, 4).font = Font(bold=True)
    ws.cell(ri_tong, 6).value = t_sl
    ws.cell(ri_tong, 8).value = round(t_ban)
    ws.cell(ri_tong, 12).value = t_sl_gan
    ws.cell(ri_tong, 14).value = round(t_kho)
    for c in (6, 8, 12, 14):
        ws.cell(ri_tong, c).font = Font(bold=True)
        if c in (8, 14):
            ws.cell(ri_tong, c).number_format = "#,##0"
    conn = db()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst = _chuan_mst(comp["mst"]) if comp else str(cid)
    fname = f"KetXuatGiaThanh_{mst}_{datetime.datetime.now().strftime('%d%m%Y')}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    import shutil
    for d in (_get_desktop_dir(),
              (_du_lieu_cty_path(cid) and os.path.dirname(_du_lieu_cty_path(cid)))):
        if d and os.path.isdir(d):
            try:
                shutil.copy(path, os.path.join(d, fname))
            except Exception:
                pass
    return _resp_xuat(path, fname, {"X-So-Dong": str(len(rows))})


# ============================================================
#  KẾT NỐI TRỰC TIẾP CSDL MISA SME (SQL Server) — GIAI ĐOẠN CHỈ ĐỌC
#  ⚠ CẢNH BÁO: Ghi thẳng vào DB của MISA KHÔNG được MISA hỗ trợ, rủi ro hỏng
#  sổ sách / mất bảo hành. Giai đoạn này CHỈ ĐỌC (khám phá cấu trúc bảng) để
#  lập bản đồ. Mọi thao tác GHI sẽ chỉ làm sau khi đã sao lưu + thử trên bản
#  sao. Cấu hình kết nối lưu trong dữ liệu công ty (máy local).
# ============================================================
def _misa_sql_drivers():
    try:
        import pyodbc
        return [d for d in pyodbc.drivers() if "SQL Server" in d]
    except Exception:
        return []

def _misa_sql_cfg(cid):
    data = _doc_du_lieu_cty(cid)
    return data.get("misa_sql") or {}

def _misa_sql_connect(cid, cfg=None, database=None, timeout=8, readonly=True):
    """Mở kết nối pyodbc tới SQL Server của MISA. Mặc định ApplicationIntent
    ReadOnly (chỉ đọc). Ném HTTPException với thông báo rõ nếu lỗi."""
    try:
        import pyodbc
    except Exception:
        raise HTTPException(400, "Chưa cài thư viện pyodbc. Đóng phần mềm, chạy lại start.bat để tự cài.")
    cfg = cfg or _misa_sql_cfg(cid)
    drivers = _misa_sql_drivers()
    if not drivers:
        raise HTTPException(400, "Máy chưa có 'ODBC Driver for SQL Server'. Tải & cài "
                                 "'ODBC Driver 17 (hoặc 18) for SQL Server' của Microsoft rồi thử lại.")
    server = str(cfg.get("server") or "").strip()
    if not server:
        raise HTTPException(400, "Thiếu 'server' (tên máy chủ SQL). Ví dụ: .\\MISASME2023 hoặc localhost\\MISASME2022")
    db = database or str(cfg.get("database") or "").strip()
    # ưu tiên driver hiện đại (18/17) rồi mới tới Native Client / 'SQL Server'
    _uu_tien = ["ODBC Driver 18 for SQL Server", "ODBC Driver 17 for SQL Server",
                "ODBC Driver 13 for SQL Server", "SQL Server Native Client 11.0",
                "SQL Server Native Client 10.0", "SQL Server"]
    drv = next((d for d in _uu_tien if d in drivers), sorted(drivers, reverse=True)[0])
    parts = ["DRIVER={%s}" % drv, "SERVER=%s" % server]
    if db:
        parts.append("DATABASE=%s" % db)
    if cfg.get("trusted"):
        parts.append("Trusted_Connection=yes")
    else:
        parts.append("UID=%s" % (cfg.get("user") or ""))
        parts.append("PWD=%s" % (cfg.get("password") or ""))
    parts.append("TrustServerCertificate=yes")
    # KHÔNG dùng ApplicationIntent=ReadOnly (một số cấu hình/driver cũ gây lỗi
    # kết nối). An toàn "chỉ đọc" đảm bảo bằng việc chỉ chạy lệnh SELECT.
    try:
        conn = pyodbc.connect(";".join(parts) + ";", timeout=timeout)
        conn.autocommit = True
        return conn
    except Exception as e:
        msg = str(e)
        # Kết nối TỚI ĐƯỢC server nhưng SAI đăng nhập (18456) -> hướng dẫn đúng cách
        if "18456" in msg or "Login failed" in msg:
            # 4060 = đăng nhập OK nhưng không có quyền mở đúng database này
            if "4060" in msg or "Cannot open database" in msg:
                raise HTTPException(
                    400,
                    "Đăng nhập được nhưng tài khoản này KHÔNG có quyền mở database '%s' (lỗi 4060). "
                    "Tài khoản Windows thường không được MISA cấp quyền trên database của nó. "
                    "Hãy dùng tài khoản 'sa': đổi Xác thực = 'Tài khoản SQL', User = sa, kèm mật khẩu sa "
                    "của MISA (tài khoản sa có toàn quyền, mở được mọi database). Nếu chưa biết mật khẩu "
                    "sa, hỏi người cài MISA hoặc xem trong cấu hình kết nối của MISA."
                    % (db or "?"))
            u = cfg.get("user") or ""
            raise HTTPException(
                400,
                "Kết nối được SQL Server nhưng SAI TÀI KHOẢN/MẬT KHẨU"
                + (" (đăng nhập '%s' không hợp lệ)" % u if u else "")
                + ". Lưu ý: MISA KHÔNG dùng Mã số thuế làm tài khoản SQL. Hãy thử: "
                "① Xác thực = 'Windows (Trusted)' để trống user/mật khẩu; hoặc "
                "② nhập đúng tài khoản 'sa' của MISA.")
        if "Cannot open database" in msg or "4060" in msg:
            raise HTTPException(400, "Không mở được database đã chọn (không có quyền hoặc sai tên DB). "
                                     "Thử dùng tài khoản 'sa'. Chi tiết: %s" % msg[:200])
        raise HTTPException(400, "Không kết nối được SQL Server: %s. Kiểm tra tên server/instance, "
                                 "đã bật TCP/IP + SQL Browser, và tài khoản/mật khẩu." % (msg[:300]))


@app.get("/api/misa-sql/trang-thai/{cid}")
def misa_sql_trang_thai(cid: int):
    """Cho biết máy có driver ODBC chưa + cấu hình đã lưu (ẩn mật khẩu)."""
    cfg = _misa_sql_cfg(cid)
    return {"drivers": _misa_sql_drivers(),
            "cfg": {"server": cfg.get("server", ""), "database": cfg.get("database", ""),
                    "trusted": bool(cfg.get("trusted")), "user": cfg.get("user", ""),
                    "co_mat_khau": bool(cfg.get("password"))}}


@app.post("/api/misa-sql/luu-cau-hinh/{cid}")
async def misa_sql_luu_cau_hinh(cid: int, request: Request):
    """Lưu cấu hình kết nối SQL của MISA cho công ty (local)."""
    body = await request.json()
    data = _doc_du_lieu_cty(cid)
    cu = data.get("misa_sql") or {}
    cfg = {
        "server": str(body.get("server") or "").strip(),
        "database": str(body.get("database") or "").strip(),
        "trusted": bool(body.get("trusted")),
        "user": str(body.get("user") or "").strip(),
        # giữ mật khẩu cũ nếu FE gửi rỗng (không bắt gõ lại mỗi lần)
        "password": (body.get("password") if body.get("password") else cu.get("password", "")),
    }
    data["misa_sql"] = cfg
    _ghi_du_lieu_cty(cid, data)
    return {"ok": True}


@app.post("/api/misa-sql/test/{cid}")
def misa_sql_test(cid: int):
    """Thử kết nối + trả phiên bản SQL Server và danh sách database (để chọn
    đúng DB của công ty trong MISA)."""
    conn = _misa_sql_connect(cid, database="master")
    try:
        cur = conn.cursor()
        ver = cur.execute("SELECT @@VERSION").fetchval()
        dbs = [r[0] for r in cur.execute(
            "SELECT name FROM sys.databases WHERE database_id > 4 ORDER BY name").fetchall()]
        return {"ok": True, "version": str(ver)[:120], "databases": dbs}
    finally:
        conn.close()


@app.post("/api/misa-sql/bang/{cid}")
def misa_sql_bang(cid: int, database: str = "", loc: str = ""):
    """Liệt kê bảng (+ số dòng ước lượng) của 1 database. loc: lọc theo tên."""
    if not database:
        raise HTTPException(400, "Chưa chọn database")
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        sql = (
            "SELECT t.name, ISNULL(SUM(p.rows),0) AS so_dong "
            "FROM sys.tables t "
            "LEFT JOIN sys.partitions p ON p.object_id=t.object_id AND p.index_id IN (0,1) "
            "GROUP BY t.name")
        rows = cur.execute(sql).fetchall()
        out = [{"bang": r[0], "so_dong": int(r[1])} for r in rows]
        if loc:
            lo = loc.lower()
            out = [x for x in out if lo in x["bang"].lower()]
        out.sort(key=lambda x: x["bang"].lower())
        return {"ok": True, "database": database, "so_bang": len(out), "bang": out}
    finally:
        conn.close()


@app.post("/api/misa-sql/cot/{cid}")
def misa_sql_cot(cid: int, database: str = "", bang: str = ""):
    """Mô tả cột của 1 bảng (tên, kiểu, cho null, khóa chính)."""
    if not (database and bang):
        raise HTTPException(400, "Thiếu database/bảng")
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        cols = cur.execute("""
            SELECT c.COLUMN_NAME, c.DATA_TYPE, c.CHARACTER_MAXIMUM_LENGTH,
                   c.IS_NULLABLE, c.COLUMN_DEFAULT
            FROM INFORMATION_SCHEMA.COLUMNS c
            WHERE c.TABLE_NAME=? ORDER BY c.ORDINAL_POSITION""", bang).fetchall()
        pk = set()
        try:
            for r in cur.execute("""
                SELECT k.COLUMN_NAME FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS t
                JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE k ON k.CONSTRAINT_NAME=t.CONSTRAINT_NAME
                WHERE t.TABLE_NAME=? AND t.CONSTRAINT_TYPE='PRIMARY KEY'""", bang).fetchall():
                pk.add(r[0])
        except Exception:
            pass
        out = [{"cot": r[0], "kieu": r[1], "dai": r[2], "null": (r[3] == "YES"),
                "mac_dinh": (str(r[4]) if r[4] is not None else ""), "khoa_chinh": (r[0] in pk)}
               for r in cols]
        return {"ok": True, "database": database, "bang": bang, "cot": out}
    finally:
        conn.close()


@app.post("/api/misa-sql/xem/{cid}")
def misa_sql_xem(cid: int, database: str = "", bang: str = "", n: int = 20):
    """Xem thử tối đa N dòng đầu của bảng (CHỈ ĐỌC) để hiểu dữ liệu mẫu."""
    if not (database and bang):
        raise HTTPException(400, "Thiếu database/bảng")
    if not bang.replace("_", "").replace(" ", "").isalnum():
        raise HTTPException(400, "Tên bảng không hợp lệ")
    n = max(1, min(int(n or 20), 100))
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        cur.execute("SELECT TOP (%d) * FROM [%s]" % (n, bang))
        cols = [d[0] for d in cur.description]
        rows = []
        for r in cur.fetchall():
            rows.append([(v.isoformat() if hasattr(v, "isoformat") else
                          (str(v) if v is not None else "")) for v in r])
        return {"ok": True, "cot": cols, "rows": rows, "so_dong": len(rows)}
    finally:
        conn.close()


@app.post("/api/misa-sql/liet-ke-reftype/{cid}")
def misa_sql_liet_ke_reftype(cid: int, database: str = "", master_table: str = "PUVoucher",
                             tu_khoa: str = ""):
    """Liệt kê loại chứng từ (SYSRefType) — CHỈ ĐỌC. Dùng khi phần mềm báo
    'không dò được loại chứng từ MISA phù hợp' để xem chính xác tên các loại
    chứng từ MISA đang có. Mặc định lọc theo master_table (vd PUVoucher);
    nếu truyền tu_khoa (vd 'dịch vụ') thì tìm theo TÊN trên MỌI bảng đích,
    trả kèm MasterTableName — để biết loại đó thực sự thuộc bảng nào (có
    thể KHÁC PUVoucher, vd PUService cho chứng từ mua dịch vụ)."""
    database = (database or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    tu_khoa = (tu_khoa or "").strip()
    if not database:
        raise HTTPException(400, "Thiếu database")
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        if tu_khoa:
            rows = cur.execute(
                "SELECT RefType, RefTypeName, MasterTableName FROM SYSRefType "
                "WHERE RefTypeName LIKE ? ORDER BY MasterTableName, RefTypeName",
                "%" + tu_khoa + "%").fetchall()
            ket = [{"reftype": r[0], "ten": r[1], "bang": r[2]} for r in rows]
            return {"ok": True, "tu_khoa": tu_khoa, "so_luong": len(ket), "danh_sach": ket}
        master_table = (master_table or "PUVoucher").strip()
        rows = cur.execute(
            "SELECT RefType, RefTypeName FROM SYSRefType WHERE MasterTableName=? "
            "ORDER BY RefTypeName", master_table).fetchall()
        ket = [{"reftype": r[0], "ten": r[1]} for r in rows]
        return {"ok": True, "master_table": master_table, "so_luong": len(ket), "danh_sach": ket}
    finally:
        conn.close()


def _misa_sql_doc_schema(cid, database):
    """Đọc toàn bộ cấu trúc (bảng -> danh sách cột) của database (chỉ đọc)."""
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        rows = cur.execute(
            "SELECT TABLE_NAME, COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA='dbo' ORDER BY TABLE_NAME, ORDINAL_POSITION").fetchall()
        tables = {}
        for t, c in rows:
            tables.setdefault(t, []).append(c)
        return tables
    finally:
        conn.close()

def _misa_sql_fingerprint(tables):
    import hashlib
    s = "|".join(t + ":" + ",".join(tables[t]) for t in sorted(tables))
    return (hashlib.sha1(s.encode("utf-8")).hexdigest(),
            len(tables), sum(len(v) for v in tables.values()))


@app.post("/api/misa-sql/schema-check/{cid}")
def misa_sql_schema_check(cid: int, database: str = ""):
    """So sánh cấu trúc MISA hiện tại với mốc đã lưu. Nếu chưa có mốc -> lưu
    làm mốc. Nếu khác (MISA cập nhật thêm/bớt bảng/cột) -> báo có thay đổi để
    người dùng đồng bộ lại. Chỉ đọc INFORMATION_SCHEMA."""
    if not database:
        raise HTTPException(400, "Chưa chọn database")
    tables = _misa_sql_doc_schema(cid, database)
    fp, nt, nc = _misa_sql_fingerprint(tables)
    data = _doc_du_lieu_cty(cid)
    snap = data.get("misa_sql_schema") or {}
    now = datetime.datetime.now().isoformat()
    if snap.get("database") != database or not snap.get("hash"):
        data["misa_sql_schema"] = {"database": database, "hash": fp, "so_bang": nt,
                                   "so_cot": nc, "tables": tables, "updated_at": now}
        _ghi_du_lieu_cty(cid, data)
        return {"trang_thai": "moc_moi", "changed": False, "so_bang": nt, "so_cot": nc,
                "updated_at": now}
    if snap.get("hash") == fp:
        return {"trang_thai": "khong_doi", "changed": False, "so_bang": nt, "so_cot": nc,
                "updated_at": snap.get("updated_at", "")}
    old = snap.get("tables") or {}
    them = sorted(set(tables) - set(old))
    mat = sorted(set(old) - set(tables))
    doi = sorted(t for t in (set(tables) & set(old)) if tables[t] != old.get(t))
    return {"trang_thai": "da_doi", "changed": True, "so_bang": nt, "so_cot": nc,
            "bang_them": them[:80], "bang_mat": mat[:80], "bang_doi_cot": doi[:80],
            "moc_updated_at": snap.get("updated_at", "")}


@app.post("/api/misa-sql/schema-dongbo/{cid}")
def misa_sql_schema_dongbo(cid: int, database: str = ""):
    """Cập nhật mốc cấu trúc = cấu trúc MISA HIỆN TẠI (sau khi người dùng xác
    nhận đồng bộ theo phiên bản MISA mới)."""
    if not database:
        raise HTTPException(400, "Chưa chọn database")
    tables = _misa_sql_doc_schema(cid, database)
    fp, nt, nc = _misa_sql_fingerprint(tables)
    data = _doc_du_lieu_cty(cid)
    data["misa_sql_schema"] = {"database": database, "hash": fp, "so_bang": nt,
                               "so_cot": nc, "tables": tables,
                               "updated_at": datetime.datetime.now().isoformat()}
    _ghi_du_lieu_cty(cid, data)
    return {"ok": True, "so_bang": nt, "so_cot": nc}


# Danh sách TỪ KHÓA các bảng cần lấy để dựng chức năng ghi vào MISA (danh mục
# + chứng từ). Dùng để lọc nhanh khi xuất cấu trúc.
_MISA_BANG_QUAN_TRONG = [
    "InventoryItem", "Unit", "InventoryItemCategory", "InventoryItemUnitConvert",
    "Stock", "AccountObject", "AccountObjectCategory", "Account",
    "FixedAsset", "FixedAssetCategory", "Tool", "InstrumentTool",
    "PUVoucher", "PUInvoice", "SAVoucher", "SAInvoice", "INVoucher", "INWard",
    "OUTWard", "GLVoucher", "Voucher", "Invoice", "GeneralLedger"]

@app.post("/api/misa-sql/xuat-cau-truc/{cid}")
def misa_sql_xuat_cau_truc(cid: int, database: str = "", loc: str = "", n_mau: int = 2,
                           tat_ca: int = 0):
    """Xuất cấu trúc bảng MISA ra 1 FILE text để gửi cho người phát triển dựng
    chức năng ghi. tat_ca=1 -> TẤT CẢ bảng; loc có -> bảng khớp lọc; else ->
    các bảng theo danh sách từ khóa quan trọng. Kèm tối đa n_mau dòng mẫu. Chỉ đọc."""
    if not database:
        raise HTTPException(400, "Chưa chọn database")
    n_mau = max(0, min(int(n_mau or 0), 5))
    conn = _misa_sql_connect(cid, database=database)
    try:
        cur = conn.cursor()
        # cột + kiểu
        col_rows = cur.execute(
            "SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE, CHARACTER_MAXIMUM_LENGTH, IS_NULLABLE "
            "FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_SCHEMA='dbo' "
            "ORDER BY TABLE_NAME, ORDINAL_POSITION").fetchall()
        cols_by = {}
        for t, c, dt, ln, nu in col_rows:
            cols_by.setdefault(t, []).append((c, dt, ln, nu))
        # khóa chính
        pk_by = {}
        try:
            for t, c in cur.execute(
                "SELECT tc.TABLE_NAME, kcu.COLUMN_NAME "
                "FROM INFORMATION_SCHEMA.TABLE_CONSTRAINTS tc "
                "JOIN INFORMATION_SCHEMA.KEY_COLUMN_USAGE kcu ON kcu.CONSTRAINT_NAME=tc.CONSTRAINT_NAME "
                "WHERE tc.CONSTRAINT_TYPE='PRIMARY KEY'").fetchall():
                pk_by.setdefault(t, set()).add(c)
        except Exception:
            pass
        all_names = sorted(cols_by)
        if tat_ca:
            names = all_names                       # xuất TẤT CẢ bảng, không giới hạn
        elif loc.strip():
            lo = loc.strip().lower()
            names = [t for t in all_names if lo in t.lower()][:120]
        else:
            names = [t for t in all_names
                     if any(k.lower() in t.lower() for k in _MISA_BANG_QUAN_TRONG)][:120]
        out = ["CAU TRUC BANG MISA — database: %s" % database,
               "Loc: %s | So bang xuat: %d" % (
                   "(TAT CA bang)" if tat_ca else (loc or "(bang quan trong)"), len(names)),
               "=" * 70, ""]
        for t in names:
            pk = pk_by.get(t, set())
            out.append("### BANG: %s" % t)
            for (c, dt, ln, nu) in cols_by[t]:
                out.append("   %-40s %s%s %s%s" % (
                    c, dt, ("(%s)" % ln if ln not in (None, -1) else ""),
                    "NULL" if nu == "YES" else "NOTNULL", "  [PK]" if c in pk else ""))
            if n_mau > 0 and t.replace("_", "").replace(" ", "").isalnum():
                try:
                    cur.execute("SELECT TOP (%d) * FROM [%s]" % (n_mau, t))
                    scols = [d[0] for d in cur.description]
                    out.append("   --- DU LIEU MAU ---")
                    for r in cur.fetchall():
                        kv = []
                        for cn, v in zip(scols, r):
                            sv = v.isoformat() if hasattr(v, "isoformat") else (str(v) if v is not None else "NULL")
                            if len(sv) > 60:
                                sv = sv[:60] + "..."
                            kv.append("%s=%s" % (cn, sv))
                        out.append("   • " + " | ".join(kv))
                except Exception as e:
                    out.append("   (khong doc duoc du lieu mau: %s)" % str(e)[:80])
            out.append("")
        conn2 = db()
        comp = conn2.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
        conn2.close()
        mst = _chuan_mst(comp["mst"]) if comp else str(cid)
        fname = "CauTrucMISA_%s_%s.txt" % (mst, datetime.datetime.now().strftime("%d%m%Y_%H%M"))
        base = _get_desktop_dir() or DOWNLOAD_DIR
        path = os.path.join(base, fname)
        with open(path, "w", encoding="utf-8") as f:
            f.write("\n".join(out))
        _open_file_local(path)
        return {"ok": True, "file": path, "so_bang": len(names)}
    finally:
        conn.close()


# ============================================================
#  GHI DANH MỤC HÀNG HÓA THẲNG VÀO MISA (bảng InventoryItem)
#  ⚠ Ghi vào DB MISA — chỉ chạy trên dữ liệu THỬ/đã sao lưu. Có preview trước,
#  bỏ qua mã đã tồn tại, chạy trong transaction (lỗi -> rollback, không ghi dở).
# ============================================================
# Danh mục kho -> (TK kho, TK giá vốn, TK doanh thu, tiền tố TK để dò tính chất)
# TẤT CẢ đều ghi vào bảng InventoryItem (Danh mục > Vật tư hàng hóa) với tính
# chất "Vật tư hàng hóa". TSCĐ/CCDC cũng đăng ký làm mã Vật tư hàng hóa (theo
# yêu cầu người dùng) — chỉ khác TK kho ngầm định (có thể sửa lại trong MISA).
_MISA_INV_ACC = {
    "hh":   ("1561", "632", "5111", "156%"),   # Hàng hóa
    "nvl":  ("152",  "632", "5111", "152%"),   # Nguyên vật liệu
    "tscd": ("211",  "632", "5111", "211%"),   # TSCĐ (ghi làm mã Vật tư hàng hóa)
    "ccdc": ("153",  "632", "5111", "153%"),   # CCDC (ghi làm mã Vật tư hàng hóa)
}
_MISA_INV_TEN = {"hh": "Hàng hóa", "nvl": "Nguyên vật liệu",
                 "tscd": "TSCĐ", "ccdc": "CCDC"}

def _misa_ghi_hang_hoa(cid, database, dm_rows, preview=True, loai="hh"):
    """Thêm các mã (Danh mục Hàng hóa/NVL/TSCĐ/CCDC của phần mềm) vào bảng
    InventoryItem của MISA (Danh mục > Vật tư hàng hóa), tính chất luôn là
    "Vật tư hàng hóa". Chỉ THÊM mã MỚI (bỏ qua mã đã có). Tự tạo đơn vị tính
    (Unit) nếu chưa có. preview=True: chỉ xem trước, KHÔNG ghi (rollback).
    loai: hh/nvl/tscd/ccdc -> chỉ khác TK kho ngầm định."""
    import uuid as _uuid
    inv_acc, cogs_acc, sale_acc, _acc_like = _MISA_INV_ACC.get(loai, _MISA_INV_ACC["hh"])
    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        # mã hàng đã tồn tại + TÊN trong MISA (để bỏ qua, và đối chiếu tên lệch)
        existing = {}   # ma_lower -> tên trong MISA
        for code, name in cur.execute(
                "SELECT InventoryItemCode, InventoryItemName FROM InventoryItem").fetchall():
            if code:
                existing[str(code).strip().lower()] = str(name or "")

        def _norm_ten(s):
            return " ".join(str(s or "").split()).lower()
        # đơn vị tính hiện có {tên_lower: UnitID} — BỎ QUA Unit tên RÁC (chuỗi
        # GUID thay vì tên thật, xem _misa_unit_hong) để không lỡ khớp/tái sử
        # dụng bản ghi hỏng; dvt trùng tên với Unit rác sẽ tự tạo Unit MỚI sạch.
        units = {}
        for uid, uname in cur.execute("SELECT UnitID, UnitName FROM Unit").fetchall():
            if uname and not _misa_unit_hong(uname):
                units[str(uname).strip().lower()] = uid
        # 'Tính chất' (InventoryItemType) = 1 (Vật tư, hàng hóa) cho cả 3 danh
        # mục Hàng hóa/NVL/CCDC. Trước đây dò theo dữ liệu có sẵn trong MISA
        # (mã nào cùng TK kho thì lấy tính chất đó) nhưng có công ty đã lỡ gắn
        # sai tính chất (vd Thành phẩm) cho các mã cũ trên TK 152 -> khiến mã
        # mới import theo cũng bị sai. Vì Hàng hóa/NVL/CCDC của phần mềm luôn
        # là "Vật tư, hàng hóa" nên gắn cố định, không dò nữa.
        item_type = 1

        cols = ["InventoryItemID", "InventoryItemCode", "InventoryItemName", "InventoryItemType",
                "UnitID", "InventoryAccount", "COGSAccount", "SaleAccount", "TaxRate",
                "MinimumStock", "PurchaseDiscountRate", "UnitPrice", "SalePrice1", "SalePrice2",
                "SalePrice3", "FixedSalePrice", "FixedUnitPrice", "IsUnitPriceAfterTax",
                "IsSystem", "Inactive", "IsPromotion", "VAT43Type", "CreatedDate"]
        sql_ins = ("INSERT INTO InventoryItem ([%s]) VALUES (%s)"
                   % ("],[".join(cols), ",".join(["?"] * len(cols))))
        now = datetime.datetime.now()

        ket = []
        them = trung = dv_moi = lech_ten = 0
        seen = set()
        for r in dm_rows:
            ma = str((r[0] if len(r) > 0 else "") or "").strip()
            ten = str((r[1] if len(r) > 1 else "") or "").strip()
            dvt = str((r[2] if len(r) > 2 else "") or "").strip()
            ts = r[3] if len(r) > 3 else None
            if not ma or not ten:
                continue
            k = ma.lower()
            if k in seen:
                continue
            seen.add(k)
            if k in existing:
                ten_misa = existing[k]
                lech = _norm_ten(ten) != _norm_ten(ten_misa)
                trung += 1
                if lech:
                    lech_ten += 1
                ket.append({"ma": ma, "ten": ten, "ten_misa": ten_misa, "lech_ten": lech,
                            "trang_thai": "đã có — KHÁC TÊN" if lech else "đã có (bỏ qua)"})
                continue
            # đơn vị tính -> UnitID (tạo mới nếu chưa có)
            unit_id = None
            if dvt:
                uk = dvt.lower()
                if uk in units:
                    unit_id = units[uk]
                else:
                    unit_id = str(_uuid.uuid4())
                    if not preview:
                        cur.execute("INSERT INTO Unit (UnitID, UnitName, Description, Inactive) "
                                    "VALUES (?,?,?,0)", unit_id, dvt[:20], None)
                    units[uk] = unit_id
                    dv_moi += 1
            tax = ts if isinstance(ts, (int, float)) else None
            if not preview:
                cur.execute(sql_ins,
                    str(_uuid.uuid4()), ma[:50], ten[:500], item_type,
                    unit_id, inv_acc, cogs_acc, sale_acc, tax,
                    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, "Chưa xác định", now)
            them += 1
            ket.append({"ma": ma, "ten": ten, "dvt": dvt,
                        "trang_thai": "sẽ thêm" if preview else "đã thêm"})
        if preview:
            conn.rollback()
        else:
            conn.commit()
        return {"preview": preview, "database": database, "so_them": them, "so_trung": trung,
                "so_lech_ten": lech_ten, "so_don_vi_moi": dv_moi, "item_type": item_type,
                "danh_sach": ket[:1000]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()


def _misa_quet_sua_dvt_hang_hoa(cid, database, dm_rows, preview=True):
    """Quét TOÀN BỘ InventoryItem đã có trong MISA (không chỉ mã mới import)
    để tìm + sửa ĐVT "mồ côi"/RÁC: UnitID không tồn tại trong Unit, hoặc trỏ
    tới Unit có tên là chuỗi GUID rác (_misa_unit_hong) — xác nhận thật qua
    mã HH00087-8 hiện "ĐVT chính: fac19b53" và crash InvalidCastException khi
    mở trong MISA. Đối chiếu Mã hàng với Bảng kê Danh mục hiện tại của phần
    mềm để lấy đúng tên ĐVT thật, tạo/tái dùng Unit sạch rồi UPDATE
    InventoryItem.UnitID. Không đụng tới mã hàng có ĐVT đã sạch."""
    dvt_theo_ma = {}   # ma_lower -> dvt text (từ Bảng kê Danh mục của phần mềm)
    for r in dm_rows:
        ma = str((r[0] if len(r) > 0 else "") or "").strip()
        dvt = str((r[2] if len(r) > 2 else "") or "").strip()
        if ma and dvt:
            dvt_theo_ma.setdefault(ma.lower(), dvt)
    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        unit_ten = {}   # tên sạch (lower) -> UnitID
        unit_id_sach = set()
        unit_ten_theo_id = {}   # UnitID (lower) -> tên gốc (để chẩn đoán/hiển thị)
        for u_id, u_name in cur.execute("SELECT UnitID, UnitName FROM Unit").fetchall():
            unit_ten_theo_id[str(u_id).strip().lower()] = str(u_name or "")
            if u_name and _misa_unit_hong(u_name):
                continue
            unit_id_sach.add(str(u_id).strip().lower())
            if u_name:
                unit_ten[str(u_name).strip().lower()] = u_id
        import uuid as _uuid
        them = sua = doi_chieu = 0
        ket = []
        chan_doan = []   # mẫu các mã ĐÃ đối chiếu (để soi khi quét ra 0)
        for iid, code, uid in cur.execute(
                "SELECT InventoryItemID, InventoryItemCode, UnitID FROM InventoryItem").fetchall():
            k = str(code or "").strip().lower()
            if k not in dvt_theo_ma:
                continue
            doi_chieu += 1
            uid_l = str(uid).strip().lower() if uid is not None else None
            ten_unit_misa = unit_ten_theo_id.get(uid_l, "") if uid_l else ""
            hong = uid is None or uid_l not in unit_id_sach
            if len(chan_doan) < 60:
                chan_doan.append({"ma": code, "unit_id": (str(uid) if uid is not None else None),
                                  "unit_ten_misa": ten_unit_misa, "hong": hong})
            if not hong:
                continue
            dvt_text = dvt_theo_ma[k]
            uk = dvt_text.lower()
            if uk in unit_ten:
                uid2 = unit_ten[uk]
            else:
                uid2 = str(_uuid.uuid4())
                if not preview:
                    cur.execute("INSERT INTO Unit (UnitID, UnitName, Description, Inactive) "
                                "VALUES (?,?,?,0)", uid2, dvt_text[:20], None)
                unit_ten[uk] = uid2
                unit_id_sach.add(uid2.lower())
                them += 1
            if not preview:
                cur.execute("UPDATE InventoryItem SET UnitID=? WHERE InventoryItemID=?", uid2, iid)
            sua += 1
            ket.append({"ma": code, "dvt": dvt_text, "unit_ten_cu": ten_unit_misa,
                        "trang_thai": "sẽ sửa" if preview else "đã sửa"})
        if preview:
            conn.rollback()
        else:
            conn.commit()
        return {"preview": preview, "database": database, "so_sua": sua, "so_don_vi_moi": them,
                "so_doi_chieu": doi_chieu, "chan_doan": chan_doan, "danh_sach": ket[:1000]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()


@app.post("/api/misa-sql/quet-sua-dvt/{cid}")
async def misa_sql_quet_sua_dvt(cid: int, request: Request):
    """Quét & sửa ĐVT mồ côi/rác trên TOÀN BỘ Danh mục Vật tư hàng hóa đã có
    trong MISA (không chỉ mã mới import) — xem _misa_quet_sua_dvt_hang_hoa.
    body: {rows, preview, database?} — rows = Bảng kê Danh mục hiện tại của
    phần mềm (Mã, Tên, ĐVT...) để tra đúng tên ĐVT thật."""
    body = await request.json()
    rows = body.get("rows") or []
    preview = bool(body.get("preview", True))
    database = (body.get("database") or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    if not database:
        raise HTTPException(400, "Chưa cấu hình kết nối/CSDL MISA. Mở '🗄 Kết nối CSDL MISA', "
                                 "kết nối tới dữ liệu THỬ trước.")
    if not rows:
        raise HTTPException(400, "Danh mục trống — không có mã để đối chiếu.")
    return _misa_quet_sua_dvt_hang_hoa(cid, database, rows, preview=preview)


@app.post("/api/misa-sql/import-hang-hoa/{cid}")
async def misa_sql_import_hang_hoa(cid: int, request: Request):
    """Import Danh mục Hàng hóa/NVL/TSCĐ/CCDC vào MISA (bảng InventoryItem).
    body: {rows, preview, loai, database?}. preview=true -> chỉ xem trước."""
    body = await request.json()
    rows = body.get("rows") or []
    preview = bool(body.get("preview", True))
    loai = (body.get("loai") or "hh").strip()
    if loai not in _MISA_INV_ACC:
        raise HTTPException(400, "Loại danh mục '%s' chưa hỗ trợ ghi vào MISA "
                                 "(chỉ hh/nvl/tscd/ccdc)." % loai)
    database = (body.get("database") or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    if not database:
        raise HTTPException(400, "Chưa cấu hình kết nối/CSDL MISA. Mở '🗄 Kết nối CSDL MISA', "
                                 "kết nối tới dữ liệu THỬ trước.")
    if not rows:
        raise HTTPException(400, "Danh mục trống — không có mã để import.")
    return _misa_ghi_hang_hoa(cid, database, rows, preview=preview, loai=loai)


# ============================================================
#  GHI DANH MỤC KH/NCC THẲNG VÀO MISA (bảng AccountObject)
#  Đối tượng đơn thuần (không kèm bút toán) — an toàn như InventoryItem.
#  Thu thập MST+tên từ hóa đơn mua vào (NCC) / bán ra (KH) đã lưu, MST nào
#  xuất hiện ở mua vào -> IsVendor=1, ở bán ra -> IsCustomer=1 (có thể cả 2).
# ============================================================
def _misa_khncc_chuan_mst(s):
    return str(s or "").strip().replace("-", "").replace(" ", "").replace(".", "")

def _misa_khncc_dinh_dang_mst(s):
    s = _misa_khncc_chuan_mst(s)
    if len(s) == 13 and s.isdigit():
        return s[:10] + "-" + s[10:]
    return s

def _misa_thu_thap_khncc(cid):
    """Thu thập MST + tên + vai trò (NCC/KH) từ hóa đơn mua vào/bán ra đã
    lưu. Trả list [{"mst","mst_hien","ten","ncc","kh"}], sắp theo MST."""
    conn = db()
    rows = conn.execute(
        "SELECT loai, nbmst, nbten, nmmst, raw FROM invoices WHERE company_id=?", (cid,)).fetchall()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst_cty = _misa_khncc_chuan_mst(comp["mst"]) if comp else ""
    ds = {}   # mst -> {"ten","ncc","kh"}
    for r in rows:
        if r["loai"] == "purchase":
            mst = _misa_khncc_chuan_mst(r["nbmst"])
            ten = str(r["nbten"] or "").strip()
            vai = "ncc"
        else:
            try:
                raw = json.loads(r["raw"]) if r["raw"] else {}
            except Exception:
                raw = {}
            mst = _misa_khncc_chuan_mst(r["nmmst"] or raw.get("nmmst", ""))
            ten = str(raw.get("nmten", "") or "").strip()
            vai = "kh"
        if not mst or mst == mst_cty:
            continue
        e = ds.setdefault(mst, {"ten": "", "ncc": False, "kh": False})
        e[vai] = True
        if ten and len(ten) > len(e["ten"]):
            e["ten"] = ten
    out = []
    for mst, e in sorted(ds.items()):
        if not e["ten"]:
            continue
        out.append({"mst": mst, "mst_hien": _misa_khncc_dinh_dang_mst(mst),
                    "ten": e["ten"], "ncc": e["ncc"], "kh": e["kh"]})
    return out

def _misa_branch_id(cur):
    """Chi nhánh gốc (OrganizationUnit không có ParentID) — dùng làm BranchID
    ngầm định khi ghi các bảng MISA (đa số công ty vừa/nhỏ chỉ có 1 chi nhánh)."""
    row = cur.execute(
        "SELECT TOP 1 OrganizationUnitID FROM OrganizationUnit "
        "WHERE ParentID IS NULL ORDER BY Grade").fetchone()
    return row[0] if row else None

def _misa_ghi_khncc(cid, database, preview=True):
    """Ghi Danh mục KH/NCC (thu thập từ hóa đơn mua vào/bán ra đã lưu) thẳng
    vào bảng AccountObject của MISA (Danh mục > Đối tượng). Chỉ THÊM mã
    MỚI (bỏ qua MST đã có, không ghi đè). preview=True: chỉ xem trước."""
    import uuid as _uuid
    items = _misa_thu_thap_khncc(cid)
    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        branch_id = _misa_branch_id(cur)
        existing = {}   # mst_lower -> tên trong MISA (đối chiếu theo Mã ĐT hoặc MST)
        for code, taxcode, name in cur.execute(
                "SELECT AccountObjectCode, CompanyTaxCode, AccountObjectName "
                "FROM AccountObject").fetchall():
            ten_misa = str(name or "")
            if code:
                existing[str(code).strip().lower()] = ten_misa
            if taxcode:
                existing[_misa_khncc_chuan_mst(taxcode).lower()] = ten_misa

        def _norm_ten(s):
            return " ".join(str(s or "").split()).lower()

        now = datetime.datetime.now()
        ket = []
        them = trung = lech_ten = 0
        for it in items:
            mst, mst_hien, ten = it["mst"], it["mst_hien"], it["ten"]
            k = mst.lower()
            if k in existing:
                ten_misa = existing[k]
                lech = _norm_ten(ten) != _norm_ten(ten_misa)
                trung += 1
                if lech:
                    lech_ten += 1
                ket.append({"mst": mst_hien, "ten": ten, "ten_misa": ten_misa, "lech_ten": lech,
                            "vai_tro": "+".join(x for x in (["NCC"] if it["ncc"] else []) + (["KH"] if it["kh"] else [])),
                            "trang_thai": "đã có — KHÁC TÊN" if lech else "đã có (bỏ qua)"})
                continue
            is_vendor = 1 if it["ncc"] else 0
            is_customer = 1 if it["kh"] else 0
            taxcode = mst_hien if (mst.isdigit() and len(mst) != 12) else None
            if not preview:
                cur.execute(
                    "INSERT INTO AccountObject ([AccountObjectID],[AccountObjectCode],"
                    "[AccountObjectName],[CompanyTaxCode],[IsVendor],[IsCustomer],"
                    "[AccountObjectType],[Inactive],[BranchID],[CreatedDate]) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?)",
                    str(_uuid.uuid4()), mst_hien[:50], ten[:400], taxcode,
                    is_vendor, is_customer, 0, 0, branch_id, now)
            them += 1
            ket.append({"mst": mst_hien, "ten": ten,
                        "vai_tro": "+".join(x for x in (["NCC"] if it["ncc"] else []) + (["KH"] if it["kh"] else [])),
                        "trang_thai": "sẽ thêm" if preview else "đã thêm"})
        if preview:
            conn.rollback()
        else:
            conn.commit()
        return {"preview": preview, "database": database, "so_them": them, "so_trung": trung,
                "so_lech_ten": lech_ten, "danh_sach": ket[:1000]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()


@app.post("/api/misa-sql/import-khncc/{cid}")
def misa_sql_import_khncc(cid: int, preview: int = 1, database: str = ""):
    """Ghi Danh mục KH/NCC thẳng vào MISA (bảng AccountObject).
    preview=1 -> chỉ xem trước, không ghi."""
    database = (database or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    if not database:
        raise HTTPException(400, "Chưa cấu hình kết nối/CSDL MISA. Mở '🗄 Kết nối CSDL MISA', "
                                 "kết nối tới dữ liệu THỬ trước.")
    return _misa_ghi_khncc(cid, database, preview=bool(preview))


# ============================================================
#  GHI CHỨNG TỪ MUA HÀNG THẲNG VÀO MISA (PUVoucher + PUVoucherDetail)
#  ⚠ RỦI RO CAO NHẤT từ trước tới giờ — đây là chứng từ kế toán thật (Nợ
#  152/156/242/211/6xx, Có 111/112/331, thuế GTGT đầu vào...). LUÔN ghi ở
#  trạng thái CHƯA GHI SỔ (IsPostedFinance=IsPostedManagement=0), KHÔNG
#  đụng tới GeneralLedger/GLVoucher — anh PHẢI mở từng chứng từ trong MISA
#  kiểm tra rồi tự bấm "Ghi sổ". Mã hàng/NCC phải ĐÃ CÓ sẵn trong Danh mục
#  MISA (từ các bước import trước) — dòng/chứng từ nào thiếu sẽ bị BỎ QUA,
#  không tự đoán/tạo mới. Dữ liệu lấy từ Bảng kê đầu vào ĐÃ LƯU, dùng lại
#  đúng logic đã kiểm chứng khi xuất Excel theo mẫu MISA (_gen_mua_hang_*).
# ============================================================
_MUA_TEN = {"nk": "Mua hàng nhập kho", "kqk": "Mua hàng không qua kho",
            "dv": "Mua hàng dịch vụ"}
# Dấu hiệu NỘI BỘ đánh dấu chứng từ do chính phần mềm này ghi — lưu ở
# CustomField10 (không hiển thị trên diễn giải/JournalMemo, để diễn giải
# đọc tự nhiên như người dùng tự nhập tay: "Mua hàng - NCC - Số HĐ Ngày").
# Dùng để: (1) loại trừ khi "học mẫu" từ chứng từ thật, (2) cho phép ghi đè
# chứng từ do phần mềm tạo dù đang mang cờ đã ghi sổ.
_PM_MARK = "HDDT-AUTO"
# vị trí (0-based) các cột cần trong mảng phẳng do _gen_mua_hang_* sinh ra
_MUA_COT = {
    "nk": dict(doc=6, ngayct=5, sohd=10, mst=12, ten_ncc=13, ma=19, ten=20,
               dvt=25, sl=26, dgia=27, tt=28, no=23, co=24, ts=34, tthue=35,
               tk_thue=38, tkdu_thue=37, nhtg=41, nts=42, nthue=43, ntk=44,
               la_nk_col=1, kho=21),
    "kqk": dict(doc=6, ngayct=5, sohd=9, mst=11, ten_ncc=12, ma=17, ten=18,
                dvt=21, sl=22, dgia=23, tt=24, no=19, co=20, ts=30, tthue=31,
                tk_thue=34, tkdu_thue=33, nhtg=36, nts=37, nthue=38, ntk=39,
                la_nk_col=1, kho=None),
    "dv": dict(doc=6, ngayct=5, sohd=33, mst=7, ten_ncc=8, ma=14, ten=15,
               dvt=19, sl=20, dgia=21, tt=22, no=16, co=17, ts=27, tthue=28,
               tk_thue=30, tkdu_thue=None, nhtg=None, nts=None, nthue=None,
               ntk=None, la_nk_col=None, kho=None),
}

def _misa_pu_reftype(cur, tu_khoa, co_tk=None, master_table="PUVoucher"):
    """Chọn RefType phù hợp nhất trong SYSRefType (mặc định MasterTableName=
    'PUVoucher'; 'Mua hàng dịch vụ' dùng bảng RIÊNG 'PUService', khác hẳn
    PUVoucher — truyền master_table='PUService' cho loại đó) theo từ khóa
    loại chứng từ (vd 'nhập kho'/'không qua kho'/'dịch vụ') + gợi ý phương
    thức thanh toán theo TK Có (111->tiền mặt, 112->chuyển khoản/ngân hàng,
    khác->công nợ). Trả (RefType, RefTypeName) hoặc (None, None)."""
    rows = cur.execute(
        "SELECT RefType, RefTypeName FROM SYSRefType WHERE MasterTableName=?",
        master_table).fetchall()
    if not rows:
        return None, None
    pt_kw = ()
    if co_tk:
        if co_tk.startswith("111"):
            pt_kw = ("tiền mặt",)
        elif co_tk.startswith("112"):
            pt_kw = ("chuyển khoản", "ngân hàng")
        else:
            pt_kw = ("công nợ", "chưa thanh toán", "chưa trả")

    def diem(name):
        n = (name or "").lower()
        if not all(k in n for k in tu_khoa):
            return -1
        return 1 + (1 if any(k in n for k in pt_kw) else 0)

    best = max(rows, key=lambda r: diem(r[1]))
    if diem(best[1]) < 0:
        return None, None
    return int(best[0]), best[1]

def _num0(v):
    return v if isinstance(v, (int, float)) else 0

# Mẫu ĐẦY ĐỦ mọi cột của PUVoucher/PUVoucherDetail (trừ EditVersion — rowversion
# tự sinh, không được phép ghi). LUÔN liệt kê hết mọi cột (kể cả để NULL tường
# minh) thay vì bỏ qua cột không dùng — vì nhiều cột (đặc biệt các cột *Account
# FK sang bảng Account) có RÀNG BUỘC DEFAULT khác NULL trên CSDL thật của MISA;
# nếu bỏ qua, SQL Server tự điền giá trị default đó và có thể VI PHẠM FOREIGN
# KEY (vd cột SpecialConsumeTaxAccount) dù cột này NULL được. Ghi tường minh
# NULL để chắc chắn không dính default lạ.
_PU_HEADER_DEFAULT = {
    "RefID": None, "BranchID": None, "RefDate": None, "PostedDate": None,
    "CABARefDate": None, "CABAPostedDate": None, "RefType": None,
    "RefNoFinance": None, "RefNoManagement": None, "CABARefNoManagement": None,
    "CABARefNoFinance": None, "IsPostedFinance": 0, "IsPostedManagement": 0,
    "IncludeInvoice": 0, "PUInvoiceRefID": None, "AccountObjectID": None,
    "AccountObjectName": None, "AccountObjectAddress": None,
    "AccountObjectBankAccount": None, "AccountObjectBankName": None,
    "AccountObjectContactName": None, "IdentificationNumber": None,
    "IssueDate": None, "IssueBy": None, "Receiver": None, "JournalMemo": None,
    "EmployeeID": None, "DocumentIncluded": None, "CABAJournalMemo": None,
    "CABADocumentIncluded": None, "BankAccountID": None, "BankName": None,
    "PaymentTermID": None, "DueTime": None, "DueDate": None,
    "CurrencyID": "VND", "ExchangeRate": 1,
    "TotalAmountOC": 0, "TotalAmount": 0, "TotalImportTaxAmountOC": 0,
    "TotalImportTaxAmount": 0, "TotalVATAmountOC": 0, "TotalVATAmount": 0,
    "TotalDiscountAmountOC": 0, "TotalDiscountAmount": 0, "TotalFreightAmount": 0,
    "TotalInwardAmount": 0, "TotalSpecialConsumeTaxAmountOC": 0,
    "TotalSpecialConsumeTaxAmount": 0, "TotalCustomBeforeAmount": 0,
    "DisplayOnBook": 1, "IsPaid": 0, "IsPostedCashBookFinance": 0,
    "IsPostedCashBookManagement": 0, "CashBookPostedDate": None,
    "IsPostedInventoryBookFinance": 0, "IsPostedInventoryBookManagement": 0,
    "InventoryPostedDate": None, "RefOrder": 0, "CreatedDate": None,
    "CreatedBy": None, "ModifiedDate": None, "ModifiedBy": None,
    "CustomField1": None, "CustomField2": None, "CustomField3": None,
    "CustomField4": None, "CustomField5": None, "CustomField6": None,
    "CustomField7": None, "CustomField8": None, "CustomField9": None,
    "CustomField10": None, "CABAAmountOC": 0, "CABAAmount": 0,
    "INRefOrder": None, "IsPULotVoucher": 0, "AccountObjectAddressOther": None,
    "AccountObjectIdentificationNumberOther": None, "IsConvertVAT": None,
    "RefIDMshop": None, "RefNoMshop": None,
}
_PU_DETAIL_DEFAULT = {
    "RefDetailID": None, "RefID": None, "InventoryItemID": None,
    "Description": None, "StockID": None, "DebitAccount": None,
    "CreditAccount": None, "UnitID": None, "Quantity": 0, "UnitPrice": 0,
    "AmountOC": 0, "Amount": 0, "DiscountRate": 0, "DiscountAmountOC": 0,
    "DiscountAmount": 0, "FreightAmount": 0, "InwardAmount": 0,
    "PUContractID": None, "LotNo": None, "ExpiryDate": None,
    "MainUnitID": None, "MainConvertRate": 1, "ExchangeRateOperator": "*",
    "MainQuantity": 0, "MainUnitPrice": 0, "VATRate": None,
    "VATAmountOC": 0, "VATAmount": 0, "VATAccount": None,
    "DeductionDebitAccount": None, "PurchasePurposeID": None,
    "ExpenseItemID": None, "OrganizationUnitID": None, "JobID": None,
    "ProjectWorkID": None, "OrderID": None, "ContractID": None,
    "ListItemID": None, "PUOrderRefID": None, "PUOrderRefDetailID": None,
    "INProductionOrderRefID": None, "UnResonableCost": 0,
    "FOBAmountOC": 0, "FOBAmount": 0, "ImportChargeAmount": 0,
    "ImportTaxRatePrice": 0, "ImportTaxRate": 0, "ImportTaxAmountOC": 0,
    "ImportTaxAmount": 0, "ImportTaxAccount": None,
    "SpecialConsumeTaxRate": 0, "SpecialConsumeTaxAmountOC": 0,
    "SpecialConsumeTaxAmount": 0, "SpecialConsumeTaxAccount": None,
    "SortOrder": 0, "CustomField1": None, "CustomField2": None,
    "CustomField3": None, "CustomField4": None, "CustomField5": None,
    "CustomField6": None, "CustomField7": None, "CustomField8": None,
    "CustomField9": None, "CustomField10": None,
    "PUInvoiceRefID": None, "BudgetItemID": None, "ProductionID": None,
    "EnvironmentalTaxAmount": 0, "EnvironmentalTaxAccount": None,
    "InventoryResaleTypeID": 0, "EnvironmentalTaxAmountOC": 0,
    "VATDescription": None, "AccountObjectID": None,
    "TaxAccountObjectID": None, "TaxAccountObjectName": None,
    "TaxAccountObjectAddress": None, "TaxAccountObjectTaxCode": None,
    "InvDate": None, "InvNo": None, "InvSeries": None, "InvTemplateNo": None,
    "CashOutAmountFinance": 0, "CashOutDiffAmountFinance": 0,
    "CashOutVATAmountFinance": 0, "CashOutDiffVATAmountFinance": 0,
    "CashOutDiffAccountNumberFinance": None,
    "CashOutAmountManagement": 0, "CashOutDiffAmountManagement": 0,
    "CashOutVATAmountManagement": 0, "CashOutDiffVATAmountManagement": 0,
    "CashOutDiffAccountNumberManagement": None,
    # các giá trị 0 (không phải 1/NULL) — đối chiếu từng cột với dòng chi tiết
    # của chứng từ MISA THẬT cho thấy MISA để 0
    "CashOutExchangeRateFinance": 0, "CashOutExchangeRateManagement": 0,
    "AllocationRate": 0, "AllocationRateImport": 0,
    "DateEnoughTaxPayment": None, "UnitPriceAfterTax": 0,
    "ImportChargeExchangeRate": 0, "ImportTaxRatePriceOC": 0,
    "ImportChargeBeforeCustomAmountOC": 0,
    "ImportChargeBeforeCustomAmountMainCurrency": 0,
    "AllocationRateImportOriginCurrency": 0,
    "ImportChargeBeforeCustomAmountAllocated": 0,
    "EInvoiceItemName": None, "PanelLengthQuantity": 0,
    "PanelWidthQuantity": 0, "PanelHeightQuantity": 0,
    "PanelRadiusQuantity": 0, "PanelQuantity": 0,
    "LOANAgreementID": None, "InvestmentProjectID": None,
    "DeductionsTaxAmountOC": 0, "DeductionsTaxAmount": 0,
    "VATRate406": None, "VATRateOther": None,
    "AntiDumpingTaxRate": 0, "AntiDumpingTaxAmountOC": 0,
    "AntiDumpingTaxAmount": 0, "AntiDumpingTaxAccount": None,
    "PUContractDetailID": None,
}
# Bảng HÓA ĐƠN mua hàng (PUInvoice/PUInvoiceDetail) — tạo kèm khi ghi chứng từ
# Mua hàng "Nhận kèm hóa đơn": chứng từ MISA thật có IncludeInvoice=1 nhưng
# PUVoucher.PUInvoiceRefID vẫn NULL — hóa đơn liên kết CHIỀU NGƯỢC LẠI qua
# PUInvoiceDetail.PUVoucherRefID/PUVoucherRefDetailID. Liệt kê đủ mọi cột
# (trừ EditVersion rowversion) như bài học FK-default ở PUVoucher.
_PU_INV_DEFAULT = {
    "RefID": None, "BranchID": None, "RefType": None, "RefDate": None,
    "PostedDate": None, "RefNoFinance": None, "RefNoManagement": None,
    "IsPostedFinance": 0, "IsPostedManagement": 0, "IsImportPurchase": 0,
    "IncludeInvoice": 1, "AccountObjectID": None, "AccountObjectName": None,
    "AccountObjectAddress": None, "AccountObjectTaxCode": None,
    "JournalMemo": None, "EmployeeID": None, "InvTemplateNo": None,
    "InvDate": None, "InvSeries": None, "InvNo": None, "CurrencyID": "VND",
    "ExchangeRate": 1, "TotalTurnoverAmountOC": 0, "TotalTurnoverAmount": 0,
    "TotalVATAmountOC": 0, "TotalVATAmount": 0, "IsPaid": 0,
    "IsSummaryBySameInventoryItem": 0, "DisplayOnBook": None, "RefOrder": 0,
    "CreatedDate": None, "CreatedBy": None, "ModifiedDate": None, "ModifiedBy": None,
    "CustomField1": None, "CustomField2": None, "CustomField3": None,
    "CustomField4": None, "CustomField5": None, "CustomField6": None,
    "CustomField7": None, "CustomField8": None, "CustomField9": None,
    "CustomField10": None, "DueDate": None, "TransactionID": None,
    "SellerTaxCode": None, "EInvoiceType": None, "IsImportEInvoice": 0,
}
_PU_INV_DET_DEFAULT = {
    "RefDetailID": None, "RefID": None, "InventoryItemID": None, "Description": None,
    "DebitAccount": None, "CreditAccount": None, "TurnoverAmountOC": 0,
    "TurnoverAmount": 0, "VATRate": None, "VATAmountOC": 0, "VATAmount": 0,
    "PurchasePurposeID": None, "PUVoucherRefID": None, "PUVoucherRefDetailID": None,
    "PUContractID": None, "ListItemID": None, "ExpenseItemID": None,
    "OrganizationUnitID": None, "JobID": None, "ProjectWorkID": None,
    "OrderID": None, "ContractID": None, "UnResonableCost": 0, "SortOrder": 0,
    "PUVoucherRefNoFinance": None, "PUVoucherRefNoManagement": None,
    "CustomField1": None, "CustomField2": None, "CustomField3": None,
    "CustomField4": None, "CustomField5": None, "CustomField6": None,
    "CustomField7": None, "CustomField8": None, "CustomField9": None,
    "CustomField10": None, "VATDescription": None, "PUOrderRefID": None,
    "PUOrderRefDetailID": None, "InvestmentProjectID": None, "VATRateOther": None,
}
# "Mua hàng dịch vụ" KHÔNG dùng PUVoucher — nó có bảng RIÊNG PUService/
# PUServiceDetail (xác nhận từ cấu trúc CSDL thật + đối chiếu SYSRefType:
# không có loại chứng từ nào chứa "dịch vụ" trong MasterTableName='PUVoucher').
# Không có PUInvoice liên kết — thông tin hóa đơn (InvNo/InvDate/NCC thuế)
# nằm THẲNG trên từng dòng PUServiceDetail. Không có cột Kho/ĐVT chính quy
# đổi (dịch vụ không nhập kho). Mẫu đầy đủ cột (trừ EditVersion) như trên.
_PU_SERVICE_DEFAULT = {
    "RefID": None, "BranchID": None, "RefDate": None, "PostedDate": None,
    "RefType": None, "RefNoFinance": None, "RefNoManagement": None,
    "IsPostedFinance": 0, "IsPostedManagement": 0, "IsFreightService": 0,
    "AccountObjectID": None, "AccountObjectName": None, "AccountObjectAddress": None,
    "AccountObjectBankAccount": None, "AccountObjectBankName": None,
    "AccountObjectContactname": None, "IdentificationNumber": None,
    "IssueDate": None, "IssueBy": None, "JournalMemo": None, "EmployeeID": None,
    "DocumentIncluded": None, "BankAccountID": None, "BankName": None,
    "PaymentTermID": None, "DueTime": None, "DueDate": None,
    "CurrencyID": "VND", "ExchangeRate": 1,
    "TotalAmountOC": 0, "TotalAmount": 0, "TotalVATAmountOC": 0, "TotalVATAmount": 0,
    "TotalDiscountAmountOC": 0, "TotalDiscountAmount": 0,
    "DisplayOnBook": 1, "IsPaid": 0, "IsPostedCashBookFinance": 0,
    "IsPostedCashBookManagement": 0, "CashBookPostedDate": None,
    "RefOrder": 0, "CreatedDate": None, "CreatedBy": None,
    "ModifiedDate": None, "ModifiedBy": None,
    "CustomField1": None, "CustomField2": None, "CustomField3": None,
    "CustomField4": None, "CustomField5": None, "CustomField6": None,
    "CustomField7": None, "CustomField8": None, "CustomField9": None,
    "CustomField10": None, "CABAAmountOC": 0, "CABAAmount": 0,
    "PUInvoiceRefID": None, "IncludeInvoice": 0,
    "AccountObjectAddressOther": None, "AccountObjectIdentificationNumberOther": None,
    "TransactionID": None, "SellerTaxCode": None, "EInvoiceType": None,
    "IsImportEInvoice": 0,
}
_PU_SERVICE_DET_DEFAULT = {
    "RefDetailID": None, "RefID": None, "InventoryItemID": None, "Description": None,
    "DebitAccount": None, "CreditAccount": None, "UnitID": None,
    "Quantity": 0, "UnitPrice": 0, "AmountOC": 0, "Amount": 0,
    "DiscountRate": 0, "DiscountAmountOC": 0, "DiscountAmount": 0, "InwardAmount": 0,
    "PUContractID": None, "VATRate": None, "VATAmountOC": 0, "VATAmount": 0,
    "VATAccount": None, "InvTemplateNo": None, "InvSeries": None,
    "InvDate": None, "InvNo": None, "PurchasePurposeID": None,
    "TaxAccountObjectID": None, "TaxAccountObjectName": None,
    "TaxAccountObjectTaxCode": None, "TaxAccountObjectAddress": None,
    "BudgetItemID": None, "ExpenseItemID": None, "OrganizationUnitID": None,
    "AccountObjectID": None, "JobID": None, "ProjectWorkID": None,
    "OrderID": None, "ContractID": None, "ListItemID": None, "UnResonableCost": 0,
    "PUOrderRefDetailID": None, "PUOrderRefID": None, "SortOrder": 0,
    "CustomField1": None, "CustomField2": None, "CustomField3": None,
    "CustomField4": None, "CustomField5": None, "CustomField6": None,
    "CustomField7": None, "CustomField8": None, "CustomField9": None,
    "CustomField10": None, "VATDescription": None,
    "CashOutAmountFinance": 0, "CashOutDiffAmountFinance": 0,
    "CashOutVATAmountFinance": 0, "CashOutDiffVATAmountFinance": 0,
    "CashOutDiffAccountNumberFinance": None,
    "CashOutAmountManagement": 0, "CashOutDiffAmountManagement": 0,
    "CashOutVATAmountManagement": 0, "CashOutDiffVATAmountManagement": 0,
    "CashOutDiffAccountNumberManagement": None,
    "CashOutExchangeRateFinance": 0, "CashOutExchangeRateManagement": 0,
    "EInvoiceItemName": None, "LOANAgreementID": None, "InvestmentProjectID": None,
    "DeductionsTaxAmountOC": 0, "DeductionsTaxAmount": 0,
    "VATRate406": None, "VATRateOther": None,
}

def _misa_tu_kiem_tra_muahang(cur, ref_ids, branch_id, bang_chinh, ket):
    """TỰ KIỂM TRA sau khi ghi Mua hàng: chạy lại đúng view + BỘ LỌC THẬT của
    màn hình MISA "Mua hàng hóa, dịch vụ" (bắt được từ câu lệnh MISA thực thi:
    View_PUVoucherService WHERE PostedDate trong kỳ + BranchID + DisplayOnBook
    IN (0,2)). Khi có chứng từ KHÔNG lọt bộ lọc, TỰ CHẨN ĐOÁN đích danh lý do
    từng chứng từ: (a) không có mặt trong view (view JOIN rớt — vd thiếu hóa
    đơn kèm/loại chứng từ), hay (b) có trong view nhưng trượt 1 trong 3 điều
    kiện lọc — kèm giá trị PostedDate/BranchID/DisplayOnBook THẬT đã ghi để
    so. bang_chinh: 'PUVoucher' (nk/kqk) hoặc 'PUService' (dv). Gắn cờ
    view_thay vào từng dòng ket. Trả dict tu_kiem_tra."""
    try:
        ph = ",".join(["?"] * len(ref_ids))
        thay = cur.execute(
            "SELECT RefID FROM View_PUVoucherService WHERE RefID IN (%s) "
            "AND PostedDate IS NOT NULL AND BranchID=? "
            "AND (DisplayOnBook = 0 OR DisplayOnBook = 2)" % ph,
            ref_ids + [branch_id]).fetchall()
        thay_ids = {str(r[0]).upper() for r in thay}
        for x in ket:
            if x.get("ref_id"):
                x["view_thay"] = str(x["ref_id"]).upper() in thay_ids
        tkt = {
            "view": "View_PUVoucherService (kèm đúng bộ lọc thật của màn hình MISA: "
                    "PostedDate + chi nhánh + DisplayOnBook 0/2)",
            "so_ghi": len(ref_ids), "so_view_thay": len(thay),
            "ket_luan": ("Chứng từ LỌT QUA đúng bộ lọc màn hình MISA -> sẽ hiện "
                        "trên lưới (bấm 'Lấy dữ liệu' trên MISA)."
                        if len(thay) == len(ref_ids) else
                        "Chứng từ KHÔNG lọt qua bộ lọc màn hình MISA — xem chẩn "
                        "đoán từng dòng bên dưới."),
        }
        if len(thay) != len(ref_ids):
            # chẩn đoán: chứng từ có MẶT trong view không (không kèm lọc)?
            trong_view = set()
            try:
                trong_view = {str(r[0]).upper() for r in cur.execute(
                    "SELECT RefID FROM View_PUVoucherService WHERE RefID IN (%s)" % ph,
                    ref_ids).fetchall()}
            except Exception:
                pass
            # giá trị THẬT đã ghi của 3 cột bị lọc
            dong_loi = []
            try:
                for rid, rn, pd, br, dob in cur.execute(
                        "SELECT RefID, RefNoManagement, PostedDate, BranchID, DisplayOnBook "
                        "FROM %s WHERE RefID IN (%s)" % (bang_chinh, ph),
                        ref_ids).fetchall():
                    k = str(rid).upper()
                    if k in thay_ids:
                        continue
                    dong_loi.append({
                        "so_ct": rn,
                        "trong_view": k in trong_view,
                        "posted_date": (pd.strftime("%d/%m/%Y")
                                        if hasattr(pd, "strftime") else str(pd or "NULL")),
                        "dung_chi_nhanh": (str(br).upper() == str(branch_id).upper()
                                           if br is not None and branch_id is not None else False),
                        "display_on_book": dob,
                    })
            except Exception:
                pass
            tkt["dong_loi"] = dong_loi[:50]
            tkt["goi_y"] = ("Chứng từ KHÔNG có mặt trong View_PUVoucherService -> "
                            "view JOIN rớt (vd bảng hóa đơn kèm/loại chứng từ)."
                            if dong_loi and not any(d["trong_view"] for d in dong_loi) else
                            "Chứng từ CÓ trong view nhưng trượt bộ lọc — so 3 cột "
                            "PostedDate / đúng chi nhánh / DisplayOnBook ở bảng dưới.")
        return tkt
    except Exception as e:
        return {"loi": "Không tự kiểm tra được view: %s" % str(e)[:200]}

def _misa_ghi_mua_hang(cid, database, loai, preview=True, ghi_de=False):
    """Ghi chứng từ Mua hàng NHẬP KHO/KHÔNG QUA KHO (loai: nk/kqk) thẳng vào
    PUVoucher — xem cảnh báo ở đầu file. loai='dv' (Mua hàng dịch vụ) KHÔNG
    dùng hàm này nữa — xem _misa_ghi_mua_hang_dv() (bảng RIÊNG PUService, xác
    nhận qua cấu trúc CSDL thật). Dữ liệu lấy từ Bảng kê đầu vào ĐÃ LƯU
    (nhap_lieu 'in'). ghi_de=True: gỡ các chứng từ CHƯA GHI SỔ trùng số (do
    phần mềm tạo, có thể sai loại CT) rồi ghi lại — KHÔNG đụng chứng từ đã
    ghi sổ."""
    import uuid as _uuid
    if loai not in _MUA_COT:
        raise HTTPException(400, "Loại '%s' chưa hỗ trợ (chỉ nk/kqk/dv)." % loai)
    cfg = _MUA_COT[loai]
    dl = nhap_lieu_get(cid, "in")
    header, rows = dl.get("header") or [], dl.get("rows") or []
    if not rows:
        raise HTTPException(400, "Chưa có Bảng kê đầu vào đã lưu — Import & Lưu trước.")
    if loai == "nk":
        flat = _gen_mua_hang_nk(cid, header, rows)
    elif loai == "kqk":
        flat = _gen_mua_hang_kqk(cid, header, rows)
    else:
        flat = _gen_mua_hang_dv(cid, header, rows)
    if not flat:
        return {"preview": preview, "database": database, "so_chungtu": 0, "so_dong": 0,
                "so_trung": 0, "so_bo_qua_ncc": 0, "so_bo_qua_mahang": 0, "danh_sach": [],
                "ghi_chu": "Không có dòng nào phù hợp (%s) trong Bảng kê đầu vào." % _MUA_TEN[loai]}
    # gom các dòng theo Số chứng từ / Số phiếu nhập -> 1 nhóm = 1 PUVoucher
    groups, order = {}, []
    for r in flat:
        doc = str(r[cfg["doc"]] or "").strip()
        if not doc:
            continue
        if doc not in groups:
            groups[doc] = []
            order.append(doc)
        groups[doc].append(r)

    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        branch_id = _misa_branch_id(cur)
        ncc = {}   # mst/mã ĐT (lower) -> (AccountObjectID, tên trong MISA)
        for aid, taxcode, code, name in cur.execute(
                "SELECT AccountObjectID, CompanyTaxCode, AccountObjectCode, AccountObjectName "
                "FROM AccountObject").fetchall():
            nm = str(name or "")
            if taxcode:
                ncc[_misa_khncc_chuan_mst(taxcode).lower()] = (aid, nm)
            if code:
                ncc[str(code).strip().lower()] = (aid, nm)
        hang = {}  # mã hàng (lower) -> (InventoryItemID, UnitID, tên trong MISA)
        for iid, code, uid, ten_h in cur.execute(
                "SELECT InventoryItemID, InventoryItemCode, UnitID, InventoryItemName "
                "FROM InventoryItem").fetchall():
            if code:
                hang[str(code).strip().lower()] = (iid, uid, str(ten_h or ""))
        # danh mục TÀI KHOẢN thật của MISA — các cột TK trên PUVoucherDetail
        # có FK sang Account.AccountNumber; xem _misa_tk_fallback
        tk_set = set()
        try:
            for (an,) in cur.execute("SELECT AccountNumber FROM Account").fetchall():
                if an:
                    tk_set.add(str(an).strip())
        except Exception:
            pass
        # ĐƠN VỊ TÍNH: InventoryItem.UnitID có thể "MỒ CÔI" (GUID không còn
        # trong bảng Unit — lỗi dữ liệu danh mục cũ) khiến MISA hiện GUID thô
        # ở cột ĐVT trên chứng từ. Nạp bảng Unit để kiểm tra và TỰ SỬA: tra/
        # tạo Unit theo tên ĐVT trong Bảng kê rồi cập nhật lại danh mục hàng.
        unit_ten = {}    # tên ĐVT (lower) -> UnitID
        unit_ids = set()   # CHỈ chứa Unit tên SẠCH — Unit tên rác coi như "mồ côi" để bị sửa
        try:
            for u_id, u_name in cur.execute("SELECT UnitID, UnitName FROM Unit").fetchall():
                if u_name and _misa_unit_hong(u_name):
                    continue
                unit_ids.add(str(u_id).strip().lower())
                if u_name:
                    unit_ten[str(u_name).strip().lower()] = u_id
        except Exception:
            pass
        so_dvt_sua = 0

        def sua_dvt(r, iid, uid):
            """Trả UnitID dùng được cho dòng chứng từ; UnitID mồ côi thì thay
            bằng Unit thật theo tên ĐVT của Bảng kê (tạo mới nếu chưa có) và
            sửa luôn InventoryItem để hết tận gốc."""
            nonlocal so_dvt_sua
            if not unit_ids or uid is None or str(uid).strip().lower() in unit_ids:
                return uid
            import uuid as _uuid2
            dvt_text = (str(r[cfg["dvt"]] or "").strip()
                        if cfg.get("dvt") is not None else "")
            uk = dvt_text.lower()
            if uk and uk in unit_ten:
                uid2 = unit_ten[uk]
            elif dvt_text:
                uid2 = str(_uuid2.uuid4())
                if not preview:
                    cur.execute("INSERT INTO Unit (UnitID, UnitName, Description, Inactive) "
                                "VALUES (?,?,?,0)", uid2, dvt_text[:20], None)
                unit_ten[uk] = uid2
                unit_ids.add(uid2.lower())
            else:
                uid2 = None
            if not preview:
                try:
                    cur.execute("UPDATE InventoryItem SET UnitID=? WHERE InventoryItemID=?",
                               uid2, iid)
                except Exception:
                    pass
            so_dvt_sua += 1
            return uid2
        # chứng từ đã có trong MISA (dò theo Số chứng từ = RefNoManagement) —
        # KHÔNG đụng chứng từ đã ghi sổ CỦA NGƯỜI DÙNG; riêng chứng từ do CHÍNH
        # PHẦN MỀM tạo (nhận diện qua CustomField10=_PM_MARK) thì luôn cho phép
        # ghi đè kể cả khi đang mang cờ "đã ghi sổ".
        reftype_ten = dict(cur.execute(
            "SELECT RefType, RefTypeName FROM SYSRefType WHERE MasterTableName='PUVoucher'").fetchall())
        # kho (Stock) — điền StockID cho dòng nhập kho theo cột "Kho" của form.
        # Kho THIẾU trên MISA sẽ được TỰ TẠO (như Unit) để cột Kho trên chứng
        # từ nhập kho không bị trống — trước đây chỉ cảnh báo rồi để trống nên
        # chứng từ hiện thiếu Kho (và có thể báo "Failed to enable constraints").
        kho_map = {}
        try:
            for sid, scode in cur.execute("SELECT StockID, StockCode FROM Stock").fetchall():
                if scode:
                    kho_map[str(scode).strip().lower()] = sid
        except Exception:
            pass
        kho_moi = set()   # mã kho vừa tự tạo (để báo lại)
        now_stock = datetime.datetime.now()

        def tra_kho(kho_text):
            """StockID theo mã kho; tự tạo Stock nếu chưa có (không để trống)."""
            kc = str(kho_text or "").strip()
            if not kc:
                return None
            kl = kc.lower()
            if kl in kho_map:
                return kho_map[kl]
            sid = str(_uuid.uuid4())
            if not preview:
                try:
                    cur.execute(
                        "INSERT INTO Stock (StockID, StockCode, StockName, Description, "
                        "Inactive, CreatedDate, CreatedBy, InventoryAccount, BranchID) "
                        "VALUES (?,?,?,?,0,?,?,?,?)",
                        sid, kc[:20], kc[:128], None, now_stock, hoc_nguoi_tao, None, branch_id)
                except Exception:
                    return None
            kho_map[kl] = sid
            kho_moi.add(kc.upper())
            return sid
        # loại chứng từ của bảng HÓA ĐƠN mua hàng (PUInvoice) — để tạo hóa đơn
        # kèm theo ("Nhận kèm hóa đơn"); ưu tiên tên có chữ "mua"
        inv_reftype = None
        try:
            _inv_rows = cur.execute(
                "SELECT RefType, RefTypeName FROM SYSRefType "
                "WHERE MasterTableName='PUInvoice'").fetchall()
            if _inv_rows:
                inv_reftype = next((int(r[0]) for r in _inv_rows
                                    if "mua" in str(r[1] or "").lower()), int(_inv_rows[0][0]))
        except Exception:
            pass
        posted_refno = set()
        unposted_docs = {}   # k_doc -> {"refids": [...], "reftype_ten": name}
        for refid, rn, rt, pf, pm, mark in cur.execute(
                "SELECT RefID, RefNoManagement, RefType, ISNULL(IsPostedFinance,0), "
                "ISNULL(IsPostedManagement,0), ISNULL(CustomField10,'') FROM PUVoucher").fetchall():
            if not rn:
                continue
            k = str(rn).strip().lower()
            la_cua_minh = mark == _PM_MARK
            if (pf or pm) and not la_cua_minh:
                posted_refno.add(k)
            else:
                d = unposted_docs.setdefault(k, {"refids": [], "reftype_ten": reftype_ten.get(rt)})
                d["refids"].append(refid)
        # HỌC "mẫu" từ chứng từ Mua hàng MISA THẬT (do người dùng/MISA tạo,
        # đang HIỂN THỊ được) — CHỈ học từ chứng từ THẬT, KHÔNG học từ bản ghi
        # do PHẦN MỀM tạo (CustomField10=_PM_MARK) để tránh vòng lặp học lại
        # chính bản ghi của mình.
        # - RefType: khớp theo loại (mua + nhập kho/không qua kho/dịch vụ).
        # - DisplayOnBook: lấy giá trị PHỔ BIẾN NHẤT của chứng từ THẬT (bất kỳ
        #   loại nào) vì đây là thiết lập hiển thị chung, quyết định chứng từ có
        #   hiện trong danh sách hay không.
        tu_khoa_loai = {"nk": ["mua", "nhập kho"], "kqk": ["mua", "không qua kho"],
                        "dv": ["mua", "dịch vụ"]}[loai]
        loai_ct_dang_co = []
        mau_that = []   # chẩn đoán: mẫu chứng từ THẬT (để đối chiếu)
        hoc = None      # RefType học theo loại
        hoc_dob = None  # DisplayOnBook phổ biến của chứng từ THẬT
        hoc_branch = None      # BranchID phổ biến của chứng từ THẬT
        hoc_branch_ten = None  # tên chi nhánh tương ứng (để hiển thị)
        hoc_nguoi_tao = None   # CreatedBy phổ biến của chứng từ THẬT
        max_reforder = 0       # RefOrder lớn nhất đang có (để ghi tiếp theo, không để 0)
        try:
            _dem_loai, _dem_dob, _dem_branch, _dem_nguoi = {}, {}, {}, {}
            for rt, rn, dob, inc, mark, rnm, pf, pm, br, brn, cb, ro in cur.execute(
                    "SELECT pv.RefType, rt.RefTypeName, pv.DisplayOnBook, pv.IncludeInvoice, "
                    "ISNULL(pv.CustomField10,''), pv.RefNoManagement, "
                    "ISNULL(pv.IsPostedFinance,0), ISNULL(pv.IsPostedManagement,0), "
                    "pv.BranchID, ou.OrganizationUnitName, pv.CreatedBy, ISNULL(pv.RefOrder,0) "
                    "FROM PUVoucher pv JOIN SYSRefType rt ON rt.RefType=pv.RefType "
                    "LEFT JOIN OrganizationUnit ou ON ou.OrganizationUnitID=pv.BranchID").fetchall():
                max_reforder = max(max_reforder, ro or 0)
                if mark == _PM_MARK:
                    continue   # bỏ qua bản ghi do phần mềm tạo
                key = (int(rt), rn)
                _dem_loai[key] = _dem_loai.get(key, 0) + 1
                if dob is not None:
                    _dem_dob[dob] = _dem_dob.get(dob, 0) + 1
                if br is not None:
                    _dem_branch[br] = _dem_branch.get(br, [0, brn])
                    _dem_branch[br][0] += 1
                if cb:
                    _dem_nguoi[cb] = _dem_nguoi.get(cb, 0) + 1
                if len(mau_that) < 5:
                    mau_that.append({"so_ct": rnm, "loai": rn, "display_on_book": dob,
                                     "include_invoice": inc, "da_ghi_so": bool(pf or pm),
                                     "chi_nhanh": brn})
            for (rt, rn), c in sorted(_dem_loai.items(), key=lambda kv: -kv[1]):
                loai_ct_dang_co.append({"ten": rn, "so": c})
                n = (rn or "").lower()
                if hoc is None and all(k in n for k in tu_khoa_loai):
                    hoc = {"reftype": rt, "refname": rn}
            if _dem_dob:
                hoc_dob = max(_dem_dob.items(), key=lambda kv: kv[1])[0]
            if _dem_branch:
                hoc_branch, (_, hoc_branch_ten) = max(
                    _dem_branch.items(), key=lambda kv: kv[1][0])
            if _dem_nguoi:
                hoc_nguoi_tao = max(_dem_nguoi.items(), key=lambda kv: kv[1])[0]
        except Exception:
            pass
        # PurchasePurposeID (Nhóm HHDV mua vào) — chứng từ thật LUÔN có; học từ
        # dòng chi tiết thật, fallback tra bảng PurchasePurpose theo mã '1'
        # (Hàng hóa, dịch vụ trong nước — đúng giá trị cột "Nhóm HHDV mua vào"
        # của form Excel mẫu).
        hoc_purpose = None
        try:
            row_pp = cur.execute(
                "SELECT TOP 1 d.PurchasePurposeID FROM PUVoucherDetail d "
                "JOIN PUVoucher pv ON pv.RefID=d.RefID "
                "WHERE ISNULL(pv.CustomField10,'') <> ? "
                "AND d.PurchasePurposeID IS NOT NULL", _PM_MARK).fetchone()
            if row_pp:
                hoc_purpose = row_pp[0]
            if hoc_purpose is None:
                row_pp = cur.execute(
                    "SELECT TOP 1 PurchasePurposeID FROM PurchasePurpose "
                    "WHERE PurchasePurposeCode='1'").fetchone()
                if row_pp:
                    hoc_purpose = row_pp[0]
        except Exception:
            pass
        # Ưu tiên chi nhánh (BranchID) học được từ chứng từ Mua hàng MISA THẬT
        # — nếu công ty có nhiều chi nhánh, màn hình lưới thường lọc theo chi
        # nhánh đang chọn nên ghi SAI chi nhánh sẽ khiến chứng từ "biến mất".
        if hoc_branch is not None:
            branch_id = hoc_branch

        now = datetime.datetime.now()
        ket = []
        them_ct = them_dong = trung = bo_ncc = bo_mahang = go = so_ngay_loi = so_tien_0 = 0
        so_hoa_don = 0        # số hóa đơn (PUInvoice) tạo kèm chứng từ
        thieu_kho = set()     # mã kho trong form nhưng CHƯA có trong MISA
        bo_tk = 0             # chứng từ bỏ qua vì TK Nợ/Có không có trong MISA
        tk_thay = set()       # "2421→242": TK đã tự rút về TK cha đang tồn tại

        def tra_tk(tk, bat_buoc=False):
            """Đối chiếu 1 số TK với danh mục Account thật; ghi nhận thay thế."""
            t = str(tk or "").strip()
            if not t or not tk_set:
                return t or None
            t2 = _misa_tk_fallback(t, tk_set)
            if t2 and t2 != t:
                tk_thay.add("%s→%s" % (t, t2))
            return t2
        for doc in order:
            lines = groups[doc]
            k_doc = doc.strip().lower()
            if k_doc in posted_refno:
                trung += 1
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "đã ghi sổ trong MISA (bỏ qua)"})
                continue
            if k_doc in unposted_docs and not ghi_de:
                trung += 1
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "loai_ct_misa": unposted_docs[k_doc]["reftype_ten"],
                            "trang_thai": "đã có (chưa ghi sổ, bỏ qua)"})
                continue
            first = lines[0]
            mst = str(first[cfg["mst"]] or "").strip()
            mst_k = mst.lower()
            if mst_k not in ncc:
                bo_ncc += 1
                ket.append({"so_ct": doc, "so_dong": len(lines), "ncc_mst": mst,
                            "trang_thai": "bỏ qua — NCC (MST %s) chưa có trong MISA" % mst})
                continue
            ghi_de_ct = k_doc in unposted_docs   # chứng từ cũ (chưa ghi sổ) cần gỡ ghi lại
            if ghi_de_ct and not preview:
                for rid in unposted_docs[k_doc]["refids"]:
                    # dò HÓA ĐƠN đã tạo kèm (link 2 chiều: PUInvoiceDetail.
                    # PUVoucherRefID và PUVoucher.PUInvoiceRefID) để gỡ cùng
                    inv_ids = set()
                    try:
                        inv_ids.update(r[0] for r in cur.execute(
                            "SELECT DISTINCT RefID FROM PUInvoiceDetail "
                            "WHERE PUVoucherRefID=?", rid).fetchall())
                        row_h = cur.execute(
                            "SELECT PUInvoiceRefID FROM PUVoucher WHERE RefID=?", rid).fetchone()
                        if row_h and row_h[0]:
                            inv_ids.add(row_h[0])
                    except Exception:
                        pass
                    # gỡ CHỨNG TỪ TRƯỚC (PUVoucher.PUInvoiceRefID trỏ tới hóa
                    # đơn nên phải gỡ chứng từ rồi mới gỡ được hóa đơn — ngược
                    # với thứ tự lúc ghi)
                    cur.execute("DELETE FROM PUVoucherDetailCost WHERE RefID=?", rid)
                    cur.execute("DELETE FROM PUVoucherDetail WHERE RefID=?", rid)
                    cur.execute("DELETE FROM PUVoucher WHERE RefID=?", rid)
                    for ivid in inv_ids:
                        try:
                            cur.execute("DELETE FROM PUInvoiceDetail WHERE RefID=?", ivid)
                            cur.execute("DELETE FROM PUInvoice WHERE RefID=?", ivid)
                        except Exception:
                            pass
            acc_obj_id, ten_ncc_misa = ncc[mst_k]
            valid_lines = []
            for r in lines:
                ma = str(r[cfg["ma"]] or "").strip()
                mk = ma.lower()
                if mk not in hang:
                    bo_mahang += 1
                    continue
                valid_lines.append((r, hang[mk]))
            if not valid_lines:
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "bỏ qua — tất cả dòng đều thiếu mã hàng trong MISA"})
                continue
            # TK Nợ/Có phải tồn tại trong danh mục Account MISA (FK) — TK
            # không rút được về TK cha thì bỏ qua chứng từ, báo rõ lý do.
            if tk_set:
                tk_loi = set()
                for r, _h in valid_lines:
                    for acc in (str(r[cfg["no"]] or "").strip(),
                                str(r[cfg["co"]] or "").strip()):
                        if acc and _misa_tk_fallback(acc, tk_set) is None:
                            tk_loi.add(acc)
                if tk_loi:
                    bo_tk += 1
                    ket.append({"so_ct": doc, "so_dong": len(lines),
                                "trang_thai": "bỏ qua — TK %s không có trong danh mục "
                                              "tài khoản MISA" % ", ".join(sorted(tk_loi))})
                    continue
            ngay_dt = None
            ngay_str = str(first[cfg["ngayct"]] or "").strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    ngay_dt = datetime.datetime.strptime(ngay_str, fmt)
                    break
                except Exception:
                    pass
            # KHÔNG đọc được ngày từ dữ liệu -> rơi về NGÀY HÔM NAY, dễ khiến
            # chứng từ nằm NGOÀI khoảng "Kỳ" đang lọc trên màn hình MISA (vd
            # dữ liệu là hóa đơn năm 2025 nhưng ghi nhầm thành ngày hôm nay) -
            # đánh dấu lại để báo rõ trong kết quả.
            ngay_loi = ngay_dt is None
            ngay_dt = ngay_dt or now
            co_tk = str(first[cfg["co"]] or "").strip()
            # Ưu tiên "mẫu" học được từ chứng từ MISA đang hiển thị được (chắc
            # chắn đúng). Nếu công ty chưa có chứng từ loại này thì mới dò
            # SYSRefType theo từ khóa ("mua" + loại) làm dự phòng.
            if hoc:
                ref_type, ref_type_ten = hoc["reftype"], hoc["refname"]
            else:
                tu_khoa = {"nk": ["mua", "nhập kho"], "kqk": ["mua", "không qua kho"],
                           "dv": ["mua", "dịch vụ"]}[loai]
                ref_type, ref_type_ten = _misa_pu_reftype(cur, tu_khoa, co_tk)
            if not ref_type:
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "bỏ qua — không dò được loại chứng từ MISA phù hợp"})
                continue
            ref_id = str(_uuid.uuid4())
            # sinh sẵn RefID hóa đơn kèm để link từ header + TỪNG DÒNG chi tiết
            # (chứng từ NHẬP KHO thật: cả PUVoucher.PUInvoiceRefID lẫn
            # PUVoucherDetail.PUInvoiceRefID đều trỏ tới hóa đơn)
            inv_id = str(_uuid.uuid4()) if inv_reftype else None
            total_amount = total_vat = total_import_tax = 0
            detail_rows = []
            for idx, (r, (iid, uid, ten_h)) in enumerate(valid_lines, 1):
                uid = sua_dvt(r, iid, uid)
                sl = _num0(r[cfg["sl"]])
                dgia = _num0(r[cfg["dgia"]])
                tt = _num0(r[cfg["tt"]])
                ts = _num0(r[cfg["ts"]]) if cfg["ts"] is not None else None
                tthue = _num0(r[cfg["tthue"]])
                # mọi cột TK đều đối chiếu danh mục Account (FK) — TK không
                # có tiểu khoản tự rút về TK cha (vd 2421 -> 242)
                tk_thue = tra_tk(r[cfg["tk_thue"]])
                tkdu = (tra_tk(r[cfg["tkdu_thue"]])
                        if cfg["tkdu_thue"] is not None else None)
                no_acc = tra_tk(r[cfg["no"]])
                co_acc = tra_tk(r[cfg["co"]])
                nk_tg = _num0(r[cfg["nhtg"]]) if cfg["nhtg"] is not None else 0
                nk_ts = _num0(r[cfg["nts"]]) if cfg["nts"] is not None else 0
                nk_thue = _num0(r[cfg["nthue"]]) if cfg["nthue"] is not None else 0
                nk_tk = (tra_tk(r[cfg["ntk"]])
                         if (cfg["ntk"] is not None and nk_thue) else None)
                total_amount += tt
                total_vat += tthue
                total_import_tax += nk_thue
                # Kho: tra StockID theo cột "Kho" của form (vd 'HH'/'NVL') —
                # thiếu kho là 1 nguyên nhân MISA báo "Failed to enable
                # constraints" khi mở chứng từ nhập kho
                stock_id = None
                if cfg.get("kho") is not None:
                    stock_id = tra_kho(r[cfg["kho"]])
                mo_ta = ten_h or str(r[cfg["ten"]] or "")
                detail_rows.append(dict(_PU_DETAIL_DEFAULT, **{
                    "RefDetailID": str(_uuid.uuid4()), "RefID": ref_id, "InventoryItemID": iid,
                    "Description": mo_ta, "DebitAccount": no_acc,
                    "CreditAccount": co_acc, "UnitID": uid, "Quantity": sl, "UnitPrice": dgia,
                    "AmountOC": tt, "Amount": tt, "MainQuantity": sl, "MainUnitPrice": dgia,
                    # MainUnitID = ĐVT chính (bằng chính ĐVT khi không có quy đổi)
                    # + NCC trên từng dòng + Kho — khớp form nhập tay của MISA
                    "MainUnitID": uid, "StockID": stock_id, "AccountObjectID": acc_obj_id,
                    # giá trị NHẬP KHO/FOB = thành tiền — CHỈ áp dụng cho loại
                    # thực sự nhập kho (có Kho); "không qua kho"/"dịch vụ" theo
                    # đúng tên gọi KHÔNG có hàng nhập kho nên để 0 (mặc định).
                    "InwardAmount": tt if cfg.get("kho") is not None else 0,
                    "FOBAmountOC": tt if cfg.get("kho") is not None else 0,
                    "FOBAmount": tt if cfg.get("kho") is not None else 0,
                    # NCC thuế, nhóm HHDV mua vào, diễn giải thuế, link hóa đơn
                    # kèm trên TỪNG dòng — áp dụng chung mọi loại
                    "TaxAccountObjectID": acc_obj_id, "PurchasePurposeID": hoc_purpose,
                    "VATDescription": ("Thuế GTGT - %s" % mo_ta)[:255],
                    "PUInvoiceRefID": inv_id,
                    "VATRate": ts, "VATAmountOC": tthue, "VATAmount": tthue,
                    "VATAccount": tk_thue, "DeductionDebitAccount": tkdu,
                    "ImportTaxRatePrice": nk_tg, "ImportTaxRate": nk_ts,
                    "ImportTaxAmountOC": nk_thue, "ImportTaxAmount": nk_thue,
                    "ImportTaxAccount": nk_tk, "SortOrder": idx - 1,
                    "InvNo": str(r[cfg["sohd"]] or "")[:25] if cfg["sohd"] is not None else None,
                    "InvDate": ngay_dt, "InvSeries": "", "InvTemplateNo": "",
                }))
            # DisplayOnBook BẮT BUỘC phải lọt bộ lọc thật của màn hình MISA
            # "Mua hàng hóa, dịch vụ": DisplayOnBook IN (0, 2) — bắt được từ
            # câu lệnh MISA thực thi. Bài học thực tế: khi công ty không còn
            # chứng từ thật nào để học (hoc_dob=None), mặc định cũ 3 làm chứng
            # từ TRƯỢT bộ lọc và biến mất khỏi lưới (tự kiểm tra chỉ ra đích
            # danh cột này). Mặc định 0 = hiện trên cả 2 sổ; giá trị học được
            # cũng phải thuộc (0,2) mới dùng.
            dob = hoc_dob if hoc_dob in (0, 2) else 0
            ten_ncc_dg = ten_ncc_misa or str(first[cfg["ten_ncc"]] or "")
            so_hd_dg = str(first[cfg["sohd"]] or "").strip() if cfg["sohd"] is not None else ""
            dien_giai = ("Mua hàng - %s - %s %s" %
                        (ten_ncc_dg, so_hd_dg, ngay_dt.strftime("%d/%m/%Y"))).strip()[:500]
            header_cols = dict(_PU_HEADER_DEFAULT, **{
                "RefID": ref_id, "BranchID": branch_id, "RefDate": ngay_dt,
                # PostedDate (ngày HẠCH TOÁN) BẮT BUỘC phải có: bắt câu lệnh thật
                # của MISA cho thấy màn hình lưới lọc "Kỳ" theo PostedDate
                # (PostedDate BETWEEN ...) chứ KHÔNG theo RefDate — để NULL là
                # chứng từ bị loại khỏi mọi màn hình dù các cột khác đúng hết.
                "PostedDate": ngay_dt, "CABAPostedDate": ngay_dt,
                "CABARefDate": ngay_dt,   # chứng từ THẬT luôn có CABARefDate = RefDate
                "RefType": ref_type, "RefNoFinance": doc[:20], "RefNoManagement": doc[:20],
                # IncludeInvoice=1 ("Nhận kèm hóa đơn") + tạo bản ghi hóa đơn
                # PUInvoice/PUInvoiceDetail bên dưới. Chứng từ NHẬP KHO thật link
                # hóa đơn ở CẢ 2 CHIỀU: PUVoucher(+Detail).PUInvoiceRefID trỏ đến
                # hóa đơn, và PUInvoiceDetail.PUVoucherRefID trỏ ngược về chứng từ.
                "IncludeInvoice": 1, "PUInvoiceRefID": inv_id,
                "DisplayOnBook": dob, "AccountObjectID": acc_obj_id,
                # JournalMemo = "Diễn giải" (form nhập kho); CABAJournalMemo =
                # "Lý do chi" trên tab Phiếu chi (biến thể Thanh toán ngay/Tiền
                # mặt của Mua hàng không qua kho) — cùng nội dung để cả 2 chỗ
                # đều có, không còn ô "Lý do chi" trống.
                "AccountObjectName": ten_ncc_dg, "JournalMemo": dien_giai,
                "CABAJournalMemo": dien_giai,
                "TotalAmountOC": total_amount, "TotalAmount": total_amount,
                "TotalImportTaxAmountOC": total_import_tax, "TotalImportTaxAmount": total_import_tax,
                "TotalVATAmountOC": total_vat, "TotalVATAmount": total_vat,
                "CreatedDate": now, "INRefOrder": now,
                # Khớp mẫu chứng từ THẬT: CreatedBy/ModifiedBy = người dùng phổ biến
                # (thay vì để trống), ModifiedDate = CreatedDate (mới tạo = mới sửa),
                # RefOrder tiếp nối số lớn nhất đang có (thay vì luôn 0), IsConvertVAT=False.
                "CreatedBy": hoc_nguoi_tao, "ModifiedBy": hoc_nguoi_tao, "ModifiedDate": now,
                "RefOrder": max_reforder + them_ct + 1, "IsConvertVAT": False,
                # Dấu hiệu NỘI BỘ (không hiển thị) để nhận diện chứng từ do
                # phần mềm này ghi — xem định nghĩa _PM_MARK.
                "CustomField10": _PM_MARK,
            })
            # HÓA ĐƠN kèm chứng từ ("Nhận kèm hóa đơn"): 1 chứng từ = 1 hóa đơn
            # (mỗi doc đã gộp theo đúng 1 số hóa đơn), chi tiết trỏ ngược về
            # chứng từ qua PUVoucherRefID/PUVoucherRefDetailID.
            inv_header = None
            if inv_reftype:
                inv_no = (str(first[cfg["sohd"]] or "")[:25] or None) \
                    if cfg.get("sohd") is not None else None
                inv_header = dict(_PU_INV_DEFAULT, **{
                    "RefID": inv_id, "BranchID": branch_id, "RefType": inv_reftype,
                    "RefDate": ngay_dt, "PostedDate": ngay_dt,
                    "RefNoFinance": doc[:20], "RefNoManagement": doc[:20],
                    "AccountObjectID": acc_obj_id, "AccountObjectName": ten_ncc_dg,
                    "AccountObjectTaxCode": mst[:50] or None,
                    "JournalMemo": dien_giai,
                    "InvNo": inv_no, "InvDate": ngay_dt,
                    "TotalTurnoverAmountOC": total_amount, "TotalTurnoverAmount": total_amount,
                    "TotalVATAmountOC": total_vat, "TotalVATAmount": total_vat,
                    "DisplayOnBook": dob, "RefOrder": max_reforder + them_ct + 1,
                    "CreatedDate": now, "CreatedBy": hoc_nguoi_tao,
                    "ModifiedDate": now, "ModifiedBy": hoc_nguoi_tao,
                    "CustomField10": _PM_MARK,
                })
            if not preview:
                # THỨ TỰ GHI theo chiều khóa ngoại: PUInvoice trước (PUVoucher.
                # PUInvoiceRefID trỏ tới nó), rồi PUVoucher -> PUVoucherDetail,
                # cuối cùng PUInvoiceDetail (trỏ ngược về chứng từ + dòng).
                if inv_header:
                    ic = list(inv_header.keys())
                    cur.execute("INSERT INTO PUInvoice ([%s]) VALUES (%s)" %
                               ("],[".join(ic), ",".join(["?"] * len(ic))),
                               [inv_header[c] for c in ic])
                hc = list(header_cols.keys())
                cur.execute("INSERT INTO PUVoucher ([%s]) VALUES (%s)" %
                           ("],[".join(hc), ",".join(["?"] * len(hc))),
                           [header_cols[c] for c in hc])
                for d in detail_rows:
                    dc = list(d.keys())
                    cur.execute("INSERT INTO PUVoucherDetail ([%s]) VALUES (%s)" %
                               ("],[".join(dc), ",".join(["?"] * len(dc))),
                               [d[c] for c in dc])
                if inv_header:
                    for d in detail_rows:
                        idet = dict(_PU_INV_DET_DEFAULT, **{
                            "RefDetailID": str(_uuid.uuid4()), "RefID": inv_id,
                            "InventoryItemID": d["InventoryItemID"],
                            "Description": d["Description"],
                            "DebitAccount": d["DebitAccount"],
                            "CreditAccount": d["CreditAccount"],
                            "TurnoverAmountOC": d["AmountOC"], "TurnoverAmount": d["Amount"],
                            "VATRate": d["VATRate"], "VATAmountOC": d["VATAmountOC"],
                            "VATAmount": d["VATAmount"], "SortOrder": d["SortOrder"],
                            "PUVoucherRefID": ref_id,
                            "PUVoucherRefDetailID": d["RefDetailID"],
                            "PUVoucherRefNoFinance": doc[:20],
                            "PUVoucherRefNoManagement": doc[:20],
                        })
                        idc = list(idet.keys())
                        cur.execute("INSERT INTO PUInvoiceDetail ([%s]) VALUES (%s)" %
                                   ("],[".join(idc), ",".join(["?"] * len(idc))),
                                   [idet[c] for c in idc])
            if inv_header:
                so_hoa_don += 1
            them_ct += 1
            them_dong += len(detail_rows)
            if ghi_de_ct:
                go += 1
            if ngay_loi:
                so_ngay_loi += 1
            tien_0 = total_amount == 0
            if tien_0:
                so_tien_0 += 1
            st = ("sẽ ghi đè" if preview else "đã ghi đè") if ghi_de_ct \
                else ("sẽ thêm" if preview else "đã thêm (CHƯA ghi sổ)")
            ket.append({"so_ct": doc, "ncc": ten_ncc_misa, "so_dong": len(detail_rows),
                        "tong_tien": total_amount, "tien_thue": total_vat, "ref_type": ref_type,
                        "loai_ct_misa": ref_type_ten, "trang_thai": st, "ref_id": ref_id,
                        "ngay_ct": ngay_dt.strftime("%d/%m/%Y"), "ngay_loi": ngay_loi,
                        "ngay_goc": ngay_str, "tien_0": tien_0})
        tong_pu = cur.execute("SELECT COUNT(*) FROM PUVoucher").fetchone()[0]
        if preview:
            conn.rollback()
        else:
            conn.commit()
        # TỰ KIỂM TRA: sau khi ghi thật, chạy lại đúng view MISA dùng cho màn
        # hình "Mua hàng hóa, dịch vụ" với ĐÚNG BỘ LỌC THẬT của MISA (bắt được
        # từ câu lệnh MISA thực thi): PostedDate trong kỳ + BranchID + DisplayOnBook
        # IN (0,2) — trước đây chỉ dò RefID nên "thấy" cả chứng từ mà màn hình
        # MISA sẽ lọc mất (vd PostedDate NULL), gây kết luận sai.
        tu_kiem_tra = None
        if not preview and them_ct:
            ref_ids = [x["ref_id"] for x in ket if x.get("ref_id")]
            if ref_ids:
                tu_kiem_tra = _misa_tu_kiem_tra_muahang(cur, ref_ids, branch_id,
                                                        "PUVoucher", ket)
        for x in ket:
            x.pop("ref_id", None)
        return {"preview": preview, "database": database, "so_chungtu": them_ct,
                "so_dong": them_dong, "so_trung": trung, "so_ghi_de": go,
                "so_bo_qua_ncc": bo_ncc, "so_bo_qua_mahang": bo_mahang,
                "so_ngay_loi": so_ngay_loi, "so_tien_0": so_tien_0,
                "so_hoa_don": so_hoa_don, "thieu_kho": sorted(thieu_kho),
                "kho_moi": sorted(kho_moi),
                "so_bo_qua_tk": bo_tk, "tk_thay": sorted(tk_thay),
                "so_dvt_sua": so_dvt_sua,
                "tong_trong_bang": tong_pu, "loai_ct_dang_co": loai_ct_dang_co,
                "tu_kiem_tra": tu_kiem_tra,
                "hoc_mau": (hoc["refname"] if hoc else None),
                "hoc_display_on_book": hoc_dob, "hoc_chi_nhanh": hoc_branch_ten,
                "mau_that": mau_that,
                "danh_sach": ket[:500]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()


def _misa_ghi_mua_hang_dv(cid, database, preview=True, ghi_de=False):
    """Ghi chứng từ MUA HÀNG DỊCH VỤ thẳng vào MISA — bảng RIÊNG PUService/
    PUServiceDetail (KHÔNG phải PUVoucher — xác nhận qua cấu trúc CSDL thật:
    SYSRefType không có loại chứng từ nào chứa "dịch vụ" dưới MasterTableName
    ='PUVoucher'). Không có PUInvoice liên kết — thông tin hóa đơn nằm thẳng
    trên PUServiceDetail. Không có Kho/ĐVT quy đổi. Cùng nguyên tắc an toàn
    như _misa_ghi_mua_hang(): CHƯA GHI SỔ, không đụng chứng từ đã ghi sổ của
    người dùng, ghi_de=True chỉ gỡ chứng từ do chính phần mềm tạo."""
    import uuid as _uuid
    cfg = _MUA_COT["dv"]
    dl = nhap_lieu_get(cid, "in")
    header, rows = dl.get("header") or [], dl.get("rows") or []
    if not rows:
        raise HTTPException(400, "Chưa có Bảng kê đầu vào đã lưu — Import & Lưu trước.")
    flat = _gen_mua_hang_dv(cid, header, rows)
    if not flat:
        return {"preview": preview, "database": database, "so_chungtu": 0, "so_dong": 0,
                "so_trung": 0, "so_bo_qua_ncc": 0, "so_bo_qua_mahang": 0, "danh_sach": [],
                "ghi_chu": "Không có dòng nào phù hợp (%s) trong Bảng kê đầu vào." % _MUA_TEN["dv"]}
    groups, order = {}, []
    for r in flat:
        doc = str(r[cfg["doc"]] or "").strip()
        if not doc:
            continue
        if doc not in groups:
            groups[doc] = []
            order.append(doc)
        groups[doc].append(r)

    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        branch_id = _misa_branch_id(cur)
        ncc = {}
        for aid, taxcode, code, name in cur.execute(
                "SELECT AccountObjectID, CompanyTaxCode, AccountObjectCode, AccountObjectName "
                "FROM AccountObject").fetchall():
            nm = str(name or "")
            if taxcode:
                ncc[_misa_khncc_chuan_mst(taxcode).lower()] = (aid, nm)
            if code:
                ncc[str(code).strip().lower()] = (aid, nm)
        hang = {}
        for iid, code, uid, ten_h in cur.execute(
                "SELECT InventoryItemID, InventoryItemCode, UnitID, InventoryItemName "
                "FROM InventoryItem").fetchall():
            if code:
                hang[str(code).strip().lower()] = (iid, uid, str(ten_h or ""))
        # danh mục TÀI KHOẢN thật (FK trên các cột TK) — xem _misa_tk_fallback
        tk_set = set()
        try:
            for (an,) in cur.execute("SELECT AccountNumber FROM Account").fetchall():
                if an:
                    tk_set.add(str(an).strip())
        except Exception:
            pass
        # ĐVT mồ côi — kiểm tra + tự sửa như _misa_ghi_mua_hang
        unit_ten = {}
        unit_ids = set()
        try:
            for u_id, u_name in cur.execute("SELECT UnitID, UnitName FROM Unit").fetchall():
                unit_ids.add(str(u_id).strip().lower())
                if u_name:
                    unit_ten[str(u_name).strip().lower()] = u_id
        except Exception:
            pass
        so_dvt_sua = 0

        def sua_dvt(r, iid, uid):
            nonlocal so_dvt_sua
            if not unit_ids or uid is None or str(uid).strip().lower() in unit_ids:
                return uid
            dvt_text = (str(r[cfg["dvt"]] or "").strip()
                        if cfg.get("dvt") is not None else "")
            uk = dvt_text.lower()
            if uk and uk in unit_ten:
                uid2 = unit_ten[uk]
            elif dvt_text:
                uid2 = str(_uuid.uuid4())
                if not preview:
                    cur.execute("INSERT INTO Unit (UnitID, UnitName, Description, Inactive) "
                                "VALUES (?,?,?,0)", uid2, dvt_text[:20], None)
                unit_ten[uk] = uid2
                unit_ids.add(uid2.lower())
            else:
                uid2 = None
            if not preview:
                try:
                    cur.execute("UPDATE InventoryItem SET UnitID=? WHERE InventoryItemID=?",
                               uid2, iid)
                except Exception:
                    pass
            so_dvt_sua += 1
            return uid2
        reftype_ten = dict(cur.execute(
            "SELECT RefType, RefTypeName FROM SYSRefType WHERE MasterTableName='PUService'").fetchall())
        posted_refno = set()
        unposted_docs = {}
        for refid, rn, rt, pf, pm, mark in cur.execute(
                "SELECT RefID, RefNoManagement, RefType, ISNULL(IsPostedFinance,0), "
                "ISNULL(IsPostedManagement,0), ISNULL(CustomField10,'') FROM PUService").fetchall():
            if not rn:
                continue
            k = str(rn).strip().lower()
            la_cua_minh = mark == _PM_MARK
            if (pf or pm) and not la_cua_minh:
                posted_refno.add(k)
            else:
                d = unposted_docs.setdefault(k, {"refids": [], "reftype_ten": reftype_ten.get(rt)})
                d["refids"].append(refid)
        tu_khoa_loai = ["mua", "dịch vụ"]
        loai_ct_dang_co = []
        mau_that = []
        hoc = None
        hoc_dob = None
        hoc_branch = None
        hoc_branch_ten = None
        hoc_nguoi_tao = None
        max_reforder = 0
        try:
            _dem_loai, _dem_dob, _dem_branch, _dem_nguoi = {}, {}, {}, {}
            for rt, rn, dob, inc, mark, rnm, pf, pm, br, brn, cb, ro in cur.execute(
                    "SELECT ps.RefType, rt.RefTypeName, ps.DisplayOnBook, ps.IncludeInvoice, "
                    "ISNULL(ps.CustomField10,''), ps.RefNoManagement, "
                    "ISNULL(ps.IsPostedFinance,0), ISNULL(ps.IsPostedManagement,0), "
                    "ps.BranchID, ou.OrganizationUnitName, ps.CreatedBy, ISNULL(ps.RefOrder,0) "
                    "FROM PUService ps JOIN SYSRefType rt ON rt.RefType=ps.RefType "
                    "LEFT JOIN OrganizationUnit ou ON ou.OrganizationUnitID=ps.BranchID").fetchall():
                max_reforder = max(max_reforder, ro or 0)
                if mark == _PM_MARK:
                    continue
                key = (int(rt), rn)
                _dem_loai[key] = _dem_loai.get(key, 0) + 1
                if dob is not None:
                    _dem_dob[dob] = _dem_dob.get(dob, 0) + 1
                if br is not None:
                    _dem_branch[br] = _dem_branch.get(br, [0, brn])
                    _dem_branch[br][0] += 1
                if cb:
                    _dem_nguoi[cb] = _dem_nguoi.get(cb, 0) + 1
                if len(mau_that) < 5:
                    mau_that.append({"so_ct": rnm, "loai": rn, "display_on_book": dob,
                                     "include_invoice": inc, "da_ghi_so": bool(pf or pm),
                                     "chi_nhanh": brn})
            for (rt, rn), c in sorted(_dem_loai.items(), key=lambda kv: -kv[1]):
                loai_ct_dang_co.append({"ten": rn, "so": c})
                n = (rn or "").lower()
                if hoc is None and all(k in n for k in tu_khoa_loai):
                    hoc = {"reftype": rt, "refname": rn}
            if _dem_dob:
                hoc_dob = max(_dem_dob.items(), key=lambda kv: kv[1])[0]
            if _dem_branch:
                hoc_branch, (_, hoc_branch_ten) = max(
                    _dem_branch.items(), key=lambda kv: kv[1][0])
            if _dem_nguoi:
                hoc_nguoi_tao = max(_dem_nguoi.items(), key=lambda kv: kv[1])[0]
        except Exception:
            pass
        hoc_purpose = None
        try:
            row_pp = cur.execute(
                "SELECT TOP 1 d.PurchasePurposeID FROM PUServiceDetail d "
                "JOIN PUService ps ON ps.RefID=d.RefID "
                "WHERE ISNULL(ps.CustomField10,'') <> ? "
                "AND d.PurchasePurposeID IS NOT NULL", _PM_MARK).fetchone()
            if row_pp:
                hoc_purpose = row_pp[0]
            if hoc_purpose is None:
                row_pp = cur.execute(
                    "SELECT TOP 1 PurchasePurposeID FROM PurchasePurpose "
                    "WHERE PurchasePurposeCode='1'").fetchone()
                if row_pp:
                    hoc_purpose = row_pp[0]
        except Exception:
            pass
        if hoc_branch is not None:
            branch_id = hoc_branch

        now = datetime.datetime.now()
        ket = []
        them_ct = them_dong = trung = bo_ncc = bo_mahang = go = so_ngay_loi = so_tien_0 = 0
        bo_tk = 0
        tk_thay = set()

        def tra_tk(tk):
            t = str(tk or "").strip()
            if not t or not tk_set:
                return t or None
            t2 = _misa_tk_fallback(t, tk_set)
            if t2 and t2 != t:
                tk_thay.add("%s→%s" % (t, t2))
            return t2
        for doc in order:
            lines = groups[doc]
            k_doc = doc.strip().lower()
            if k_doc in posted_refno:
                trung += 1
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "đã ghi sổ trong MISA (bỏ qua)"})
                continue
            if k_doc in unposted_docs and not ghi_de:
                trung += 1
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "loai_ct_misa": unposted_docs[k_doc]["reftype_ten"],
                            "trang_thai": "đã có (chưa ghi sổ, bỏ qua)"})
                continue
            first = lines[0]
            mst = str(first[cfg["mst"]] or "").strip()
            mst_k = mst.lower()
            if mst_k not in ncc:
                bo_ncc += 1
                ket.append({"so_ct": doc, "so_dong": len(lines), "ncc_mst": mst,
                            "trang_thai": "bỏ qua — NCC (MST %s) chưa có trong MISA" % mst})
                continue
            ghi_de_ct = k_doc in unposted_docs
            if ghi_de_ct and not preview:
                for rid in unposted_docs[k_doc]["refids"]:
                    cur.execute("DELETE FROM PUServiceDetail WHERE RefID=?", rid)
                    cur.execute("DELETE FROM PUService WHERE RefID=?", rid)
            acc_obj_id, ten_ncc_misa = ncc[mst_k]
            valid_lines = []
            for r in lines:
                ma = str(r[cfg["ma"]] or "").strip()
                mk = ma.lower()
                if mk not in hang:
                    bo_mahang += 1
                    continue
                valid_lines.append((r, hang[mk]))
            if not valid_lines:
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "bỏ qua — tất cả dòng đều thiếu mã hàng trong MISA"})
                continue
            if tk_set:
                tk_loi = set()
                for r, _h in valid_lines:
                    for acc in (str(r[cfg["no"]] or "").strip(),
                                str(r[cfg["co"]] or "").strip()):
                        if acc and _misa_tk_fallback(acc, tk_set) is None:
                            tk_loi.add(acc)
                if tk_loi:
                    bo_tk += 1
                    ket.append({"so_ct": doc, "so_dong": len(lines),
                                "trang_thai": "bỏ qua — TK %s không có trong danh mục "
                                              "tài khoản MISA" % ", ".join(sorted(tk_loi))})
                    continue
            ngay_dt = None
            ngay_str = str(first[cfg["ngayct"]] or "").strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    ngay_dt = datetime.datetime.strptime(ngay_str, fmt)
                    break
                except Exception:
                    pass
            ngay_loi = ngay_dt is None
            ngay_dt = ngay_dt or now
            co_tk = str(first[cfg["co"]] or "").strip()
            if hoc:
                ref_type, ref_type_ten = hoc["reftype"], hoc["refname"]
            else:
                ref_type, ref_type_ten = _misa_pu_reftype(cur, tu_khoa_loai, co_tk,
                                                          master_table="PUService")
            if not ref_type:
                ket.append({"so_ct": doc, "so_dong": len(lines),
                            "trang_thai": "bỏ qua — không dò được loại chứng từ MISA phù hợp"})
                continue
            ref_id = str(_uuid.uuid4())
            total_amount = total_vat = 0
            detail_rows = []
            for idx, (r, (iid, uid, ten_h)) in enumerate(valid_lines, 1):
                uid = sua_dvt(r, iid, uid)
                sl = _num0(r[cfg["sl"]])
                dgia = _num0(r[cfg["dgia"]])
                tt = _num0(r[cfg["tt"]])
                ts = _num0(r[cfg["ts"]]) if cfg["ts"] is not None else None
                tthue = _num0(r[cfg["tthue"]])
                # TK đối chiếu danh mục Account (FK), tự rút về TK cha nếu thiếu
                tk_thue = tra_tk(r[cfg["tk_thue"]])
                no_acc = tra_tk(r[cfg["no"]])
                co_acc = tra_tk(r[cfg["co"]])
                total_amount += tt
                total_vat += tthue
                mo_ta = ten_h or str(r[cfg["ten"]] or "")
                detail_rows.append(dict(_PU_SERVICE_DET_DEFAULT, **{
                    "RefDetailID": str(_uuid.uuid4()), "RefID": ref_id, "InventoryItemID": iid,
                    "Description": mo_ta, "DebitAccount": no_acc, "CreditAccount": co_acc,
                    "UnitID": uid, "Quantity": sl, "UnitPrice": dgia,
                    "AmountOC": tt, "Amount": tt,
                    "VATRate": ts, "VATAmountOC": tthue, "VATAmount": tthue,
                    "VATAccount": tk_thue,
                    # thông tin hóa đơn + NCC thuế nằm THẲNG trên dòng (không
                    # có PUInvoice liên kết như PUVoucher)
                    "InvNo": str(r[cfg["sohd"]] or "")[:25] if cfg["sohd"] is not None else None,
                    "InvDate": ngay_dt,
                    "TaxAccountObjectID": acc_obj_id, "TaxAccountObjectName": ten_ncc_misa,
                    "TaxAccountObjectTaxCode": mst[:50] or None,
                    "AccountObjectID": acc_obj_id, "PurchasePurposeID": hoc_purpose,
                    "VATDescription": ("Thuế GTGT - %s" % mo_ta)[:255],
                    "SortOrder": idx,
                }))
            # DisplayOnBook phải thuộc (0,2) — xem giải thích ở _misa_ghi_mua_hang
            dob = hoc_dob if hoc_dob in (0, 2) else 0
            ten_ncc_dg = ten_ncc_misa or str(first[cfg["ten_ncc"]] or "")
            so_hd_dg = str(first[cfg["sohd"]] or "").strip() if cfg["sohd"] is not None else ""
            dien_giai = ("Mua hàng - %s - %s %s" %
                        (ten_ncc_dg, so_hd_dg, ngay_dt.strftime("%d/%m/%Y"))).strip()[:500]
            header_cols = dict(_PU_SERVICE_DEFAULT, **{
                "RefID": ref_id, "BranchID": branch_id, "RefDate": ngay_dt,
                "PostedDate": ngay_dt, "RefType": ref_type,
                "RefNoFinance": doc[:20], "RefNoManagement": doc[:20],
                # IncludeInvoice=1 nhưng PUInvoiceRefID vẫn NULL — chứng từ
                # dịch vụ THẬT không liên kết PUInvoice, thông tin hóa đơn
                # đã nằm thẳng trên từng dòng PUServiceDetail (xem trên).
                "IncludeInvoice": 1, "DisplayOnBook": dob,
                "AccountObjectID": acc_obj_id, "AccountObjectName": ten_ncc_dg,
                "JournalMemo": dien_giai,
                "TotalAmountOC": total_amount, "TotalAmount": total_amount,
                "TotalVATAmountOC": total_vat, "TotalVATAmount": total_vat,
                "CreatedDate": now, "CreatedBy": hoc_nguoi_tao,
                "ModifiedDate": now, "ModifiedBy": hoc_nguoi_tao,
                "RefOrder": max_reforder + them_ct + 1,
                "CustomField10": _PM_MARK,
            })
            if not preview:
                hc = list(header_cols.keys())
                cur.execute("INSERT INTO PUService ([%s]) VALUES (%s)" %
                           ("],[".join(hc), ",".join(["?"] * len(hc))),
                           [header_cols[c] for c in hc])
                for d in detail_rows:
                    dc = list(d.keys())
                    cur.execute("INSERT INTO PUServiceDetail ([%s]) VALUES (%s)" %
                               ("],[".join(dc), ",".join(["?"] * len(dc))),
                               [d[c] for c in dc])
            them_ct += 1
            them_dong += len(detail_rows)
            if ghi_de_ct:
                go += 1
            if ngay_loi:
                so_ngay_loi += 1
            tien_0 = total_amount == 0
            if tien_0:
                so_tien_0 += 1
            st = ("sẽ ghi đè" if preview else "đã ghi đè") if ghi_de_ct \
                else ("sẽ thêm" if preview else "đã thêm (CHƯA ghi sổ)")
            ket.append({"so_ct": doc, "ncc": ten_ncc_misa, "so_dong": len(detail_rows),
                        "tong_tien": total_amount, "tien_thue": total_vat, "ref_type": ref_type,
                        "loai_ct_misa": ref_type_ten, "trang_thai": st, "ref_id": ref_id,
                        "ngay_ct": ngay_dt.strftime("%d/%m/%Y"), "ngay_loi": ngay_loi,
                        "ngay_goc": ngay_str, "tien_0": tien_0})
        tong_pu = cur.execute("SELECT COUNT(*) FROM PUService").fetchone()[0]
        if preview:
            conn.rollback()
        else:
            conn.commit()
        tu_kiem_tra = None
        if not preview and them_ct:
            ref_ids = [x["ref_id"] for x in ket if x.get("ref_id")]
            if ref_ids:
                tu_kiem_tra = _misa_tu_kiem_tra_muahang(cur, ref_ids, branch_id,
                                                        "PUService", ket)
        for x in ket:
            x.pop("ref_id", None)
        return {"preview": preview, "database": database, "so_chungtu": them_ct,
                "so_dong": them_dong, "so_trung": trung, "so_ghi_de": go,
                "so_bo_qua_ncc": bo_ncc, "so_bo_qua_mahang": bo_mahang,
                "so_ngay_loi": so_ngay_loi, "so_tien_0": so_tien_0,
                "so_hoa_don": 0, "thieu_kho": [],
                "so_bo_qua_tk": bo_tk, "tk_thay": sorted(tk_thay),
                "so_dvt_sua": so_dvt_sua,
                "tong_trong_bang": tong_pu, "loai_ct_dang_co": loai_ct_dang_co,
                "tu_kiem_tra": tu_kiem_tra,
                "hoc_mau": (hoc["refname"] if hoc else None),
                "hoc_display_on_book": hoc_dob, "hoc_chi_nhanh": hoc_branch_ten,
                "mau_that": mau_that,
                "danh_sach": ket[:500]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()


@app.post("/api/misa-sql/import-mua-hang/{cid}")
def misa_sql_import_mua_hang(cid: int, loai: str, preview: int = 1, database: str = "",
                            ghi_de: int = 0):
    """Ghi chứng từ Mua hàng (loai=nk/kqk/dv) thẳng vào MISA. nk/kqk ghi vào
    PUVoucher; dv (Mua hàng dịch vụ) ghi vào bảng RIÊNG PUService (khác hẳn
    PUVoucher — xác nhận qua SYSRefType thật). preview=1 -> chỉ xem trước,
    không ghi. LUÔN ghi ở trạng thái CHƯA GHI SỔ. ghi_de=1 -> gỡ chứng từ
    chưa ghi sổ trùng số rồi ghi lại."""
    database = (database or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    if not database:
        raise HTTPException(400, "Chưa cấu hình kết nối/CSDL MISA. Mở '🗄 Kết nối CSDL MISA', "
                                 "kết nối tới dữ liệu THỬ trước.")
    if loai == "dv":
        return _misa_ghi_mua_hang_dv(cid, database, preview=bool(preview), ghi_de=bool(ghi_de))
    return _misa_ghi_mua_hang(cid, database, loai, preview=bool(preview), ghi_de=bool(ghi_de))


# ============ GHI TĂNG TSCĐ / CCDC THẲNG VÀO MISA (FixedAsset / SUIncrement) ==
# KHÁC Mua hàng: đây KHÔNG phải "chứng từ" (không có SYSRefType/loại chứng từ
# trên form Excel mẫu) mà là GHI TĂNG DANH MỤC tài sản/CCDC — xem yêu cầu của
# anh: "không phải ghi tăng chứng từ mà chỉ là ghi tăng danh mục tài sản và
# ccdc". Dữ liệu lấy ĐÚNG công thức đã dùng cho file Excel mẫu MISA
# (_gen_ghi_tang_tscd/_gen_ghi_tang_ccdc — đã đối chiếu byte-for-byte với
# form Excel MISA thật ở phần trước). CSDL công ty này CHƯA có bản ghi
# FixedAsset/SUIncrement nào (2 bảng đang TRỐNG) nên KHÔNG có "mẫu thật" để
# học quy ước như Mua hàng — thay vì đoán bừa danh sách cột, DÒ TÊN CỘT THẬT
# qua sys.columns lúc chạy (không hard-code) rồi mới liệt kê ĐỦ MỌI CỘT THẬT
# khi INSERT (đúng bài học từ Mua hàng: bỏ sót cột dễ dính DEFAULT constraint
# lạ, vi phạm FK ẩn). Với từng trường cần điền dữ liệu thật, thử NHIỀU TÊN
# ỨNG VIÊN hợp lý — cột nào không khớp thì bỏ qua (giữ giá trị mặc định theo
# kiểu dữ liệu), KHÔNG làm hỏng câu lệnh. Không có cột CustomField trên 2
# bảng này (khác PUVoucher/PUService) nên KHÔNG dùng được _PM_MARK; thay vào
# đó chỉ dựa vào cờ ĐÃ GHI SỔ (IsPostedFinance/IsPostedManagement, nếu có) để
# quyết định ghi_de — bản ghi ĐÃ GHI SỔ không bao giờ bị đụng, y hệt nguyên
# tắc an toàn của Mua hàng.
def _misa_cot_bang_that(cur, table):
    """Trả {tên cột viết thường: (tên cột thật, kiểu SQL)} của bảng THẬT
    trong CSDL MISA đang kết nối — loại cột rowversion (timestamp, vd
    EditVersion), cột computed và cột identity (không được phép ghi tay)."""
    try:
        rows = cur.execute(
            "SELECT c.name, ty.name FROM sys.columns c "
            "JOIN sys.types ty ON ty.user_type_id = c.user_type_id "
            "WHERE c.object_id = OBJECT_ID(?) AND ty.name <> 'timestamp' "
            "AND c.is_computed = 0 AND c.is_identity = 0 "
            "ORDER BY c.column_id", table).fetchall()
    except Exception:
        return {}
    return {str(n).strip().lower(): (n, t) for n, t in rows}

def _misa_gia_tri_mac_dinh(sqltype):
    """Giá trị mặc định AN TOÀN theo kiểu SQL khi liệt kê đủ mọi cột: số/bit
    -> 0 (tránh NULL trên cột NOT NULL không có default), còn lại -> NULL
    tường minh (tránh dính DEFAULT constraint lạ — xem lý do ở đầu mục Mua
    hàng)."""
    t = (sqltype or "").lower()
    if t in ("bit", "int", "smallint", "tinyint", "bigint", "money",
             "smallmoney", "decimal", "numeric", "float", "real"):
        return 0
    return None

def _misa_chon_cot(cols_that, *ung_vien):
    """Trả TÊN CỘT THẬT đầu tiên (trong danh sách ứng viên hợp lý, không
    phân biệt hoa/thường) đang thật sự tồn tại trên bảng; None nếu không cột
    nào khớp (khi đó trường tương ứng được bỏ qua, không đoán bừa tên cột)."""
    for uv in ung_vien:
        hit = cols_that.get(uv.lower())
        if hit:
            return hit[0]
    return None

def _misa_gan(row, cols_that, gt, *ung_vien):
    ten = _misa_chon_cot(cols_that, *ung_vien)
    if ten:
        row[ten] = gt

def _misa_doc_ngay(s):
    s = str(s or "").strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(s, fmt)
        except Exception:
            pass
    return None

def _snum(v):
    """Ép về số, chấp nhận cả CHUỖI số ('24' -> 24) — dữ liệu DM lưu qua lưới
    trình duyệt hay thành chuỗi; _num0 thuần sẽ trả 0 làm mất giá trị (vd Số
    kỳ phân bổ CCDC ghi vào MISA thành 0 dù DM ghi 24)."""
    v = _to_num(v)
    return v if isinstance(v, (int, float)) else 0

def _misa_unit_hong(ten):
    """Nhận diện tên ĐVT (Unit.UnitName) là RÁC — thật ra chứa (một phần)
    chuỗi GUID thay vì tên đơn vị thật (vd 'fac19b53-57' thay vì 'PCS'/'Cái').
    Xác nhận qua thực tế: mã hàng hiện "ĐVT chính: fac19b53" trong MISA và
    crash "InvalidCastException" khi mở — Unit đứng sau UnitID này bị hỏng
    tên. Tên ĐVT thật (chữ Việt có dấu, hoặc PCS/KG/Cái/Hộp...) không bao
    giờ khớp mẫu toàn hex+gạch ngang này nên phân biệt an toàn. Lưu ý mẫu
    phải chấp nhận cả GUID bị CẮT CỤT ở 20 ký tự của cột UnitName (vd
    'fac19b53-57bb-4995-9' — đuôi sau dấu gạch chỉ còn 1 ký tự)."""
    import re as _re_uh
    t = str(ten or "").strip()
    return bool(t) and bool(_re_uh.match(r"^[0-9a-f]{6,}(-[0-9a-f]{1,})*-?$", t, _re_uh.IGNORECASE))

def _misa_tk_fallback(tk, tk_set):
    """Đối chiếu số TÀI KHOẢN với danh mục TK thật của MISA (bảng Account —
    các cột TK trên PUVoucherDetail có FOREIGN KEY sang đó, ghi TK không tồn
    tại là dính lỗi FK_..._DebitAccount). TK không có thì RÚT VỀ TK CHA gần
    nhất đang tồn tại (vd công ty không mở tiểu khoản 2421 -> dùng 242, đúng
    tình huống thật: mua CCDC Nợ 2421 báo lỗi FK còn TSCĐ Nợ 2111 thì có TK
    nên ghi được). Trả TK dùng được (>=3 số) hoặc None nếu không rút được."""
    t = str(tk or "").strip()
    if not t:
        return None
    while len(t) >= 3:
        if t in tk_set:
            return t
        t = t[:-1]
    return None

def _misa_dvsd_map(cur):
    """Bảng mã 'Đơn vị sử dụng'/'Đối tượng phân bổ' (vd 'VP') -> ID — xác
    nhận qua cấu trúc CSDL thật: MISA KHÔNG có bảng Department riêng, chỉ có
    OrganizationUnit. Công ty 1 chi nhánh/không có phòng ban thì bảng này rỗng
    (không có mã 'VP') — bên gọi hàm PHẢI tự fallback về branch_id (đơn vị tổ
    chức gốc) khi tra không ra, KHÔNG được để trống (cột thường NOT NULL)."""
    m = {}
    try:
        for oid, code in cur.execute(
                "SELECT OrganizationUnitID, OrganizationUnitCode FROM OrganizationUnit").fetchall():
            if code:
                m[str(code).strip().upper()] = oid
    except Exception:
        pass
    return m

# RefType CHUẨN của dòng SỔ CÁI ghi tăng (xác nhận từ dữ liệu thật công ty:
# FixedAssetLedger.RefType=250, SupplyLedger.RefType=450). Đây là giá trị mà
# màn hình lưới "Ghi tăng" lọc theo — KHÔNG học được thì phải dùng hằng số
# này, KHÔNG để None/0 (bằng 0 = "Phi chứng từ" -> lưới Ghi tăng bỏ qua).
# Bài học: khi người dùng XÓA hết bản ghi mẫu (kể cả bản nhập tay), phép học
# RefType trả rỗng -> nếu không có fallback cứng, dòng sổ cái ghi sai RefType
# và biến mất khỏi lưới Ghi tăng dù vẫn hiện ở Sổ theo dõi/Sổ tài sản.
_FA_LEDGER_REFTYPE = 250   # Ghi tăng TSCĐ
_SU_LEDGER_REFTYPE = 450   # Ghi tăng CCDC

def _misa_hoc_reftype(cur, table, tu_khoa):
    """Học RefType + DisplayOnBook PHỔ BIẾN NHẤT từ bản ghi ĐANG CÓ Ý NGHĨA
    (RefType<>0 — RefType=0 là "Phi chứng từ", giá trị mặc định rỗng của SQL
    Server, KHÔNG phải loại chứng từ thật; bài học từ lần ghi thử đầu tiên:
    để RefType=0 khiến MISA không hiện bản ghi trên lưới dù mọi cột khác đều
    đúng). Bản ghi lỗi cũ do phần mềm ghi (RefType=0) tự động bị loại khỏi
    phép học này. Nếu bảng chưa có bản ghi nào RefType<>0 thì dò theo từ khóa
    trong SYSRefType (_misa_pu_reftype). Trả (RefType, DisplayOnBook) — cả 2
    có thể None nếu không dò được gì."""
    hoc_rt, hoc_dob = None, None
    try:
        cnt_rt, cnt_dob = {}, {}
        for rt, dob in cur.execute(
                "SELECT RefType, DisplayOnBook FROM %s WHERE RefType<>0" % table).fetchall():
            cnt_rt[rt] = cnt_rt.get(rt, 0) + 1
            if dob is not None:
                cnt_dob[dob] = cnt_dob.get(dob, 0) + 1
        if cnt_rt:
            hoc_rt = max(cnt_rt.items(), key=lambda kv: kv[1])[0]
        if cnt_dob:
            hoc_dob = max(cnt_dob.items(), key=lambda kv: kv[1])[0]
    except Exception:
        pass
    if hoc_rt is None:
        rt, _rtn = _misa_pu_reftype(cur, tu_khoa, master_table=table)
        hoc_rt = rt
    return hoc_rt, hoc_dob

def _misa_ghi_tang_tscd(cid, database, preview=True, ghi_de=False):
    """Ghi tăng TSCĐ thẳng vào bảng FixedAsset — mỗi dòng Excel = 1 tài sản =
    1 dòng FixedAsset (KHÔNG gộp nhiều dòng như Mua hàng — Ghi tăng TSCĐ
    không phải chứng từ nhiều dòng, mỗi tài sản độc lập). Nhận diện tài sản
    ĐÃ GHI vào MISA qua Mã tài sản (FixedAssetCode, theo đúng dãy mã tự sinh
    TSCD00001... của phần mềm). RefType BẮT BUỘC phải là loại "Ghi tăng" thật
    (RefType=0 = "Phi chứng từ" — để mặc định 0 khiến MISA KHÔNG hiện bản ghi
    trên lưới "Ghi tăng" dù mọi cột khác đều đúng — xác nhận qua lần ghi thử
    đầu tiên: 11 dòng ghi thành công nhưng lưới MISA rỗng)."""
    import uuid as _uuid
    header, rows = _doc_nhap_lieu(cid, "in")
    out = _gen_ghi_tang_tscd(cid, header, rows)
    if not out:
        return {"preview": preview, "database": database, "so_them": 0, "so_trung": 0,
                "danh_sach": [],
                "ghi_chu": "Chưa có tài sản cố định nào (Nợ 2111/211) trong Bảng kê đầu vào."}
    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        branch_id = _misa_branch_id(cur)
        cols = _misa_cot_bang_that(cur, "FixedAsset")
        if not cols:
            raise HTTPException(400, "Không tìm thấy bảng FixedAsset trong CSDL MISA đang kết nối.")
        # Màn hình lưới "Tài sản cố định > Ghi tăng" KHÔNG đọc bảng FixedAsset
        # (sổ tài sản) mà đọc bảng FixedAssetLedger — bản ghi "chứng từ ghi
        # tăng" có RefNo/RefDate/PostedDate (NOTNULL) đúng các cột Số chứng
        # từ/Ngày ghi tăng của lưới; khung "Thiết lập phân bổ" đọc bảng
        # FixedAssetDetailAllocation. Xác nhận qua thử nghiệm thật: chỉ ghi
        # FixedAsset -> lưới trống; bản ghi nhập tay trên MISA hiện bình
        # thường vì MISA tạo đủ cả 3 bảng.
        cols_led = _misa_cot_bang_that(cur, "FixedAssetLedger")
        cols_alloc = _misa_cot_bang_that(cur, "FixedAssetDetailAllocation")
        hoc_rt, hoc_dob = _misa_hoc_reftype(cur, "FixedAsset", ["tăng"])
        # RefType của dòng sổ cái ghi tăng — học riêng từ FixedAssetLedger
        # (có thể khác RefType trên FixedAsset); fallback dùng chung hoc_rt.
        led_rt = None
        try:
            _cnt_led = {}
            for (rt,) in cur.execute(
                    "SELECT RefType FROM FixedAssetLedger WHERE RefType<>0").fetchall():
                _cnt_led[rt] = _cnt_led.get(rt, 0) + 1
            if _cnt_led:
                led_rt = max(_cnt_led.items(), key=lambda kv: kv[1])[0]
        except Exception:
            pass
        # Ưu tiên RefType học từ SỔ CÁI ghi tăng: trước phiên bản này phần
        # mềm chưa từng ghi FixedAssetLedger nên dòng nào trong đó cũng là
        # THẬT (nhập tay trên MISA) — đáng tin hơn phép đếm đa số trên
        # FixedAsset (đang bị chính các dòng cũ do phần mềm ghi lấn át).
        if led_rt is not None:
            hoc_rt = led_rt
        # Fallback CỨNG: khi KHÔNG học/dò được gì (vd người dùng xóa hết mẫu),
        # dùng RefType chuẩn 250 thay vì để None/0 — dòng sổ cái ghi tăng ghi
        # RefType=0 sẽ biến mất khỏi lưới "Ghi tăng" (xem _FA_LEDGER_REFTYPE).
        # Chỉ là DỰ PHÒNG: nếu học/dò được (led_rt/hoc_rt) thì luôn ưu tiên.
        if hoc_rt is None:
            hoc_rt = _FA_LEDGER_REFTYPE
        # Học thêm quy ước từ dòng FixedAsset THẬT — nhận diện qua CreatedBy
        # có giá trị (MISA luôn đóng dấu người tạo; các dòng do phần mềm ghi
        # để trống). Ưu tiên hơn phép đếm đa số hoc_dob (bị dòng cũ của phần
        # mềm lấn át).
        hoc_fa_that = None   # (DisplayOnBook, IsEnoughVoucher, CreatedBy)
        try:
            hoc_fa_that = cur.execute(
                "SELECT TOP 1 DisplayOnBook, IsEnoughVoucher, CreatedBy "
                "FROM FixedAsset WHERE ISNULL(CreatedBy,'')<>''").fetchone()
        except Exception:
            pass
        # RefOrderInSubSystem của sổ cái — dòng thật đánh số 1,2,3... riêng
        # trong phân hệ TSCĐ (khác RefOrder chung); ghi tiếp nối MAX hiện có.
        max_ro_sub = 0
        try:
            r_sub = cur.execute(
                "SELECT ISNULL(MAX(RefOrderInSubSystem),0) FROM FixedAssetLedger").fetchone()
            max_ro_sub = (r_sub[0] or 0) if r_sub else 0
        except Exception:
            pass
        # ObjectType của dòng phân bổ — học từ bản ghi thật (nhập tay); chưa
        # có thì đoán 1 (đơn vị tổ chức).
        hoc_objtype = None
        try:
            row_ot = cur.execute(
                "SELECT TOP 1 ObjectType FROM FixedAssetDetailAllocation "
                "WHERE ObjectType IS NOT NULL").fetchone()
            if row_ot:
                hoc_objtype = row_ot[0]
        except Exception:
            pass
        # FixedAssetLedgerID là int PK — nếu KHÔNG phải identity (còn xuất
        # hiện trong danh sách cột ghi được) thì phải tự cấp MAX+1.
        led_id_tay = "fixedassetledgerid" in cols_led
        next_led_id = 1
        if led_id_tay:
            try:
                next_led_id = (cur.execute(
                    "SELECT ISNULL(MAX(FixedAssetLedgerID),0) FROM FixedAssetLedger"
                ).fetchone()[0] or 0) + 1
            except Exception:
                pass
        # DisplayOnBook: chứng từ ghi tăng THẬT (GTTS00001 nhập tay) có
        # DisplayOnBook=0; mặc định 0 (không phải 3) và chỉ nhận giá trị học
        # được nếu thuộc (0,2) — cùng bài học DisplayOnBook bên Mua hàng: giá
        # trị 3 làm bản ghi trượt bộ lọc màn hình và biến mất.
        dob = hoc_dob if hoc_dob in (0, 2) else 0
        cat_map = {}
        try:
            for catid, code in cur.execute(
                    "SELECT FixedAssetCategoryID, FixedAssetCategoryCode "
                    "FROM FixedAssetCategory").fetchall():
                if code:
                    cat_map[str(code).strip()] = catid
        except Exception:
            pass
        dvsd_map = _misa_dvsd_map(cur)
        ma_col = _misa_chon_cot(cols, "FixedAssetCode")
        id_col = _misa_chon_cot(cols, "FixedAssetID", "RefID")
        pf_col = _misa_chon_cot(cols, "IsPostedFinance")
        pm_col = _misa_chon_cot(cols, "IsPostedManagement")
        ro_col = _misa_chon_cot(cols, "RefOrder")
        existed = {}
        if ma_col and id_col:
            try:
                sel = "SELECT [%s],[%s]" % (id_col, ma_col)
                sel += (",ISNULL([%s],0)" % pf_col) if pf_col else ",0"
                sel += (",ISNULL([%s],0)" % pm_col) if pm_col else ",0"
                sel += " FROM FixedAsset"
                for rid, ma, pf, pm in cur.execute(sel).fetchall():
                    if ma:
                        existed[str(ma).strip().lower()] = {"id": rid, "posted": bool(pf or pm)}
            except Exception:
                pass
        max_reforder = 0
        if ro_col:
            try:
                row_mo = cur.execute("SELECT MAX([%s]) FROM FixedAsset" % ro_col).fetchone()
                max_reforder = (row_mo[0] or 0) if row_mo else 0
            except Exception:
                pass
        now = datetime.datetime.now()
        them = trung = go = 0
        ket = []
        for r in out:
            (ma, ten, loai, dtpb, e, ngct, ngkh, tkng, tkkh, ng_gia, gt_kh, dvt_tg,
             tg_sd, ty_le_kh, kh_thang, hao_mon, dtpb2, ty_le_pb, tkcp, _kmcp, _mts) = \
                (list(r) + [""] * 21)[:21]
            k = str(ma).strip().lower()
            ghi_de_ts = False
            if k in existed:
                # Mã TSCD00001, TSCD00002... là DÃY DO PHẦN MỀM TỰ SINH nên
                # bản ghi trùng mã chính là bản ghi phần mềm đã ghi lần trước
                # — cho phép ghi đè kể cả khi đang mang cờ đã ghi sổ (giống
                # nguyên tắc _PM_MARK bên Mua hàng; bản ghi người dùng tự tạo
                # trên MISA mang mã kiểu GTTS00001 không bao giờ trùng dãy này).
                if not ghi_de:
                    trung += 1
                    ket.append({"ma": ma, "ten": ten,
                                "trang_thai": "đã có (bỏ qua — bấm Ghi đè để ghi lại)"})
                    continue
                ghi_de_ts = True
                if not preview:
                    old_id = existed[k]["id"]
                    # gỡ các bảng vệ tinh TRƯỚC (dòng sổ cái ghi tăng, thiết
                    # lập phân bổ, các bảng chi tiết) rồi mới gỡ FixedAsset.
                    for tbl in ("FixedAssetLedger", "FixedAssetDetailAllocation",
                               "FixedAssetDetail", "FixedAssetDetailAccessory",
                               "FixedAssetDetailBoardDelivery", "FixedAssetDetailSource"):
                        try:
                            cur.execute("DELETE FROM %s WHERE FixedAssetID=?" % tbl, old_id)
                        except Exception:
                            pass
                    try:
                        cur.execute("DELETE FROM FixedAsset WHERE [%s]=?" % id_col, old_id)
                    except Exception:
                        pass
            ngay_dt = _misa_doc_ngay(ngct) or now
            ngay_kh_dt = _misa_doc_ngay(ngkh) or ngay_dt
            ng_gia_n = _snum(ng_gia)
            hao_mon_n = _snum(hao_mon)
            tg_sd_n = _snum(tg_sd)
            dvt_tg_n = _snum(dvt_tg)
            kh_thang_n = _snum(kh_thang)
            catid = cat_map.get(str(loai).strip()) or (
                next(iter(cat_map.values())) if cat_map else None)
            dvid = dvsd_map.get(str(dtpb).strip().upper()) or branch_id
            # Dòng THẬT (nhập tay trên MISA): RefID = RefDetailID =
            # FixedAssetID — cả 3 dùng CHUNG 1 GUID của tài sản. Bắt chước
            # y hệt (trước đây RefID sinh GUID riêng + RefDetailID NULL —
            # nghi phạm chính khiến lưới Ghi tăng không hiện dòng nào).
            fa_id = str(_uuid.uuid4())
            # Tỷ lệ KH tháng: dòng thật = 100/số tháng (vd 60 tháng -> 1.67%)
            # — KHÔNG phải giá trị 100 của cột "Tỷ lệ tính KH tháng (%)"
            # trong form Excel.
            ty_le_thang = round(100.0 / tg_sd_n, 2) if tg_sd_n else _snum(ty_le_kh)
            row = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols.values()}
            _misa_gan(row, cols, fa_id, "FixedAssetID")
            _misa_gan(row, cols, fa_id, "RefID")
            _misa_gan(row, cols, branch_id, "BranchID")
            _misa_gan(row, cols, hoc_rt, "RefType")
            _misa_gan(row, cols, str(ma)[:25], "FixedAssetCode")
            _misa_gan(row, cols, str(ten)[:128], "FixedAssetName")
            if catid is not None:
                _misa_gan(row, cols, catid, "FixedAssetCategoryID")
            _misa_gan(row, cols, dvid, "OrganizationUnitID", "DepartmentID", "UsingDepartmentID")
            _misa_gan(row, cols, str(e)[:20], "RefNo", "RefNoManagement")
            _misa_gan(row, cols, ngay_dt, "RefDate")
            _misa_gan(row, cols, ngay_kh_dt, "DepreciationDate", "StartDepreciationDate")
            _misa_gan(row, cols, str(tkng or ""), "OrgPriceAccount", "OriginalPriceAccount")
            _misa_gan(row, cols, str(tkkh or ""), "DepreciationAccount", "AccumDepreciationAccount")
            _misa_gan(row, cols, 1, "Quantity")
            _misa_gan(row, cols, ng_gia_n, "OrgPrice", "OriginalPrice")
            _misa_gan(row, cols, ng_gia_n, "DepreciationAmount", "DepreciationValue",
                     "DepreciationCost")
            _misa_gan(row, cols, tg_sd_n, "LifeTime")
            _misa_gan(row, cols, tg_sd_n, "LifeTimeRemaining")
            _misa_gan(row, cols, tg_sd_n, "LifeTimeInMonth")
            _misa_gan(row, cols, tg_sd_n, "LifeTimeRemainingInMonth")
            _misa_gan(row, cols, dvt_tg_n, "LifeTimeUnit", "LifeTimeType")
            _misa_gan(row, cols, dvt_tg_n, "LifeTimeRemainingUnit")
            _misa_gan(row, cols, ty_le_thang, "DepreciationRateMonth", "DepreciationRate")
            _misa_gan(row, cols, kh_thang_n, "MonthlyDepreciationAmount",
                     "DepreciationAmountMonth")
            _misa_gan(row, cols, kh_thang_n * 12, "YearlyDepreciationAmount")
            _misa_gan(row, cols, hao_mon_n, "AccumDepreciationAmount",
                     "AccumulatedDepreciationAmount")
            _misa_gan(row, cols, ng_gia_n - hao_mon_n, "RemainingAmount")
            _misa_gan(row, cols, _snum(ty_le_pb), "AllocationRate")
            _misa_gan(row, cols, str(tkcp or ""), "ExpenseAccount", "CostAccount")
            # GHI SỔ LUÔN theo yêu cầu: ghi tăng danh mục TSCĐ không sinh bút
            # toán Nợ/Có mới (hóa đơn mua đã hạch toán Nợ 211 ở Mua hàng) nên
            # đặt cờ đã ghi sổ ngay, khỏi phải mở từng bản ghi bấm Ghi sổ.
            _misa_gan(row, cols, 1, "IsPostedFinance")
            _misa_gan(row, cols, 1, "IsPostedManagement")
            _misa_gan(row, cols, hoc_fa_that[1] if hoc_fa_that else True, "IsEnoughVoucher")
            _misa_gan(row, cols, hoc_fa_that[0] if hoc_fa_that and hoc_fa_that[0] is not None
                     else dob, "DisplayOnBook")
            _misa_gan(row, cols, max_reforder + them + 1, "RefOrder")
            _misa_gan(row, cols, 0, "Inactive")
            _misa_gan(row, cols, ("Ghi tăng tài sản - %s" % ten)[:500], "JournalMemo", "Reason")
            _misa_gan(row, cols, now, "CreatedDate")
            _misa_gan(row, cols, now, "ModifiedDate")
            if hoc_fa_that and hoc_fa_that[2]:
                _misa_gan(row, cols, hoc_fa_that[2], "CreatedBy")
                _misa_gan(row, cols, hoc_fa_that[2], "ModifiedBy")
            if not preview:
                cs = list(row.keys())
                cur.execute("INSERT INTO FixedAsset ([%s]) VALUES (%s)" %
                           ("],[".join(cs), ",".join(["?"] * len(cs))),
                           [row[c] for c in cs])
                # Dòng SỔ CÁI ghi tăng (FixedAssetLedger) — chính là dòng mà
                # màn hình lưới "Ghi tăng" hiển thị. RefDate + PostedDate đều
                # NOTNULL và là cột lưới lọc theo "Kỳ".
                if cols_led:
                    lrow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_led.values()}
                    if led_id_tay:
                        _misa_gan(lrow, cols_led, next_led_id, "FixedAssetLedgerID")
                        next_led_id += 1
                    # dòng thật: RefID = RefDetailID = FixedAssetID (cùng GUID)
                    _misa_gan(lrow, cols_led, fa_id, "RefID")
                    _misa_gan(lrow, cols_led, fa_id, "RefDetailID")
                    _misa_gan(lrow, cols_led, fa_id, "FixedAssetID")
                    _misa_gan(lrow, cols_led, led_rt if led_rt is not None else hoc_rt, "RefType")
                    _misa_gan(lrow, cols_led, str(e)[:20], "RefNo")
                    _misa_gan(lrow, cols_led, ngay_dt, "RefDate")
                    _misa_gan(lrow, cols_led, ngay_dt, "PostedDate")
                    _misa_gan(lrow, cols_led, dvid, "OrganizationUnitID")
                    _misa_gan(lrow, cols_led, tg_sd_n, "LifeTimeInMonth")
                    _misa_gan(lrow, cols_led, tg_sd_n, "LifeTimeRemainingInMonth")
                    _misa_gan(lrow, cols_led, ty_le_thang, "DepreciationRateMonth")
                    _misa_gan(lrow, cols_led, kh_thang_n, "MonthlyDepreciationAmount")
                    _misa_gan(lrow, cols_led, _snum(gt_kh) or ng_gia_n, "DepreciationAmount")
                    _misa_gan(lrow, cols_led, hao_mon_n, "AccumDepreciationAmount")
                    _misa_gan(lrow, cols_led, hao_mon_n, "TotalDepreciationAmount")
                    _misa_gan(lrow, cols_led, ng_gia_n - hao_mon_n, "RemainingAmount")
                    # dòng thật: OriginDepreciationAmount = giá trị tính KH
                    _misa_gan(lrow, cols_led, _snum(gt_kh) or ng_gia_n,
                             "OriginDepreciationAmount")
                    _misa_gan(lrow, cols_led, str(tkkh or ""), "DepreciationAccount")
                    _misa_gan(lrow, cols_led, str(tkng or ""), "OrgPriceAccount")
                    _misa_gan(lrow, cols_led, ng_gia_n, "OrgPrice")
                    _misa_gan(lrow, cols_led, ("Ghi tăng tài sản - %s" % ten)[:500], "JournalMemo")
                    _misa_gan(lrow, cols_led, branch_id, "BranchID")
                    _misa_gan(lrow, cols_led, max_reforder + them + 1, "RefOrder")
                    # dòng thật đánh số riêng trong phân hệ TSCĐ: 1,2,3...
                    _misa_gan(lrow, cols_led, max_ro_sub + them + 1, "RefOrderInSubSystem")
                    _misa_gan(lrow, cols_led, str(ma)[:25], "FixedAssetCode")
                    _misa_gan(lrow, cols_led, str(ten)[:128], "FixedAssetName")
                    if catid is not None:
                        _misa_gan(lrow, cols_led, catid, "FixedAssetCategoryID")
                    lc = list(lrow.keys())
                    cur.execute("INSERT INTO FixedAssetLedger ([%s]) VALUES (%s)" %
                               ("],[".join(lc), ",".join(["?"] * len(lc))),
                               [lrow[c] for c in lc])
                # Thiết lập PHÂN BỔ (FixedAssetDetailAllocation) — khung
                # "Thiết lập phân bổ" dưới lưới Ghi tăng: Đối tượng PB +
                # Tỷ lệ PB + TK chi phí.
                if cols_alloc:
                    arow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_alloc.values()}
                    _misa_gan(arow, cols_alloc, str(_uuid.uuid4()), "FixedAssetDetailID")
                    _misa_gan(arow, cols_alloc, fa_id, "FixedAssetID")
                    _misa_gan(arow, cols_alloc, 0, "SortOrder")
                    _misa_gan(arow, cols_alloc, dvid, "ObjectID")
                    _misa_gan(arow, cols_alloc, hoc_objtype if hoc_objtype is not None else 1,
                             "ObjectType")
                    _misa_gan(arow, cols_alloc, _snum(ty_le_pb) or 100, "AllocationRate")
                    _misa_gan(arow, cols_alloc, str(tkcp or ""), "CostAccount")
                    ac = list(arow.keys())
                    cur.execute("INSERT INTO FixedAssetDetailAllocation ([%s]) VALUES (%s)" %
                               ("],[".join(ac), ",".join(["?"] * len(ac))),
                               [arow[c] for c in ac])
            them += 1
            if ghi_de_ts:
                go += 1
            ket.append({"ma": ma, "ten": ten, "nguyen_gia": ng_gia_n,
                        "trang_thai": ("sẽ ghi đè" if preview else "đã ghi đè") if ghi_de_ts
                                    else ("sẽ thêm" if preview else "đã thêm (ĐÃ ghi sổ)")})
        if preview:
            conn.rollback()
        else:
            conn.commit()
        return {"preview": preview, "database": database, "so_them": them, "so_trung": trung,
                "so_ghi_de": go, "danh_sach": ket[:500]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()

def _misa_ghi_tang_ccdc(cid, database, preview=True, ghi_de=False):
    """Ghi tăng CCDC thẳng vào MISA — CẤU TRÚC THẬT (đối chiếu cấu trúc CSDL
    thật, KHÁC hẳn suy đoán ban đầu): SUIncrement là bảng PHẲNG, mỗi DÒNG = 1
    CCDC (giống FixedAsset — KHÔNG phải chứng từ nhiều dòng kiểu PUVoucher).
    Phân bổ theo ĐƠN VỊ SỬ DỤNG nằm ở bảng riêng SUIncrementDetailDepartment,
    phân bổ theo ĐỐI TƯỢNG/TK CHI PHÍ nằm ở bảng riêng
    SUIncrementDetailAllocation — SUIncrementDetail chỉ là mô tả phụ
    (Description/NumberNo), không chứa dữ liệu chính. Nhận diện CCDC ĐÃ GHI
    qua Mã CCDC (SupplyCode, theo đúng dãy mã tự sinh CCDC00001... của phần
    mềm). RefType BẮT BUỘC là loại "Ghi tăng" thật (giống FixedAsset — xem
    _misa_ghi_tang_tscd)."""
    import uuid as _uuid
    header, rows = _doc_nhap_lieu(cid, "in")
    out = _gen_ghi_tang_ccdc(cid, header, rows)
    if not out:
        return {"preview": preview, "database": database, "so_them": 0, "so_trung": 0,
                "danh_sach": [],
                "ghi_chu": "Chưa có CCDC nào (Nợ 2421/242) trong Bảng kê đầu vào."}
    conn = _misa_sql_connect(cid, database=database)
    conn.autocommit = False
    try:
        cur = conn.cursor()
        branch_id = _misa_branch_id(cur)
        cols = _misa_cot_bang_that(cur, "SUIncrement")
        cols_dep = _misa_cot_bang_that(cur, "SUIncrementDetailDepartment")
        cols_alloc = _misa_cot_bang_that(cur, "SUIncrementDetailAllocation")
        cols_det = _misa_cot_bang_that(cur, "SUIncrementDetail")
        # Màn hình lưới "Công cụ dụng cụ > Ghi tăng" KHÔNG đọc SUIncrement
        # (bảng đó là "Sổ theo dõi CCDC") mà đọc bảng sổ cái SupplyLedger —
        # y hệt bài học FixedAssetLedger bên TSCĐ: chỉ ghi SUIncrement thì
        # CCDC hiện ở Sổ theo dõi nhưng lưới Ghi tăng trống.
        cols_led = _misa_cot_bang_that(cur, "SupplyLedger")
        if not cols:
            raise HTTPException(400, "Không tìm thấy bảng SUIncrement trong CSDL MISA đang kết nối.")
        # DỌN RÁC: các dòng chi tiết/sổ cái CCDC MỒ CÔI (SupplyID không còn
        # trong SUIncrement) — di chứng của các lần import cũ mà ghi đè chưa
        # gỡ hết (đối chiếu dữ liệu thật: SUIncrementDetailDepartment còn hàng
        # loạt SupplyID không khớp SUIncrement/SupplyLedger nào). Gỡ để dữ
        # liệu sạch, tránh lưới Ghi tăng nhiễu.
        if not preview:
            for _tbl in ("SupplyLedger", "SUIncrementDetailAllocation",
                        "SUIncrementDetailDepartment", "SUIncrementDetail",
                        "SUIncrementDetailSource"):
                try:
                    cur.execute("DELETE FROM %s WHERE SupplyID NOT IN "
                               "(SELECT SupplyID FROM SUIncrement)" % _tbl)
                except Exception:
                    pass
        hoc_rt, hoc_dob = _misa_hoc_reftype(cur, "SUIncrement", ["tăng"])
        # RefType sổ cái — học riêng từ SupplyLedger (bản ghi nhập tay trên
        # MISA), ưu tiên hơn phép đếm trên SUIncrement (bị dòng cũ do phần
        # mềm ghi lấn át) — như bài TSCĐ.
        led_rt = None
        try:
            _cnt_led = {}
            for (rt,) in cur.execute(
                    "SELECT RefType FROM SupplyLedger WHERE RefType<>0").fetchall():
                _cnt_led[rt] = _cnt_led.get(rt, 0) + 1
            if _cnt_led:
                led_rt = max(_cnt_led.items(), key=lambda kv: kv[1])[0]
        except Exception:
            pass
        if led_rt is not None:
            hoc_rt = led_rt
        # Fallback CỨNG như TSCĐ (chỉ dự phòng, ưu tiên học/dò được trước):
        # SupplyLedger dùng RefType 450 khi không biết gì — không để dòng sổ
        # cái CCDC biến mất khỏi lưới "Ghi tăng".
        if hoc_rt is None:
            hoc_rt = _SU_LEDGER_REFTYPE
        led_id_tay = "supplyledgerid" in cols_led
        next_led_id = 1
        if led_id_tay:
            try:
                next_led_id = (cur.execute(
                    "SELECT ISNULL(MAX(SupplyLedgerID),0) FROM SupplyLedger"
                ).fetchone()[0] or 0) + 1
            except Exception:
                pass
        max_ro_sub = 0
        try:
            r_sub = cur.execute(
                "SELECT ISNULL(MAX(RefOrderInSubSystem),0) FROM SupplyLedger").fetchone()
            max_ro_sub = (r_sub[0] or 0) if r_sub else 0
        except Exception:
            pass
        if hoc_rt is None:
            return {"preview": preview, "database": database, "so_them": 0, "so_trung": 0,
                    "khong_do_loai": True,
                    "danh_sach": [{"ma": r[0], "ten": r[1],
                                  "trang_thai": "bỏ qua — không dò được loại chứng từ MISA phù hợp"}
                                 for r in out]}
        # DisplayOnBook BẮT BUỘC thuộc (0,2): lưới "Công cụ dụng cụ > Ghi
        # tăng" lọc DisplayOnBook IN (0,2) — giá trị 3 (mặc định cũ / học lại
        # từ chính bản ghi phần mềm) bị loại khỏi lưới dù vẫn hiện ở Sổ theo
        # dõi. ĐÂY là nguyên nhân CCDC không hiện Ghi tăng dù mọi thứ khác
        # đúng (cùng bài học DisplayOnBook bên Mua hàng/TSCĐ).
        dob = hoc_dob if hoc_dob in (0, 2) else 0
        dvsd_map = _misa_dvsd_map(cur)
        # ObjectType dòng phân bổ — học từ bản ghi thật (CCDC trước, rồi tới
        # TSCĐ vì 2 màn hình dùng chung khái niệm Đối tượng phân bổ).
        hoc_objtype = None
        for tbl_ot in ("SUIncrementDetailAllocation", "FixedAssetDetailAllocation"):
            try:
                row_ot = cur.execute(
                    "SELECT TOP 1 ObjectType FROM %s WHERE ObjectType IS NOT NULL" % tbl_ot
                ).fetchone()
                if row_ot:
                    hoc_objtype = row_ot[0]
                    break
            except Exception:
                pass
        ma_col = _misa_chon_cot(cols, "SupplyCode")
        id_col = _misa_chon_cot(cols, "SupplyID", "RefID")
        pf_col = _misa_chon_cot(cols, "IsPostedFinance")
        pm_col = _misa_chon_cot(cols, "IsPostedManagement")
        ro_col = _misa_chon_cot(cols, "RefOrder")
        existed = {}
        if ma_col and id_col:
            try:
                sel = "SELECT [%s],[%s]" % (id_col, ma_col)
                sel += (",ISNULL([%s],0)" % pf_col) if pf_col else ",0"
                sel += (",ISNULL([%s],0)" % pm_col) if pm_col else ",0"
                sel += " FROM SUIncrement"
                for rid, ma, pf, pm in cur.execute(sel).fetchall():
                    if ma:
                        existed[str(ma).strip().lower()] = {"id": rid, "posted": bool(pf or pm)}
            except Exception:
                pass
        max_reforder = 0
        if ro_col:
            try:
                row_mo = cur.execute("SELECT MAX([%s]) FROM SUIncrement" % ro_col).fetchone()
                max_reforder = (row_mo[0] or 0) if row_mo else 0
            except Exception:
                pass
        now = datetime.datetime.now()
        them = trung = go = 0
        ket = []
        for r in out:
            (ma, ten, _loai, ly_do, e, ngct, tk_cho_pb, dvt, sl, dgia, tt, so_ky, tien_ky,
             dvsd, sl_dvsd, dtpb, ty_le_pb, tkcp, _kmcp) = (list(r) + [""] * 19)[:19]
            k = str(ma).strip().lower()
            ghi_de_su = False
            if k in existed:
                # Mã CCDC00001... là dãy do phần mềm tự sinh — bản ghi trùng
                # mã là của phần mềm, cho ghi đè kể cả đang mang cờ đã ghi sổ
                # (xem giải thích ở _misa_ghi_tang_tscd).
                if not ghi_de:
                    trung += 1
                    ket.append({"ma": ma, "ten": ten,
                                "trang_thai": "đã có (bỏ qua — bấm Ghi đè để ghi lại)"})
                    continue
                ghi_de_su = True
                if not preview:
                    old_id = existed[k]["id"]
                    for tbl in ("SupplyLedger", "SUIncrementDetailAllocation",
                               "SUIncrementDetailDepartment", "SUIncrementDetail",
                               "SUIncrementDetailSource"):
                        try:
                            cur.execute("DELETE FROM %s WHERE SupplyID=?" % tbl, old_id)
                        except Exception:
                            pass
                    try:
                        cur.execute("DELETE FROM SUIncrement WHERE [%s]=?" % id_col, old_id)
                    except Exception:
                        pass
            ngay_dt = _misa_doc_ngay(ngct) or now
            supply_id = str(_uuid.uuid4())
            dvid = dvsd_map.get(str(dvsd).strip().upper()) or branch_id
            objid = dvsd_map.get(str(dtpb).strip().upper()) or branch_id
            tt_n = _snum(tt)
            sl_n = _snum(sl) or 1
            so_ky_n = _snum(so_ky)
            row = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols.values()}
            _misa_gan(row, cols, supply_id, "SupplyID", "RefID")
            _misa_gan(row, cols, str(ma)[:25], "SupplyCode")
            _misa_gan(row, cols, str(ten)[:255], "SupplyName")
            _misa_gan(row, cols, branch_id, "BranchID")
            _misa_gan(row, cols, hoc_rt, "RefType")
            _misa_gan(row, cols, ngay_dt, "RefDate")
            _misa_gan(row, cols, str(e)[:20], "RefNo")
            # GHI SỔ LUÔN theo yêu cầu (như _misa_ghi_tang_tscd)
            _misa_gan(row, cols, 1, "IsPostedManagement")
            _misa_gan(row, cols, 1, "IsPostedFinance")
            _misa_gan(row, cols, str(dvt or "")[:20], "Unit")
            _misa_gan(row, cols, sl_n, "Quantity")
            _misa_gan(row, cols, _snum(dgia), "UnitPrice")
            _misa_gan(row, cols, tt_n, "Amount")
            _misa_gan(row, cols, so_ky_n, "AllocationTime")
            _misa_gan(row, cols, so_ky_n, "RemainingAllocationTime")
            _misa_gan(row, cols, 0, "AllocatedAmount")
            _misa_gan(row, cols, tt_n, "RemaingAmount")
            _misa_gan(row, cols, _snum(tien_ky), "TermlyAllocationAmount")
            _misa_gan(row, cols, str(tk_cho_pb or ""), "AllocationAccount")
            _misa_gan(row, cols, dob, "DisplayOnBook")
            _misa_gan(row, cols, max_reforder + them + 1, "RefOrder")
            _misa_gan(row, cols, now, "CreatedDate")
            _misa_gan(row, cols, now, "ModifiedDate")
            _misa_gan(row, cols, str(ly_do or "")[:255], "ReasonIncrement")
            _misa_gan(row, cols, False, "SuspendAllocate")
            # 1 DÒNG CHI TIẾT của increment có 1 SupplyDetailID DUY NHẤT, dùng
            # CHUNG cho cả 3 bảng con (SUIncrementDetail "Mô tả", ...Department
            # "Đơn vị sử dụng", ...Allocation "Phân bổ"); dòng sổ cái
            # SupplyLedger.RefDetailID trỏ về chính SupplyDetailID này. Đối
            # chiếu dữ liệu thật: SupplyLedger.RefDetailID KHÁC SupplyID và
            # KHÔNG khớp SUIncrementDetail.SupplyDetailID khi 3 bảng dùng ID
            # RIÊNG -> lưới Ghi tăng JOIN theo RefDetailID bị trượt, mất dòng.
            # Nay 3 bảng + ledger dùng CHUNG det_id nên mọi đường JOIN đều khớp.
            det_id = str(_uuid.uuid4())
            if not preview:
                cs = list(row.keys())
                cur.execute("INSERT INTO SUIncrement ([%s]) VALUES (%s)" %
                           ("],[".join(cs), ",".join(["?"] * len(cs))),
                           [row[c] for c in cs])
                if cols_dep:
                    drow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_dep.values()}
                    _misa_gan(drow, cols_dep, det_id, "SupplyDetailID")
                    _misa_gan(drow, cols_dep, supply_id, "SupplyID")
                    _misa_gan(drow, cols_dep, dvid, "OrganizationUnitID")
                    _misa_gan(drow, cols_dep, 0, "SortOrder")
                    _misa_gan(drow, cols_dep, _snum(sl_dvsd) or sl_n, "Quantity")
                    _misa_gan(drow, cols_dep, _snum(dgia), "UnitPrice")
                    _misa_gan(drow, cols_dep, tt_n, "Amount")
                    _misa_gan(drow, cols_dep, so_ky_n, "AllocationTime")
                    _misa_gan(drow, cols_dep, so_ky_n, "RemainingAllocationTime")
                    _misa_gan(drow, cols_dep, 0, "AllocatedAmount")
                    dc = list(drow.keys())
                    cur.execute("INSERT INTO SUIncrementDetailDepartment ([%s]) VALUES (%s)" %
                               ("],[".join(dc), ",".join(["?"] * len(dc))),
                               [drow[c] for c in dc])
                if cols_alloc:
                    arow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_alloc.values()}
                    _misa_gan(arow, cols_alloc, det_id, "SupplyDetailID")
                    _misa_gan(arow, cols_alloc, supply_id, "SupplyID")
                    _misa_gan(arow, cols_alloc, 0, "SortOrder")
                    _misa_gan(arow, cols_alloc, objid, "ObjectID")
                    _misa_gan(arow, cols_alloc, hoc_objtype if hoc_objtype is not None else 1,
                             "ObjectType")
                    _misa_gan(arow, cols_alloc, _snum(ty_le_pb), "AllocationRate")
                    _misa_gan(arow, cols_alloc, str(tkcp or ""), "CostAccount")
                    ac = list(arow.keys())
                    cur.execute("INSERT INTO SUIncrementDetailAllocation ([%s]) VALUES (%s)" %
                               ("],[".join(ac), ",".join(["?"] * len(ac))),
                               [arow[c] for c in ac])
                if cols_det:
                    xrow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_det.values()}
                    _misa_gan(xrow, cols_det, det_id, "SupplyDetailID")
                    _misa_gan(xrow, cols_det, supply_id, "SupplyID")
                    _misa_gan(xrow, cols_det, str(ten)[:255], "Description")
                    _misa_gan(xrow, cols_det, 0, "SortOrder")
                    xc = list(xrow.keys())
                    cur.execute("INSERT INTO SUIncrementDetail ([%s]) VALUES (%s)" %
                               ("],[".join(xc), ",".join(["?"] * len(xc))),
                               [xrow[c] for c in xc])
                # Dòng SỔ CÁI ghi tăng CCDC (SupplyLedger) — chính là dòng mà
                # màn hình lưới "Ghi tăng" hiển thị (RefDate/PostedDate lọc
                # theo Kỳ). RefID/RefDetailID = SupplyID theo mẫu FixedAsset.
                if cols_led:
                    lrow = {real: _misa_gia_tri_mac_dinh(t) for real, t in cols_led.values()}
                    if led_id_tay:
                        _misa_gan(lrow, cols_led, next_led_id, "SupplyLedgerID")
                        next_led_id += 1
                    _misa_gan(lrow, cols_led, supply_id, "RefID")
                    # RefDetailID = SupplyDetailID CHUNG của dòng chi tiết (khớp
                    # cả 3 bảng con) — không phải SupplyID, không phải ID riêng
                    _misa_gan(lrow, cols_led, det_id, "RefDetailID")
                    _misa_gan(lrow, cols_led, supply_id, "SupplyID")
                    _misa_gan(lrow, cols_led, hoc_rt, "RefType")
                    _misa_gan(lrow, cols_led, str(e)[:20], "RefNo")
                    _misa_gan(lrow, cols_led, ngay_dt, "RefDate")
                    _misa_gan(lrow, cols_led, ngay_dt, "PostedDate")
                    _misa_gan(lrow, cols_led, ("Ghi tăng CCDC - %s" % ten)[:500], "JournalMemo")
                    _misa_gan(lrow, cols_led, str(ten)[:255], "Description")
                    _misa_gan(lrow, cols_led, so_ky_n, "IncrementAllocationTime")
                    _misa_gan(lrow, cols_led, sl_n, "IncrementQuantity")
                    _misa_gan(lrow, cols_led, tt_n, "IncrementAmount")
                    _misa_gan(lrow, cols_led, _snum(tien_ky), "TermlyAllocationAmount")
                    _misa_gan(lrow, cols_led, branch_id, "BranchID")
                    _misa_gan(lrow, cols_led, 0, "SortOrder")
                    # dòng thật để RefOrder=0 (khác FixedAssetLedger) — bắt chước
                    _misa_gan(lrow, cols_led, 0, "RefOrder")
                    _misa_gan(lrow, cols_led, str(ma)[:25], "SupplyCode")
                    _misa_gan(lrow, cols_led, str(ten)[:255], "SupplyName")
                    _misa_gan(lrow, cols_led, dvid, "OrganizationUnitID")
                    _misa_gan(lrow, cols_led, max_ro_sub + them + 1, "RefOrderInSubSystem")
                    lc = list(lrow.keys())
                    cur.execute("INSERT INTO SupplyLedger ([%s]) VALUES (%s)" %
                               ("],[".join(lc), ",".join(["?"] * len(lc))),
                               [lrow[c] for c in lc])
            them += 1
            if ghi_de_su:
                go += 1
            ket.append({"ma": ma, "ten": ten, "thanh_tien": tt_n,
                        "trang_thai": ("sẽ ghi đè" if preview else "đã ghi đè") if ghi_de_su
                                    else ("sẽ thêm" if preview else "đã thêm (ĐÃ ghi sổ)")})
        if preview:
            conn.rollback()
        else:
            conn.commit()
        return {"preview": preview, "database": database, "so_them": them, "so_trung": trung,
                "so_ghi_de": go, "danh_sach": ket[:500]}
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(400, "Lỗi khi ghi vào MISA (đã hoàn tác, không ghi gì): %s" % str(e)[:400])
    finally:
        conn.close()

@app.post("/api/misa-sql/import-ghitang/{cid}")
def misa_sql_import_ghitang(cid: int, loai: str, preview: int = 1, database: str = "",
                            ghi_de: int = 0):
    """Ghi tăng TSCĐ (loai='tscd' -> bảng FixedAsset) / CCDC (loai='ccdc' ->
    bảng SUIncrement/SUIncrementDetail) thẳng vào MISA, đúng dữ liệu đã dùng
    cho file Excel mẫu Ghi Tăng TSCĐ/CCDC. Đây là GHI TĂNG DANH MỤC tài sản/
    CCDC (không phải chứng từ nhiều loại như Mua hàng). preview=1 -> chỉ xem
    trước, không ghi. Ghi ĐÃ GHI SỔ luôn (theo yêu cầu — ghi tăng danh mục
    không sinh bút toán mới), mỗi tài sản/CCDC 1 số chứng từ riêng = mã.
    ghi_de=1 -> gỡ bản ghi trùng mã (dãy mã do phần mềm tự sinh nên trùng mã
    = bản ghi của phần mềm, kể cả đang mang cờ đã ghi sổ) rồi ghi lại."""
    database = (database or "").strip() or (_misa_sql_cfg(cid).get("database") or "")
    if not database:
        raise HTTPException(400, "Chưa cấu hình kết nối/CSDL MISA. Mở '🗄 Kết nối CSDL MISA', "
                                 "kết nối tới dữ liệu THỬ trước.")
    if loai == "ccdc":
        return _misa_ghi_tang_ccdc(cid, database, preview=bool(preview), ghi_de=bool(ghi_de))
    if loai == "tscd":
        return _misa_ghi_tang_tscd(cid, database, preview=bool(preview), ghi_de=bool(ghi_de))
    raise HTTPException(400, "Loại '%s' chưa hỗ trợ (chỉ tscd/ccdc)." % loai)


@app.get("/api/danh-muc-ncc/{cid}")
def danh_muc_ncc(cid: int):
    """Tạo file Danh mục KH/NCC từ dữ liệu hóa đơn mua vào + bán ra.
    Lấy MST + tên từ Chi tiết MUA VÀO (nbmst/nbten) và BK Bán ra (nmmst/nmten).
    Lọc trùng MST, loại MST công ty mình. Trả về file Excel theo form mẫu MISA."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = db()
    rows = conn.execute(
        "SELECT loai, nbmst, nbten, nmmst, raw FROM invoices WHERE company_id=?", (cid,)).fetchall()
    comp = conn.execute("SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()
    conn.close()
    mst_cty = str(comp["mst"] or "").strip() if comp else ""

    def chuan_hoa_mst(s):
        """Chuẩn hóa MST: bỏ khoảng trắng, dấu gạch, chỉ giữ chữ số và chữ cái."""
        s = str(s or "").strip().replace("-", "").replace(" ", "").replace(".", "")
        return s

    def dinh_dang_mst(s):
        """MST 13 số (đơn vị trực thuộc) -> 10 số + '-' + 3 số: 0312320573003
        -> 0312320573-003. MST khác giữ nguyên."""
        s = chuan_hoa_mst(s)
        if len(s) == 13 and s.isdigit():
            return s[:10] + "-" + s[10:]
        return s

    # thu thập MST + tên (unique, lọc trùng)
    ds = {}  # {mst_chuan: ten}
    for r in rows:
        if r["loai"] == "purchase":
            mst = chuan_hoa_mst(r["nbmst"])
            ten = str(r["nbten"] or "").strip()
        else:
            try:
                raw = json.loads(r["raw"]) if r["raw"] else {}
            except Exception:
                raw = {}
            mst = chuan_hoa_mst(r["nmmst"] or raw.get("nmmst", ""))
            ten = str(raw.get("nmten", "") or "").strip()
        # bỏ qua: trống, MST công ty mình
        if not mst or mst == chuan_hoa_mst(mst_cty):
            continue
        # chỉ lấy lần đầu (tên dài hơn ưu tiên)
        if mst not in ds or (len(ten) > len(ds[mst])):
            ds[mst] = ten

    # ===== ĐỌC THÊM TỪ NHẬP LIỆU (bảng kê đã import vào nhap_lieu) =====
    try:
        conn2 = db()
        nl_rows = conn2.execute(
            "SELECT loai, header_json, rows_json FROM nhap_lieu WHERE company_id=?", (cid,)).fetchall()
        conn2.close()
        for nl in nl_rows:
            try:
                hdr = json.loads(nl["header_json"]) if nl["header_json"] else []
                data = json.loads(nl["rows_json"]) if nl["rows_json"] else []
            except Exception:
                continue
            hdr_low = [str(h or "").strip().lower() for h in hdr]
            # tìm cột MST và tên
            ci_mst = ci_ten = -1
            for i, h in enumerate(hdr_low):
                if ci_mst < 0 and ("mst" in h and ("bán" in h or "mua" in h or "ncc" in h or "cung" in h)):
                    ci_mst = i
                if ci_ten < 0 and ("người" in h and ("bán" in h or "mua" in h)):
                    ci_ten = i
            # fallback: cột có "mst" hoặc "mã số thuế"
            if ci_mst < 0:
                for i, h in enumerate(hdr_low):
                    if "mst" in h or "mã số thuế" in h:
                        ci_mst = i; break
            if ci_ten < 0:
                for i, h in enumerate(hdr_low):
                    if "tên" in h and ("bán" in h or "mua" in h or "ncc" in h or "khách" in h):
                        ci_ten = i; break
            if ci_mst < 0:
                continue
            for row in data:
                if ci_mst >= len(row):
                    continue
                mst = chuan_hoa_mst(row[ci_mst])
                ten = str(row[ci_ten] if ci_ten >= 0 and ci_ten < len(row) else "").strip()
                if not mst or mst == chuan_hoa_mst(mst_cty):
                    continue
                if mst not in ds or (ten and len(ten) > len(ds.get(mst, ""))):
                    ds[mst] = ten
    except Exception:
        pass

    # sắp xếp theo MST
    items = sorted(ds.items(), key=lambda x: x[0])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh mục NCC"
    # header
    headers = ["Là tổ chức/cá nhân", "Là khách hàng", "Mã nhà cung cấp (*)",
               "Tên nhà cung cấp (*)", "Địa chỉ", "Mã số thuế",
               "Điện thoại", "Fax", "Email", "Website", "Nhóm KH/NCC"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2E5C8A")

    for i, (mst, ten) in enumerate(items, 2):
        mst_hien = dinh_dang_mst(mst)              # 13 số -> 0312320573-003
        ws.cell(i, 1).value = 0                    # A: Là tổ chức/cá nhân = 0
        ws.cell(i, 2).value = 1                    # B: Là khách hàng = 1
        ws.cell(i, 3).value = mst_hien             # C: Mã NCC = MST (có gạch nếu 13 số)
        ws.cell(i, 3).number_format = "@"          # ép kiểu chữ để giữ số 0 đầu + dấu '-'
        ws.cell(i, 4).value = ten                  # D: Tên NCC
        # F: Mã số thuế = MST nếu là số hợp lệ và KHÔNG phải CCCD 12 số
        ws.cell(i, 6).value = mst_hien if (mst.isdigit() and len(mst) != 12) else ""
        ws.cell(i, 6).number_format = "@"

    # format
    from openpyxl.utils import get_column_letter
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 30
    ws.column_dimensions["F"].width = 20
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    fname = f"DanhMuc_KHNCC_{mst_cty}.xlsx"
    path = os.path.join(DOWNLOAD_DIR, fname)
    wb.save(path)
    desktop = _get_desktop_dir()
    if desktop and os.path.isdir(desktop):
        try:
            import shutil
            shutil.copy(path, os.path.join(desktop, fname))
        except Exception:
            pass
    return _resp_xuat(path, fname)


@app.post("/api/import-excel/{cid}")
async def import_excel(cid: int, request: Request, ky: str = ""):
    """
    Import file Excel đã kiểm tra/điều chỉnh. Đọc sheet 'Chi tiết MUA VÀO' và
    'Chi tiết BÁN RA', tính tổng doanh số + thuế (bán ra tách theo thuế suất),
    lưu vào imported_data. Sau khi import, xuất XML và tạm tính VAT dùng số này.
    """
    import openpyxl, io as _io
    form = await request.form()
    up = form.get("file")
    if up is None:
        raise HTTPException(400, "Chưa chọn file Excel")
    content = await up.read()
    try:
        wb = openpyxl.load_workbook(_io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Không đọc được file Excel: {e}")

    def _sach(s):
        # bỏ khoảng trắng thường + non-breaking space, phòng file bị chỉnh tay
        return str(s or "").replace("\xa0", " ").strip().lower()

    def col_idx(ws, ten, row=1):
        for c in range(1, ws.max_column + 1):
            if _sach(ws.cell(row, c).value) == ten.lower():
                return c
        return None

    mua_ds = mua_thue = 0
    mua_ds_nk = mua_thue_nk = 0     # TRONG ĐÓ: phần hàng NHẬP KHẨU (tờ khai NK) -> [23a]/[24a]
    mua_rows = 0
    mua_sheet_found = "BK Mua vào" in wb.sheetnames
    ban_sheet_found = "BK Bán ra" in wb.sheetnames
    ban = {"0": {"ds": 0, "thue": 0}, "5": {"ds": 0, "thue": 0},
           "8": {"ds": 0, "thue": 0}, "10": {"ds": 0, "thue": 0}}

    def num(v):
        try:
            return float(v)
        except Exception:
            return 0

    def norm_ts(v):
        s = str(v or "").replace("%", "").strip()
        if s in ("0", "5", "8", "10"):
            return s
        return None

    _TTHAI_MA_TU_MOTA = {v.lower(): k for k, v in TTHAI_DESC.items()}

    def _ma_trang_thai(mota_hoac_ma):
        """'Hóa đơn mới' -> '1' (chấp nhận cả khi ô đã ghi sẵn mã số)."""
        s = str(mota_hoac_ma or "").strip()
        if s in TTHAI_DESC:
            return s
        return _TTHAI_MA_TU_MOTA.get(s.lower(), "1")

    def _ngay_iso(v):
        """'dd/mm/yyyy' (hoặc đã là yyyy-mm-dd) -> 'yyyy-mm-dd'. '' nếu không đọc được."""
        s = str(v or "").strip()
        if not s:
            return ""
        if "-" in s and len(s.split("-")[0]) == 4:
            return s.split("T")[0].split()[0]
        try:
            d, m, y = s.split("/")
            return f"{int(y):04d}-{int(m):02d}-{int(d):02d}"
        except Exception:
            return ""

    mst_cty = str((conn := db()).execute(
        "SELECT mst FROM companies WHERE id=?", (cid,)).fetchone()["mst"] or "").strip()
    conn.close()
    mua_invoices = []   # hóa đơn MUA VÀO đọc được đầy đủ (để GHI ĐÈ bảng invoices)
    ban_invoices = []   # hóa đơn BÁN RA đọc được đầy đủ

    # ===== MUA VÀO: đọc từ sheet 'BK Mua vào' =====
    # Dò dòng tiêu đề ở vài dòng đầu (không giả định luôn ở dòng 1) — cùng cách
    # làm với BÁN RA bên dưới, để không bỏ sót khi file có thêm dòng tiêu đề/khoảng trắng.
    if mua_sheet_found:
        ws = wb["BK Mua vào"]
        c_ds = c_thue = c_kh = c_tt = None
        c_shdon = c_ngay = c_mst = c_ten = c_tt_bso = None
        hdr_row = 1
        for r in range(1, min(ws.max_row, 15) + 1):
            for c in range(1, ws.max_column + 1):
                v = _sach(ws.cell(r, c).value)
                if v == "doanh số mua chưa thuế":
                    c_ds = c; hdr_row = r
                elif v == "thuế gtgt":
                    c_thue = c; hdr_row = max(hdr_row, r)
                elif v in ("ký hiệu", "ki hieu"):
                    c_kh = c
                elif v in ("trạng thái", "trang thai"):
                    c_tt = c
                elif v in ("số hoá đơn", "số hóa đơn", "so hoa don"):
                    c_shdon = c
                elif v in ("ngày lập", "ngay lap"):
                    c_ngay = c
                elif v in ("mst người bán", "mst nguoi ban"):
                    c_mst = c
                elif v in ("tên người bán", "ten nguoi ban"):
                    c_ten = c
                elif v in ("tổng thanh toán", "tong thanh toan"):
                    c_tt_bso = c
            if c_ds:
                break
        for r in range(hdr_row + 1, ws.max_row + 1):
            full = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))
            if not full.strip():
                continue
            kh = _sach(ws.cell(r, c_kh).value) if c_kh else ""
            low_full = full.lower()
            # Bỏ dòng TỔNG CỘNG / Tổng nhóm — nhận biết bằng Ký hiệu ĐỂ TRỐNG
            # (dòng dữ liệu luôn có Ký hiệu). TUYỆT ĐỐI không lọc theo chữ 'tổng'
            # chung chung như trước, vì sẽ bỏ nhầm hóa đơn của NCC có TÊN chứa
            # 'TỔNG' (vd 'TỔNG CÔNG TY HÀNG KHÔNG VIỆT NAM', 'TỔNG CÔNG TY CP
            # BƯU CHÍNH VIETTEL') -> thiếu doanh số/thuế mua vào.
            if (c_kh and not kh) or "tổng cộng" in low_full or "tổng nhóm" in low_full:
                continue
            got = False
            row_ds = row_thue = 0
            if c_ds:
                row_ds = num(ws.cell(r, c_ds).value)
                mua_ds += row_ds
                got = True
            if c_thue:
                row_thue = num(ws.cell(r, c_thue).value)
                mua_thue += row_thue
                got = True
            # Nhận biết dòng HÀNG NHẬP KHẨU (tờ khai NK): Ký hiệu = 'TKNK' hoặc
            # Trạng thái ghi 'Tờ khai nhập khẩu' -> tách riêng cho [23a]/[24a].
            tt_stt = _sach(ws.cell(r, c_tt).value) if c_tt else ""
            if got and ("tknk" in kh or "nhập khẩu" in tt_stt or "nhap khau" in tt_stt):
                mua_ds_nk += row_ds
                mua_thue_nk += row_thue
            if got:
                mua_rows += 1
                # ĐỦ Ký hiệu + Số hóa đơn -> lưu thành 1 dòng hóa đơn đầy đủ để
                # GHI ĐÈ bảng invoices (không chỉ cộng số tổng như trước) — để
                # "Xuất Excel tổng hợp" dùng ĐÚNG dữ liệu vừa import, đúng kỳ.
                if c_shdon:
                    shdon_v = str(ws.cell(r, c_shdon).value or "").strip()
                    if shdon_v:
                        ngay_iso = _ngay_iso(ws.cell(r, c_ngay).value) if c_ngay else ""
                        tt_bso = num(ws.cell(r, c_tt_bso).value) if c_tt_bso else (row_ds + row_thue)
                        mua_invoices.append({
                            "khmshdon": "1", "khhdon": ws.cell(r, c_kh).value or "",
                            "shdon": shdon_v, "tdlap": ngay_iso,
                            "nbmst": str(ws.cell(r, c_mst).value or "").strip() if c_mst else "",
                            "nbten": ws.cell(r, c_ten).value or "" if c_ten else "",
                            "nmmst": mst_cty,
                            "tgtcthue": row_ds, "tgtthue": row_thue, "tgtttbso": tt_bso or (row_ds + row_thue),
                            "tthai": _ma_trang_thai(ws.cell(r, c_tt).value) if c_tt else "1",
                        })

    # ===== BÁN RA: đọc từ sheet 'BK Bán ra' (tách nhóm theo dòng tiêu đề) =====
    if "BK Bán ra" in wb.sheetnames:
        ws = wb["BK Bán ra"]
        c_ds = col_idx(ws, "Doanh số bán chưa thuế")
        c_thue = col_idx(ws, "Thuế GTGT")
        c_khms = col_idx(ws, "Ký hiệu mẫu")
        c_kh_b = col_idx(ws, "Ký hiệu HĐ") or col_idx(ws, "Ký hiệu")
        c_shdon_b = col_idx(ws, "Số hóa đơn") or col_idx(ws, "Số Hoá Đơn")
        c_ngay_b = col_idx(ws, "Ngày lập")
        c_mst_b = col_idx(ws, "MST người mua")
        c_ten_b = col_idx(ws, "Tên người mua")
        c_tt_b = col_idx(ws, "Trạng thái")
        # nếu không tìm thấy header (do có tiêu đề phía trên), dò theo cột cố định
        if not c_ds:
            # tìm dòng header chứa "Doanh số bán chưa thuế"
            for r in range(1, min(ws.max_row, 30) + 1):
                for c in range(1, ws.max_column + 1):
                    v = str(ws.cell(r, c).value or "")
                    if "Doanh số bán chưa thuế" in v:
                        c_ds = c
                    if v.strip() == "Thuế GTGT":
                        c_thue = c
                if c_ds:
                    break
        cur_nhom = None
        ban_rows = 0
        for r in range(1, ws.max_row + 1):
            # gộp text cả dòng để phát hiện tiêu đề nhóm / dòng tổng
            full = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))
            low = full.lower()
            head3 = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 4))
            # dòng tiêu đề nhóm thuế suất
            if "thuế suất" in head3.lower() or "%" in head3 or "không chịu thuế" in head3.lower():
                if "0%" in head3: cur_nhom = "0"
                elif "5%" in head3: cur_nhom = "5"
                elif "8%" in head3: cur_nhom = "8"
                elif "10%" in head3: cur_nhom = "10"
                elif "không chịu" in head3.lower(): cur_nhom = None
                continue
            # BỎ dòng tổng (Tổng nhóm / TỔNG CỘNG) — dùng CỤM TỪ cụ thể, KHÔNG
            # lọc theo chữ 'tổng' chung chung để tránh bỏ nhầm hóa đơn của khách
            # hàng có TÊN chứa 'TỔNG' (vd 'TỔNG CÔNG TY ...').
            if "tổng cộng" in low or "tổng nhóm" in low:
                continue
            ds = num(ws.cell(r, c_ds).value) if c_ds else 0
            th = num(ws.cell(r, c_thue).value) if c_thue else 0
            if (ds or th) and cur_nhom in ban:
                ban[cur_nhom]["ds"] += ds
                ban[cur_nhom]["thue"] += th
                ban_rows += 1
                if c_shdon_b:
                    shdon_v = str(ws.cell(r, c_shdon_b).value or "").strip()
                    if shdon_v:
                        ngay_iso = _ngay_iso(ws.cell(r, c_ngay_b).value) if c_ngay_b else ""
                        ban_invoices.append({
                            "khmshdon": str(ws.cell(r, c_khms).value or "1").strip() if c_khms else "1",
                            "khhdon": ws.cell(r, c_kh_b).value or "" if c_kh_b else "",
                            "shdon": shdon_v, "tdlap": ngay_iso,
                            "nmmst": str(ws.cell(r, c_mst_b).value or "").strip() if c_mst_b else "",
                            "nmten": ws.cell(r, c_ten_b).value or "" if c_ten_b else "",
                            "nbmst": mst_cty,
                            "tgtcthue": ds, "tgtthue": th, "tgtttbso": ds + th,
                            "tthai": _ma_trang_thai(ws.cell(r, c_tt_b).value) if c_tt_b else "1",
                        })
    else:
        ban_rows = 0

    # ===== GHI ĐÈ bảng 'invoices' (danh sách chi tiết dùng cho Xuất Excel tổng
    # hợp) bằng ĐÚNG dữ liệu vừa import — trước đây import chỉ lưu SỐ TỔNG
    # (imported_data), không đụng tới 'invoices', nên sau khi import "Xuất
    # Excel tổng hợp" (đọc từ 'invoices') vẫn hiện dữ liệu TRA CỨU/IMPORT CŨ
    # của kỳ trước, không phải dữ liệu vừa import. File "đã kiểm tra" là bản
    # CHÍNH THỨC/CUỐI CÙNG của kỳ đang import -> THAY THẾ TOÀN BỘ danh sách cũ
    # (giống hệt cách 1 lượt tra cứu mới thay thế dữ liệu cũ), KHÔNG cộng dồn
    # lẫn với kỳ trước — CHỈ áp dụng riêng cho loại (mua/bán) đọc được đủ Ký
    # hiệu + Số hóa đơn; loại không đọc được cột này thì GIỮ NGUYÊN dữ liệu cũ
    # (file cũ/thiếu cột thì vẫn chỉ lưu số tổng như trước, không đụng invoices).
    so_dong_ghi_de = 0
    if mua_invoices or ban_invoices:
        conn_gd = db()
        if mua_invoices:
            conn_gd.execute(
                "DELETE FROM invoices WHERE company_id=? AND loai='purchase'", (cid,))
            for inv in mua_invoices:
                conn_gd.execute("""
                    INSERT OR REPLACE INTO invoices
                    (company_id, loai, he_thong, nbmst, nbten, nmmst,
                     khmshdon, khhdon, shdon, tdlap, tgtcthue, tgtthue, tgtttbso, tthai, raw)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (cid, "purchase", "import", inv["nbmst"], inv["nbten"], inv["nmmst"],
                      inv["khmshdon"], inv["khhdon"], inv["shdon"], inv["tdlap"],
                      inv["tgtcthue"], inv["tgtthue"], inv["tgtttbso"], inv["tthai"],
                      json.dumps({"tthai": inv["tthai"], "nguon": "import_excel"}, ensure_ascii=False)))
                so_dong_ghi_de += 1
        if ban_invoices:
            conn_gd.execute(
                "DELETE FROM invoices WHERE company_id=? AND loai='sold'", (cid,))
            for inv in ban_invoices:
                conn_gd.execute("""
                    INSERT OR REPLACE INTO invoices
                    (company_id, loai, he_thong, nbmst, nbten, nmmst,
                     khmshdon, khhdon, shdon, tdlap, tgtcthue, tgtthue, tgtttbso, tthai, raw)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (cid, "sold", "import", inv["nbmst"], "", inv["nmmst"],
                      inv["khmshdon"], inv["khhdon"], inv["shdon"], inv["tdlap"],
                      inv["tgtcthue"], inv["tgtthue"], inv["tgtttbso"], inv["tthai"],
                      json.dumps({"tthai": inv["tthai"], "nguon": "import_excel"}, ensure_ascii=False)))
                so_dong_ghi_de += 1
        conn_gd.commit()
        conn_gd.close()

    conn = db()
    conn.execute("""
        INSERT INTO imported_data (company_id, ky, mua_ds, mua_thue,
            mua_ds_nk, mua_thue_nk,
            ban_ds_0, ban_ds_5, ban_thue_5, ban_ds_8, ban_thue_8,
            ban_ds_10, ban_thue_10, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ON CONFLICT(company_id, ky) DO UPDATE SET
            mua_ds=excluded.mua_ds, mua_thue=excluded.mua_thue,
            mua_ds_nk=excluded.mua_ds_nk, mua_thue_nk=excluded.mua_thue_nk,
            ban_ds_0=excluded.ban_ds_0, ban_ds_5=excluded.ban_ds_5,
            ban_thue_5=excluded.ban_thue_5, ban_ds_8=excluded.ban_ds_8,
            ban_thue_8=excluded.ban_thue_8, ban_ds_10=excluded.ban_ds_10,
            ban_thue_10=excluded.ban_thue_10, updated_at=excluded.updated_at
    """, (cid, ky, round(mua_ds), round(mua_thue),
          round(mua_ds_nk), round(mua_thue_nk), round(ban["0"]["ds"]),
          round(ban["5"]["ds"]), round(ban["5"]["thue"]),
          round(ban["8"]["ds"]), round(ban["8"]["thue"]),
          round(ban["10"]["ds"]), round(ban["10"]["thue"]),
          datetime.datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return {
        "ok": True, "ky": ky,
        "mua_ds": round(mua_ds), "mua_thue": round(mua_thue), "mua_rows": mua_rows,
        "mua_ds_nk": round(mua_ds_nk), "mua_thue_nk": round(mua_thue_nk),
        "mua_sheet_found": mua_sheet_found,
        "ban_thue": round(sum(b["thue"] for b in ban.values())),
        "ban_8_ds": round(ban["8"]["ds"]), "ban_rows": ban_rows,
        "ban_sheet_found": ban_sheet_found,
        "so_dong_ghi_de": so_dong_ghi_de,
    }


def _get_imported(cid, ky=""):
    """Lấy dữ liệu đã import ĐÚNG KỲ. Nếu kỳ này chưa import -> None (dùng tra cứu).
    Chỉ khi không truyền kỳ (ky='') mới lấy bản import mới nhất."""
    conn = db()
    row = None
    if ky:
        # chỉ lấy import KHỚP ĐÚNG kỳ đang xem
        row = conn.execute("SELECT * FROM imported_data WHERE company_id=? AND ky=?",
                           (cid, ky)).fetchone()
    else:
        # không có kỳ -> lấy bản mới nhất (vd gọi không kèm kỳ)
        row = conn.execute(
            "SELECT * FROM imported_data WHERE company_id=? ORDER BY updated_at DESC LIMIT 1",
            (cid,)).fetchone()
    conn.close()
    return row


@app.get("/api/imported/{cid}")
def get_imported_api(cid: int, ky: str = ""):
    """Trả dữ liệu đã import để màn hình hiển thị (nếu có)."""
    row = _get_imported(cid, ky)
    if not row:
        return {"has_import": False}
    d = dict(row)
    mua_ds = _to_num(d["mua_ds"]) or 0
    mua_thue = _to_num(d["mua_thue"]) or 0
    ban_ds = sum((_to_num(d[k]) or 0) for k in
                 ["ban_ds_0", "ban_ds_5", "ban_ds_8", "ban_ds_10"])
    ban_thue = sum((_to_num(d[k]) or 0) for k in
                   ["ban_thue_5", "ban_thue_8", "ban_thue_10"])
    return {
        "has_import": True, "ky": d["ky"],
        "mua_ds": round(mua_ds), "mua_thue": round(mua_thue),
        "ban_ds": round(ban_ds), "ban_thue": round(ban_thue),
        "updated_at": d["updated_at"],
    }


@app.delete("/api/imported/{cid}")
def del_imported_api(cid: int):
    """Xóa dữ liệu import -> quay lại dùng dữ liệu tra cứu."""
    conn = db()
    conn.execute("DELETE FROM imported_data WHERE company_id=?", (cid,))
    conn.commit()
    conn.close()
    return {"ok": True}


def _khoang_ngay_ky(ky):
    """'MM/YYYY' hoặc 'QX/YYYY' -> (d_tu, d_den) của đúng kỳ đó (tháng thì
    trọn tháng, quý thì trọn 3 tháng). (None, None) nếu ky không hợp lệ."""
    import calendar as _cal
    ky = (ky or "").strip()
    if not ky or "/" not in ky:
        return None, None
    try:
        a, b = ky.split("/", 1)
        a = a.strip().upper()
        yyyy = int(b.strip())
        if a.startswith("Q") and a[1:].isdigit():
            q = max(1, min(4, int(a[1:])))
            mm1 = (q - 1) * 3 + 1
            mm2 = mm1 + 2
            d_tu = datetime.date(yyyy, mm1, 1)
            d_den = datetime.date(yyyy, mm2, _cal.monthrange(yyyy, mm2)[1])
        else:
            mm = int(a)
            d_tu = datetime.date(yyyy, mm, 1)
            d_den = datetime.date(yyyy, mm, _cal.monthrange(yyyy, mm)[1])
        return d_tu, d_den
    except Exception:
        return None, None


def _tong_thue_nk_tokhai(cid, ky):
    """Tổng thuế GTGT hàng nhập khẩu (bảng tokhai_nhap — import RIÊNG từ tờ
    khai hải quan, KHÔNG phải import bảng kê Excel) của ĐÚNG kỳ đang tính."""
    d_tu, d_den = _khoang_ngay_ky(ky)
    if not d_tu:
        return 0
    conn = db()
    tk_rows = conn.execute(
        "SELECT ngay_dk, items_json FROM tokhai_nhap WHERE company_id=?", (cid,)).fetchall()
    conn.close()
    tong = 0
    for tkr in tk_rows:
        s = str(tkr["ngay_dk"] or "").split("T")[0]
        try:
            d = datetime.datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            continue
        if not (d_tu <= d <= d_den):
            continue
        try:
            its = json.loads(tkr["items_json"]) if tkr["items_json"] else []
        except Exception:
            its = []
        for it in its:
            tong += round(it.get("tien_thue_gtgt", 0) or 0)
    return tong


@app.get("/api/vat-tmtinh/{cid}")
def vat_tam_tinh(cid: int, ky: str = "", du_dau_ky: float = None):
    """
    Tạm tính thuế VAT trong kỳ:
      - vat_mua: tổng VAT đầu vào (từ hóa đơn mua, loại HĐ thay thế/hủy)
      - vat_ban: tổng VAT đầu ra (từ hóa đơn bán)
      - du_dau_ky: số thuế còn được khấu trừ kỳ trước chuyển sang (người dùng nhập
        hoặc tự lấy 'du_cuoi_ky' của kỳ trước nếu đã lưu)
    Công thức: chenh = vat_ban - vat_mua - du_dau_ky
      chenh > 0 -> phải nộp = chenh ; du_cuoi_ky = 0
      chenh <= 0 -> phải nộp = 0 ; du_cuoi_ky = -chenh (chuyển kỳ sau)
    """
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    rows = conn.execute("SELECT * FROM invoices WHERE company_id=?", (cid,)).fetchall()
    if not comp:
        conn.close()
        raise HTTPException(404, "Không tìm thấy công ty")

    def loai_bo(r):
        try:
            raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            raw = {}
        tt = str(raw.get("tthai", r["tthai"]) or "").strip()
        if tt in ("4", "6"):
            return True
        # loại HĐ không đủ điều kiện cấp mã
        kq = str(raw.get("ttxly", "") or "").strip().lower()
        kq_mota = _mo_ta_ket_qua(raw.get("ttxly", "")).lower()
        if kq == "4" or "không đủ điều kiện" in kq or "không đủ điều kiện" in kq_mota:
            return True
        return False

    # Thuế GTGT hàng NHẬP KHẨU từ bảng tokhai_nhap (import RIÊNG tờ khai hải
    # quan) — TRƯỚC ĐÂY hàm này hoàn toàn không cộng khoản này, nên sau khi
    # import tờ khai nhập khẩu, phần tạm tính VẪN GIỮ NGUYÊN như chưa import.
    thue_nk_rieng = _tong_thue_nk_tokhai(cid, ky)

    # ƯU TIÊN dữ liệu đã import từ Excel (nếu có); nếu không, dùng dữ liệu tra cứu
    imp = _get_imported(cid, ky)
    if imp:
        vat_mua = _to_num(imp["mua_thue"]) or 0
        # Nếu file Excel bảng kê ĐÃ có sẵn dòng hàng nhập khẩu (ký hiệu TKNK) thì
        # mua_thue đã bao gồm thuế NK rồi — KHÔNG cộng thêm lần nữa (giống hệt
        # cách tính của tờ khai chính thức ở /api/xuat-to-khai) để tránh trùng.
        imp_nk_thue = _to_num(imp["mua_thue_nk"]) or 0 if "mua_thue_nk" in imp.keys() else 0
        if not imp_nk_thue:
            vat_mua += thue_nk_rieng
        vat_ban = (_to_num(imp["ban_thue_5"]) or 0) + (_to_num(imp["ban_thue_8"]) or 0) \
                  + (_to_num(imp["ban_thue_10"]) or 0)
        nguon = "import"
    else:
        vat_mua = vat_ban = 0
        for r in rows:
            if loai_bo(r):
                continue
            if r["loai"] == "purchase":
                vat_mua += _to_num(r["tgtthue"]) or 0
            elif r["loai"] == "sold":
                vat_ban += _to_num(r["tgtthue"]) or 0
        # Hóa đơn tra cứu từ trang Thuế KHÔNG bao giờ có tờ khai nhập khẩu (đây
        # là dữ liệu hải quan, nhập riêng) -> luôn cộng thêm.
        vat_mua += thue_nk_rieng
        nguon = "tra cứu"

    # số dư đầu kỳ: ưu tiên giá trị truyền vào; nếu None thì lấy du_cuoi_ky kỳ trước
    if du_dau_ky is None:
        prev = None
        prev_ky = _ky_lien_truoc(ky)   # hỗ trợ cả 'MM/YYYY' và 'QX/YYYY'
        if prev_ky:
            prev = conn.execute(
                "SELECT du_cuoi_ky FROM vat_balance WHERE company_id=? AND ky=?",
                (cid, prev_ky)).fetchone()
        du_dau_ky = (prev["du_cuoi_ky"] if prev and prev["du_cuoi_ky"] else 0) or 0
        # nếu kỳ này đã lưu rồi, lấy lại số dư đầu kỳ đã lưu
        cur = conn.execute(
            "SELECT du_dau_ky FROM vat_balance WHERE company_id=? AND ky=?",
            (cid, ky)).fetchone()
        if cur and cur["du_dau_ky"] is not None:
            du_dau_ky = cur["du_dau_ky"]

    chenh = vat_ban - vat_mua - (du_dau_ky or 0)
    phai_nop = chenh if chenh > 0 else 0
    du_cuoi_ky = -chenh if chenh <= 0 else 0
    conn.close()
    return {
        "ky": ky, "vat_mua": round(vat_mua), "vat_ban": round(vat_ban),
        "du_dau_ky": round(du_dau_ky or 0),
        "phai_nop": round(phai_nop), "du_cuoi_ky": round(du_cuoi_ky),
        "nguon": nguon,
    }


@app.post("/api/vat-tmtinh/{cid}")
def vat_luu(cid: int, data: dict = Body(...)):
    """Lưu số dư VAT của kỳ (gọi khi bấm Save)."""
    ky = data.get("ky", "")
    if not ky:
        raise HTTPException(400, "Thiếu kỳ kê khai")
    conn = db()
    conn.execute("""
        INSERT INTO vat_balance (company_id, ky, du_dau_ky, vat_mua, vat_ban,
                                 phai_nop, du_cuoi_ky, updated_at)
        VALUES (?,?,?,?,?,?,?,?)
        ON CONFLICT(company_id, ky) DO UPDATE SET
            du_dau_ky=excluded.du_dau_ky, vat_mua=excluded.vat_mua,
            vat_ban=excluded.vat_ban, phai_nop=excluded.phai_nop,
            du_cuoi_ky=excluded.du_cuoi_ky, updated_at=excluded.updated_at
    """, (cid, ky, data.get("du_dau_ky", 0), data.get("vat_mua", 0),
          data.get("vat_ban", 0), data.get("phai_nop", 0),
          data.get("du_cuoi_ky", 0), datetime.datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.get("/api/export-htkk/{cid}")
def export_htkk(cid: int, ky: str = "", nguoi_ky: str = "", tu: str = "", den: str = ""):
    """
    Tạo file XML tờ khai 01/GTGT (TT80/2021) để nhập vào HTKK.
    Hàng bán ra 8% -> đưa vào nhóm 10% (ct32/ct33) + Phụ lục NQ142/2024.
    ky: 'MM/YYYY' (vd '05/2026'); nếu trống lấy từ hóa đơn.
    """
    import html as _html
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    rows = conn.execute("SELECT * FROM invoices WHERE company_id=?", (cid,)).fetchall()
    if not comp:
        conn.close()
        raise HTTPException(404, "Không tìm thấy công ty")

    # Việc 1: lưu tên người ký vào công ty; nếu không truyền thì dùng cái đã lưu
    try:
        saved_nk = comp["nguoi_ky"] if "nguoi_ky" in comp.keys() else ""
    except Exception:
        saved_nk = ""
    if nguoi_ky:
        conn.execute("UPDATE companies SET nguoi_ky=? WHERE id=?", (nguoi_ky, cid))
        conn.commit()
    else:
        nguoi_ky = saved_nk or ""
    conn.close()

    def status_loai_bo(r):
        try:
            raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            raw = {}
        tt = str(raw.get("tthai", r["tthai"]) or "").strip()
        if tt in ("4", "6"):  # đã bị thay thế / hủy
            return True
        kq = str(raw.get("ttxly", "") or "").strip().lower()
        kq_mota = _mo_ta_ket_qua(raw.get("ttxly", "")).lower()
        if kq == "4" or "không đủ điều kiện" in kq or "không đủ điều kiện" in kq_mota:
            return True
        return False

    mst_cty = str(comp["mst"] or "").strip()

    # ===== XÁC ĐỊNH KỲ KÊ KHAI + RANH GIỚI NGÀY *TRƯỚC* khi cộng dồn hóa đơn =====
    # QUAN TRỌNG: trước đây hàm này cộng dồn TẤT CẢ hóa đơn/tờ khai NK của công
    # ty (không lọc theo ngày) rồi mới xác định kỳ ở CUỐI hàm -> tờ khai kỳ nào
    # cũng vô tình cộng nhầm cả hóa đơn của kỳ khác (kể cả những năm trước) còn
    # lưu trong CSDL, khiến số liệu luôn SAI/CAO HƠN thực tế. Giờ xác định kỳ
    # (và khoảng ngày dd/mm/yyyy tương ứng) TRƯỚC, rồi chỉ cộng những hóa đơn có
    # ngày lập (tdlap) THỰC SỰ nằm trong đúng khoảng đó.
    import calendar

    def _ngay_hdon(tdlap):
        """'tdlap' dạng 'YYYY-MM-DD...' (hoặc có hậu tố giờ) -> date, None nếu hỏng."""
        s = str(tdlap or "").split("T")[0]
        try:
            return datetime.datetime.strptime(s, "%Y-%m-%d").date()
        except Exception:
            return None

    ky_auto = ""
    if not ky or "/" not in ky:
        for r in rows:
            if r["loai"] == "sold" and not status_loai_bo(r):
                d = _ngay_hdon(r["tdlap"])
                if d:
                    ky_auto = f"{d.month:02d}/{d.year}"
                    break
        ky = ky or ky_auto or "01/2026"

    # ky có thể là 'MM/YYYY' hoặc 'QX/YYYY' (kỳ quý) -> tách an toàn, không crash
    try:
        mm, yyyy = _ky_ve_thang(ky)
    except Exception:
        ky = ky_auto or "01/2026"
        mm, yyyy = _ky_ve_thang(ky)
    last_day = calendar.monthrange(int(yyyy), int(mm))[1]

    # ===== NHẬN BIẾT KỲ QUÝ: nếu khoảng ngày tra cứu (tu/den do FE gửi) trải
    # đúng 1 quý -> dùng ĐÚNG khoảng ngày đó làm ranh giới lọc (không chỉ để
    # LABEL); nếu FE không gửi tu/den thì mặc định lọc theo cả THÁNG của ky =====
    la_quy = False
    quy_so = 0
    tu_ngay_kkhai = f"01/{mm}/{yyyy}"
    den_ngay_kkhai = f"{last_day}/{mm}/{yyyy}"
    d_tu = datetime.date(int(yyyy), int(mm), 1)
    d_den = datetime.date(int(yyyy), int(mm), last_day)
    if tu and den:
        try:
            d1 = datetime.datetime.strptime(tu, "%d/%m/%Y").date()
            d2 = datetime.datetime.strptime(den, "%d/%m/%Y").date()
            songay = (d2 - d1).days
            # quý: bắt đầu tháng 1/4/7/10, kéo dài ~3 tháng
            if d1.month in (1, 4, 7, 10) and d1.day == 1 and 85 <= songay <= 95:
                la_quy = True
                quy_so = (d1.month - 1) // 3 + 1
                yyyy = str(d1.year)
            tu_ngay_kkhai = d1.strftime("%d/%m/%Y")
            den_ngay_kkhai = d2.strftime("%d/%m/%Y")
            d_tu, d_den = d1, d2
        except Exception:
            pass

    def trong_ky(tdlap):
        d = _ngay_hdon(tdlap)
        return bool(d and d_tu <= d <= d_den)

    # ----- Tổng MUA VÀO (PHẢI KHỚP với sheet 'BK Mua vào' trong file Excel) -----
    # Dùng ĐÚNG cách tính của BK Mua vào:
    #  - CHỈ lấy hóa đơn có ngày lập THUỘC ĐÚNG KỲ đang kê khai
    #  - chỉ loại HĐ thay thế (4) / xóa bỏ (6); KHÔNG loại 'không đủ điều kiện'
    #    (để [23]/[24] khớp bảng kê — vẫn liệt kê đủ HĐ mua vào)
    #  - loại HĐ lẫn của công ty khác (MST người mua khác MST công ty)
    #  - HKD: nếu tgtcthue=0 nhưng tgtttbso>0 -> lấy tgtttbso (doanh số)
    mua_ds = mua_thue = 0
    for r in rows:
        if r["loai"] != "purchase":
            continue
        if not trong_ky(r["tdlap"]):
            continue
        try:
            _raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            _raw = {}
        _tt = str(_raw.get("tthai", r["tthai"]) or "").strip()
        if _tt in ("4", "6"):
            continue
        _nm = str(r["nmmst"] or "").strip()
        if _nm and _nm != mst_cty:
            continue
        ds = _to_num(r["tgtcthue"]) or 0
        if not ds:
            ds = _to_num(r["tgtttbso"]) or 0
        mua_ds += ds
        mua_thue += _to_num(r["tgtthue"]) or 0

    # ----- BÁN RA: tách theo nhóm thuế suất từ chi tiết (CHỈ hóa đơn thuộc kỳ) -----
    # Cần đọc chi tiết để biết hàng nào 8% (NQ142). Dùng detail_json/file nếu có.
    ban_theo_ts = {"0": {"ds": 0, "thue": 0}, "5": {"ds": 0, "thue": 0},
                   "8": {"ds": 0, "thue": 0}, "10": {"ds": 0, "thue": 0},
                   "KCT": {"ds": 0, "thue": 0}}
    save_dir = (comp["save_dir"] or "").strip() if comp else ""

    # Xây index file 1 lần (tránh os.walk lặp lại gây chậm)
    _fidx = {}
    if save_dir and os.path.isdir(save_dir):
        for rootdir, _d, files in os.walk(save_dir):
            for fn in files:
                low = fn.lower()
                if not (low.endswith(".zip") or low.endswith(".xml")):
                    continue
                nm = fn.rsplit(".", 1)[0]; parts = nm.split("_")
                if len(parts) >= 2:
                    _fidx.setdefault((parts[0], parts[1].lstrip("0") or "0"),
                                     os.path.join(rootdir, fn))

    def get_summary(r):
        """CHỈ dùng dữ liệu đã có (không gọi mạng để tránh treo)."""
        try:
            dj = r["detail_json"] if "detail_json" in r.keys() else None
        except Exception:
            dj = None
        if dj:
            try:
                return _summary_from_detail_json(json.loads(dj))
            except Exception:
                pass
        fp = _fidx.get((str(r["khhdon"] or ""), str(r["shdon"] or "").lstrip("0") or "0"))
        if fp:
            try:
                with open(fp, "rb") as f:
                    return _parse_invoice_summary(f.read())
            except Exception:
                pass
        return None

    def suy_nhom_thue(ds, thue):
        """Khi không có chi tiết, suy nhóm thuế suất từ tỷ lệ thuế/doanh số."""
        if not ds:
            return "10"
        ty_le = (thue or 0) / ds * 100
        if ty_le < 1:
            return "0"
        for muc in ("5", "8", "10"):
            if abs(ty_le - float(muc)) < 1.5:
                return muc
        return "10"

    for r in rows:
        if r["loai"] != "sold" or status_loai_bo(r):
            continue
        if not trong_ky(r["tdlap"]):
            continue
        info = get_summary(r)
        if info and info.get("theo_ts"):
            for k, v in info["theo_ts"].items():
                tgt = k if k in ban_theo_ts else "10"
                ban_theo_ts[tgt]["ds"] += v["ds"]
                ban_theo_ts[tgt]["thue"] += v["thue"]
        else:
            ds = _to_num(r["tgtcthue"]) or 0
            thue = _to_num(r["tgtthue"]) or 0
            nhom = suy_nhom_thue(ds, thue)
            ban_theo_ts[nhom]["ds"] += ds
            ban_theo_ts[nhom]["thue"] += thue

    # ƯU TIÊN dữ liệu đã import từ Excel (nếu có, đúng kỳ) -> ghi đè số liệu tra cứu
    imp = _get_imported(cid, ky)
    imp_nk_ds = imp_nk_thue = 0     # phần hàng NHẬP KHẨU đã nằm SẴN trong mua_ds/mua_thue của file import
    if imp:
        mua_ds = _to_num(imp["mua_ds"]) or 0
        mua_thue = _to_num(imp["mua_thue"]) or 0
        # phần hàng nhập khẩu (tờ khai NK) đã được cộng SẴN trong tổng mua vào
        # của 'BK Mua vào' -> chỉ tách ra để điền [23a]/[24a], KHÔNG cộng thêm lần nữa.
        try:
            imp_nk_ds = _to_num(imp["mua_ds_nk"]) or 0
            imp_nk_thue = _to_num(imp["mua_thue_nk"]) or 0
        except Exception:
            imp_nk_ds = imp_nk_thue = 0
        ban_theo_ts = {
            "0": {"ds": _to_num(imp["ban_ds_0"]) or 0, "thue": 0},
            "5": {"ds": _to_num(imp["ban_ds_5"]) or 0, "thue": _to_num(imp["ban_thue_5"]) or 0},
            "8": {"ds": _to_num(imp["ban_ds_8"]) or 0, "thue": _to_num(imp["ban_thue_8"]) or 0},
            "10": {"ds": _to_num(imp["ban_ds_10"]) or 0, "thue": _to_num(imp["ban_thue_10"]) or 0},
            "KCT": {"ds": 0, "thue": 0},
        }

    # ===== SỐ DƯ ĐẦU KỲ [22]: lấy từ tạm tính VAT đã lưu (vat_balance) =====
    ct22_val = 0
    try:
        conn2 = db()
        ky_tim = (f"Q{quy_so}/{yyyy}" if la_quy else ky)
        vb = conn2.execute(
            "SELECT du_dau_ky FROM vat_balance WHERE company_id=? AND ky=?",
            (cid, ky_tim)).fetchone()
        if not vb:
            vb = conn2.execute(
                "SELECT du_dau_ky FROM vat_balance WHERE company_id=? AND ky=?",
                (cid, ky)).fetchone()
        conn2.close()
        if vb and vb["du_dau_ky"]:
            ct22_val = round(_to_num(vb["du_dau_ky"]) or 0)
    except Exception:
        pass

    # ----- Tính các chỉ tiêu tờ khai -----
    # ===== TỜ KHAI NHẬP KHẨU: tổng trị giá tính thuế GTGT + thuế GTGT hàng NK
    # CỦA ĐÚNG KỲ (theo ngày đăng ký tờ khai, ngay_dk) — trước đây cộng TẤT CẢ
    # tờ khai nhập khẩu từng nhập cho công ty, không lọc theo kỳ, nên tờ khai
    # GTGT kỳ nào cũng bị cộng nhầm cả tờ khai NK của kỳ/năm khác. =====
    tk_ds_nk = tk_thue_nk = 0
    try:
        conn_tk = db()
        tk_rows = conn_tk.execute(
            "SELECT ngay_dk, items_json FROM tokhai_nhap WHERE company_id=?", (cid,)).fetchall()
        conn_tk.close()
        for tkr in tk_rows:
            if not trong_ky(tkr["ngay_dk"]):
                continue
            try:
                its = json.loads(tkr["items_json"]) if tkr["items_json"] else []
            except Exception:
                its = []
            for it in its:
                tk_ds_nk += round(it.get("tri_gia_gtgt", 0) or 0)
                tk_thue_nk += round(it.get("tien_thue_gtgt", 0) or 0)
    except Exception:
        pass

    # Hàng 8% bản chất là 10% được giảm -> gộp vào nhóm 10% (ct32/ct33)
    ds_8 = ban_theo_ts["8"]["ds"]
    thue_8 = ban_theo_ts["8"]["thue"]
    # nhóm 10% trên tờ khai = hàng 10% thật + hàng 8% (tính theo thuế GỐC 10%)
    ds_10 = ban_theo_ts["10"]["ds"] + ds_8
    thue_10 = ban_theo_ts["10"]["thue"] + thue_8
    ds_5 = ban_theo_ts["5"]["ds"]; thue_5 = ban_theo_ts["5"]["thue"]
    ds_0 = ban_theo_ts["0"]["ds"]

    ct23 = round(mua_ds); ct24 = round(mua_thue); ct25 = ct24
    ct30 = round(ds_5); ct31 = round(thue_5)
    ct32 = round(ds_10); ct33 = round(thue_10)
    ct27 = ct30 + ct32; ct28 = ct31 + ct33
    ct29 = round(ds_0)
    ct34 = ct27 + ct29; ct35 = ct28
    # NQ142: thuế được giảm = doanh số 8% × 2%
    thue_duoc_giam = round(ds_8 * 0.02)

    # ===== DÙNG TEMPLATE XML THẬT, chỉ thay giá trị các thẻ (giữ nguyên cấu trúc) =====
    import re as _re
    tpl_path = os.path.join(BASE_DIR, "templates", "htkk_01gtgt_template.xml")
    with open(tpl_path, encoding="utf-8-sig") as f:
        xml = f.read()

    def set_tag(xml, tag, value):
        """Thay nội dung <tag>...</tag> bằng value (chỉ thẻ đầu tiên khớp)."""
        pat = _re.compile(r"(<" + tag + r">)(.*?)(</" + tag + r">)", _re.DOTALL)
        return pat.sub(lambda m: m.group(1) + str(value) + m.group(3), xml, count=1)

    def esc(s):
        return _html.escape(str(s or ""))

    # Thông tin chung - kỳ kê khai (quý hoặc tháng)
    if la_quy:
        xml = set_tag(xml, "kieuKy", "Q")   # Q = kê khai theo QUÝ -> HTKK tự hiển thị "Quý"
        xml = set_tag(xml, "kyKKhai", f"{quy_so}/{yyyy}")  # chỉ số quý, KHÔNG kèm 'Q' (tránh 'Quý Q2')
    else:
        xml = set_tag(xml, "kieuKy", "M")   # M = theo tháng
        xml = set_tag(xml, "kyKKhai", esc(ky))
    xml = set_tag(xml, "kyKKhaiTuNgay", tu_ngay_kkhai)
    xml = set_tag(xml, "kyKKhaiDenNgay", den_ngay_kkhai)
    xml = set_tag(xml, "ngayLapTKhai", datetime.date.today().isoformat())
    xml = set_tag(xml, "mst", esc(comp["mst"]))
    xml = set_tag(xml, "tenNNT", esc(comp["ten"]))
    if nguoi_ky:
        xml = set_tag(xml, "nguoiKy", esc(nguoi_ky))
        xml = set_tag(xml, "ngayKy", datetime.date.today().isoformat())

    # Các chỉ tiêu (giữ nguyên những thẻ khác như ct22, ct36... = giá trị template
    # hoặc tính lại). Ở đây ta điền số liệu kỳ này:
    # ct23/24 = TỔNG mua vào (gồm cả hàng nhập khẩu); ct23a/24a = TRONG ĐÓ hàng nhập khẩu.
    if imp and imp_nk_ds:
        # File import Excel: sheet 'BK Mua vào' đã bao gồm SẴN các dòng tờ khai
        # nhập khẩu (Ký hiệu 'TKNK'/Trạng thái 'Tờ khai nhập khẩu') trong TỔNG
        # mua vào -> [23]/[24] giữ nguyên tổng, [23a]/[24a] chỉ TÁCH riêng phần
        # nhập khẩu ra để khai đúng chỉ tiêu, TUYỆT ĐỐI không cộng thêm lần nữa.
        ct23 = round(mua_ds)
        ct24 = round(mua_thue)
        ct23a = round(imp_nk_ds)      # giá trị HHDV nhập khẩu (chưa thuế GTGT)
        ct24a = round(imp_nk_thue)    # thuế GTGT hàng nhập khẩu
    else:
        # Tra cứu / file import không có dòng NK: tờ khai NK (nếu có) được nhập
        # RIÊNG (bảng tokhai_nhap) và CHƯA nằm trong mua_ds -> cộng thêm vào tổng.
        ct23a = tk_ds_nk
        ct24a = tk_thue_nk
        ct23 = round(mua_ds) + tk_ds_nk
        ct24 = round(mua_thue) + tk_thue_nk
    ct25 = ct24
    ct30 = round(ds_5); ct31 = round(thue_5)
    ds_10_tk = ban_theo_ts["10"]["ds"] + ds_8
    thue_10_tk = ban_theo_ts["10"]["thue"] + thue_8
    ct32 = round(ds_10_tk); ct33 = round(thue_10_tk)
    ct27 = ct30 + ct32; ct28 = ct31 + ct33
    ct29 = round(ds_0)
    ct34 = ct27 + ct29; ct35 = ct28
    ct36 = ct35 - ct25          # thuế GTGT phát sinh trong kỳ
    # ct22 = thuế còn được khấu trừ kỳ trước chuyển sang (số dư đầu kỳ)
    ct22 = ct22_val
    # Theo mẫu 01/GTGT (TT80):
    # [40a] = ([36]-[22]+[37]-[38]-[39a]) nếu ≥0 -> thuế phải nộp của hoạt động SXKD trong kỳ
    # [40]  = [40a] - [40b]  (ở đây 40b=0 -> [40]=[40a])
    # [41]  = phần âm (thuế chưa khấu trừ hết, chuyển kỳ sau) -> [43]=[41]-[42]
    ct37 = 0; ct38 = 0; ct39a = 0; ct40b = 0; ct42 = 0
    con_lai = ct36 - ct22 + ct37 - ct38 - ct39a
    ct40a = max(con_lai, 0)      # thuế phải nộp của SXKD trong kỳ
    ct40 = max(ct40a - ct40b, 0) # thuế còn phải nộp trong kỳ
    ct41 = max(-con_lai, 0)      # thuế chưa khấu trừ hết kỳ này
    ct43 = max(ct41 - ct42, 0)   # còn được khấu trừ chuyển kỳ sau

    for tag, val in [
        ("ct21", 0), ("ct22", ct22),
        ("ct23", ct23), ("ct24", ct24),
        ("ct23a", ct23a), ("ct24a", ct24a),
        ("ct25", ct25), ("ct26", 0),
        ("ct27", ct27), ("ct28", ct28), ("ct29", ct29),
        ("ct30", ct30), ("ct31", ct31),
        ("ct32", ct32), ("ct33", ct33), ("ct32a", 0),
        ("ct34", ct34), ("ct35", ct35),
        ("ct36", ct36), ("ct37", ct37), ("ct38", ct38),
        ("ct39a", ct39a), ("ct40a", ct40a), ("ct40b", ct40b), ("ct40", ct40),
        ("ct41", ct41), ("ct42", ct42), ("ct43", ct43),
    ]:
        xml = set_tag(xml, tag, val)

    # Phụ lục NQ142: điền số liệu hàng 8%
    xml = set_tag(xml, "giaTriHHDV", round(ds_8))
    xml = set_tag(xml, "thueGTGTDuocGiam", thue_duoc_giam)
    xml = set_tag(xml, "tongCongGiaTriHHDV", round(ds_8))
    xml = set_tag(xml, "tongCongThueGTGTDuocGiam", thue_duoc_giam)
    xml = set_tag(xml, "ct9", thue_duoc_giam)
    # tên hàng hóa bán ra trong phụ lục (thẻ tenHHDV - chỉ thẻ trong HH_DV_BanRaTrongKy)
    xml = _re.sub(r"(<tenHHDV>)(.*?)(</tenHHDV>)",
                  r"\g<1>Hàng hóa, dịch vụ thuế suất 8%\g<3>", xml, count=1)

    # phần kỳ trong tên file: quý -> Q{n}{yyyy}, tháng -> M{mm}{yyyy}
    if la_quy:
        ky_fname = f"Q{quy_so}{yyyy}"
    else:
        ky_fname = f"M{mm}{yyyy}"
    # MST trong tên file HTKK dùng dạng 13 số (MST 10 số + '000' cho trụ sở chính)
    mst_file = str(comp["mst"] or "").strip()
    if len(mst_file) == 10:
        mst_file = mst_file + "000"
    fname = f"{mst_file}-01_GTGT_TT80-{ky_fname}-L00.xml"
    path = os.path.join(DOWNLOAD_DIR, fname)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\ufeff" + xml)
    # Mặc định lưu ra DESKTOP cho dễ tìm; lưu thêm vào thư mục công ty (nếu có)
    open_path = path
    desktop = _get_desktop_dir()
    if desktop and os.path.isdir(desktop):
        try:
            dest = os.path.join(desktop, fname)
            with open(dest, "w", encoding="utf-8") as f:
                f.write("\ufeff" + xml)
            open_path = dest
        except Exception:
            pass
    save_dir2 = (comp["save_dir"] or "").strip() if comp else ""
    if save_dir2 and os.path.isdir(save_dir2):
        try:
            dest2 = os.path.join(save_dir2, fname)
            with open(dest2, "w", encoding="utf-8") as f:
                f.write("\ufeff" + xml)
        except Exception:
            pass
    _open_file_local(open_path)
    return {"ok": True, "fname": fname, "path": open_path}


def _ky_ve_thang(ky):
    """Tách kỳ 'MM/YYYY' hoặc 'QX/YYYY' -> (mm:str 2 số, yyyy:str).
    Với kỳ quý, lấy THÁNG ĐẦU của quý làm mm (chỉ dùng để tính ngày mặc định
    của tờ khai; KHÔNG dùng để tra cứu/so khớp dữ liệu đã lưu theo kỳ)."""
    a, b = ky.split("/", 1)
    a = a.strip().upper()
    yyyy = int(b.strip())
    if a.startswith("Q") and a[1:].isdigit():
        q = max(1, min(4, int(a[1:])))
        mm = (q - 1) * 3 + 1
    else:
        mm = int(a)
    return f"{mm:02d}", str(yyyy)


def _ky_lien_truoc(ky):
    """Trả về CHUỖI kỳ liền trước 'ky', GIỮ NGUYÊN kiểu kỳ (tháng hay quý):
    'MM/YYYY' -> tháng trước; 'QX/YYYY' -> quý trước. '' nếu không hợp lệ."""
    ky = (ky or "").strip()
    if not ky or "/" not in ky:
        return ""
    a, b = ky.split("/", 1)
    a = a.strip().upper()
    try:
        y = int(b.strip())
    except Exception:
        return ""
    if a.startswith("Q") and a[1:].isdigit():
        q = int(a[1:])
        pq = 4 if q <= 1 else q - 1
        py = y - 1 if q <= 1 else y
        return f"Q{pq}/{py}"
    try:
        m = int(a)
    except Exception:
        return ""
    pm = 12 if m <= 1 else m - 1
    py = y - 1 if m <= 1 else y
    return f"{pm:02d}/{py}"


def _parse_thue_suat(ts_raw):
    """Trả về tỷ lệ thuế (0.08 cho 8%) từ chuỗi thuế suất.
    Hỗ trợ: '8%', '8', 'KHAC:5.26%', 'KHAC:5.26', '5.26%'.
    Trả None nếu không xác định được (KCT, KHÔNG, rỗng, hoặc không có số)."""
    s = str(ts_raw or "").strip().upper()
    if not s or s in ("KCT", "KHÔNG", "KO", "0%", "0"):
        return 0.0 if s in ("0%", "0") else None
    # lấy phần sau dấu ':' nếu có (KHAC:5.26%)
    if ":" in s:
        s = s.split(":", 1)[1]
    s = s.replace("%", "").replace(",", ".").strip()
    try:
        v = float(s)
        return v / 100
    except Exception:
        return None


def _thue_theo_cong_thue(it, items, r):
    """Khi không parse được thuế suất của 1 dòng hàng, lấy tiền thuế theo số
    tổng của cổng thuế (tgtthue) chia theo tỷ lệ thành tiền của dòng đó."""
    tong_thue_hd = _to_num(r["tgtthue"]) or 0
    if not tong_thue_hd:
        return 0
    tong_tt = sum((_to_num(x.get("thtien")) or 0) for x in items) or 1
    ds = _to_num(it.get("thtien")) or 0
    return round(tong_thue_hd * (ds / tong_tt))


@app.get("/api/export-excel/{cid}")
def export_excel(cid: int):
    import time as _tx
    _t0 = _tx.time()
    def _tlog(m):
        try: print(f"[xuat-excel +{_tx.time()-_t0:.1f}s] {m}", flush=True)
        except Exception: pass
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    rows = conn.execute(
        "SELECT * FROM invoices WHERE company_id=? ORDER BY loai, tdlap DESC",
        (cid,)).fetchall()

    # ===== NẠP TRƯỚC CHI TIẾT SONG SONG (tăng tốc xuất Excel) =====
    # Chỉ nạp hóa đơn CHƯA có detail_json VÀ chưa có file đã tải (ƯU TIÊN TUYỆT ĐỐI
    # file XML/ZIP đã tải sẵn trên máy — KHÔNG gọi mạng lại cho hóa đơn đã có file).
    import concurrent.futures as _cf
    client0 = CLIENTS.get(cid)
    save_dir0 = (comp["save_dir"] or "").strip() if comp else ""

    # XÂY INDEX FILE 1 LẦN — dùng chung cho cả bước kiểm tra "đã có file" lẫn
    # lúc dựng sheet sau này (trước đây quét thư mục 2 LẦN riêng biệt).
    _file_index = {}
    if save_dir0 and os.path.isdir(save_dir0):
        for rootdir, _d, files in os.walk(save_dir0):
            for fn in files:
                low = fn.lower()
                if not (low.endswith(".zip") or low.endswith(".xml")):
                    continue
                name = fn.rsplit(".", 1)[0]
                parts = name.split("_")
                if len(parts) >= 2:
                    f_khh = parts[0]
                    f_sho = parts[1].lstrip("0") or "0"
                    key = (f_khh, f_sho)
                    path = os.path.join(rootdir, fn)
                    prev = _file_index.get(key)
                    # ƯU TIÊN file .xml (đọc trực tiếp); chỉ giữ .zip khi chưa có .xml
                    if prev is None or (prev.lower().endswith(".zip") and low.endswith(".xml")):
                        _file_index[key] = path
    have_file = set(_file_index.keys())
    _tlog(f"tim thay {len(have_file)} file hoa don da tai tren may (thu muc: {save_dir0 or '(chưa đặt)'})")

    def find_invoice_file(r):
        khh = str(r["khhdon"] or "").strip()
        sho = str(r["shdon"] or "").strip().lstrip("0") or "0"
        return _file_index.get((khh, sho))

    def _doc_file_hoadon(fpath):
        """Đọc + parse 1 file XML/ZIP hóa đơn đã tải. Trả về (items, summary)."""
        try:
            with open(fpath, "rb") as f:
                data = f.read()
            if fpath.lower().endswith(".zip"):
                import zipfile as _zf, io as _io2
                try:
                    z = _zf.ZipFile(_io2.BytesIO(data))
                    xn = next((n for n in z.namelist()
                               if n.lower().endswith(".xml")), None)
                    if xn:
                        data = z.read(xn)
                except Exception:
                    pass
            return _parse_xml_invoice(data), _parse_invoice_summary(data)
        except Exception:
            return [], None

    can_nap = []
    da_co_file_rows = []
    for r in rows:
        if r["detail_json"]:
            continue
        khh = str(r["khhdon"] or ""); sho = str(r["shdon"] or "").lstrip("0") or "0"
        if (khh, sho) in have_file:
            da_co_file_rows.append(r)
            continue   # đã có file trên máy -> get_invoice_items sẽ đọc file, KHÔNG cần mạng
        can_nap.append(dict(r))
    _tlog(f"{len(da_co_file_rows)} hóa đơn dùng được file đã tải, {len(can_nap)} hóa đơn thiếu cả detail lẫn file")

    def _tai_1(rr, cho_khi_429=False, max_retry=1):
        if client0._token_dead:
            return rr["id"], None   # phiên đã hết -> khỏi thử, trả nhanh
        ht0 = rr["he_thong"] or "query"
        for ht in [ht0, ("sco-query" if ht0 == "query" else "query")]:
            try:
                d = client0.get_detail(rr["nbmst"], rr["khhdon"],
                                       rr["khmshdon"], rr["shdon"], ht,
                                       max_retry=max_retry, cho_khi_429=cho_khi_429)
                if d and (d.get("hdhhdvu") or d.get("nbmst")):
                    return rr["id"], json.dumps(d, ensure_ascii=False)
            except Exception:
                pass
        return rr["id"], None

    def _nap_song_song(ds_rows, nhan, cho_khi_429=False, workers=None):
        """Nạp chi tiết qua mạng SONG SONG cho 1 danh sách hóa đơn, lưu DB
        1 lần. Trả về số hóa đơn lấy được.
        cho_khi_429=True -> CHỜ theo Retry-After khi bị giới hạn tốc độ thay
        vì bỏ ngay (dùng cho lượt cuối, đảm bảo KHÔNG bỏ sót hóa đơn nào)."""
        if not (ds_rows and client0 and client0.token and not client0._token_dead):
            return 0
        results_map = {}
        _tlog(f"{nhan} {len(ds_rows)} hóa đơn qua mạng (song song)...")
        if workers is None:
            # số luồng song song: KHÔNG chờ (429 bỏ ngay) nên có thể chạy nhiều luồng hơn
            workers = {"fast": 20, "balanced": 12, "safe": 6}.get(CURRENT_SPEED, 12)
        max_retry = 2 if cho_khi_429 else 1

        def _goi(rr):
            return _tai_1(rr, cho_khi_429=cho_khi_429, max_retry=max_retry)

        with _cf.ThreadPoolExecutor(max_workers=workers) as ex:
            for inv_id, dj in ex.map(_goi, ds_rows):
                if dj:
                    results_map[inv_id] = dj
        cn = db()
        for inv_id, dj in results_map.items():
            cn.execute("UPDATE invoices SET detail_json=? WHERE id=?", (dj, inv_id))
        cn.commit(); cn.close()
        return len(results_map)

    if can_nap:
        if client0 and client0.token and not client0._token_dead:
            so_lay_duoc = _nap_song_song(can_nap, "thử nạp nhanh (không chờ khi bị giới hạn tốc độ)")
            con_thieu = len(can_nap) - so_lay_duoc
            if client0._token_dead:
                _tlog(f"phiên đăng nhập đã hết hạn giữa chừng — dừng nạp qua mạng, còn {con_thieu} hóa đơn chưa có chi tiết")
            elif con_thieu:
                _tlog(f"lấy được {so_lay_duoc}/{len(can_nap)}; còn {con_thieu} hóa đơn bị giới hạn tốc độ — "
                     f"chạy lại 'Kết xuất Excel' sau ít phút để lấy nốt (đã lưu tiến độ)")
        else:
            why = ("chưa đăng nhập/phiên đã hết" if not (client0 and client0.token) else "phiên đã hết hạn")
            _tlog(f"bỏ qua nạp mạng cho {len(can_nap)} hóa đơn ({why}) — sẽ hiện placeholder, "
                 f"đăng nhập lại rồi xuất lại để lấy nốt")

    # ===== KIỂM TRA NHANH file "đã có" có THỰC SỰ ĐỌC ĐƯỢC không (offline,
    # không cần mạng) — hóa đơn nào file lỗi/rỗng (hỏng, sai định dạng...)
    # thì gộp NẠP LẠI QUA MẠNG SONG SONG luôn ở đây. TRƯỚC ĐÂY việc này để
    # tới lúc dựng sheet mới phát hiện rồi gọi mạng TUẦN TỰ TỪNG CÁI 1 — với
    # hàng trăm hóa đơn, đây chính là nguyên nhân "đã nạp xong chi tiết" mà
    # vẫn treo rất lâu ở bước dựng sheet.
    _items_cache = {}
    can_nap2 = []
    if da_co_file_rows:
        for r in da_co_file_rows:
            fpath = find_invoice_file(r)
            items, summary = (_doc_file_hoadon(fpath) if fpath else ([], None))
            if items:
                key = (r["khhdon"], r["shdon"], r["nbmst"], r["loai"])
                _items_cache[key] = (items, summary)
            else:
                can_nap2.append(dict(r))
        if can_nap2:
            so_lay_duoc2 = _nap_song_song(can_nap2, "file đã tải nhưng đọc lỗi/rỗng — nạp lại")
            if so_lay_duoc2 < len(can_nap2):
                _tlog(f"⚠ {len(can_nap2) - so_lay_duoc2}/{len(can_nap2)} hóa đơn có file trên máy "
                     f"nhưng ĐỌC LỖI và nạp lại qua mạng cũng không được")

    # ===== LƯỢT CUỐI: ĐẢM BẢO LẤY ĐỦ, không để hóa đơn nào hiện placeholder
    # "chưa lấy được chi tiết" chỉ vì bị giới hạn tốc độ (429) ở 2 lượt nhanh
    # phía trên. 2 lượt trên bỏ ngay khi gặp 429 (để không làm chậm cả quá
    # trình) — số hóa đơn CÒN THIẾU lúc này thường đã nhỏ, nên lượt cuối này
    # CHỜ đúng theo Retry-After của trang Thuế (chấp nhận chậm hơn) với ít
    # luồng song song hơn, để lấy cho BẰNG ĐƯỢC thay vì bỏ cuộc.
    ds_can_kiem_tra = can_nap + can_nap2
    if ds_can_kiem_tra and client0 and client0.token and not client0._token_dead:
        ids_can_kiem_tra = [r["id"] for r in ds_can_kiem_tra]
        ph = ",".join("?" * len(ids_can_kiem_tra))
        con_thieu_rows = conn.execute(
            f"SELECT * FROM invoices WHERE id IN ({ph}) AND (detail_json IS NULL OR detail_json='')",
            ids_can_kiem_tra).fetchall()
        if con_thieu_rows:
            _tlog(f"còn {len(con_thieu_rows)} hóa đơn chưa lấy được — thử LẦN CUỐI "
                 f"(chờ đúng thời gian giới hạn tốc độ của trang Thuế, để không bỏ sót)...")
            so_lay_duoc3 = _nap_song_song(con_thieu_rows, "thử lại lần cuối (chờ khi bị giới hạn tốc độ)",
                                          cho_khi_429=True, workers=4)
            if so_lay_duoc3 < len(con_thieu_rows):
                _tlog(f"⚠ vẫn còn {len(con_thieu_rows) - so_lay_duoc3} hóa đơn KHÔNG lấy được chi tiết "
                     f"dù đã chờ thử lại — có thể phiên đăng nhập hết hạn giữa chừng, đăng nhập lại rồi "
                     f"xuất lại để lấy nốt")
            else:
                _tlog("đã lấy đủ chi tiết cho toàn bộ hóa đơn còn thiếu ở lượt cuối")

    # đọc lại rows để có detail_json mới nhất (từ các lượt nạp mạng ở trên)
    rows = conn.execute(
        "SELECT * FROM invoices WHERE company_id=? ORDER BY loai, tdlap DESC",
        (cid,)).fetchall()
    conn.close()
    _tlog(f"xong nap chi tiet ({len(rows)} hoa don) -> bat dau dung sheet")

    wb = openpyxl.Workbook()
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial")
    hdr_fill = PatternFill("solid", fgColor="2E5C8A")

    def style_header(ws, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(1, c)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def autofit(ws):
        # CHỈ lấy mẫu tiêu đề + ~60 dòng đầu để tính độ rộng cột.
        # (Quét toàn bộ bằng ws.columns rất chậm với bảng vài nghìn dòng.)
        widths = {}
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
            for c in row:
                if c.value is not None:
                    l = len(str(c.value))
                    if l > widths.get(c.column, 0):
                        widths[c.column] = l
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = min(w + 2, 50)

    # Định dạng SỐ cho các cột tiền (có dấu phân cách hàng nghìn)
    NUM_FMT = "#,##0"
    TIEN_COLS = {"thành tiền", "tiền thuế gtgt", "thuế gtgt", "tổng thanh toán",
                 "doanh số mua chưa thuế", "doanh số bán chưa thuế", "đơn giá",
                 "số lượng", "thành tiền chưa vat", "tổng tiền", "tổng cộng",
                 "chưa vat (chi tiết)", "chưa vat (bảng kê)", "lệch chưa vat",
                 "vat (chi tiết)", "vat (bảng kê)", "lệch vat", "thuế", "thành tiền chưa thuế",
                 "trị giá tính thuế nk", "tiền thuế nk"}

    def format_so(ws, header_row=None):
        """Áp định dạng số cho các cột tiền. Tự tìm (các) dòng header nếu cần."""
        # tìm tất cả dòng có chứa tên cột tiền (BK Bán ra có nhiều cụm header)
        header_rows = []
        if header_row:
            header_rows = [header_row]
        else:
            for r in range(1, min(ws.max_row, 200) + 1):
                for c in range(1, ws.max_column + 1):
                    if str(ws.cell(r, c).value or "").strip().lower() in TIEN_COLS:
                        header_rows.append(r); break
        if not header_rows:
            return
        num_cols = set()
        for hr in header_rows:
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(hr, c).value or "").strip().lower() in TIEN_COLS:
                    num_cols.add(c)
        for r in range(1, ws.max_row + 1):
            if r in header_rows:
                continue
            for c in num_cols:
                cell = ws.cell(r, c)
                v = cell.value
                if isinstance(v, (int, float)):
                    cell.number_format = NUM_FMT
                elif isinstance(v, str) and v.strip() and v.replace(".", "").replace("-", "").isdigit():
                    try:
                        cell.value = float(v) if "." in v else int(v)
                        cell.number_format = NUM_FMT
                    except Exception:
                        pass

    # ===== Chuẩn bị: hàm tìm/đọc chi tiết hóa đơn =====
    # (_file_index / find_invoice_file / _items_cache đã được xây + nạp trước
    # ở bước "NẠP TRƯỚC CHI TIẾT SONG SONG" phía trên — dùng lại ở đây,
    # KHÔNG quét lại thư mục hay xóa cache đã có.)
    client = CLIENTS.get(cid)  # client đang đăng nhập (nếu có) để tải lại khi thiếu file

    def get_invoice_items(r):
        """Lấy danh sách mặt hàng của 1 hóa đơn:
        1) từ chi tiết đã lưu trong DB -> 2) từ file XML đã tải ->
        3) gọi endpoint detail (thử cả query và sco-query).
        Lấy được thì LƯU VÀO DB để lần sau khỏi gọi lại.
        Trả về (items, summary)."""
        key = (r["khhdon"], r["shdon"], r["nbmst"], r["loai"])
        if key in _items_cache:
            return _items_cache[key]
        items = []
        summary = None

        # (0) chi tiết đã lưu trong DB (cột detail_json)
        try:
            dj = r["detail_json"] if "detail_json" in r.keys() else None
        except Exception:
            dj = None
        if dj:
            try:
                detail = json.loads(dj)
                items = _parse_detail_json(detail)
                summary = _summary_from_detail_json(detail)
            except Exception:
                items = []; summary = None

        # (1) file đã tải (XML, hoặc giải nén XML từ trong ZIP)
        if not items:
            fpath = find_invoice_file(r)
            if fpath:
                items, summary = _doc_file_hoadon(fpath)

        # (2) gọi detail JSON — thử cả hệ thống đã lưu và hệ thống còn lại
        # (đã nạp song song trước -> ở đây chỉ vớt nhanh, KHÔNG chờ nếu bị giới hạn
        # tốc độ hay phiên đã hết, để không làm chậm cả quá trình dựng sheet)
        if not items and client and client.token and not client._token_dead:
            ht0 = r["he_thong"] or "query"
            for ht in [ht0, ("sco-query" if ht0 == "query" else "query")]:
                try:
                    detail = client.get_detail(
                        r["nbmst"], r["khhdon"], r["khmshdon"], r["shdon"], ht,
                        max_retry=1, cho_khi_429=False)
                    if detail and (detail.get("hdhhdvu") or detail.get("nbmst")):
                        items = _parse_detail_json(detail)
                        summary = _summary_from_detail_json(detail)
                        # LƯU vào DB để lần sau khỏi gọi lại
                        try:
                            cn = db()
                            cn.execute("UPDATE invoices SET detail_json=? WHERE id=?",
                                       (json.dumps(detail, ensure_ascii=False), r["id"]))
                            cn.commit(); cn.close()
                        except Exception:
                            pass
                        break
                    time.sleep(SP()["file"])
                except Exception:
                    pass
        _items_cache[key] = (items, summary)
        return items, summary

    detail_headers = ["Ký hiệu", "Số HĐ", "Ngày", "Người bán", "MST bán",
                      "Người mua", "MST mua", "STT", "Mã vt",
                      "Tên hàng hóa/dịch vụ",
                      "ĐVT", "Số lượng", "Đơn giá", "Thành tiền",
                      "Thuế suất", "Tiền thuế GTGT",
                      "Trạng thái", "Kết quả"]

    # Lưu tổng theo từng hóa đơn (để đối chiếu việc 4)
    ct_totals = {"purchase": {}, "sold": {}}

    def phan_bo_chiet_khau(items):
        """Xử lý dòng hàng hóa đơn MUA VÀO — gồm CÁC LOẠI CHIẾT KHẤU THƯƠNG MẠI:

        LOẠI 1 — Chiết khấu trên TỪNG DÒNG (STCKhau>0 ngay trên dòng hàng):
          TRỪ chiết khấu vào THÀNH TIỀN của chính dòng đó
          (net = thành tiền gộp - chiết khấu), giữ DƯƠNG, KHÔNG tách dòng âm.
          (vd HĐ 116058 - H&M: dòng TChat=3 vừa là hàng vừa có STCKhau)

        LOẠI 2 — Dòng CHIẾT KHẤU THƯƠNG MẠI RIÊNG (TChat=3, KHÔNG có STCKhau,
          thành tiền là số tiền chiết khấu): PHÂN BỔ trừ đều vào thành tiền các
          dòng hàng cùng thuế suất, tính lại tiền thuế; KHÔNG hiện dòng riêng.
          (vd HĐ 8 - chiết khấu 10%)

        LOẠI 3 — Dòng thành tiền ÂM sẵn (điều chỉnh giảm/CK ghi âm): giữ ÂM.

        LOẠI 4 — Giảm 20% tỷ lệ % của HKD theo NQ204 (TChat=4, thtien=0,
          'Đã giảm X ... 20% ... tỷ lệ %'): phân bổ X đều vào THÀNH TIỀN các
          dòng hàng theo tỷ lệ. (vd HĐ 91)

        Ghi chú thuần khác (TChat=4, thtien=0): BỎ QUA."""

        import re as _re_hkd

        def _vn_money(s):
            """Parse số tiền dạng chữ VN trong ghi chú:
            '7.200,00'->7200, '2.970'->2970, '1.192.800'->1192800.
            (chấm = phân cách nghìn, phẩy = thập phân)"""
            m = _re_hkd.search(r'\d[\d.,]*', str(s or ""))
            if not m:
                return 0
            t = m.group(0)
            if ',' in t:                       # có phẩy -> phẩy là thập phân
                t = t.replace('.', '').replace(',', '.')
            else:                              # chỉ có chấm -> phân cách nghìn
                t = t.replace('.', '')
            try:
                return round(float(t))
            except Exception:
                return 0

        def _net_sau_ck(it):
            """Thành tiền sau khi trừ chiết khấu dòng (STCKhau). Tự nhận biết
            thành tiền XML là gộp hay đã net để không trừ 2 lần."""
            tt = _to_num(it.get("thtien")) or 0
            ck = _to_num(it.get("stckhau")) or 0
            if not isinstance(tt, (int, float)):
                return tt
            if not isinstance(ck, (int, float)) or ck <= 0:
                return tt
            sl = _to_num(it.get("sluong")) or 0
            dg = _to_num(it.get("dgia")) or 0
            gross = sl * dg if (isinstance(sl, (int, float)) and isinstance(dg, (int, float))
                                and sl and dg) else None
            if gross is not None:
                if abs(tt - gross) <= 1:               # thtien = gộp -> trừ CK
                    return round(tt - ck)
                if abs(tt - (gross - ck)) <= 1:         # thtien đã = net -> giữ
                    return round(tt)
            return round(tt - ck)                       # mặc định: coi là gộp

        # --- Pass 1: phát hiện ghi chú HKD NQ204 (TChat=4, thtien=0) ---
        so_giam_hkd = 0
        for it in items:
            tchat = str(it.get("tchat", "") or "")
            tt = _to_num(it.get("thtien")) or 0
            if tchat == "4" and (not tt or tt == 0):
                ten_l = str(it.get("ten_hang", "") or "").lower()
                if (("giảm" in ten_l or "giam" in ten_l) and "20" in ten_l
                        and ("tỷ lệ" in ten_l or "ty le" in ten_l
                             or "204" in ten_l or "gtgt" in ten_l)):
                    so_giam_hkd = _vn_money(it.get("ten_hang"))
                    break

        def _norm_rate(ts):
            """Chuẩn hóa thuế suất để gom nhóm phân bổ chiết khấu."""
            return str(ts or "").strip().upper()

        # --- Pass 1b: gom dòng CHIẾT KHẤU THƯƠNG MẠI RIÊNG (LOẠI 2) theo thuế suất ---
        # dòng TChat=3, KHÔNG có STCKhau, thành tiền là số tiền chiết khấu
        # -> sẽ PHÂN BỔ TRỪ vào thành tiền các dòng hàng (KHÔNG hiện dòng riêng).
        ck_rieng = {}          # thuế suất -> tổng tiền chiết khấu
        ck_rieng_items = {}    # thuế suất -> list các dòng CK gốc (dùng khi HĐ chỉ có dòng CK, không có dòng hàng để phân bổ)
        skip_ck_rieng = set()  # id() các dòng CK riêng để bỏ qua khi dựng
        for it in items:
            tchat = str(it.get("tchat", "") or "")
            tt = _to_num(it.get("thtien")) or 0
            ck = _to_num(it.get("stckhau")) or 0
            if tchat == "4" and (not tt or tt == 0):
                continue
            if not isinstance(tt, (int, float)) or tt <= 0:
                continue
            if tchat == "3" and not (isinstance(ck, (int, float)) and ck > 0):
                rate = _norm_rate(it.get("tsuat"))
                ck_rieng[rate] = ck_rieng.get(rate, 0) + abs(tt)
                ck_rieng_items.setdefault(rate, []).append(it)
                skip_ck_rieng.add(id(it))

        # --- Pass 2: dựng các dòng hàng (đã trừ CK dòng), bỏ ghi chú + CK riêng ---
        out = []
        for it in items:
            tchat = str(it.get("tchat", "") or "")
            tt = _to_num(it.get("thtien")) or 0
            # ghi chú thuần không có thành tiền -> bỏ (kể cả NQ204 đã xử lý ở Pass 1)
            if tchat == "4" and (not tt or tt == 0):
                continue
            # dòng chiết khấu thương mại riêng -> đã gom ở Pass 1b, KHÔNG hiện
            if id(it) in skip_ck_rieng:
                continue
            ck = _to_num(it.get("stckhau")) or 0
            h = dict(it)
            if isinstance(tt, (int, float)) and tt < 0:
                # LOẠI 3: dòng điều chỉnh giảm sẵn ÂM -> giữ ÂM cả thành tiền + thuế
                h["thtien"] = -abs(tt)
                dg = _to_num(h.get("dgia")) or 0
                if isinstance(dg, (int, float)):
                    h["dgia"] = -abs(dg)
                thue_goc = _to_num(h.get("tien_thue"))
                if (thue_goc is not None and str(h.get("tien_thue")).strip() != ""
                        and _to_num(thue_goc) != 0):
                    h["tien_thue"] = -abs(_to_num(thue_goc))
                ten = str(h.get("ten_hang", "") or "")
                if "204" in ten or "nq" in ten.lower():
                    h["_la_nq204"] = True
                else:
                    h["_la_ck"] = True
            else:
                # LOẠI 1: dòng hàng dương -> trừ chiết khấu dòng (nếu có) vào thành tiền
                if isinstance(ck, (int, float)) and ck > 0:
                    h["thtien"] = _net_sau_ck(it)
                    h["_co_ck_dong"] = True
            out.append(h)

        # --- Pass 3: phân bổ giảm HKD NQ204 vào thành tiền các dòng dương ---
        if so_giam_hkd > 0:
            idxs = [i for i, h in enumerate(out)
                    if isinstance(_to_num(h.get("thtien")), (int, float))
                    and (_to_num(h.get("thtien")) or 0) > 0]
            tong = sum((_to_num(out[i].get("thtien")) or 0) for i in idxs)
            if tong > 0:
                allocated = 0
                for k, i in enumerate(idxs):
                    tt_i = _to_num(out[i].get("thtien")) or 0
                    if k == len(idxs) - 1:
                        giam = so_giam_hkd - allocated
                    else:
                        giam = round(tt_i / tong * so_giam_hkd)
                        allocated += giam
                    out[i]["thtien"] = round(tt_i - giam)

        # --- Pass 4: phân bổ CHIẾT KHẤU THƯƠNG MẠI RIÊNG vào thành tiền (LOẠI 2) ---
        # Trừ đều vào các dòng hàng DƯƠNG cùng thuế suất (nếu không có thì mọi dòng
        # dương). Tính lại tiền thuế GTGT của dòng được phân bổ = thành tiền * thuế suất.
        for rate_key, tong_ck in ck_rieng.items():
            if not tong_ck or tong_ck <= 0:
                continue
            idxs = [i for i, h in enumerate(out)
                    if (_to_num(h.get("thtien")) or 0) > 0
                    and _norm_rate(h.get("tsuat")) == rate_key]
            if not idxs:
                idxs = [i for i, h in enumerate(out)
                        if (_to_num(h.get("thtien")) or 0) > 0]
            tong_tt = sum((_to_num(out[i].get("thtien")) or 0) for i in idxs)
            # HĐ CHỈ CÓ (các) dòng chiết khấu, KHÔNG có dòng hàng nào để phân bổ vào
            # (vd hóa đơn điều chỉnh giảm doanh số/chiết khấu riêng, 1 dòng duy nhất)
            # -> ghi nhận TRỰC TIẾP (các) dòng chiết khấu đó dưới dạng ÂM, không bỏ mất.
            if not idxs or tong_tt <= 0:
                rate_num = _parse_thue_suat(rate_key)
                for it_ck in ck_rieng_items.get(rate_key, []):
                    h = dict(it_ck)
                    tt_goc = abs(_to_num(h.get("thtien")) or 0)
                    h["thtien"] = -tt_goc
                    dg = _to_num(h.get("dgia")) or 0
                    if isinstance(dg, (int, float)) and dg:
                        h["dgia"] = -abs(dg)
                    if rate_num is not None and rate_num > 0:
                        h["tien_thue"] = round(-tt_goc * rate_num)
                    else:
                        h["tien_thue"] = 0
                    h["_la_ck"] = True
                    out.append(h)
                continue
            allocated = 0
            for k, i in enumerate(idxs):
                tt_i = _to_num(out[i].get("thtien")) or 0
                if k == len(idxs) - 1:
                    giam = tong_ck - allocated
                else:
                    giam = round(tt_i / tong_tt * tong_ck)
                    allocated += giam
                net_i = round(tt_i - giam)
                out[i]["thtien"] = net_i
                # tính lại tiền thuế theo thành tiền mới (dòng CK riêng đã bị bỏ)
                rate = _parse_thue_suat(str(out[i].get("tsuat", "") or ""))
                if rate is not None and rate > 0:
                    out[i]["tien_thue"] = round(net_i * rate)
                elif rate == 0.0:
                    out[i]["tien_thue"] = 0
        return out

    def _fmt_ngay(s):
        """yyyy-mm-dd hoặc yyyy-mm-ddThh -> dd/mm/yyyy."""
        s = (s or "").split("T")[0]
        if "-" in s:
            p = s.split("-")
            if len(p) == 3:
                return f"{p[2]}/{p[1]}/{p[0]}"
        if "/" in s:  # đã đúng định dạng
            return s
        return s

    def _sort_key_ngay(r):
        nd = (r["tdlap"] or "").split("T")[0]
        return nd  # yyyy-mm-dd so sánh chuỗi = đúng thứ tự thời gian

    map_no_ht = _get_map_no(cid)  # {mst: tk_no} đã học -> tự điền cột Nợ

    def build_detail_sheet(sheet_name, loai):
        ws = wb.create_sheet(sheet_name)
        # Việc 9: Chi tiết MUA VÀO bỏ Người mua/MST mua; BÁN RA bỏ Người bán/MST bán
        if loai == "purchase":
            headers = ["Ký hiệu", "Số HĐ", "Ngày", "Người bán", "MST bán",
                       "STT", "Mã vt", "Tên hàng hóa/dịch vụ", "ĐVT",
                       "Số lượng", "Đơn giá", "Thành tiền",
                       "Thuế suất", "Tiền thuế GTGT", "Trạng thái", "Kết quả",
                       "Trị giá tính thuế NK", "Thuế suất NK", "Tiền thuế NK",
                       "Nợ", "Có"]   # cột T = Nợ, U = Có
        else:
            headers = ["Ký hiệu", "Số HĐ", "Ngày", "Người mua", "MST mua",
                       "STT", "Mã vt", "Tên hàng hóa/dịch vụ", "ĐVT",
                       "Số lượng", "Đơn giá", "Thành tiền",
                       "Thuế suất", "Tiền thuế GTGT", "Trạng thái", "Kết quả"]
        ws.append(headers)
        style_header(ws, len(headers))

        def append_row(vals, no_tk="", co_tk=""):
            """Append 1 dòng; với MUA VÀO thêm cột Nợ (T) + Có (U)."""
            vals = list(vals)
            if loai == "purchase":
                while len(vals) < 19:   # đệm cho hết cột S (thuế NK)
                    vals.append("")
                vals.append(no_tk)      # T = Nợ
                vals.append(co_tk)      # U = Có
            ws.append(vals)

        # lọc + SORT theo ngày TĂNG DẦN (thấp -> cao)
        loai_rows = [r for r in rows if r["loai"] == loai]

        # AN TOÀN: với hóa đơn MUA VÀO, loại bỏ hóa đơn mà MST người mua (nmmst)
        # KHÁC MST công ty đang xem (đó là hóa đơn lẫn của công ty khác, vd hóa đơn
        # xuất khẩu của bên bán bị cổng thuế trả nhầm). Giữ lại nếu nmmst trống/khớp
        # — so khớp theo MST GỐC 10 số (bỏ hậu tố chi nhánh '-001'...), vì hóa đơn
        # có thể ghi MST đơn vị trực thuộc (13 số) trong khi công ty đăng ký MST
        # gốc (10 số) hoặc ngược lại — TRƯỚC ĐÂY so khớp CHUỖI Y HỆT nên loại nhầm
        # hóa đơn hợp lệ của chính công ty, khiến VAT/doanh số mua vào ở Excel
        # thấp hơn số liệu tra cứu (KPI trang chính/Tạm tính VAT không lọc MST).
        if loai == "purchase":
            mst_cty_goc = _chuan_mst(comp["mst"])[:10]
            def _hop_le_mua(r):
                nm = str(r["nmmst"] or "").strip()
                # nmmst trống -> giữ (một số HĐ không ghi MST mua)
                if not nm:
                    return True
                return _chuan_mst(nm)[:10] == mst_cty_goc
            loai_rows = [r for r in loai_rows if _hop_le_mua(r)]
        # BÁN RA: KHÔNG lọc theo MST người bán nữa — trang Thuế tra cứu "bán ra"
        # đã CHỈ trả về đúng hóa đơn của công ty đang đăng nhập (không có rủi ro
        # lẫn hóa đơn công ty khác như bên mua vào), nên lọc thêm ở đây chỉ có
        # hại: hộ/cá nhân kinh doanh (HKD) đôi khi có MST trên hóa đơn (nbmst,
        # vd mã số thuế cá nhân) KHÁC với MST đăng ký trong phần mềm (vd mã số
        # thuế sàn TMĐT dùng để đăng nhập) — trước đây bị lọc SẠCH TOÀN BỘ hóa
        # đơn bán ra dù đã tra cứu/tải được, nhìn như phần mềm "không ghi nhận".

        loai_rows.sort(key=_sort_key_ngay)

        ngay_col = 3  # cột Ngày để format dd/mm/yyyy
        for r in loai_rows:
            try:
                raw = json.loads(r["raw"]) if r["raw"] else {}
            except Exception:
                raw = {}
            tthai_ma = str(raw.get("tthai", r["tthai"]) or "").strip()
            tt = _mo_ta_trang_thai(raw.get("tthai", r["tthai"]))
            kq = _mo_ta_ket_qua(raw.get("ttxly", ""))
            ttxly_raw = str(raw.get("ttxly", "") or "").strip()

            # Loại hóa đơn 'đã bị thay thế' (mã 4), hủy/xóa bỏ (6),
            # và 'không đủ điều kiện cấp mã' (kết quả kiểm tra ttxly=4)
            if (tthai_ma in ("4", "6")
                    or tt in ("Hóa đơn đã bị thay thế", "Hóa đơn hủy",
                              "Hóa đơn đã bị xóa bỏ", "Hóa đơn xóa bỏ")):
                continue
            if ttxly_raw == "4" or "không đủ điều kiện" in kq.lower():
                continue

            items, _summary = get_invoice_items(r)
            ikey = (str(r["khhdon"]), str(r["shdon"]).lstrip("0") or "0")
            ngay_fmt = _fmt_ngay(r["tdlap"])

            # Hạch toán MUA VÀO: Có theo tổng HĐ (>=5tr->331, <5tr->1111); Nợ học theo MST
            no_r = co_r = ""
            if loai == "purchase":
                co_r = _co_theo_tong(r["tgtttbso"])
                no_r = map_no_ht.get(_chuan_mst(r["nbmst"]), "")

            if not items:
                dang_nhap_ok = bool(client and client.token and not getattr(client, "_token_dead", False))
                tong_tien_hd = _to_num(r["tgtcthue"]) or _to_num(r["tgtttbso"]) or 0
                # Hóa đơn KHÔNG lấy được chi tiết dòng hàng (không phải do lỗi
                # mạng/đăng nhập — ĐÃ đăng nhập thành công) NHƯNG có số tiền
                # thật từ danh sách tra cứu -> thường là "Hóa đơn bán hàng" của
                # hộ/cá nhân kinh doanh (không tách VAT, TCT không trả chi tiết
                # dòng hàng cho loại này). Ghi nhận CẢ HÓA ĐƠN thành 1 dòng
                # KHÔNG CHỊU THUẾ (KCT) thay vì để trống/báo "chưa lấy được chi
                # tiết" — báo lỗi đó SAI vì không phải do mạng, thử lại cũng
                # không bao giờ có chi tiết dòng hàng cho loại hóa đơn này.
                if loai == "sold" and dang_nhap_ok and tong_tien_hd:
                    nmten_raw = raw.get("nmten", "") or raw.get("nmtnmua", "") or ""
                    append_row([r["khhdon"], r["shdon"], ngay_fmt,
                                nmten_raw, r["nmmst"], 1, "",
                                "(Cả hóa đơn — không tách dòng hàng)", "",
                                "", "", tong_tien_hd, "KCT", 0, tt, kq])
                    cur = ct_totals[loai].setdefault(ikey, {"ds": 0, "thue": 0})
                    cur["ds"] += tong_tien_hd
                    continue
                if client and getattr(client, "_token_dead", False):
                    ly_do = "(chưa lấy được chi tiết — phiên đăng nhập đã hết, đăng nhập lại rồi xuất lại)"
                elif not (client and client.token):
                    ly_do = "(chưa có file XML đã tải — đăng nhập rồi xuất lại để lấy qua mạng)"
                else:
                    ly_do = "(chưa lấy được chi tiết — có thể do giới hạn tốc độ, xuất lại sau ít phút)"
                nmten_raw = raw.get("nmten", "") or raw.get("nmtnmua", "") or ""
                if loai == "purchase":
                    append_row([r["khhdon"], r["shdon"], ngay_fmt,
                                r["nbten"], r["nbmst"], "", "", ly_do, "",
                                "", "", "", "", r["tgtttbso"],
                                tt, kq], no_r, co_r)
                else:
                    append_row([r["khhdon"], r["shdon"], ngay_fmt,
                                nmten_raw, r["nmmst"], "", "", ly_do, "",
                                "", "", "", "", r["tgtttbso"],
                                tt, kq])
                cur = ct_totals[loai].setdefault(ikey, {"ds": 0, "thue": 0})
                cur["ds"] += _to_num(r["tgtcthue"]) or 0
                cur["thue"] += _to_num(r["tgtthue"]) or 0
                continue

            if loai == "purchase":
                items = phan_bo_chiet_khau(items)

            _TS_KHONG_THUE = ("", "KCT", "KKKNT", "KHTKKNT", "KO", "KHÔNG",
                              "KHONG", "0%", "0")
            # --- Pass 1: tính từng dòng (đánh dấu dòng thuế suất KHAC cần phân bổ) ---
            dong_list = []
            for it in items:
                ds = _to_num(it.get("thtien")) or 0
                ts_raw = str(it.get("tsuat", "") or "").strip()
                thue_goc_raw = it.get("tien_thue")
                ts_upper = ts_raw.upper().replace(" ", "")
                phan_bo = False
                # Thuế suất hiển thị: ô trống / KCT / KKKNT -> "0%" — RIÊNG dòng
                # "Tổng tiền phí" (_la_phi) hiện đúng chữ "KCT" theo yêu cầu,
                # không quy về "0%" như các dòng hàng thông thường.
                if it.get("_la_phi"):
                    ts_hien = "KCT"
                    tien_thue = 0
                elif ts_upper in ("", "KCT", "KKKNT", "KHTKKNT", "KO", "KHÔNG", "KHONG"):
                    ts_hien = "0%"
                    tien_thue = 0
                # Có tiền thuế GỐC trên hóa đơn → dùng (kể cả = 0)
                elif thue_goc_raw is not None and str(thue_goc_raw).strip() != "":
                    ts_hien = ts_raw
                    tien_thue = round(_to_num(thue_goc_raw))
                # Biết thuế suất → tính
                else:
                    ts_hien = ts_raw
                    rate = _parse_thue_suat(ts_raw)
                    if rate is not None and rate > 0:
                        tien_thue = round(ds * rate) if isinstance(ds, (int, float)) else 0
                    else:
                        # Thuế suất KHAC/lẻ không tính được -> sẽ PHÂN BỔ theo tổng
                        # tiền thuế của hóa đơn (giữ nguyên hiển thị, vd "KHAC")
                        tien_thue = 0
                        if ts_upper not in _TS_KHONG_THUE:
                            phan_bo = True
                ten = it.get("ten_hang", "")
                if it.get("_la_nq204"):
                    ten += " (Giảm NQ204 - ghi âm)"
                elif it.get("_la_ck"):
                    ten += " (Chiết khấu TM - ghi âm)"
                # Đơn giá = Thành tiền / Số lượng (luôn đồng bộ sau khi trừ chiết khấu)
                sl_num = _to_num(it.get("sluong"))
                if isinstance(sl_num, (int, float)) and sl_num and isinstance(ds, (int, float)):
                    dg_val = ds / sl_num
                    dgia_out = int(dg_val) if dg_val == int(dg_val) else round(dg_val, 2)
                else:
                    dgia_out = _to_num(it.get("dgia"))
                # cột người: purchase -> người bán; sold -> người mua
                if loai == "purchase":
                    nguoi = it.get("ten_nban", "") or r["nbten"]
                    mst = it.get("mst_nban", "") or r["nbmst"]
                else:
                    nguoi = it.get("ten_nmua", "") or raw.get("nmten", "") or ""
                    mst = it.get("mst_nmua", "") or r["nmmst"]
                dong_list.append({
                    "ds": ds, "ts_hien": ts_hien, "tien_thue": tien_thue,
                    "phan_bo": phan_bo, "ten": ten, "dgia": dgia_out,
                    "sl": sl_num, "nguoi": nguoi, "mst": mst, "it": it})

            # --- Phân bổ tiền thuế cho dòng thuế suất KHAC để TỔNG khớp hóa đơn ---
            idx_pb = [i for i, d in enumerate(dong_list) if d["phan_bo"]]
            if idx_pb:
                tong_thue_hd = round(_to_num(r["tgtthue"]) or 0)
                da_gan = sum(d["tien_thue"] for d in dong_list if not d["phan_bo"])
                con_lai = tong_thue_hd - da_gan
                tong_ds = sum((dong_list[i]["ds"] or 0) for i in idx_pb)
                if con_lai > 0 and tong_ds:
                    allocated = 0
                    for k, i in enumerate(idx_pb):
                        if k == len(idx_pb) - 1:
                            dong_list[i]["tien_thue"] = con_lai - allocated
                        else:
                            t = round(dong_list[i]["ds"] / tong_ds * con_lai)
                            dong_list[i]["tien_thue"] = t
                            allocated += t

            # --- Pass 2: ghi ra sheet ---
            for d in dong_list:
                it = d["it"]
                ds = d["ds"]
                dvt_out, sl_out, dgia_out = it.get("dvt", ""), d["sl"], d["dgia"]
                # HĐ hạch toán Nợ 6427: khi import phần mềm cần ĐVT=MHDV, SL=1,
                # Đơn giá = Thành tiền/Số lượng (=Thành tiền vì SL=1)
                if loai == "purchase" and str(no_r or "").strip() == "6427":
                    dvt_out, sl_out = "MHDV", 1
                    dgia_out = (ds / sl_out) if isinstance(ds, (int, float)) else ds
                append_row([
                    r["khhdon"], r["shdon"],
                    ngay_fmt, d["nguoi"], d["mst"],
                    it.get("stt", ""), it.get("ma_vt", ""),
                    d["ten"], dvt_out,
                    sl_out, dgia_out,
                    ds, d["ts_hien"], d["tien_thue"],
                    tt, kq,
                ], no_r, co_r)
                cur = ct_totals[loai].setdefault(ikey, {"ds": 0, "thue": 0})
                cur["ds"] += ds if isinstance(ds, (int, float)) else 0
                cur["thue"] += d["tien_thue"] if isinstance(d["tien_thue"], (int, float)) else 0

        # ===== TỜ KHAI NHẬP KHẨU (chỉ cho MUA VÀO) =====
        if loai == "purchase":
            conn_tk = db()
            tk_rows = conn_tk.execute(
                "SELECT * FROM tokhai_nhap WHERE company_id=? ORDER BY ngay_dk", (cid,)).fetchall()
            conn_tk.close()
            for tkr in tk_rows:
                try:
                    tk_items = json.loads(tkr["items_json"]) if tkr["items_json"] else []
                except Exception:
                    tk_items = []
                ngay_tk = tkr["ngay_dk"] or ""
                if ngay_tk and "-" in ngay_tk:
                    p = ngay_tk.split("-"); ngay_tk = f"{p[2]}/{p[1]}/{p[0]}"
                for idx, it in enumerate(tk_items, 1):
                    ds_tk = it.get("tri_gia_gtgt", 0) or 0
                    thue_tk = it.get("tien_thue_gtgt", 0) or 0
                    sl_tk = it.get("sluong", 0) or 0
                    # đơn giá = thành tiền / số lượng
                    dgia_tk = round(ds_tk / sl_tk) if sl_tk else 0
                    # A..P: như hóa đơn thường; Q,R,S = thuế NK; T = Nợ, U = Có
                    co_tk = _co_theo_tong(round(ds_tk) + round(thue_tk))
                    append_row([
                        "TKNK", tkr["so_tk"], ngay_tk,
                        tkr["nguoi_xk"], tkr["nguoi_xk"],          # Người bán & MST bán = tên người XK
                        idx, "", it.get("ten", ""), it.get("dvt", ""),
                        sl_tk, dgia_tk,                            # số lượng, đơn giá
                        round(ds_tk), it.get("ts_gtgt", ""), round(thue_tk),
                        "Tờ khai nhập khẩu", "",
                        round(it.get("tri_gia_nk", 0) or 0),       # Q: trị giá tính thuế NK
                        it.get("ts_nk", ""),                       # R: thuế suất NK
                        round(it.get("tien_thue_nk", 0) or 0),     # S: tiền thuế NK
                    ], "", co_tk)
                    # cộng vào tổng đối chiếu (theo số tờ khai)
                    ikey = (str("TKNK"), str(tkr["so_tk"]))
                    cur = ct_totals["purchase"].setdefault(ikey, {"ds": 0, "thue": 0})
                    cur["ds"] += round(ds_tk)
                    cur["thue"] += round(thue_tk)

        autofit(ws)
        format_so(ws)
        ws.freeze_panes = "C2"
        # Việc 9: thêm Filter kéo từ đầu đến cuối
        if ws.max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"

    build_detail_sheet("Chi tiết MUA VÀO", "purchase")
    build_detail_sheet("Chi tiết BÁN RA", "sold")
    _tlog("xong 2 sheet chi tiet -> dung BK + Doi chieu")

    # ===== SHEET BẢNG KÊ CHUẨN (theo mẫu Thông tư) =====
    from openpyxl.styles import Border, Side, Alignment as _Al
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def la_hd_loai_bo(raw, r):
        """Loại bỏ hóa đơn:
        - 'Hóa đơn đã bị thay thế' (mã 4) và 'Hóa đơn hủy' (mã 6)
        - Kết quả kiểm tra = 'Hóa đơn không đủ điều kiện cấp mã'
        GIỮ 'Hóa đơn thay thế' (mã 2), 'điều chỉnh' (mã 3), 'mới' (mã 1)..."""
        tt = str(raw.get("tthai", r["tthai"]) or "").strip()
        mota = _mo_ta_trang_thai(raw.get("tthai", r["tthai"]))
        if tt in ("4", "6"):
            return True
        if mota in ("Hóa đơn đã bị thay thế", "Hóa đơn hủy",
                     "Hóa đơn đã bị xóa bỏ", "Hóa đơn xóa bỏ"):
            return True
        # loại HĐ không đủ điều kiện cấp mã (kết quả kiểm tra)
        kq = str(raw.get("ttxly", "") or "").strip().lower()
        kq_mota = _mo_ta_ket_qua(raw.get("ttxly", "")).lower()
        if kq == "4" or "không đủ điều kiện" in kq or "không đủ điều kiện" in kq_mota:
            return True
        return False

    # Lưu tổng BK theo hóa đơn để đối chiếu (việc 4)
    bk_totals = {"purchase": {}, "sold": {}}

    mst_cty_bk_goc = _chuan_mst(comp["mst"])[:10]
    def _hd_dung_cty(r, loai):
        """Loại hóa đơn lẫn của công ty khác: mua vào -> nmmst phải cùng MST GỐC
        (10 số, bỏ hậu tố chi nhánh) với công ty. BÁN RA: KHÔNG lọc theo MST
        người bán (nbmst) — xem giải thích ở build_detail_sheet phía trên
        (trang Thuế đã tự lọc đúng công ty đăng nhập; hộ/cá nhân kinh doanh có
        thể có MST trên hóa đơn khác MST đăng ký trong phần mềm, lọc thêm ở
        đây gây mất trắng hóa đơn bán ra)."""
        if loai == "purchase":
            nm = str(r["nmmst"] or "").strip()
            return (not nm) or _chuan_mst(nm)[:10] == mst_cty_bk_goc
        return True

    # ----- BẢNG KÊ MUA VÀO (mỗi hóa đơn 1 dòng + cột Mặt hàng) -----
    ws = wb.create_sheet("BK Mua vào")
    hdr1 = ["STT", "Ký hiệu", "Số Hoá Đơn", "Ngày lập",
            "Tên người bán", "MST người bán", "Mặt hàng", "Thuế suất GTGT",
            "Doanh số mua chưa thuế", "Thuế GTGT", "Tổng thanh toán",
            "Trạng thái"]
    ws.append(hdr1)
    style_header(ws, len(hdr1))
    stt = 0
    tong_ds_mua = tong_thue_mua = 0
    for r in rows:
        if r["loai"] != "purchase":
            continue
        if not _hd_dung_cty(r, "purchase"):   # loại HĐ lẫn của công ty khác
            continue
        try:
            raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            raw = {}
        if la_hd_loai_bo(raw, r):   # việc 3: chỉ loại HĐ thay thế / xóa bỏ
            continue
        stt += 1
        ngay = (r["tdlap"] or "").split("T")[0]
        if ngay and "-" in ngay:
            y, m, d = ngay.split("-"); ngay = f"{d}/{m}/{y}"
        ds = _to_num(r["tgtcthue"]) or 0
        thue = _to_num(r["tgtthue"]) or 0
        # HKD: nếu tgtcthue=0 nhưng tgtttbso>0 -> lấy thành tiền (tổng thanh toán)
        if not ds:
            ds = _to_num(r["tgtttbso"]) or 0
        # mặt hàng đầu tiên + thuế suất (loại dòng "Tổng tiền phí" ra khỏi
        # phần gộp này vì nó đã được ghi thành 1 dòng riêng ở dưới)
        _items, _sm = get_invoice_items(r)
        _items_hh = [it for it in _items if not it.get("_la_phi")]
        la_ck_hd = False
        if _items_hh:
            def _is_ck(it):
                """Dòng CHIẾT KHẤU THUẦN (cần ghi âm cả hóa đơn nếu mọi dòng đều vậy).
                LƯU Ý: dòng TChat=3 NHƯNG có chiết khấu dòng (STCKhau>0) là HÀNG
                bán có giảm giá (net dương) -> KHÔNG phải dòng CK thuần."""
                tc = str(it.get("tchat", "") or "")
                if tc == "4": return False   # ghi chú -> không phải CK
                tt = _to_num(it.get("thtien")) or 0
                ck = _to_num(it.get("stckhau")) or 0
                # có chiết khấu dòng -> là hàng có giảm giá, KHÔNG phải CK thuần
                if isinstance(ck, (int, float)) and ck > 0:
                    return False
                # thành tiền âm -> dòng giảm thực sự
                if isinstance(tt, (int, float)) and tt < 0:
                    return True
                # TChat=3 không có STCKhau -> dòng chiết khấu TM riêng
                if tc == "3":
                    return True
                ten = str(it.get("ten_hang", "") or "").lower()
                return "chiết khấu" in ten or "chiet khau" in ten
            items_ko_gc = [it for it in _items_hh if str(it.get("tchat","") or "")!="4"]
            if items_ko_gc and all(_is_ck(it) for it in items_ko_gc):
                la_ck_hd = True
        if la_ck_hd:
            ds = -abs(ds) if isinstance(ds, (int, float)) else ds
            thue = -abs(thue) if isinstance(thue, (int, float)) else thue
        mat_hang = ""
        thue_suat = ""
        if _items_hh:
            mat_hang = _items_hh[0].get("ten_hang", "")
            # thuế suất từ summary (nếu có nhiều mức thì gộp); lấy mức đầu tiên có
            ts_set = []
            for it in _items_hh:
                tsv = str(it.get("tsuat", "") or "").strip()
                if tsv and tsv not in ts_set:
                    ts_set.append(tsv)
            thue_suat = ", ".join(ts_set) if ts_set else ""
            if la_ck_hd:
                mat_hang += " (Chiết khấu TM - ghi âm)"
            elif len(_items_hh) > 1:
                mat_hang += " ..."
        ws.append([stt, r["khhdon"], r["shdon"], ngay, r["nbten"], r["nbmst"],
                   mat_hang, thue_suat, ds, thue, _to_num(r["tgtttbso"]),
                   _mo_ta_trang_thai(raw.get("tthai", r["tthai"]))])
        tong_ds_mua += ds if isinstance(ds, (int, float)) else 0
        tong_thue_mua += thue if isinstance(thue, (int, float)) else 0
        ikey = (str(r["khhdon"]), str(r["shdon"]).lstrip("0") or "0")
        bk_totals["purchase"][ikey] = {"ds": ds, "thue": thue}
        # dòng riêng "Tổng tiền phí" (thu hộ/lệ phí ngoài DSHHDVu, không nằm
        # trong TgTCThue) — khớp với dòng đã thêm ở "Chi tiết MUA VÀO"
        for it_phi in _items:
            if not it_phi.get("_la_phi"):
                continue
            tien_phi = _to_num(it_phi.get("thtien")) or 0
            if not tien_phi:
                continue
            stt += 1
            ws.append([stt, r["khhdon"], r["shdon"], ngay, r["nbten"], r["nbmst"],
                       "Tổng tiền phí", "KCT", tien_phi, 0, "", ""])
            tong_ds_mua += tien_phi
            bk_totals["purchase"][ikey]["ds"] += tien_phi

    # ===== TỜ KHAI NHẬP KHẨU: mỗi tờ khai gộp thành 1 dòng trong BK Mua vào =====
    conn_tk2 = db()
    tk_rows2 = conn_tk2.execute(
        "SELECT * FROM tokhai_nhap WHERE company_id=? ORDER BY ngay_dk", (cid,)).fetchall()
    conn_tk2.close()
    for tkr in tk_rows2:
        try:
            tk_items = json.loads(tkr["items_json"]) if tkr["items_json"] else []
        except Exception:
            tk_items = []
        if not tk_items:
            continue
        ds_tk_tong = sum(round(it.get("tri_gia_gtgt", 0) or 0) for it in tk_items)
        thue_tk_tong = sum(round(it.get("tien_thue_gtgt", 0) or 0) for it in tk_items)
        ngay_tk = tkr["ngay_dk"] or ""
        if ngay_tk and "-" in ngay_tk:
            p = ngay_tk.split("-"); ngay_tk = f"{p[2]}/{p[1]}/{p[0]}"
        stt += 1
        # gộp tên hàng: lấy tên dòng đầu + "(N dòng hàng)"
        ten_gop = (tk_items[0].get("ten", "") or "")[:40]
        if len(tk_items) > 1:
            ten_gop += f" ... ({len(tk_items)} dòng hàng)"
        ts_set = []
        for it in tk_items:
            tsv = str(it.get("ts_gtgt", "") or "").strip()
            if tsv and tsv not in ts_set:
                ts_set.append(tsv)
        ws.append([stt, "TKNK", tkr["so_tk"], ngay_tk, tkr["nguoi_xk"], tkr["nguoi_xk"],
                   ten_gop, ", ".join(ts_set), ds_tk_tong, thue_tk_tong,
                   _to_num(ds_tk_tong + thue_tk_tong), "Tờ khai nhập khẩu"])
        tong_ds_mua += ds_tk_tong
        tong_thue_mua += thue_tk_tong
        # key đối chiếu khớp với Chi tiết (TKNK + số tờ khai)
        bk_totals["purchase"][("TKNK", str(tkr["so_tk"]))] = {"ds": ds_tk_tong, "thue": thue_tk_tong}

    # dòng tổng (Doanh số giờ ở cột I=9, Thuế cột J=10)
    ws.append(["", "", "", "", "", "", "", "TỔNG CỘNG",
               _to_num(tong_ds_mua), _to_num(tong_thue_mua), "", ""])
    for c in range(1, 13):
        ws.cell(ws.max_row, c).font = Font(bold=True, color="C00000")
    autofit(ws)
    format_so(ws)
    ws.freeze_panes = "C2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr1))}{ws.max_row}"

    # ----- BẢNG KÊ BÁN RA (form BKDAURA chuẩn 01-1/GTGT, tách theo thuế suất) -----
    from openpyxl.styles import Border, Side, Alignment as _Al
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    bold = Font(bold=True, name="Times New Roman")
    center = _Al(horizontal="center", vertical="center", wrap_text=True)

    ws = wb.create_sheet("BK Bán ra")
    sold_rows = [r for r in rows if r["loai"] == "sold" and _hd_dung_cty(r, "sold")]

    # Bỏ phần tiêu đề dài (gây lệch cột) — header bắt đầu ngay dòng 1
    hdr2 = ["STT", "Ký hiệu mẫu", "Ký hiệu HĐ", "Số hóa đơn", "Ngày lập",
            "Tên người mua", "MST người mua", "Mặt hàng",
            "Doanh số bán chưa thuế", "Thuế GTGT", "Trạng thái", "Kết quả"]
    ws.append(hdr2)
    hrow = ws.max_row
    for c in range(1, len(hdr2) + 1):
        cell = ws.cell(hrow, c)
        cell.font = Font(bold=True, color="FFFFFF", name="Times New Roman")
        cell.fill = PatternFill("solid", fgColor="2E5C8A")
        cell.alignment = center
        cell.border = border

    # Gom hóa đơn bán ra theo nhóm thuế suất (loại HĐ bị thay thế - việc 3)
    groups = {"KCT": [], "0": [], "5": [], "8": [], "10": [], "KHAC": []}
    for r in sold_rows:
        try:
            raw = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            raw = {}
        if la_hd_loai_bo(raw, r):   # việc 3: chỉ bỏ HĐ thay thế / xóa bỏ
            continue
        tt = _mo_ta_trang_thai(raw.get("tthai", r["tthai"]))
        kq = _mo_ta_ket_qua(raw.get("ttxly", ""))
        items, info = get_invoice_items(r)
        ikey = (str(r["khhdon"]), str(r["shdon"]).lstrip("0") or "0")
        if not info or not info.get("theo_ts"):
            ds_fallback = _to_num(r["tgtcthue"]) or 0
            thue_fallback = _to_num(r["tgtthue"]) or 0
            dang_nhap_ok = bool(client and client.token and not getattr(client, "_token_dead", False))
            # Không lấy được chi tiết dòng hàng nhưng ĐÃ đăng nhập thành công
            # và có số tiền thật -> hóa đơn bán hàng của hộ/cá nhân kinh doanh
            # (không tách VAT) -> xếp vào nhóm KCT (đúng bản chất), không phải
            # "Khác/chưa lấy được file" (dễ hiểu nhầm là lỗi mạng, thử lại được).
            g = "KCT" if (dang_nhap_ok and ds_fallback) else "KHAC"
            groups[g].append((r, raw, tt, kq, None, None))
            bk_totals["sold"][ikey] = {"ds": ds_fallback, "thue": thue_fallback}
            continue
        bt = bk_totals["sold"].setdefault(ikey, {"ds": 0, "thue": 0})
        for key, val in info["theo_ts"].items():
            g = key if key in groups else "KHAC"
            groups[g].append((r, raw, tt, kq, info, (key, val)))
            bt["ds"] += val["ds"]; bt["thue"] += val["thue"]

    nhom_label = {
        "KCT": "1. Hàng hóa, dịch vụ không chịu thuế GTGT (KCT)",
        "0": "2. Hàng hóa, dịch vụ chịu thuế suất 0%",
        "5": "3. Hàng hóa, dịch vụ chịu thuế suất 5%",
        "8": "4. Hàng hóa, dịch vụ chịu thuế suất 8%",
        "10": "5. Hàng hóa, dịch vụ chịu thuế suất 10%",
        "KHAC": "6. Khác / chưa lấy được file XML",
    }

    stt = 0
    grand_ds = grand_thue = 0
    for key in ["KCT", "0", "5", "8", "10", "KHAC"]:
        glist = groups[key]
        if not glist:
            continue
        ws.append([nhom_label[key]])
        ws.cell(ws.max_row, 1).font = bold
        sub_ds = sub_thue = 0
        for (r, raw, tt, kq, info, tsdata) in glist:
            stt += 1
            ngay = (r["tdlap"] or "").split("T")[0]
            if "-" in ngay:
                y, m, d = ngay.split("-"); ngay = f"{d}/{m}/{y}"
            if info and tsdata:
                _k, val = tsdata
                ds = val["ds"]; thue = val["thue"]
                ws.append([stt, info["khmshdon"], info["khhdon"], info["shdon"],
                           ngay, info["ten_nmua"] or "Khách lẻ", info["mst_nmua"],
                           info["mat_hang"], _to_num(ds), _to_num(thue), tt, kq])
            else:
                ds = _to_num(r["tgtcthue"]) or 0
                thue = _to_num(r["tgtthue"]) or 0
                dang_nhap_ok_row = bool(client and client.token and not getattr(client, "_token_dead", False))
                mat_hang_txt = ("(Cả hóa đơn — không tách dòng hàng)" if (dang_nhap_ok_row and ds)
                                else "(chưa lấy được file XML)")
                ws.append([stt, "1", r["khhdon"], r["shdon"], ngay,
                           "", r["nmmst"], mat_hang_txt,
                           ds, thue, tt, kq])
            sub_ds += ds if isinstance(ds, (int, float)) else 0
            sub_thue += thue if isinstance(thue, (int, float)) else 0
        ws.append(["", "", "", "", "", "", "", "Tổng nhóm",
                   _to_num(sub_ds), _to_num(sub_thue), "", ""])
        for c in range(1, 13):
            ws.cell(ws.max_row, c).font = bold
        grand_ds += sub_ds; grand_thue += sub_thue

    ws.append([])
    ws.append(["", "", "", "", "", "", "", "TỔNG CỘNG",
               _to_num(grand_ds), _to_num(grand_thue), "", ""])
    for c in range(1, 13):
        ws.cell(ws.max_row, c).font = Font(bold=True, color="C00000", name="Times New Roman")
    autofit(ws)
    format_so(ws)
    # thu nhỏ cột A (STT) và B (Ký hiệu mẫu) - giá trị ngắn
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 11
    # cố định phần header (dòng 1) để cuộn vẫn thấy cột
    ws.freeze_panes = f"A{hrow + 1}"
    # Việc 9: thêm Filter từ dòng header tới cuối
    if ws.max_row > hrow:
        ws.auto_filter.ref = f"A{hrow}:{get_column_letter(ws.max_column)}{ws.max_row}"

    # ===== SHEET ĐỐI CHIẾU (việc 4): so Chi tiết vs Bảng kê =====
    ws = wb.create_sheet("Đối chiếu")
    bold = Font(bold=True)
    ws.append(["BẢNG ĐỐI CHIẾU TỔNG HỢP (tự cập nhật khi sửa số liệu các sheet)"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, color="2E5C8A", size=13)
    ws.append([])
    # Thông tin công ty
    ws.append(["Thông tin công ty"])
    ws.cell(ws.max_row, 1).font = Font(bold=True, size=11)
    ws.append([comp["ten"] or ""])
    ws.append([comp["mst"] or ""])
    # lấy địa chỉ từ hóa đơn bán ra (nbdchi = địa chỉ người bán = địa chỉ công ty)
    diachi = ""
    try:
        for r_dc in rows:
            if r_dc["loai"] == "sold" and r_dc["raw"]:
                raw_dc = json.loads(r_dc["raw"])
                dc = raw_dc.get("nbdchi", "") or raw_dc.get("dchi", "") or ""
                if dc and len(dc) > 5:
                    diachi = dc; break
    except Exception:
        pass
    if diachi:
        ws.append([diachi])
    ws.append([])

    # Cấu trúc cột:
    #  Chi tiết MUA VÀO/BÁN RA: Thành tiền = N(14), Tiền thuế = P(16)
    #  BK Mua vào: Doanh số = H(8), Thuế = I(9)
    #  BK Bán ra:  Doanh số = I(9), Thuế = J(10)
    CT_MUA, CT_BAN = "'Chi tiết MUA VÀO'", "'Chi tiết BÁN RA'"
    BK_MUA, BK_BAN = "'BK Mua vào'", "'BK Bán ra'"

    def add_row(nhom, chi_tieu, formula, ktra_label=None, ktra_formula=None):
        ws.append([nhom, chi_tieu, None, None, ktra_label, None])
        rr = ws.max_row
        ws.cell(rr, 3).value = formula
        if ktra_formula:
            ws.cell(rr, 6).value = ktra_formula
        return rr

    # ===== KHỐI MUA VÀO (đặt trên) =====
    ws.append(["KHỐI MUA VÀO"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="8a6d1f")
    # BK Mua vào: chỉ SUM dòng có STT (cột A) -> loại dòng TỔNG CỘNG
    r_bk_mua_ds = add_row("BK Mua vào", "Doanh số mua chưa có thuế",
                          f'=SUMIF({BK_MUA}!A:A,">0",{BK_MUA}!I:I)')
    r_bk_mua_thue = add_row("", "Thuế GTGT", f'=SUMIF({BK_MUA}!A:A,">0",{BK_MUA}!J:J)')
    r_ct_mua_ds = add_row("Chi tiết MUA VÀO", "Doanh số mua chưa có thuế",
                          f"=SUM({CT_MUA}!L:L)")
    r_ct_mua_thue = add_row("", "Thuế GTGT", f"=SUM({CT_MUA}!N:N)")
    ws.cell(r_bk_mua_ds, 5).value = "Kiểm tra Hàng hoá"
    ws.cell(r_bk_mua_ds, 6).value = f"=C{r_bk_mua_ds}-C{r_ct_mua_ds}"
    ws.cell(r_bk_mua_thue, 5).value = "Kiểm tra VAT"
    ws.cell(r_bk_mua_thue, 6).value = f"=C{r_bk_mua_thue}-C{r_ct_mua_thue}"
    ws.append([])

    # ===== KHỐI BÁN RA (đặt dưới) =====
    ws.append(["KHỐI BÁN RA"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="1F6B4A")
    # BK Bán ra: chỉ SUM dòng có STT (cột A là số) -> loại dòng "Tổng nhóm"/"TỔNG CỘNG"
    r_bk_ban_ds = add_row("BK Bán ra", "Tổng doanh thu hàng hoá, dịch vụ bán ra:",
                          f'=SUMIF({BK_BAN}!A:A,">0",{BK_BAN}!I:I)')
    r_bk_ban_thue = add_row("", "Tổng thuế GTGT của hàng hóa, dịch vụ bán ra:",
                            f'=SUMIF({BK_BAN}!A:A,">0",{BK_BAN}!J:J)')
    # Chi tiết BÁN RA: SUM theo dòng có STT (cột H)
    r_ct_ban_ds = add_row("Chi tiết BÁN RA", "Tổng tiền chưa thuế",
                          f"=SUM({CT_BAN}!L:L)")
    r_ct_ban_thue = add_row("", "Tổng tiền thuế", f"=SUM({CT_BAN}!N:N)")
    # dòng kiểm tra
    ws.cell(r_bk_ban_ds, 5).value = "Kiểm tra Hàng hoá"
    ws.cell(r_bk_ban_ds, 6).value = f"=C{r_bk_ban_ds}-C{r_ct_ban_ds}"
    ws.cell(r_bk_ban_thue, 5).value = "Kiểm tra VAT"
    ws.cell(r_bk_ban_thue, 6).value = f"=C{r_bk_ban_thue}-C{r_ct_ban_thue}"
    ws.append([])

    # ===== KẾT LUẬN =====
    ws.append(["KẾT LUẬN ĐỐI CHIẾU"]); ws.cell(ws.max_row, 1).font = Font(bold=True, size=12, color="2E5C8A")
    kr = ws.max_row
    ws.append(["", "Bán ra:", None])
    rr = ws.max_row
    ws.cell(rr, 3).value = (f'=IF(AND(ABS(C{r_bk_ban_ds}-C{r_ct_ban_ds})<100,ABS(C{r_bk_ban_thue}-C{r_ct_ban_thue})<100),'
                            f'"✓ KHỚP (BK Bán ra = Chi tiết Bán ra)",'
                            f'"⚠ LỆCH: hàng hoá "&TEXT(C{r_bk_ban_ds}-C{r_ct_ban_ds},"#,##0")&", VAT "&TEXT(C{r_bk_ban_thue}-C{r_ct_ban_thue},"#,##0"))')
    ws.cell(rr, 3).font = bold
    ws.append(["", "Mua vào:", None])
    rr = ws.max_row
    ws.cell(rr, 3).value = (f'=IF(AND(ABS(C{r_bk_mua_ds}-C{r_ct_mua_ds})<100,ABS(C{r_bk_mua_thue}-C{r_ct_mua_thue})<100),'
                            f'"✓ KHỚP (BK Mua vào = Chi tiết Mua vào)",'
                            f'"⚠ LỆCH: hàng hoá "&TEXT(C{r_bk_mua_ds}-C{r_ct_mua_ds},"#,##0")&", VAT "&TEXT(C{r_bk_mua_thue}-C{r_ct_mua_thue},"#,##0"))')
    ws.cell(rr, 3).font = bold

    # ===== DANH SÁCH HÓA ĐƠN LỆCH (để kiểm tra từng hóa đơn) =====
    # cấu hình cột để SUMIFS (ký hiệu + số HĐ → doanh số, thuế) theo từng sheet
    lech_cfg = {
        "sold": {"ct_sheet": "'Chi tiết BÁN RA'", "ct_kh": "A", "ct_key": "B", "ct_ds": "L", "ct_thue": "N",
                 "bk_sheet": "'BK Bán ra'", "bk_kh": "B", "bk_key": "D", "bk_ds": "I", "bk_thue": "J"},
        "purchase": {"ct_sheet": "'Chi tiết MUA VÀO'", "ct_kh": "A", "ct_key": "B", "ct_ds": "L", "ct_thue": "N",
                     "bk_sheet": "'BK Mua vào'", "bk_kh": "B", "bk_key": "C", "bk_ds": "I", "bk_thue": "J"},
    }

    def them_ds_lech(tieu_de, loai):
        cfg = lech_cfg[loai]
        ct = ct_totals[loai]; bk = bk_totals[loai]
        all_keys = sorted(set(ct.keys()) | set(bk.keys()),
                          key=lambda x: (x[0], int(x[1]) if x[1].isdigit() else 0))
        lech = []
        for k in all_keys:
            cc = ct.get(k, {"ds": 0, "thue": 0})
            bb = bk.get(k, {"ds": 0, "thue": 0})
            ld = round((cc["ds"] or 0) - (bb["ds"] or 0))
            lt = round((cc["thue"] or 0) - (bb["thue"] or 0))
            if abs(ld) >= 100 or abs(lt) >= 100:
                lech.append(k)
        ws.append([])
        ws.append([tieu_de])
        ws.cell(ws.max_row, 1).font = Font(bold=True, size=11, color="C00000")
        if not lech:
            ws.append(["", "✓ Không có hóa đơn nào lệch (>= 100đ)"])
            ws.cell(ws.max_row, 2).font = Font(color="1F6B4A")
            return
        ws.append(["Ký hiệu", "Số HĐ", "Chưa VAT (Chi tiết)", "Chưa VAT (Bảng kê)",
                   "Lệch chưa VAT", "VAT (Chi tiết)", "VAT (Bảng kê)", "Lệch VAT", "Ghi chú"])
        hr = ws.max_row
        for c in range(1, 10):
            ws.cell(hr, c).font = Font(bold=True, color="FFFFFF")
            ws.cell(hr, c).fill = PatternFill("solid", fgColor="C0392B")
        for k in lech:
            ghi = ""
            if k in ct and k not in bk:
                ghi = "Chỉ có ở Chi tiết (HĐ bị thay thế/không vào BK)"
            elif k in bk and k not in ct:
                ghi = "Chỉ có ở Bảng kê"
            ws.append([k[0], k[1], None, None, None, None, None, None, ghi])
            rr = ws.max_row
            # CÔNG THỨC SUMIFS: dùng cả Ký hiệu + Số HĐ để không bị trùng
            ws.cell(rr, 3).value = (f"=SUMIFS({cfg['ct_sheet']}!{cfg['ct_ds']}:{cfg['ct_ds']},"
                                    f"{cfg['ct_sheet']}!{cfg['ct_kh']}:{cfg['ct_kh']},$A{rr},"
                                    f"{cfg['ct_sheet']}!{cfg['ct_key']}:{cfg['ct_key']},$B{rr})")
            ws.cell(rr, 4).value = (f"=SUMIFS({cfg['bk_sheet']}!{cfg['bk_ds']}:{cfg['bk_ds']},"
                                    f"{cfg['bk_sheet']}!{cfg['bk_kh']}:{cfg['bk_kh']},$A{rr},"
                                    f"{cfg['bk_sheet']}!{cfg['bk_key']}:{cfg['bk_key']},$B{rr})")
            ws.cell(rr, 5).value = f"=C{rr}-D{rr}"
            ws.cell(rr, 6).value = (f"=SUMIFS({cfg['ct_sheet']}!{cfg['ct_thue']}:{cfg['ct_thue']},"
                                    f"{cfg['ct_sheet']}!{cfg['ct_kh']}:{cfg['ct_kh']},$A{rr},"
                                    f"{cfg['ct_sheet']}!{cfg['ct_key']}:{cfg['ct_key']},$B{rr})")
            ws.cell(rr, 7).value = (f"=SUMIFS({cfg['bk_sheet']}!{cfg['bk_thue']}:{cfg['bk_thue']},"
                                    f"{cfg['bk_sheet']}!{cfg['bk_kh']}:{cfg['bk_kh']},$A{rr},"
                                    f"{cfg['bk_sheet']}!{cfg['bk_key']}:{cfg['bk_key']},$B{rr})")
            ws.cell(rr, 8).value = f"=F{rr}-G{rr}"
            for c in range(1, 9):
                ws.cell(rr, c).font = Font(color="C00000")
                if c >= 3:
                    ws.cell(rr, c).number_format = "#,##0"
            ws.cell(rr, 9).font = Font(color="C00000")

    them_ds_lech("CHI TIẾT HÓA ĐƠN LỆCH - BÁN RA (kiểm tra & điều chỉnh):", "sold")
    them_ds_lech("CHI TIẾT HÓA ĐƠN LỆCH - MUA VÀO (kiểm tra & điều chỉnh):", "purchase")

    # định dạng số cột C và F (kể cả ô công thức SUMIF/SUM ra số lẻ)
    for r in range(1, ws.max_row + 1):
        for cc in (3, 6):
            v = ws.cell(r, cc).value
            if v is None:
                continue
            # ô kết luận dạng chữ (IF...TEXT) thì bỏ qua
            if isinstance(v, str) and v.startswith("=") and "TEXT" in v:
                continue
            if isinstance(v, (int, float)) or (isinstance(v, str) and v.startswith("=")):
                ws.cell(r, cc).number_format = "#,##0"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.freeze_panes = "C1"  # cố định cột tới cột 2 (A,B) để dễ xem

    wb.remove(wb["Sheet"])
    # Sắp xếp lại thứ tự sheet: Đối chiếu; Chi tiết MUA VÀO; BK Mua vào;
    # Chi tiết BÁN RA; BK Bán ra
    thu_tu = ["Đối chiếu", "Chi tiết MUA VÀO", "BK Mua vào",
              "Chi tiết BÁN RA", "BK Bán ra"]
    try:
        wb._sheets.sort(key=lambda s: thu_tu.index(s.title)
                        if s.title in thu_tu else 999)
        wb.active = 0
    except Exception:
        pass
    # tên file kèm kỳ (nếu có) để các kỳ không ghi đè nhau trên Desktop
    ky_fname = ""
    try:
        for r in rows:
            nd = (r["tdlap"] or "").split("T")[0]
            if "-" in nd:
                y, m, _d = nd.split("-"); ky_fname = f"_T{int(m):02d}{y}"; break
    except Exception:
        pass
    fname_x = f"BangKe_HoaDon_{comp['mst']}{ky_fname}.xlsx"
    # chỉ tạo DUY NHẤT 1 file BangKe_HoaDon (không tạo thêm file TongHop)
    path = os.path.join(DOWNLOAD_DIR, fname_x)
    _tlog("bat dau ghi file Excel ra dia...")
    wb.save(path)
    _tlog("DA GHI XONG file Excel")
    import shutil
    open_path = path
    # Mặc định lưu ra DESKTOP cho dễ tìm
    desktop = _get_desktop_dir()
    if desktop and os.path.isdir(desktop):
        try:
            dest = os.path.join(desktop, fname_x)
            shutil.copy(path, dest)
            open_path = dest
        except Exception:
            pass
    # lưu thêm vào thư mục công ty (nếu có cấu hình)
    save_dir3 = (comp["save_dir"] or "").strip() if comp else ""
    if save_dir3 and os.path.isdir(save_dir3):
        try:
            shutil.copy(path, os.path.join(save_dir3, fname_x))
        except Exception:
            pass
    _open_file_local(open_path)
    return _resp_xuat(path, fname_x)


# ---------- ĐỌC FILE XML -> XUẤT EXCEL CHI TIẾT ----------
def _extract_invoice_xml(data_bytes):
    """
    Nhận bytes của file người dùng upload. File có thể là:
      - file invoice.zip của TCT (chứa invoice.xml bên trong) — kể cả khi đặt đuôi .xml
      - file XML thuần
    Trả về bytes của nội dung XML hóa đơn.
    """
    import io as _io
    import zipfile as _zip
    # File zip bắt đầu bằng 'PK'
    if data_bytes[:2] == b"PK":
        try:
            zf = _zip.ZipFile(_io.BytesIO(data_bytes))
            # ưu tiên invoice.xml, nếu không có lấy file .xml đầu tiên
            names = zf.namelist()
            target = None
            for n in names:
                if n.lower().endswith("invoice.xml"):
                    target = n
                    break
            if not target:
                for n in names:
                    if n.lower().endswith(".xml"):
                        target = n
                        break
            if target:
                return zf.read(target)
        except Exception:
            pass
    return data_bytes


def _parse_detail_json(detail):
    """
    Parse JSON chi tiết hóa đơn (từ endpoint detail của TCT) thành list mặt hàng.
    Cùng định dạng output với _parse_xml_invoice để dùng chung.
    """
    rows = []
    if not detail or not isinstance(detail, dict):
        return rows
    khmshdon = str(detail.get("khmshdon", "") or "")
    khhdon = detail.get("khhdon", "") or ""
    shdon = str(detail.get("shdon", "") or "")
    ngay = detail.get("tdlap", "") or ""
    ten_nban = detail.get("nbten", "") or ""
    mst_nban = detail.get("nbmst", "") or ""
    dchi_nban = detail.get("nbdchi", "") or ""
    ten_nmua = detail.get("nmten", "") or detail.get("nmtnmua", "") or ""
    mst_nmua = detail.get("nmmst", "") or ""
    items = detail.get("hdhhdvu") or []
    for it in items:
        rows.append({
            "khmshdon": khmshdon, "khhdon": khhdon, "shdon": shdon,
            "ngay": ngay, "ten_nban": ten_nban, "mst_nban": mst_nban,
            "dchi_nban": dchi_nban, "ten_nmua": ten_nmua, "mst_nmua": mst_nmua,
            "stt": str(it.get("stt", "") or ""),
            "tchat": str(it.get("tchat", "") or ""),
            "ma_vt": it.get("mhhdvu", "") or "",
            "ten_hang": it.get("ten", "") or it.get("thhdvu", "") or "",
            "dvt": it.get("dvtinh", "") or "",
            "sluong": it.get("sluong", ""),
            "dgia": it.get("dgia", ""),
            "thtien": it.get("thtien", ""),
            "stckhau": it.get("stckhau", ""),
            "tsuat": str(it.get("ltsuat", "") or it.get("tsuat", "") or ""),
            "tien_thue": it.get("tthue", "") or it.get("tongtien_thue", ""),
            "tgtcthue": detail.get("tgtcthue", ""),
            "tgtthue": detail.get("tgtthue", ""),
            "tgtttbso": detail.get("tgtttbso", ""),
        })

    # Tổng tiền phí (không chịu thuế) — xem giải thích ở _lay_tong_tien_phi_xml
    tong_phi = _lay_tong_tien_phi_json(detail)
    if tong_phi:
        common = {"khmshdon": khmshdon, "khhdon": khhdon, "shdon": shdon,
                  "ngay": ngay, "ten_nban": ten_nban, "mst_nban": mst_nban,
                  "dchi_nban": dchi_nban, "ten_nmua": ten_nmua, "mst_nmua": mst_nmua,
                  "tgtcthue": detail.get("tgtcthue", ""), "tgtthue": detail.get("tgtthue", ""),
                  "tgtttbso": detail.get("tgtttbso", "")}
        rows.append(_dong_phi_hoa_don(tong_phi, common))
    return rows


def _summary_from_detail_json(detail):
    """Parse JSON detail -> thông tin gộp theo thuế suất (cho BK bán ra)."""
    if not detail or not isinstance(detail, dict):
        return None
    info = {
        "khmshdon": str(detail.get("khmshdon", "") or ""),
        "khhdon": detail.get("khhdon", "") or "",
        "shdon": str(detail.get("shdon", "") or ""),
        "ngay": detail.get("tdlap", "") or "",
        "ten_nban": detail.get("nbten", "") or "",
        "mst_nban": detail.get("nbmst", "") or "",
        "ten_nmua": detail.get("nmten", "") or detail.get("nmtnmua", "") or "",
        "mst_nmua": detail.get("nmmst", "") or "",
    }
    items = detail.get("hdhhdvu") or []
    ten_hangs = [it.get("ten") or it.get("thhdvu") for it in items if (it.get("ten") or it.get("thhdvu"))]
    info["mat_hang"] = (ten_hangs[0] + (" ..." if len(ten_hangs) > 1 else "")) if ten_hangs else ""

    def norm_ts(ts):
        s = str(ts or "").strip().upper()
        if s in ("", "KCT", "KHAC", "KO", "KHÔNG") or "KKK" in s or "KCT" in s:
            return "KCT"
        s2 = s.replace("%", "").strip()
        if s2 in ("0", "5", "8", "10"):
            return s2
        return s or "KHAC"

    theo_ts = {}
    # ưu tiên phần tổng hợp thuế suất nếu có
    ltsuat = detail.get("thttltsuat") or detail.get("hdhhdvu_ltsuat") or []
    if ltsuat and isinstance(ltsuat, list):
        for l in ltsuat:
            key = norm_ts(l.get("tsuat"))
            ds = _to_num(l.get("thtien")) or 0
            thue = _to_num(l.get("tthue")) or 0
            cur = theo_ts.setdefault(key, {"ds": 0, "thue": 0})
            cur["ds"] += ds if isinstance(ds, (int, float)) else 0
            cur["thue"] += thue if isinstance(thue, (int, float)) else 0
    else:
        for it in items:
            key = norm_ts(it.get("ltsuat") or it.get("tsuat"))
            ds = _to_num(it.get("thtien")) or 0
            try:
                rate = float(str(it.get("ltsuat") or it.get("tsuat") or "0").replace("%", "")) / 100
            except Exception:
                rate = 0
            thue = round(ds * rate) if isinstance(ds, (int, float)) else 0
            cur = theo_ts.setdefault(key, {"ds": 0, "thue": 0})
            cur["ds"] += ds if isinstance(ds, (int, float)) else 0
            cur["thue"] += thue
    # Tổng tiền phí (không chịu thuế) — cộng vào nhóm KCT để khớp với dòng
    # 'Tổng tiền phí' đã thêm ở _parse_detail_json (Chi tiết BÁN RA vs BK Bán ra).
    tong_phi = _lay_tong_tien_phi_json(detail)
    if tong_phi:
        cur = theo_ts.setdefault("KCT", {"ds": 0, "thue": 0})
        cur["ds"] += tong_phi
    info["theo_ts"] = theo_ts
    return info


def _lay_tong_tien_phi_xml(root):
    """Tìm 'Tổng tiền phí' trên hóa đơn — các khoản phí KHÔNG chịu VAT (vd lệ
    phí đăng kiểm, phí sân bay/thu hộ, phí dịch vụ...) hiển thị ở 1 bảng RIÊNG
    trên hóa đơn (ngoài danh sách hàng hóa/dịch vụ DSHHDVu), TRƯỚC ĐÂY phần
    mềm hoàn toàn không đọc nên bỏ sót khoản này. Khung định dạng hóa đơn điện
    tử không có 1 tên thẻ cố định duy nhất cho phần này -> dò nhiều cách.
    Trả về tổng số tiền phí (0 nếu hóa đơn không có mục này)."""
    # Cấu trúc THẬT xác nhận từ hóa đơn thật (2 hóa đơn mẫu khác nhà cung cấp,
    # cùng thống nhất 1 cấu trúc): TToan > DSLPhi > LPhi > {TLPhi (tên loại
    # phí), TPhi (số tiền)} — có thể nhiều dòng LPhi, cộng dồn hết.
    tong_dslphi = 0
    for lphi in root.findall(".//DSLPhi/LPhi"):
        el = lphi.find("TPhi")
        if el is not None and el.text:
            tong_dslphi += _to_num(el.text.strip()) or 0
    if tong_dslphi:
        return tong_dslphi
    for tag in ("TgTPhi", "TgTTPhi", "TongTienPhi"):
        el = root.find(f".//{tag}")
        if el is not None and el.text:
            v = _to_num(el.text.strip()) or 0
            if v:
                return v
    tong = 0
    for phi in (root.findall(".//DSPhi/Phi") + root.findall(".//DSKPhi/KPhi")
                + root.findall(".//DSTKhoanPhi/TKhoanPhi")):
        for tag in ("TienPhi", "TienKhoanPhi", "TPhi", "STien"):
            el = phi.find(tag)
            if el is not None and el.text:
                tong += _to_num(el.text.strip()) or 0
                break
    if tong:
        return tong
    # TTKhac chung của hóa đơn (không nằm trong từng dòng hàng) có TTruong
    # chứa chữ 'phí' -> khoản phí đặt tên tự do, không theo cấu trúc DSPhi
    for ttin in root.findall(".//TTKhac/TTin"):
        tt = ttin.find("TTruong")
        if tt is not None and tt.text and "phí" in tt.text.strip().lower():
            dl = ttin.find("DLieu")
            if dl is not None and dl.text:
                v = _to_num(dl.text.strip()) or 0
                if v:
                    return v
    return 0


def _lay_tong_tien_phi_json(detail):
    """Bản JSON của _lay_tong_tien_phi_xml (endpoint detail của TCT)."""
    if not detail or not isinstance(detail, dict):
        return 0
    # Cấu trúc THẬT (xác nhận từ XML thật, JSON dùng tên thẻ viết thường):
    # dslphi: [{tlphi (tên loại phí), tphi (số tiền)}, ...]
    lst_that = detail.get("dslphi") or []
    if isinstance(lst_that, list):
        tong_that = 0
        for it in lst_that:
            if isinstance(it, dict):
                tong_that += _to_num(it.get("tphi")) or 0
        if tong_that:
            return tong_that
    for k in ("tgtphi", "tgttphi", "tongtienphi"):
        v = _to_num(detail.get(k))
        if v:
            return v
    tong = 0
    for k in ("dsphi", "dskphi", "dstkhoanphi"):
        lst = detail.get(k) or []
        if isinstance(lst, list):
            for it in lst:
                if not isinstance(it, dict):
                    continue
                for kk in ("tienphi", "tienkhoanphi", "tphi", "stien"):
                    v = _to_num(it.get(kk))
                    if v:
                        tong += v
                        break
    if tong:
        return tong
    lst = detail.get("ttkhac") or []
    if isinstance(lst, list):
        for it in lst:
            if not isinstance(it, dict):
                continue
            truong = str(it.get("ttruong") or it.get("truong") or "").lower()
            if "phí" in truong or "phi" in truong:
                v = _to_num(it.get("dlieu") or it.get("giatri"))
                if v:
                    tong += v
    return tong


def _dong_phi_hoa_don(tong_phi, common):
    """Dựng 1 'dòng hàng' giả đại diện cho Tổng tiền phí (không chịu thuế —
    KCT) để cộng chung vào danh sách mặt hàng, hiện đúng như hóa đơn thường:
    Thành tiền = tổng tiền phí, Thuế suất = KCT, Tiền thuế GTGT = 0đ."""
    d = dict(common)
    d.update({
        "stt": "", "tchat": "", "ma_vt": "", "ten_hang": "Tổng tiền phí",
        "dvt": "", "sluong": "", "dgia": "",
        "thtien": tong_phi, "stckhau": "", "tsuat": "KCT", "tien_thue": 0,
        "_la_phi": True,   # đánh dấu để hiện đúng chữ "KCT" (không phải "0%")
    })
    return d


def _parse_xml_invoice(xml_bytes):
    """
    Đọc 1 hóa đơn (XML thuần hoặc file invoice.zip), trả về list các dòng mặt hàng.
    Cấu trúc thật: HDon>DLHDon>NDHDon>{NBan,NMua,DSHHDVu>HHDVu,TToan}
    """
    import xml.etree.ElementTree as ET
    rows = []
    xml_bytes = _extract_invoice_xml(xml_bytes)
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return rows

    def find_text(node, *tags):
        for t in tags:
            el = node.find(f".//{t}")
            if el is not None and el.text:
                return el.text.strip()
        return ""

    khmshdon = find_text(root, "KHMSHDon")
    khhdon = find_text(root, "KHHDon")
    shdon = find_text(root, "SHDon")
    ngay = find_text(root, "NLap")
    # Người bán
    nban = root.find(".//NBan")
    ten_nban = mst_nban = dchi_nban = ""
    if nban is not None:
        ten_nban = find_text(nban, "Ten")
        mst_nban = find_text(nban, "MST")
        dchi_nban = find_text(nban, "DChi")
    # Người mua
    nmua = root.find(".//NMua")
    ten_nmua = mst_nmua = ""
    if nmua is not None:
        ten_nmua = find_text(nmua, "Ten")
        mst_nmua = find_text(nmua, "MST")
    # Tổng tiền
    tgtcthue = find_text(root, "TgTCThue")
    tgtthue = find_text(root, "TgTThue")
    tgtttbso = find_text(root, "TgTTTBSo")

    # Danh sách hàng hóa dịch vụ (chỉ lấy HHDVu trong DSHHDVu)
    dshh = root.find(".//DSHHDVu")
    hh_list = dshh.findall("HHDVu") if dshh is not None else root.findall(".//HHDVu")

    def lay_ttkhac(hh, ten_truong):
        """Lấy giá trị DLieu trong TTKhac>TTin có TTruong = ten_truong."""
        for ttin in hh.findall(".//TTKhac/TTin"):
            tt = ttin.find("TTruong")
            if tt is not None and tt.text and tt.text.strip() == ten_truong:
                dl = ttin.find("DLieu")
                if dl is not None and dl.text:
                    return dl.text.strip()
        return ""

    for hh in hh_list:
        rows.append({
            "khmshdon": khmshdon, "khhdon": khhdon, "shdon": shdon,
            "ngay": ngay, "ten_nban": ten_nban, "mst_nban": mst_nban,
            "dchi_nban": dchi_nban, "ten_nmua": ten_nmua, "mst_nmua": mst_nmua,
            "stt": find_text(hh, "STT"),
            "tchat": find_text(hh, "TChat"),     # 1=hàng, 2=KM, 3=chiết khấu
            "ma_vt": find_text(hh, "MHHDVu"),
            "ten_hang": find_text(hh, "THHDVu"),
            "dvt": find_text(hh, "DVTinh"),
            "sluong": find_text(hh, "SLuong"),
            "dgia": find_text(hh, "DGia"),
            "thtien": find_text(hh, "ThTien"),
            "stckhau": find_text(hh, "STCKhau"),  # số tiền chiết khấu dòng này
            "tsuat": find_text(hh, "TSuat"),
            # tiền thuế GTGT lấy ĐÚNG từ XML (TTKhac > TongTien_Thue)
            "tien_thue": lay_ttkhac(hh, "TongTien_Thue"),
            "tgtcthue": tgtcthue, "tgtthue": tgtthue, "tgtttbso": tgtttbso,
        })

    # Tổng tiền phí (không chịu thuế) — vd lệ phí đăng kiểm, phí sân bay/thu
    # hộ... nằm RIÊNG ngoài danh sách hàng hóa dịch vụ, cộng thêm 1 dòng KCT.
    tong_phi = _lay_tong_tien_phi_xml(root)
    if tong_phi:
        common = {"khmshdon": khmshdon, "khhdon": khhdon, "shdon": shdon,
                  "ngay": ngay, "ten_nban": ten_nban, "mst_nban": mst_nban,
                  "dchi_nban": dchi_nban, "ten_nmua": ten_nmua, "mst_nmua": mst_nmua,
                  "tgtcthue": tgtcthue, "tgtthue": tgtthue, "tgtttbso": tgtttbso}
        rows.append(_dong_phi_hoa_don(tong_phi, common))
    return rows


def _parse_invoice_summary(xml_bytes):
    """
    Đọc 1 hóa đơn, trả về thông tin GỘP theo hóa đơn (cho bảng kê bán ra):
      {khmshdon, khhdon, shdon, ngay, ten_nban, mst_nban, ten_nmua, mst_nmua,
       mat_hang, theo_ts: {'8': {'ds','thue'}, 'KCT': {...}, ...}}
    Mỗi hóa đơn có thể nhiều thuế suất (lấy từ THTTLTSuat).
    """
    import xml.etree.ElementTree as ET
    xml_bytes = _extract_invoice_xml(xml_bytes)
    try:
        root = ET.fromstring(xml_bytes)
    except Exception:
        return None

    def ft(node, *tags):
        if node is None:
            return ""
        for t in tags:
            el = node.find(f".//{t}")
            if el is not None and el.text:
                return el.text.strip()
        return ""

    info = {
        "khmshdon": ft(root, "KHMSHDon"),
        "khhdon": ft(root, "KHHDon"),
        "shdon": ft(root, "SHDon"),
        "ngay": ft(root, "NLap"),
    }
    nban = root.find(".//NBan")
    info["ten_nban"] = ft(nban, "Ten")
    info["mst_nban"] = ft(nban, "MST")
    nmua = root.find(".//NMua")
    info["ten_nmua"] = ft(nmua, "Ten") or ft(nmua, "HVTNMHang")
    info["mst_nmua"] = ft(nmua, "MST")

    dshh = root.find(".//DSHHDVu")
    ten_hangs = []
    if dshh is not None:
        for hh in dshh.findall("HHDVu"):
            t = ft(hh, "THHDVu")
            if t:
                ten_hangs.append(t)
    info["mat_hang"] = (ten_hangs[0] + (" ..." if len(ten_hangs) > 1 else "")) if ten_hangs else ""

    def norm_ts(ts):
        s = str(ts or "").strip().upper()
        if s in ("", "KCT", "KHAC", "KO", "KHÔNG") or "KKK" in s or "KCT" in s:
            return "KCT"
        s2 = s.replace("%", "").strip()
        if s2 in ("0", "5", "8", "10"):
            return s2
        return s or "KHAC"

    theo_ts = {}
    lts = root.findall(".//THTTLTSuat/LTSuat")
    if lts:
        for l in lts:
            key = norm_ts(ft(l, "TSuat"))
            ds = _to_num(ft(l, "ThTien")) or 0
            thue = _to_num(ft(l, "TThue")) or 0
            cur = theo_ts.setdefault(key, {"ds": 0, "thue": 0})
            cur["ds"] += ds if isinstance(ds, (int, float)) else 0
            cur["thue"] += thue if isinstance(thue, (int, float)) else 0
    elif dshh is not None:
        for hh in dshh.findall("HHDVu"):
            key = norm_ts(ft(hh, "TSuat"))
            ds = _to_num(ft(hh, "ThTien")) or 0
            try:
                rate = float(str(ft(hh, "TSuat")).replace("%", "")) / 100
            except Exception:
                rate = 0
            thue = round(ds * rate) if isinstance(ds, (int, float)) else 0
            cur = theo_ts.setdefault(key, {"ds": 0, "thue": 0})
            cur["ds"] += ds if isinstance(ds, (int, float)) else 0
            cur["thue"] += thue
    # Tổng tiền phí (không chịu thuế) — cộng vào nhóm KCT để khớp với dòng
    # 'Tổng tiền phí' đã thêm ở _parse_xml_invoice (Chi tiết BÁN RA vs BK Bán ra).
    tong_phi = _lay_tong_tien_phi_xml(root)
    if tong_phi:
        cur = theo_ts.setdefault("KCT", {"ds": 0, "thue": 0})
        cur["ds"] += tong_phi
    info["theo_ts"] = theo_ts
    return info


@app.post("/api/parse-xml")
async def parse_xml_to_excel(request: Request):
    """
    Nhận nhiều file XML (multipart), đọc và kết xuất ra 1 file Excel chi tiết
    theo đúng định dạng từng mặt hàng.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    form = await request.form()
    files = form.getlist("files")
    if not files:
        raise HTTPException(400, "Chưa chọn file XML nào")

    all_rows = []
    for f in files:
        content = await f.read()
        all_rows.extend(_parse_xml_invoice(content))

    if not all_rows:
        raise HTTPException(400, "Không đọc được dữ liệu từ các file XML (sai định dạng?)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết hóa đơn"
    headers = ["Ký hiệu mẫu", "Số seri", "Số", "Ngày",
               "Tên người bán", "MST người bán", "Người mua", "MST người mua",
               "STT", "Tên hàng hóa/dịch vụ", "ĐVT", "Số lượng", "Đơn giá",
               "Thành tiền", "Thuế suất",
               "Tổng chưa thuế (HĐ)", "Tổng thuế (HĐ)", "Tổng thanh toán (HĐ)"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", fgColor="2E5C8A")
        cell.alignment = Alignment(horizontal="center")

    def num(v):
        try:
            return float(v)
        except Exception:
            return v

    for r in all_rows:
        ws.append([
            r["khmshdon"], r["khhdon"], r["shdon"], r["ngay"],
            r["ten_nban"], r["mst_nban"], r.get("ten_nmua", ""), r.get("mst_nmua", ""),
            r.get("stt", ""), r["ten_hang"], r["dvt"],
            num(r["sluong"]), num(r["dgia"]), num(r["thtien"]), r["tsuat"],
            num(r.get("tgtcthue", "")), num(r.get("tgtthue", "")), num(r.get("tgtttbso", "")),
        ])
    for col in ws.columns:
        w = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, 50)

    path = os.path.join(DOWNLOAD_DIR, "ChiTiet_TuXML.xlsx")
    wb.save(path)
    return _resp_xuat(path, "ChiTiet_HoaDon_TuXML.xlsx")


# ---------- MỤC 5: XEM HÓA ĐƠN DẠNG HTML (chuyển từ XML/JSON) ----------
def _render_invoice_html(inv):
    """Dựng HTML hiển thị 1 hóa đơn từ dict JSON đã lưu (field raw)."""
    def g(k, d=""):
        v = inv.get(k)
        return v if v not in (None, "") else d

    def vnd(x):
        try:
            return f"{float(x):,.0f}".replace(",", ".")
        except Exception:
            return x or ""

    items = inv.get("hdhhdvu") or []
    rows_html = ""
    for i, it in enumerate(items, 1):
        rows_html += f"""<tr>
            <td style="text-align:center">{i}</td>
            <td>{it.get('ten','') or it.get('thhdvu','')}</td>
            <td style="text-align:center">{it.get('dvtinh','')}</td>
            <td style="text-align:right">{vnd(it.get('sluong'))}</td>
            <td style="text-align:right">{vnd(it.get('dgia'))}</td>
            <td style="text-align:right">{vnd(it.get('thtien'))}</td>
        </tr>"""

    return f"""<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8">
<title>Hóa đơn {g('khhdon')}-{g('shdon')}</title>
<style>
  body{{font-family:'Times New Roman',serif;max-width:800px;margin:24px auto;color:#1a1a1a;padding:0 20px}}
  h2{{text-align:center;color:#c00;margin-bottom:4px}}
  .meta{{display:flex;justify-content:space-between;font-size:14px;margin:12px 0;flex-wrap:wrap}}
  .box{{border:1px solid #999;padding:12px 16px;border-radius:6px;margin:10px 0;font-size:14px;line-height:1.7}}
  table{{width:100%;border-collapse:collapse;margin-top:12px;font-size:13px}}
  th,td{{border:1px solid #aaa;padding:6px 8px}}
  th{{background:#f0e6d2}}
  .tongtien{{text-align:right;font-size:15px;margin-top:12px;line-height:1.8}}
  .label{{color:#666}}
</style></head><body>
<h2>HÓA ĐƠN GIÁ TRỊ GIA TĂNG</h2>
<div class="meta">
  <span>Ký hiệu: <b>{g('khmshdon')}{g('khhdon')}</b></span>
  <span>Số: <b>{g('shdon')}</b></span>
  <span>Ngày lập: <b>{(g('tdlap') or '').split('T')[0]}</b></span>
</div>
<div class="box">
  <div><span class="label">Người bán:</span> <b>{g('nbten')}</b></div>
  <div><span class="label">Mã số thuế:</span> {g('nbmst')}</div>
  <div><span class="label">Địa chỉ:</span> {g('nbdchi')}</div>
</div>
<div class="box">
  <div><span class="label">Người mua:</span> <b>{g('nmten')}</b></div>
  <div><span class="label">Mã số thuế:</span> {g('nmmst')}</div>
  <div><span class="label">Địa chỉ:</span> {g('nmdchi')}</div>
</div>
<table>
  <thead><tr><th>STT</th><th>Tên hàng hóa, dịch vụ</th><th>ĐVT</th>
  <th>Số lượng</th><th>Đơn giá</th><th>Thành tiền</th></tr></thead>
  <tbody>{rows_html or '<tr><td colspan=6 style="text-align:center">Không có chi tiết hàng hóa</td></tr>'}</tbody>
</table>
<div class="tongtien">
  <div>Cộng tiền hàng: <b>{vnd(g('tgtcthue'))}</b> đ</div>
  <div>Tiền thuế GTGT: <b>{vnd(g('tgtthue'))}</b> đ</div>
  <div>Tổng thanh toán: <b style="color:#c00;font-size:17px">{vnd(g('tgtttbso'))}</b> đ</div>
</div>
</body></html>"""


@app.get("/api/invoice-html/{invoice_id}")
def invoice_html(invoice_id: int):
    """
    Mở hóa đơn dạng HTML. Ưu tiên lấy ĐÚNG file invoice.html của Tổng cục Thuế
    (nguyên bản, không chỉnh sửa) từ file invoice.zip tải về.
    Thứ tự: (1) file zip đã lưu trong thư mục công ty -> (2) tải mới từ TCT
    -> (3) nếu không có, mới tự dựng HTML từ dữ liệu.
    """
    conn = db()
    row = conn.execute("SELECT * FROM invoices WHERE id=?", (invoice_id,)).fetchone()
    comp = conn.execute("SELECT * FROM companies WHERE id=?",
                        (row["company_id"],)).fetchone() if row else None
    conn.close()
    if not row:
        raise HTTPException(404, "Không tìm thấy hóa đơn")

    he_thong = row["he_thong"] or "query"
    base = f"{row['khhdon']}_{row['shdon']}_{row['nbmst']}"

    def html_from_zip(zip_bytes):
        try:
            zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
            for nm in zf.namelist():
                if nm.lower().endswith("invoice.html"):
                    return zf.read(nm).decode("utf-8", "replace")
        except Exception:
            pass
        return None

    raw_html = None

    # (1) Tìm file zip đã lưu trong thư mục cấu hình
    save_dir = (comp["save_dir"] or "").strip() if comp else ""
    if save_dir and os.path.isdir(save_dir):
        for rootdir, _dirs, files in os.walk(save_dir):
            for fn in files:
                if fn.startswith(base) and fn.lower().endswith(".zip"):
                    try:
                        with open(os.path.join(rootdir, fn), "rb") as f:
                            raw_html = html_from_zip(f.read())
                    except Exception:
                        pass
                    if raw_html:
                        break
            if raw_html:
                break

    # (2) Chưa có -> tải mới từ TCT (nếu đang đăng nhập)
    if not raw_html:
        client = CLIENTS.get(row["company_id"])
        if client and client.token:
            zdata = client.download_xml(row["nbmst"], row["khhdon"],
                                        row["khmshdon"], row["shdon"],
                                        row["loai"], he_thong)
            if zdata:
                raw_html = html_from_zip(zdata)

    fname = f"HoaDon_{row['khhdon']}_{row['shdon']}.html"

    # (3) Có HTML gốc của TCT -> trả nguyên bản
    if raw_html:
        return HTMLResponse(raw_html)

    # Không lấy được -> tự dựng từ dữ liệu (dự phòng)
    try:
        inv = json.loads(row["raw"]) if row["raw"] else {}
    except Exception:
        inv = {}
    client = CLIENTS.get(row["company_id"])
    if client and client.token:
        detail = client.get_detail(row["nbmst"], row["khhdon"],
                                   row["khmshdon"], row["shdon"], he_thong)
        if detail:
            inv = detail
    return HTMLResponse(_render_invoice_html(inv))


@app.get("/api/invoices-html-zip/{cid}")
def invoices_html_zip(cid: int, loai: Optional[str] = None):
    """Tải file invoice.zip chứa 1 file invoice.html gộp toàn bộ hóa đơn để đọc."""
    conn = db()
    comp = conn.execute("SELECT * FROM companies WHERE id=?", (cid,)).fetchone()
    if loai:
        rows = conn.execute(
            "SELECT * FROM invoices WHERE company_id=? AND loai=? ORDER BY tdlap DESC",
            (cid, loai)).fetchall()
    else:
        rows = conn.execute(
            "SELECT * FROM invoices WHERE company_id=? ORDER BY loai, tdlap DESC",
            (cid,)).fetchall()
    conn.close()

    if not rows:
        raise HTTPException(404, "Chưa có hóa đơn nào để xuất")

    # Gộp tất cả hóa đơn vào 1 file invoice.html (mỗi hóa đơn 1 trang, ngăn cách)
    parts = []
    for r in rows:
        try:
            inv = json.loads(r["raw"]) if r["raw"] else {}
        except Exception:
            inv = {}
        loai_txt = "MUA VÀO" if r["loai"] == "purchase" else "BÁN RA"
        # lấy phần body của từng hóa đơn
        html1 = _render_invoice_html(inv)
        body = html1.split("<body>", 1)[-1].split("</body>", 1)[0]
        parts.append(
            f'<div style="border-bottom:3px dashed #999;margin:30px 0;padding-bottom:20px">'
            f'<div style="background:#2E5C8A;color:#fff;padding:6px 12px;border-radius:6px;'
            f'display:inline-block;font-weight:bold;margin-bottom:10px">{loai_txt}</div>'
            f'{body}</div>')

    full_html = f"""<!DOCTYPE html><html lang="vi"><head><meta charset="UTF-8">
<title>Danh sách hóa đơn - {comp['ten']}</title>
<style>body{{font-family:'Times New Roman',serif;max-width:850px;margin:20px auto;padding:0 20px}}
table{{width:100%;border-collapse:collapse;font-size:13px}}th,td{{border:1px solid #aaa;padding:6px 8px}}
th{{background:#f0e6d2}}h2{{text-align:center;color:#c00}}</style></head><body>
<h1 style="text-align:center">DANH SÁCH HÓA ĐƠN - {comp['ten']}</h1>
<p style="text-align:center;color:#666">MST: {comp['mst']} | Tổng: {len(rows)} hóa đơn</p>
{''.join(parts)}
</body></html>"""

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("invoice.html", full_html)
    mem.seek(0)
    return StreamingResponse(
        mem, media_type="application/zip",
        headers={"Content-Disposition": "attachment; filename=invoice.zip"})


if __name__ == "__main__":
    import webbrowser, threading, socket

    # Tự tìm cổng trống nếu 8686 đang bị chiếm
    def find_free_port(preferred=8686):
        for port in [preferred, 8687, 8688, 8689, 8690, 8787, 8888, 9000]:
            s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            try:
                s.bind(("127.0.0.1", port))
                s.close()
                return port
            except OSError:
                s.close()
                continue
        return preferred

    PORT = find_free_port(8686)
    URL = f"http://127.0.0.1:{PORT}"

    def open_browser():
        time.sleep(1.5)
        webbrowser.open(URL)
    threading.Thread(target=open_browser, daemon=True).start()

    print("=" * 55)
    print("  PHẦN MỀM QUẢN LÝ HÓA ĐƠN ĐIỆN TỬ ĐA CÔNG TY")
    print(f"  Đang chạy tại: {URL}")
    if PORT != 8686:
        print(f"  (Cổng 8686 bận nên tự chuyển sang {PORT})")
    print("  (Đóng cửa sổ này để tắt phần mềm)")
    print("=" * 55)
    uvicorn.run(app, host="127.0.0.1", port=PORT, log_level="warning")
