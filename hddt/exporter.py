"""
Xuất dữ liệu hoá đơn ra các định dạng: JSON, CSV, Excel (.xlsx).
"""

from __future__ import annotations

import csv
import json
from typing import Any, Dict, List

# Các cột thường dùng nhất, ánh xạ mã trường -> tiêu đề tiếng Việt.
# Nếu hoá đơn có thêm trường khác, chúng vẫn được giữ trong JSON.
COLUMN_LABELS = {
    "nbmst": "MST người bán",
    "nbten": "Tên người bán",
    "nmmst": "MST người mua",
    "nmten": "Tên người mua",
    "khmshdon": "Ký hiệu mẫu số",
    "khhdon": "Ký hiệu hoá đơn",
    "shdon": "Số hoá đơn",
    "tdlap": "Ngày lập",
    "tgtcthue": "Tổng tiền chưa thuế",
    "tgtthue": "Tổng tiền thuế",
    "tgtttbso": "Tổng tiền thanh toán",
    "thlphi": "Tổng phí",
    "tthai": "Trạng thái",
    "ttxly": "Trạng thái xử lý",
    "hsgct": "Hình thức",
    "dvtte": "Đơn vị tiền tệ",
}

# Thứ tự ưu tiên các cột khi xuất
PREFERRED_ORDER = list(COLUMN_LABELS.keys())


def _ordered_fields(rows: List[Dict[str, Any]]) -> List[str]:
    """Lấy danh sách cột: ưu tiên các cột đã biết, rồi tới các cột còn lại."""
    seen = set()
    fields: List[str] = []
    for key in PREFERRED_ORDER:
        if any(key in r for r in rows):
            fields.append(key)
            seen.add(key)
    for r in rows:
        for key in r.keys():
            if key not in seen:
                fields.append(key)
                seen.add(key)
    return fields


def to_json(rows: List[Dict[str, Any]], path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)


def to_csv(rows: List[Dict[str, Any]], path: str) -> None:
    if not rows:
        # Vẫn tạo file rỗng có thể mở được
        open(path, "w", encoding="utf-8-sig").close()
        return
    fields = _ordered_fields(rows)
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writerow({k: COLUMN_LABELS.get(k, k) for k in fields})
        for r in rows:
            writer.writerow({k: _stringify(r.get(k, "")) for k in fields})


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, dict)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def to_excel(sheets: Dict[str, List[Dict[str, Any]]], path: str) -> None:
    """
    Xuất nhiều sheet ra một file Excel.
    `sheets`: dict {tên_sheet: danh_sách_hoá_đơn}.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError(
            "Cần cài openpyxl để xuất Excel: pip install openpyxl"
        ) from exc

    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, rows in sheets.items():
        # Tên sheet tối đa 31 ký tự
        ws = wb.create_sheet(title=sheet_name[:31] or "Sheet")
        if not rows:
            ws.append(["(Không có dữ liệu)"])
            continue
        fields = _ordered_fields(rows)
        header = [COLUMN_LABELS.get(k, k) for k in fields]
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([_stringify(r.get(k, "")) for k in fields])
        # Tự động giãn độ rộng cột (giới hạn để tránh quá rộng)
        for i, key in enumerate(fields, start=1):
            max_len = max(
                [len(str(COLUMN_LABELS.get(key, key)))]
                + [len(_stringify(r.get(key, ""))) for r in rows[:200]]
            )
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 50)

    if not wb.sheetnames:
        wb.create_sheet("Sheet")
    wb.save(path)
